from django.contrib import admin, messages
from django import forms
from django.core.exceptions import FieldError
from django.db.models import Count
from django.http import JsonResponse
from django.urls import path
from django.utils.html import format_html
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.db.models import Count
from django.db.models import Count, Sum, Case, When, Value, IntegerField, Min, F
from django.db.models.functions import TruncMonth
from django.template.response import TemplateResponse
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    ThematicArea,
    SkillLabTopic,
    SkillLab,
    SkillLabSession,
    SkillLabParticipantRecord,
    Skill_Lab_Mentee,
    SkillLabDashboard,
)

try:
    from hiva.admin_utils import ProvinceRestrictedAdminMixin, user_province
except Exception:
    class ProvinceRestrictedAdminMixin:
        def province_filter_kwargs(self, request):
            return {}

    def user_province(request):
        return None


# =========================================================
# Helper functions
# =========================================================
def safe_filter_by_user_province(queryset, province):
    """
    Tries common province relation paths.
    This keeps the admin stable even if related hiva models use different FK names.
    """
    if not province:
        return queryset

    possible_paths = [
        "province",
        "provincefk",
        "districtfk__provincefk",
        "facility__districtfk__provincefk",
        "facilityfk__districtfk__provincefk",
        "hfname__districtfk__provincefk",
        "hf_name__districtfk__provincefk",
    ]

    for path in possible_paths:
        try:
            return queryset.filter(**{path: province})
        except FieldError:
            continue

    return queryset


def safe_order_by(queryset, *fields):
    for field in fields:
        try:
            return queryset.order_by(field)
        except FieldError:
            continue
    return queryset


# =========================================================
# Shared media mixin for existing CSS + new cascade JS
# =========================================================
class SkillLabAdminMediaMixin:
    class Media:
        css = {
            "all": ("skilllab/admin/skilllab_admin.css",)
        }
        js = (
            "skilllab/admin/skilllab_cascade_topics_v5.js",
        )


# =========================================================
# Filters
# =========================================================
class SkillLabProvinceFilter(admin.SimpleListFilter):
    title = "Province"
    parameter_name = "province"

    def lookups(self, request, model_admin):
        from hiva.models import Province
        return [(p.pk, p.name) for p in Province.objects.all().order_by("name")]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(facility__districtfk__provincefk_id=self.value())
        return queryset


class SkillLabSessionProvinceFilter(admin.SimpleListFilter):
    title = "Province"
    parameter_name = "province"

    def lookups(self, request, model_admin):
        from hiva.models import Province
        return [(p.pk, p.name) for p in Province.objects.all().order_by("name")]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(skill_lab__facility__districtfk__provincefk_id=self.value())
        return queryset


class ParticipantProvinceFilter(admin.SimpleListFilter):
    title = "Province"
    parameter_name = "province"

    def lookups(self, request, model_admin):
        from hiva.models import Province
        return [(p.pk, p.name) for p in Province.objects.all().order_by("name")]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(session__skill_lab__facility__districtfk__provincefk_id=self.value())
        return queryset


# =========================================================
# Forms
# =========================================================
class SkillLabParticipantRecordForm(forms.ModelForm):
    class Meta:
        model = SkillLabParticipantRecord
        fields = "__all__"
        widgets = {
            "remarks": forms.Textarea(attrs={"rows": 2}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if "ls" in self.fields:
            self.fields["ls"].label = "Learning Session (LS)"
        if "mc" in self.fields:
            self.fields["mc"].label = "MODEL COMPOTENT(MC)"

        # -----------------------------------------
        # Cascading topic queryset by thematic area
        # Works for direct form and inline form prefixes.
        # -----------------------------------------
        if "topic" in self.fields:
            thematic_id = None

            # Inline field name usually looks like:
            # participant_records-0-thematic_area
            if self.data and self.prefix:
                prefixed_key = f"{self.prefix}-thematic_area"
                thematic_id = self.data.get(prefixed_key)

            # Direct change form field name
            if not thematic_id and self.data:
                thematic_id = self.data.get("thematic_area")

            # Existing saved object
            if not thematic_id and self.instance.pk and getattr(self.instance, "thematic_area_id", None):
                thematic_id = self.instance.thematic_area_id

            if thematic_id:
                try:
                    self.fields["topic"].queryset = SkillLabTopic.objects.filter(
                        thematicfk_id=int(thematic_id)
                    ).order_by("track", "seq_no", "name")
                except (TypeError, ValueError):
                    self.fields["topic"].queryset = SkillLabTopic.objects.none()
            else:
                self.fields["topic"].queryset = SkillLabTopic.objects.none()


class SkillLabSessionAdminForm(forms.ModelForm):
    class Meta:
        model = SkillLabSession
        fields = "__all__"
        widgets = {
            "objectives": forms.Textarea(attrs={"rows": 2}),
            "session_notes": forms.Textarea(attrs={"rows": 3}),
            "challenges": forms.Textarea(attrs={"rows": 2}),
            "action_points": forms.Textarea(attrs={"rows": 2}),
        }


# =========================================================
# Inline
# =========================================================
class SkillLabParticipantRecordInline(admin.TabularInline):
    model = SkillLabParticipantRecord
    form = SkillLabParticipantRecordForm
    extra = 1
    show_change_link = True
    classes = ("tabular-inline-modern",)

    # Keep thematic as autocomplete. Topic must be normal dropdown for cascade.
    autocomplete_fields = ("thematic_area",)
    raw_id_fields = ("mentee_name",)

    fields = (
        "mentee_name",
        "thematic_area",
        "topic",
        "ls",
        "mc",
        "competency_status",
        "checklist_score",
        "feedback_given",
    )

    verbose_name = "Participant Record"
    verbose_name_plural = "Participant Records"

    class Media:
        js = (
            "skilllab/admin/skilllab_cascade_topics.js",
        )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        province = user_province(request)

        # Show all mentees in user's province, not only the skill lab facility.
        if db_field.name == "mentee_name" and province and not request.user.is_superuser:
            kwargs["queryset"] = Skill_Lab_Mentee.objects.filter(
                hfname__districtfk__provincefk=province
            ).select_related("hfname", "position").order_by("firstname", "lastname")

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

# =========================================================
# Thematic Area
# =========================================================
@admin.register(ThematicArea)
class ThematicAreaAdmin(SkillLabAdminMediaMixin, admin.ModelAdmin):
    list_display = ("name", "shortname", "hqip_area")
    search_fields = ("name", "shortname", "hqip_area__name")
    list_select_related = ("hqip_area",)
    raw_id_fields = ("hqip_area",)


# =========================================================
# Skill Lab Topic
# =========================================================
@admin.register(SkillLabTopic)
class SkillLabTopicAdmin(SkillLabAdminMediaMixin, admin.ModelAdmin):
    list_display = ("name", "shortname", "thematicfk", "nameeng")
    search_fields = ("name", "shortname", "thematicfk__name", "nameeng")
    list_filter = ("thematicfk",)
    autocomplete_fields = ("thematicfk",)


# =========================================================
# Skill Lab
# =========================================================
@admin.register(SkillLab)
class SkillLabAdmin(SkillLabAdminMediaMixin, ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    list_display = (
        "name",
        "facility",
        "get_district",
        "get_province",
        "status_badge",
        "implementing_partner",
        "session_count",
    )
    search_fields = (
        "name",
        "facility__name",
        "facility__districtfk__name",
        "facility__districtfk__provincefk__name",
        "implementing_partner__name",
    )
    list_filter = (SkillLabProvinceFilter, "status", "implementing_partner")
    list_select_related = (
        "facility",
        "facility__districtfk",
        "facility__districtfk__provincefk",
        "implementing_partner",
    )
    raw_id_fields = ("facility", "implementing_partner")

    fieldsets = (
        ("MNH SKILL LAB INFORMATION", {
            "fields": ("name", "facility", "status", "implementing_partner"),
            "classes": ("wide", "skilllab-card"),
        }),
        ("Notes", {
            "fields": ("remarks",),
            "classes": ("collapse", "skilllab-card"),
        }),
    )

    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related(
            "facility",
            "facility__districtfk",
            "facility__districtfk__provincefk",
            "implementing_partner",
        ).annotate(_session_count=Count("sessions"))

        province = user_province(request)
        if province and not request.user.is_superuser:
            qs = qs.filter(facility__districtfk__provincefk=province)
        return qs

    def province_filter_kwargs(self, request):
        return {"facility__districtfk__provincefk": user_province(request)}

    @admin.display(description="District")
    def get_district(self, obj):
        return getattr(obj.facility.districtfk, "name", "-") if obj.facility else "-"

    @admin.display(description="Province")
    def get_province(self, obj):
        try:
            return obj.facility.districtfk.provincefk.name
        except Exception:
            return "-"

    @admin.display(description="Sessions")
    def session_count(self, obj):
        return getattr(obj, "_session_count", 0)

    @admin.display(description="Status")
    def status_badge(self, obj):
        css_map = {
            "ACTIVE": "sl-badge sl-badge-success",
            "INACTIVE": "sl-badge sl-badge-muted",
            "PLANNED": "sl-badge sl-badge-warning",
            "CLOSED": "sl-badge sl-badge-danger",
        }
        cls = css_map.get(obj.status, "sl-badge")
        return format_html('<span class="{}">{}</span>', cls, obj.get_status_display())


@admin.register(SkillLabSession)
class SkillLabSessionAdmin(SkillLabAdminMediaMixin, ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    form = SkillLabSessionAdminForm
    inlines = [SkillLabParticipantRecordInline]

    list_display = (
        "skill_lab",
        "get_facility",
        "get_province",
        "session_date",
        "lab_round",
        "session_type_badge",
        "mentor_name",
        "ce_checklist_applied",
        "completed_session",
        "participant_total_display",
        "duration_display",
    )

    search_fields = (
        "skill_lab__name",
        "skill_lab__facility__name",
        "skill_lab__facility__districtfk__name",
        "skill_lab__facility__districtfk__provincefk__name",
    )

    list_filter = (
        SkillLabSessionProvinceFilter,
        "session_type",
        "lab_round",
        "ce_checklist_applied",
        "completed_session",
        "planned_session",
        "followup_needed",
        "session_date",
    )

    ordering = ("-session_date",)
    save_on_top = True
    list_per_page = 50

    autocomplete_fields = ("skill_lab",)
    raw_id_fields = ("mentor_org",)

    actions = ["export_skilllab_sessions_excel"]

    readonly_fields = (
        "duration_preview",
        "participant_total_readonly",
        "created_at",
        "updated_at",
    )

    fieldsets = (
        ("Skill Lab Header Information", {
            "fields": (
                "skill_lab",
                "session_date",
                ("lab_round", "session_type"),
            ),
            "classes": ("wide", "skilllab-card"),
        }),
        ("Time and Attendance", {
            "fields": (
                ("check_in", "check_out"),
                "duration_preview",
                ("planned_session", "completed_session"),
                ("total_participants", "participant_total_readonly"),
            ),
            "classes": ("wide", "skilllab-card"),
        }),
        ("Facilitation and Follow-up", {
            "fields": (
                ("mentor_name", "mentor_org"),
                "ce_checklist_applied",
                "followup_needed",
            ),
            "classes": ("wide", "skilllab-card"),
        }),
        ("Documentation", {
            "fields": ("objectives", "session_notes", "challenges", "action_points"),
            "classes": ("collapse", "skilllab-card"),
        }),
        ("Audit Trail", {
            "fields": ("created_at", "updated_at"),
            "classes": ("collapse", "skilllab-card"),
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "topics-by-thematic/",
                self.admin_site.admin_view(self.topics_by_thematic),
                name="skilllab_topics_by_thematic",
            ),
        ]
        return custom_urls + urls

    def topics_by_thematic(self, request):
        thematic_id = request.GET.get("thematic_id")
        topics = SkillLabTopic.objects.none()

        if thematic_id:
            topics = SkillLabTopic.objects.filter(
                thematicfk_id=thematic_id
            ).order_by("track", "seq_no", "name")

        data = [{"id": topic.pk, "text": str(topic)} for topic in topics]
        return JsonResponse({"results": data})

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        province = user_province(request)

        if db_field.name in ("mentor_name", "mentor") and province and not request.user.is_superuser:
            qs = db_field.remote_field.model.objects.all()
            qs = safe_filter_by_user_province(qs, province)
            kwargs["queryset"] = safe_order_by(qs, "name", "firstname", "first_name")

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related(
            "skill_lab",
            "skill_lab__facility",
            "skill_lab__facility__districtfk",
            "skill_lab__facility__districtfk__provincefk",
            "mentor_name",
            "mentor_org",
        ).annotate(_participant_count=Count("participant_records"))

        province = user_province(request)
        if province and not request.user.is_superuser:
            qs = qs.filter(skill_lab__facility__districtfk__provincefk=province)

        return qs

    def province_filter_kwargs(self, request):
        return {"skill_lab__facility__districtfk__provincefk": user_province(request)}

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance
        actual_count = obj.participant_records.count()

        if obj.total_participants != actual_count:
            obj.total_participants = actual_count
            obj.save(update_fields=["total_participants"])
            self.message_user(
                request,
                f"Total participants updated automatically to {actual_count}.",
                level=messages.INFO,
            )

    @admin.display(description="Facility")
    def get_facility(self, obj):
        try:
            return obj.skill_lab.facility.name
        except Exception:
            return "-"

    @admin.display(description="Province")
    def get_province(self, obj):
        try:
            return obj.skill_lab.facility.districtfk.provincefk.name
        except Exception:
            return "-"

    @admin.display(description="Type")
    def session_type_badge(self, obj):
        return format_html(
            '<span class="sl-badge">{}</span>',
            obj.get_session_type_display()
        )

    @admin.display(description="Participants")
    def participant_total_display(self, obj):
        return getattr(obj, "_participant_count", 0)

    @admin.display(description="Current Participant Count")
    def participant_total_readonly(self, obj):
        if not obj.pk:
            return 0
        return obj.participant_records.count()

    @admin.display(description="Duration")
    def duration_display(self, obj):
        if obj.duration_hours is None:
            return "-"
        css = "sl-duration-ok" if obj.duration_hours <= 8 else "sl-duration-warn"
        return format_html('<span class="{}">{} hrs</span>', css, obj.duration_hours)

    @admin.display(description="Duration Preview")
    def duration_preview(self, obj):
        if not obj.pk:
            return "Will appear after save"
        if obj.duration_hours is None:
            return "-"
        return f"{obj.duration_hours} hours"

    def export_skilllab_sessions_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Skill Lab Participants"

        headers = [
            "Skill Lab",
            "Province",
            "District",
            "Facility",
            "Session Date",
            "Lab Round",
            "Check In",
            "Check Out",
            "Duration Hours",
            "Session Type",
            "Clinical Mentor",
            "Mentor Organization",
            "CE Checklist Applied",
            "Planned Session",
            "Completed Session",
            "Total Participants",
            "Participant Name",
            "Participant Facility",
            "Participant Position",
            "Thematic Area",
            "Topic",
            "LS",
            "MC",
            "Competency Status",
            "Checklist Score",
            "Feedback Given",
            "Objectives",
            "Session Notes",
            "Challenges",
            "Action Points",
            "Follow-up Needed",
            "Created At",
            "Updated At",
        ]

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="0F766E")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def yes_no(value):
            return "Yes" if value else "No"

        def date_value(value):
            return value.strftime("%Y-%m-%d") if value else ""

        def time_value(value):
            return value.strftime("%H:%M") if value else ""

        def datetime_value(value):
            return value.strftime("%Y-%m-%d %H:%M") if value else ""

        def safe_name(obj):
            return str(obj) if obj else ""

        def get_mentee_facility(mentee):
            try:
                return mentee.hfname.name if mentee and mentee.hfname else ""
            except Exception:
                return ""

        def get_mentee_position(mentee):
            try:
                return str(mentee.position) if mentee and mentee.position else ""
            except Exception:
                return ""

        for obj in queryset:
            participant_records = obj.participant_records.select_related(
                "mentee_name",
                "mentee_name__hfname",
                "mentee_name__position",
                "thematic_area",
                "topic",
            ).all()

            if not participant_records.exists():
                ws.append([
                    safe_name(obj.skill_lab),
                    self.get_province(obj),
                    obj.district.name if obj.district else "",
                    self.get_facility(obj),
                    date_value(obj.session_date),
                    obj.lab_round,
                    time_value(obj.check_in),
                    time_value(obj.check_out),
                    obj.duration_hours if obj.duration_hours is not None else "",
                    obj.get_session_type_display() if obj.session_type else "",
                    safe_name(obj.mentor_name),
                    safe_name(obj.mentor_org),
                    yes_no(obj.ce_checklist_applied),
                    yes_no(obj.planned_session),
                    yes_no(obj.completed_session),
                    obj.total_participants,
                    "", "", "", "", "", "", "", "", "", "",
                    obj.objectives or "",
                    obj.session_notes or "",
                    obj.challenges or "",
                    obj.action_points or "",
                    yes_no(obj.followup_needed),
                    datetime_value(obj.created_at),
                    datetime_value(obj.updated_at),
                ])
            else:
                for pr in participant_records:
                    ws.append([
                        safe_name(obj.skill_lab),
                        self.get_province(obj),
                        obj.district.name if obj.district else "",
                        self.get_facility(obj),
                        date_value(obj.session_date),
                        obj.lab_round,
                        time_value(obj.check_in),
                        time_value(obj.check_out),
                        obj.duration_hours if obj.duration_hours is not None else "",
                        obj.get_session_type_display() if obj.session_type else "",
                        safe_name(obj.mentor_name),
                        safe_name(obj.mentor_org),
                        yes_no(obj.ce_checklist_applied),
                        yes_no(obj.planned_session),
                        yes_no(obj.completed_session),
                        obj.total_participants,
                        safe_name(pr.mentee_name),
                        get_mentee_facility(pr.mentee_name),
                        get_mentee_position(pr.mentee_name),
                        safe_name(pr.thematic_area),
                        safe_name(pr.topic),
                        yes_no(pr.ls),
                        yes_no(pr.mc),
                        pr.competency_status or "",
                        pr.checklist_score if pr.checklist_score is not None else "",
                        yes_no(pr.feedback_given),
                        obj.objectives or "",
                        obj.session_notes or "",
                        obj.challenges or "",
                        obj.action_points or "",
                        yes_no(obj.followup_needed),
                        datetime_value(obj.created_at),
                        datetime_value(obj.updated_at),
                    ])

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = min(max_length + 3, 45)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="skill_lab_participant_level_export.xlsx"'

        wb.save(response)
        return response

    export_skilllab_sessions_excel.short_description = "Export selected Skill Lab Sessions with Participant Records to Excel"


@admin.register(SkillLabDashboard)
class SkillLabDashboardAdmin(SkillLabAdminMediaMixin, ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    change_list_template = "admin/skilllab/dashboard.html"

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return request.user.is_active and request.user.is_staff

    def get_model_perms(self, request):
        if self.has_view_permission(request):
            return {"view": True}
        return {}

    def _get_participant_model(self):
        """
        Uses your existing related_name='participant_records'.
        """
        rel = SkillLabSession._meta.get_field("participant_records")
        return rel.related_model, rel.field.name

    def _bool_sum(self, field_name):
        return Sum(
            Case(
                When(**{field_name: True}, then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            )
        )

    def changelist_view(self, request, extra_context=None):
        data = self._build_dashboard_data(request)

        if request.GET.get("export") == "1":
            return self._export_dashboard_excel(data)

        context = {
            **self.admin_site.each_context(request),
            "title": "Skill Lab Dashboard",
            **data,
        }

        return TemplateResponse(request, self.change_list_template, context)

    def _build_dashboard_data(self, request):
        ParticipantModel, session_fk = self._get_participant_model()

        session_prefix = f"{session_fk}__"

        date_lookup = f"{session_prefix}session_date"
        province_id_lookup = f"{session_prefix}skill_lab__facility__districtfk__provincefk_id"
        province_name_lookup = f"{session_prefix}skill_lab__facility__districtfk__provincefk__name"
        district_name_lookup = f"{session_prefix}skill_lab__facility__districtfk__name"
        facility_id_lookup = f"{session_prefix}skill_lab__facility_id"
        facility_name_lookup = f"{session_prefix}skill_lab__facility__name"
        hfcode_lookup = f"{session_prefix}skill_lab__facility__hfcode"

        sessions_qs = SkillLabSession.objects.select_related(
            "skill_lab",
            "skill_lab__facility",
            "skill_lab__facility__districtfk",
            "skill_lab__facility__districtfk__provincefk",
            "mentor_name",
            "mentor_org",
        )

        participants_qs = ParticipantModel.objects.select_related(
            session_fk,
            f"{session_fk}__skill_lab",
            f"{session_fk}__skill_lab__facility",
            f"{session_fk}__skill_lab__facility__districtfk",
            f"{session_fk}__skill_lab__facility__districtfk__provincefk",
            "mentee_name",
            "mentee_name__hfname",
            "mentee_name__position",
            "thematic_area",
            "topic",
        )

        province = user_province(request)
        if province and not request.user.is_superuser:
            sessions_qs = sessions_qs.filter(
                skill_lab__facility__districtfk__provincefk=province
            )
            participants_qs = participants_qs.filter(
                **{province_id_lookup: province.id}
            )

        date_from = request.GET.get("date_from", "").strip()
        date_to = request.GET.get("date_to", "").strip()
        province_id = request.GET.get("province", "").strip()
        facility_id = request.GET.get("facility", "").strip()
        session_type = request.GET.get("session_type", "").strip()
        lab_round = request.GET.get("lab_round", "").strip()

        if date_from:
            sessions_qs = sessions_qs.filter(session_date__gte=date_from)
            participants_qs = participants_qs.filter(**{f"{date_lookup}__gte": date_from})

        if date_to:
            sessions_qs = sessions_qs.filter(session_date__lte=date_to)
            participants_qs = participants_qs.filter(**{f"{date_lookup}__lte": date_to})

        if province_id:
            sessions_qs = sessions_qs.filter(
                skill_lab__facility__districtfk__provincefk_id=province_id
            )
            participants_qs = participants_qs.filter(**{province_id_lookup: province_id})

        if facility_id:
            sessions_qs = sessions_qs.filter(skill_lab__facility_id=facility_id)
            participants_qs = participants_qs.filter(**{facility_id_lookup: facility_id})

        if session_type:
            sessions_qs = sessions_qs.filter(session_type=session_type)
            participants_qs = participants_qs.filter(
                **{f"{session_prefix}session_type": session_type}
            )

        if lab_round:
            sessions_qs = sessions_qs.filter(lab_round=lab_round)
            participants_qs = participants_qs.filter(
                **{f"{session_prefix}lab_round": lab_round}
            )

        option_sessions = SkillLabSession.objects.select_related(
            "skill_lab__facility__districtfk__provincefk"
        )

        if province and not request.user.is_superuser:
            option_sessions = option_sessions.filter(
                skill_lab__facility__districtfk__provincefk=province
            )

        province_options = list(
            option_sessions.values(
                province_id=F("skill_lab__facility__districtfk__provincefk_id"),
                province=F("skill_lab__facility__districtfk__provincefk__name"),
            )
            .exclude(province_id__isnull=True)
            .distinct()
            .order_by("province")
        )

        facility_options = list(
            option_sessions.values(
                facility_id=F("skill_lab__facility_id"),
                facility=F("skill_lab__facility__name"),
            )
            .exclude(facility_id__isnull=True)
            .distinct()
            .order_by("facility")
        )

        lab_round_options = list(
            option_sessions.values_list("lab_round", flat=True)
            .distinct()
            .order_by("lab_round")
        )

        session_type_options = SkillLabSession.SESSION_TYPE_CHOICES

        kpis = participants_qs.aggregate(
            total_records=Count("id"),
            distinct_participants=Count("mentee_name", distinct=True),
            total_ls=self._bool_sum("ls"),
            total_mc=self._bool_sum("mc"),
            distinct_topics=Count("topic", distinct=True),
            distinct_facilities=Count(facility_id_lookup, distinct=True),
            distinct_sessions=Count(session_fk, distinct=True),
        )

        for key in kpis:
            kpis[key] = int(kpis[key] or 0)

        kpis["mc_minus_ls"] = kpis["total_mc"] - kpis["total_ls"]

        province_rows = list(
            participants_qs.values(
                province=F(province_name_lookup),
            )
            .annotate(
                participants=Count("mentee_name", distinct=True),
                sessions=Count(session_fk, distinct=True),
                facilities=Count(facility_id_lookup, distinct=True),
                topics=Count("topic", distinct=True),
                ls=self._bool_sum("ls"),
                mc=self._bool_sum("mc"),
                records=Count("id"),
            )
            .order_by("province")
        )

        max_value = max(
            [max(int(r["ls"] or 0), int(r["mc"] or 0)) for r in province_rows] or [1]
        )

        for r in province_rows:
            r["ls"] = int(r["ls"] or 0)
            r["mc"] = int(r["mc"] or 0)
            r["mc_minus_ls"] = r["mc"] - r["ls"]
            r["ls_width"] = round((r["ls"] / max_value) * 100, 1) if max_value else 0
            r["mc_width"] = round((r["mc"] / max_value) * 100, 1) if max_value else 0

            if r["mc_minus_ls"] > 0:
                r["interpretation"] = "More MC than LS"
            elif r["mc_minus_ls"] < 0:
                r["interpretation"] = "More LS than MC"
            else:
                r["interpretation"] = "Balanced"

        facility_rows = list(
            participants_qs.values(
                province=F(province_name_lookup),
                district=F(district_name_lookup),
                facility=F(facility_name_lookup),
                hfcode=F(hfcode_lookup),
            )
            .annotate(
                participants=Count("mentee_name", distinct=True),
                sessions=Count(session_fk, distinct=True),
                topics=Count("topic", distinct=True),
                ls=self._bool_sum("ls"),
                mc=self._bool_sum("mc"),
                records=Count("id"),
            )
            .order_by("province", "district", "facility")
        )

        for r in facility_rows:
            r["ls"] = int(r["ls"] or 0)
            r["mc"] = int(r["mc"] or 0)
            r["mc_minus_ls"] = r["mc"] - r["ls"]

            if r["mc_minus_ls"] > 0:
                r["interpretation"] = "Competency assessment instances were higher than learning session instances."
            elif r["mc_minus_ls"] < 0:
                r["interpretation"] = "Learning session instances were higher than competency assessment instances."
            else:
                r["interpretation"] = "LS and MC were balanced."

        facility_driver_rows = sorted(
            facility_rows,
            key=lambda x: abs(x["mc_minus_ls"]),
            reverse=True,
        )[:25]

        bamyan_row = None
        for r in province_rows:
            if str(r.get("province") or "").strip().lower() == "bamyan":
                bamyan_row = r
                break

        bamyan_facilities = [
            r for r in facility_rows
            if str(r.get("province") or "").strip().lower() == "bamyan"
        ]
        bamyan_facilities = sorted(
            bamyan_facilities,
            key=lambda x: x["mc_minus_ls"],
            reverse=True,
        )

        if bamyan_row and bamyan_row["mc"] > bamyan_row["ls"]:
            bamyan_narrative = (
                f"Bamyan recorded {bamyan_row['mc']} MC instances compared with "
                f"{bamyan_row['ls']} LS instances. MC was higher by "
                f"{bamyan_row['mc_minus_ls']} instances. This does not indicate a data error by itself. "
                f"LS and MC are counted at participant-topic/session level, not as unique health workers. "
                f"A single participant may be assessed across multiple topics or sessions. "
                f"Therefore, the higher MC value suggests that competency assessment activities were recorded "
                f"more frequently than learning session entries in selected Bamyan facilities."
            )
        else:
            bamyan_narrative = (
                "For the selected filters, Bamyan does not show MC higher than LS, "
                "or Bamyan is not included in the selected period."
            )

        trend_rows = list(
            participants_qs.annotate(
                month=TruncMonth(date_lookup)
            )
            .values(
                "month",
                province=F(province_name_lookup),
            )
            .annotate(
                participants=Count("mentee_name", distinct=True),
                topics=Count("topic", distinct=True),
                ls=self._bool_sum("ls"),
                mc=self._bool_sum("mc"),
            )
            .order_by("province", "month")
        )

        for r in trend_rows:
            r["month_label"] = r["month"].strftime("%b %Y") if r["month"] else "Unknown"
            r["ls"] = int(r["ls"] or 0)
            r["mc"] = int(r["mc"] or 0)
            r["mc_minus_ls"] = r["mc"] - r["ls"]

        first_session_rows = []

        first_by_province = list(
            sessions_qs.values(
                province_id=F("skill_lab__facility__districtfk__provincefk_id"),
                province=F("skill_lab__facility__districtfk__provincefk__name"),
            )
            .exclude(province_id__isnull=True)
            .annotate(first_date=Min("session_date"))
            .order_by("province")
        )

        for row in first_by_province:
            first_session = (
                sessions_qs.filter(
                    skill_lab__facility__districtfk__provincefk_id=row["province_id"],
                    session_date=row["first_date"],
                )
                .order_by("session_date", "id")
                .first()
            )

            facility = first_session.facility if first_session else None
            district = first_session.district if first_session else None

            first_session_rows.append({
                "province": row["province"],
                "first_date": row["first_date"],
                "facility": facility.name if facility else "",
                "hfcode": facility.hfcode if facility else "",
                "district": district.name if district else "",
                "mentor": str(first_session.mentor_name) if first_session and first_session.mentor_name else "",
                "session_type": first_session.session_type if first_session else "",
            })

        story_candidates = []
        for r in facility_rows:
            story_score = r["ls"] + r["mc"] + max(r["mc_minus_ls"], 0) * 2

            if r["ls"] > 0 or r["mc"] > 0:
                story_candidates.append({
                    **r,
                    "story_score": story_score,
                    "story_angle": (
                        "Competency assessment progress"
                        if r["mc_minus_ls"] > 0
                        else "Learning session implementation"
                    ),
                })

        story_candidates = sorted(
            story_candidates,
            key=lambda x: x["story_score"],
            reverse=True,
        )[:15]

    # ============================================================
    # Chart data for Django admin dashboard
    # ============================================================

        # Add MC share to province rows
        for r in province_rows:
            total_lsmc = int(r.get("ls") or 0) + int(r.get("mc") or 0)
            r["mc_share"] = round((int(r.get("mc") or 0) / total_lsmc) * 100, 1) if total_lsmc else 0

        # Add MC share to facility rows
        for r in facility_rows:
            total_lsmc = int(r.get("ls") or 0) + int(r.get("mc") or 0)
            r["mc_share"] = round((int(r.get("mc") or 0) / total_lsmc) * 100, 1) if total_lsmc else 0

        # Monthly total trend across all selected provinces/facilities
        monthly_total_rows = list(
            participants_qs.annotate(
                month=TruncMonth(date_lookup)
            )
            .values("month")
            .annotate(
                participants=Count("mentee_name", distinct=True),
                sessions=Count(session_fk, distinct=True),
                topics=Count("topic", distinct=True),
                ls=self._bool_sum("ls"),
                mc=self._bool_sum("mc"),
            )
            .order_by("month")
        )

        for r in monthly_total_rows:
            r["month_label"] = r["month"].strftime("%b %Y") if r["month"] else "Unknown"
            r["ls"] = int(r["ls"] or 0)
            r["mc"] = int(r["mc"] or 0)
            r["total"] = r["ls"] + r["mc"]
            r["mc_minus_ls"] = r["mc"] - r["ls"]

        # Top 10 facility drivers by absolute difference between LS and MC
        top_facility_drivers_chart = sorted(
            facility_rows,
            key=lambda x: abs(int(x.get("mc_minus_ls") or 0)),
            reverse=True,
        )[:10]

        # Chart-safe data only: strings and numbers
        chart_data = {
            "province_ls_mc": [
                {
                    "province": r.get("province") or "Unknown",
                    "LS": int(r.get("ls") or 0),
                    "MC": int(r.get("mc") or 0),
                }
                for r in province_rows
            ],

            "province_gap": [
                {
                    "province": r.get("province") or "Unknown",
                    "gap": int(r.get("mc_minus_ls") or 0),
                    "mc_share": float(r.get("mc_share") or 0),
                }
                for r in province_rows
            ],

            "monthly_ls_mc": [
                {
                    "month": r.get("month_label") or "Unknown",
                    "LS": int(r.get("ls") or 0),
                    "MC": int(r.get("mc") or 0),
                    "Total": int(r.get("total") or 0),
                }
                for r in monthly_total_rows
            ],

            "facility_drivers": [
                {
                    "facility": (r.get("facility") or "Unknown")[:38],
                    "LS": int(r.get("ls") or 0),
                    "MC": int(r.get("mc") or 0),
                    "gap": int(r.get("mc_minus_ls") or 0),
                }
                for r in top_facility_drivers_chart
            ],

            "bamyan_facilities": [
                {
                    "facility": (r.get("facility") or "Unknown")[:38],
                    "LS": int(r.get("ls") or 0),
                    "MC": int(r.get("mc") or 0),
                    "gap": int(r.get("mc_minus_ls") or 0),
                }
                for r in bamyan_facilities
            ],

            "story_candidates": [
                {
                    "facility": (r.get("facility") or "Unknown")[:38],
                    "story_score": int(r.get("story_score") or 0),
                    "LS": int(r.get("ls") or 0),
                    "MC": int(r.get("mc") or 0),
                }
                for r in story_candidates[:10]
            ],
        }

        # ============================================================
        # Participant Profile Table
        # Grouped by Skill Lab facility + participant
        # Shows LS topics, MC topics, graduated topics, and pending topics
        # ============================================================

        def safe_text(value):
            if value is None:
                return ""
            return str(value).strip()

        def yes_no(value):
            return "Yes" if value else "No"

        def get_participant_facility(mentee):
            try:
                return mentee.hfname.name if mentee and mentee.hfname else ""
            except Exception:
                return ""

        def get_participant_position(mentee):
            try:
                return str(mentee.position) if mentee and mentee.position else ""
            except Exception:
                return ""

        def is_topic_graduated(pr):
            """
            Graduation logic:
            1. If MC is checked, topic is treated as competency completed.
            2. If competency_status clearly says competent/completed/passed/graduated, it is graduated.
            3. If checklist_score is available and >= 80, it is graduated.
            """

            status = safe_text(getattr(pr, "competency_status", "")).lower()
            score = getattr(pr, "checklist_score", None)

            negative_terms = [
                "not competent",
                "not yet",
                "needs",
                "incomplete",
                "failed",
                "fail",
            ]

            positive_terms = [
                "competent",
                "model competent",
                "completed",
                "complete",
                "passed",
                "pass",
                "graduated",
            ]

            if any(term in status for term in negative_terms):
                return False

            if any(term in status for term in positive_terms):
                return True

            try:
                if score is not None and float(score) >= 80:
                    return True
            except Exception:
                pass

            if getattr(pr, "mc", False):
                return True

            return False


        participant_profiles = {}
        participant_detail_rows = []

        profile_records = participants_qs.order_by(
            f"{session_prefix}skill_lab__facility__districtfk__provincefk__name",
            f"{session_prefix}skill_lab__facility__districtfk__name",
            f"{session_prefix}skill_lab__facility__name",
            f"{session_prefix}session_date",
            "id",
        )

        for pr in profile_records:
            session = getattr(pr, session_fk, None)
            mentee = getattr(pr, "mentee_name", None)

            if not session:
                continue

            skill_lab_facility = session.facility
            skill_lab_district = session.district
            skill_lab_province = session.province

            participant_name = safe_text(mentee) if mentee else "Unknown participant"
            participant_id = getattr(mentee, "id", None)

            topic_name = safe_text(getattr(pr, "topic", None)) or "Unknown topic"
            thematic_name = safe_text(getattr(pr, "thematic_area", None)) or "Unknown thematic area"

            graduated = is_topic_graduated(pr)

            profile_key = (
                getattr(skill_lab_facility, "id", None),
                participant_id,
                participant_name,
            )

            if profile_key not in participant_profiles:
                participant_profiles[profile_key] = {
                    "province": safe_text(getattr(skill_lab_province, "name", "")),
                    "district": safe_text(getattr(skill_lab_district, "name", "")),
                    "skill_lab_facility": safe_text(getattr(skill_lab_facility, "name", "")),
                    "hfcode": safe_text(getattr(skill_lab_facility, "hfcode", "")),
                    "participant_name": participant_name,
                    "participant_facility": get_participant_facility(mentee),
                    "participant_position": get_participant_position(mentee),
                    "sessions": set(),
                    "topics_all": set(),
                    "ls_topics": set(),
                    "mc_topics": set(),
                    "graduated_topics": set(),
                    "needs_graduation_topics": set(),
                    "last_session_date": None,
                }

            profile = participant_profiles[profile_key]

            profile["sessions"].add(getattr(session, "id", None))
            profile["topics_all"].add(topic_name)

            session_date = getattr(session, "session_date", None)
            if session_date:
                if not profile["last_session_date"] or session_date > profile["last_session_date"]:
                    profile["last_session_date"] = session_date

            if getattr(pr, "ls", False):
                profile["ls_topics"].add(topic_name)

            if getattr(pr, "mc", False):
                profile["mc_topics"].add(topic_name)

            if graduated:
                profile["graduated_topics"].add(topic_name)

            participant_detail_rows.append({
                "province": profile["province"],
                "district": profile["district"],
                "skill_lab_facility": profile["skill_lab_facility"],
                "hfcode": profile["hfcode"],
                "participant_name": participant_name,
                "participant_facility": profile["participant_facility"],
                "participant_position": profile["participant_position"],
                "session_date": session_date,
                "lab_round": getattr(session, "lab_round", ""),
                "session_type": getattr(session, "session_type", ""),
                "thematic_area": thematic_name,
                "topic": topic_name,
                "ls": yes_no(getattr(pr, "ls", False)),
                "mc": yes_no(getattr(pr, "mc", False)),
                "competency_status": safe_text(getattr(pr, "competency_status", "")),
                "checklist_score": getattr(pr, "checklist_score", ""),
                "topic_status": "Graduated / Completed" if graduated else "Needs graduation / follow-up MC",
            })

        participant_profile_rows = []

        for _key, profile in participant_profiles.items():
            active_topics = profile["ls_topics"] | profile["mc_topics"]

            # Any LS/MC topic not graduated is considered pending graduation/follow-up
            profile["needs_graduation_topics"] = active_topics - profile["graduated_topics"]

            ls_topics = sorted(profile["ls_topics"])
            mc_topics = sorted(profile["mc_topics"])
            graduated_topics = sorted(profile["graduated_topics"])
            needs_topics = sorted(profile["needs_graduation_topics"])

            ls_count = len(ls_topics)
            mc_count = len(mc_topics)
            graduated_count = len(graduated_topics)
            needs_count = len(needs_topics)

            if needs_count > 0:
                overall_status = "Needs graduation / follow-up"
                status_badge = "warning"
            elif graduated_count > 0:
                overall_status = "Graduated / completed"
                status_badge = "success"
            else:
                overall_status = "No LS/MC topic recorded"
                status_badge = "neutral"

            participant_profile_rows.append({
                "province": profile["province"],
                "district": profile["district"],
                "skill_lab_facility": profile["skill_lab_facility"],
                "hfcode": profile["hfcode"],
                "participant_name": profile["participant_name"],
                "participant_facility": profile["participant_facility"],
                "participant_position": profile["participant_position"],
                "session_count": len([x for x in profile["sessions"] if x]),
                "topic_count": len(profile["topics_all"]),
                "ls_count": ls_count,
                "mc_count": mc_count,
                "graduated_count": graduated_count,
                "needs_count": needs_count,
                "ls_topics_text": "\n".join(ls_topics),
                "mc_topics_text": "\n".join(mc_topics),
                "graduated_topics_text": "\n".join(graduated_topics),
                "needs_topics_text": "\n".join(needs_topics),
                "last_session_date": profile["last_session_date"],
                "overall_status": overall_status,
                "status_badge": status_badge,
            })

        participant_profile_rows = sorted(
            participant_profile_rows,
            key=lambda x: (
                x["province"],
                x["district"],
                x["skill_lab_facility"],
                -x["needs_count"],
                x["participant_name"],
            ),
        )

        export_query = request.GET.copy()
        export_query["export"] = "1"

        return {
                "filters": {
                    "date_from": date_from,
                    "date_to": date_to,
                    "province": province_id,
                    "facility": facility_id,
                    "session_type": session_type,
                    "lab_round": lab_round,
                },
                "export_query": export_query.urlencode(),
                "province_options": province_options,
                "facility_options": facility_options,
                "session_type_options": session_type_options,
                "lab_round_options": lab_round_options,
                "kpis": kpis,
                "province_rows": province_rows,
                "facility_rows": facility_rows,
                "facility_driver_rows": facility_driver_rows,
                "trend_rows": trend_rows,
                "monthly_total_rows": monthly_total_rows,
                "first_session_rows": first_session_rows,
                "story_candidates": story_candidates,
                "bamyan_row": bamyan_row,
                "bamyan_facilities": bamyan_facilities,
                "bamyan_narrative": bamyan_narrative,
                "chart_data": chart_data,
                "participant_profile_rows": participant_profile_rows,
                "participant_detail_rows": participant_detail_rows,
                "methodology_note": (
                    "LS and MC figures represent participant-topic/session instances, not unique health workers. "
                    "A single participant may be included more than once across visits, topics, sessions, or competency assessments."
                ),
            }
    
    def _export_dashboard_excel(self, data):
        wb = Workbook()

        def style_sheet(ws):
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style="thin", color="D9E2F3"),
                right=Side(style="thin", color="D9E2F3"),
                top=Side(style="thin", color="D9E2F3"),
                bottom=Side(style="thin", color="D9E2F3"),
            )

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)

                for cell in col:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 45)

        ws = wb.active
        ws.title = "Province_Summary"
        ws.append([
            "Province", "Distinct Participants", "Sessions", "Facilities",
            "Topics", "LS", "MC", "MC - LS", "Interpretation",
        ])

        for r in data["province_rows"]:
            ws.append([
                r.get("province"), r.get("participants"), r.get("sessions"),
                r.get("facilities"), r.get("topics"), r.get("ls"),
                r.get("mc"), r.get("mc_minus_ls"), r.get("interpretation"),
            ])

        ws2 = wb.create_sheet("Facility_Detail")
        ws2.append([
            "Province", "District", "Facility", "HF Code",
            "Distinct Participants", "Sessions", "Topics",
            "LS", "MC", "MC - LS", "Interpretation",
        ])

        for r in data["facility_rows"]:
            ws2.append([
                r.get("province"), r.get("district"), r.get("facility"),
                r.get("hfcode"), r.get("participants"), r.get("sessions"),
                r.get("topics"), r.get("ls"), r.get("mc"),
                r.get("mc_minus_ls"), r.get("interpretation"),
            ])

        ws3 = wb.create_sheet("Monthly_Trend")
        ws3.append([
            "Province", "Month", "Distinct Participants",
            "Topics", "LS", "MC", "MC - LS",
        ])

        for r in data["trend_rows"]:
            ws3.append([
                r.get("province"), r.get("month_label"), r.get("participants"),
                r.get("topics"), r.get("ls"), r.get("mc"),
                r.get("mc_minus_ls"),
            ])

        ws4 = wb.create_sheet("First_Sessions")
        ws4.append([
            "Province", "First Session Date", "First Facility",
            "HF Code", "District", "Clinical Mentor", "Session Type",
        ])

        for r in data["first_session_rows"]:
            ws4.append([
                r.get("province"), r.get("first_date"), r.get("facility"),
                r.get("hfcode"), r.get("district"), r.get("mentor"),
                r.get("session_type"),
            ])

        ws5 = wb.create_sheet("Story_Candidates")
        ws5.append([
            "Province", "District", "Facility", "HF Code",
            "LS", "MC", "MC - LS", "Story Angle", "Story Score",
        ])

        for r in data["story_candidates"]:
            ws5.append([
                r.get("province"), r.get("district"), r.get("facility"),
                r.get("hfcode"), r.get("ls"), r.get("mc"),
                r.get("mc_minus_ls"), r.get("story_angle"),
                r.get("story_score"),
            ])

        ws6 = wb.create_sheet("Methodology_Notes")
        ws6.append(["Topic", "Explanation"])
        ws6.append(["LS and MC interpretation", data["methodology_note"]])
        ws6.append(["Bamyan explanation", data["bamyan_narrative"]])

        ws7 = wb.create_sheet("Participant_Profile")
        ws7.append([
            "Province",
            "District",
            "Skill Lab Facility",
            "HF Code",
            "Participant Name",
            "Participant Facility",
            "Participant Position",
            "Session Count",
            "Topic Count",
            "LS Topic Count",
            "MC Topic Count",
            "Graduated Topic Count",
            "Needs Graduation Topic Count",
            "LS Topics",
            "MC Topics",
            "Graduated Topics",
            "Topics Needing Graduation",
            "Last Session Date",
            "Overall Status",
        ])

        for r in data.get("participant_profile_rows", []):
            ws7.append([
                r.get("province"),
                r.get("district"),
                r.get("skill_lab_facility"),
                r.get("hfcode"),
                r.get("participant_name"),
                r.get("participant_facility"),
                r.get("participant_position"),
                r.get("session_count"),
                r.get("topic_count"),
                r.get("ls_count"),
                r.get("mc_count"),
                r.get("graduated_count"),
                r.get("needs_count"),
                r.get("ls_topics_text"),
                r.get("mc_topics_text"),
                r.get("graduated_topics_text"),
                r.get("needs_topics_text"),
                r.get("last_session_date"),
                r.get("overall_status"),
            ])


        ws8 = wb.create_sheet("Participant_Topic_Detail")
        ws8.append([
            "Province",
            "District",
            "Skill Lab Facility",
            "HF Code",
            "Participant Name",
            "Participant Facility",
            "Participant Position",
            "Session Date",
            "Lab Round",
            "Session Type",
            "Thematic Area",
            "Topic",
            "LS",
            "MC",
            "Competency Status",
            "Checklist Score",
            "Topic Status",
        ])

        for r in data.get("participant_detail_rows", []):
            ws8.append([
                r.get("province"),
                r.get("district"),
                r.get("skill_lab_facility"),
                r.get("hfcode"),
                r.get("participant_name"),
                r.get("participant_facility"),
                r.get("participant_position"),
                r.get("session_date"),
                r.get("lab_round"),
                r.get("session_type"),
                r.get("thematic_area"),
                r.get("topic"),
                r.get("ls"),
                r.get("mc"),
                r.get("competency_status"),
                r.get("checklist_score"),
                r.get("topic_status"),
            ])

        for sheet in wb.worksheets:
            style_sheet(sheet)

        timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
        filename = f"Skill_Lab_Dashboard_{timestamp}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
# =========================================================
# Participant Record
# =========================================================
@admin.register(SkillLabParticipantRecord)
class SkillLabParticipantRecordAdmin(SkillLabAdminMediaMixin, ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    form = SkillLabParticipantRecordForm

    list_display = (
        "mentee_name",
        "get_skill_lab",
        "get_session_date",
        "get_province",
        "thematic_area",
        "topic",
        "method_display",
        "competency_badge",
        "score_summary",
    )
    search_fields = (
        "mentee_name__firstname",
        "mentee_name__lastname",
        "mentee_name__fathername",
        "mentee_name__tazkiranumber",
        "session__skill_lab__name",
        "session__skill_lab__facility__name",
        "topic__name",
        "topic__nameeng",
        "thematic_area__name",
    )
    list_filter = (
        ParticipantProvinceFilter,
        "competency_status",
        "ls",
        "mc",
        "thematic_area",
    )
    list_select_related = (
        "session",
        "session__skill_lab",
        "session__skill_lab__facility",
        "session__skill_lab__facility__districtfk",
        "session__skill_lab__facility__districtfk__provincefk",
        "thematic_area",
        "topic",
        "mentee_name",
    )
    list_per_page = 50

    # Topic is normal dropdown for cascade.
    autocomplete_fields = ("session", "thematic_area")
    raw_id_fields = ("mentee_name",)

    fieldsets = (
        ("Participant Information", {
            "fields": (
                "session",
                "mentee_name",
            ),
            "classes": ("wide", "skilllab-card"),
        }),
        ("Topic and Method", {
            "fields": (
                ("thematic_area", "topic"),
                ("ls", "mc"),
                "competency_status",
            ),
            "classes": ("wide", "skilllab-card"),
        }),
        ("Assessment", {
            "fields": (
                ("pre_test_score", "post_test_score", "checklist_score"),
                ("demonstration_done", "return_demonstration_done"),
                "feedback_given",
                "next_followup_date",
            ),
            "classes": ("wide", "skilllab-card"),
        }),
        ("Remarks", {
            "fields": ("remarks",),
            "classes": ("collapse", "skilllab-card"),
        }),
    )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        province = user_province(request)

        # Show all mentees in user's province, not limited to one facility.
        if db_field.name == "mentee_name" and province and not request.user.is_superuser:
            kwargs["queryset"] = Skill_Lab_Mentee.objects.filter(
                hfname__districtfk__provincefk=province
            ).select_related("hfname", "position").order_by("firstname", "lastname")

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related(
            "session",
            "session__skill_lab",
            "session__skill_lab__facility",
            "session__skill_lab__facility__districtfk",
            "session__skill_lab__facility__districtfk__provincefk",
            "thematic_area",
            "topic",
            "mentee_name",
        )
        province = user_province(request)
        if province and not request.user.is_superuser:
            qs = qs.filter(session__skill_lab__facility__districtfk__provincefk=province)
        return qs

    def province_filter_kwargs(self, request):
        return {"session__skill_lab__facility__districtfk__provincefk": user_province(request)}

    @admin.display(description="Skill Lab")
    def get_skill_lab(self, obj):
        return obj.session.skill_lab.name if obj.session_id else "-"

    @admin.display(description="Date")
    def get_session_date(self, obj):
        return obj.session.session_date if obj.session_id else "-"

    @admin.display(description="Province")
    def get_province(self, obj):
        try:
            return obj.session.skill_lab.facility.districtfk.provincefk.name
        except Exception:
            return "-"

    @admin.display(description="Method")
    def method_display(self, obj):
        vals = []
        if obj.ls:
            vals.append("LS")
        if obj.mc:
            vals.append("MC")
        return ", ".join(vals) if vals else "-"

    @admin.display(description="Competency")
    def competency_badge(self, obj):
        css_map = {
            "NOT_STARTED": "sl-badge sl-badge-muted",
            "IN_PROGRESS": "sl-badge sl-badge-warning",
            "COMPETENT": "sl-badge sl-badge-success",
            "NEEDS_REPEAT": "sl-badge sl-badge-danger",
        }
        cls = css_map.get(obj.competency_status, "sl-badge")
        return format_html('<span class="{}">{}</span>', cls, obj.get_competency_status_display())

    @admin.display(description="Scores")
    def score_summary(self, obj):
        parts = []
        if obj.pre_test_score is not None:
            parts.append(f"Pre: {obj.pre_test_score}")
        if obj.post_test_score is not None:
            parts.append(f"Post: {obj.post_test_score}")
        if obj.checklist_score is not None:
            parts.append(f"CL: {obj.checklist_score}")
        return " | ".join(parts) if parts else "-"


# =========================================================
# Skill Lab Mentee
# =========================================================
@admin.register(Skill_Lab_Mentee)
class SkillLabMenteeAdmin(SkillLabAdminMediaMixin, ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    list_display = (
        "firstname",
        "lastname",
        "fathername",
        "position",
        "hfname",
        "get_province",
        "status",
        "tazkiranumber",
    )
    search_fields = (
        "firstname",
        "lastname",
        "fathername",
        "tazkiranumber",
        "hfname__name",
        "hfname__districtfk__name",
        "hfname__districtfk__provincefk__name",
        "position__name",
    )
    list_filter = ("status", "position")
    raw_id_fields = ("hfname", "position")

    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related(
            "hfname",
            "hfname__districtfk",
            "hfname__districtfk__provincefk",
            "position",
        )
        province = user_province(request)
        if province and not request.user.is_superuser:
            qs = qs.filter(hfname__districtfk__provincefk=province)
        return qs

    def province_filter_kwargs(self, request):
        return {"hfname__districtfk__provincefk": user_province(request)}

    @admin.display(description="Province")
    def get_province(self, obj):
        try:
            return obj.hfname.districtfk.provincefk.name
        except Exception:
            return "-"
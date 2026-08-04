from urllib.parse import urlencode
from django.db.models import Sum, Count, F
from django.db.models import Count, Max, Min, Q
import openpyxl
from datetime import datetime
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict
from django.contrib import admin
from django.db import transaction
from django.db.models import Count, Q
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils.html import format_html
from urllib.parse import urlencode
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime
from django import forms
from django.contrib import admin
from django.contrib.auth import get_user_model
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.db import connection, transaction
from django.db.models import Count, Q
from django.forms.models import BaseInlineFormSet
from django.http import HttpResponse, JsonResponse
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from .forms import AimpeeAdminForm, AimpphAdminForm
from .models import (
    HQIPAssessmentHeader,
    HQIPAssessment,
    Score,
    Criteria,
    Standards,
    Section,
    Area,
    Assessmenttype,
    Province,
    District,
    Facility,
    Facilitytype,
    Implementor,
    Assessor,
    UserProfile,
    safesurgeryclinical,
    aimpee,
    aimpph,
    Mpdsr,
    Qicdataset,
    Participantposition,
    Participanteducation,
    Trainingheader,
    Training,
    Participationtype,
    Position,
    WhoChildbirthChecklistMonthly,
    QICommittee, FacilityStaff,ShamsiMonth, ShamsiYear, Period, BaselineProgress, GregorianMonth, GregorianYear, 
    AimPPHDashboard,HQIPAssessmentDashboard,
)
from django.utils.http import urlencode
from decimal import Decimal, InvalidOperation
from django.core.exceptions import ValidationError
from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from io import BytesIO
from decimal import Decimal
from datetime import datetime, date
from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import ForeignKey
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.contrib import messages
import openpyxl
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime as py_datetime, date as py_date
from django.contrib import messages
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import unicodedata
from django.db import router, transaction

# ============================================================
# Admin Branding
# ============================================================

admin.site.site_header = "Maternal and Newborn Health Information Management System (MNHIMS)"
admin.site.site_title = "IQoC Portal"
admin.site.index_title = "M&E Data Management System"

class HQIPProvinceFilter(admin.SimpleListFilter):
    title = "Province"
    parameter_name = "province"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)

        rows = qs.values_list(
            "facilityfk__districtfk__provincefk__id",
            "facilityfk__districtfk__provincefk__name",
        ).distinct().order_by(
            "facilityfk__districtfk__provincefk__name"
        )
        return [(pid, pname) for pid, pname in rows if pid]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(
                facilityfk__districtfk__provincefk_id=self.value()
            )
        return queryset

# ============================================================
# Province helper (supports user.profile OR user.userprofile)
# ============================================================

def user_province(request):
    if request.user.is_superuser:
        return None
    profile = getattr(request.user, "profile", None) or getattr(request.user, "userprofile", None)
    return getattr(profile, "province", None)

class ProvinceRestrictedAdminMixin:
    """
    Universal restriction for province-based access.
    Subclasses must implement:
      - province_filter_kwargs(request) -> dict of filters
    """

    def province_filter_kwargs(self, request):
        raise NotImplementedError

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        prov = user_province(request)
        if prov is None:
            return qs.none()
        return qs.filter(**self.province_filter_kwargs(request))

    def _obj_in_scope(self, request, obj):
        return self.get_queryset(request).filter(pk=obj.pk).exists()

    def has_view_permission(self, request, obj=None):
        base = super().has_view_permission(request, obj=obj)
        if not base:
            return False
        if obj is None or request.user.is_superuser:
            return True
        return self._obj_in_scope(request, obj)

    def has_change_permission(self, request, obj=None):
        base = super().has_change_permission(request, obj=obj)
        if not base:
            return False
        if obj is None or request.user.is_superuser:
            return True
        return self._obj_in_scope(request, obj)

    def has_delete_permission(self, request, obj=None):
        base = super().has_delete_permission(request, obj=obj)
        if not base:
            return False
        if obj is None or request.user.is_superuser:
            return True
        return self._obj_in_scope(request, obj)

# ============================================================
# Reusable filters
# ============================================================

class ProvinceFromFacilityFilter(admin.SimpleListFilter):
    title = "Province"
    parameter_name = "province"
    province_path = None  # override in subclasses

    def lookups(self, request, model_admin):
        if not self.province_path:
            return []
        qs = model_admin.get_queryset(request)
        provinces = qs.values_list(
            f"{self.province_path}__id",
            f"{self.province_path}__name",
        ).distinct().order_by(f"{self.province_path}__name")
        return [(pid, pname) for pid, pname in provinces if pid]

    def queryset(self, request, queryset):
        if self.value() and self.province_path:
            return queryset.filter(**{f"{self.province_path}__id": self.value()})
        return queryset

class DistrictFilter(admin.SimpleListFilter):
    title = "District"
    parameter_name = "district"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)
        districts = qs.values_list(
            "aimfacilityname__districtfk__id",
            "aimfacilityname__districtfk__name",
        ).distinct()
        return [(did, dname) for did, dname in districts if did is not None]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(aimfacilityname__districtfk__id=self.value())
        return queryset
    
class WhodistrictFilter(admin.SimpleListFilter):
    title = "District"
    parameter_name = "district"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)
        districts = qs.values_list(
            "facility_name__districtfk__id",
            "facility_name__districtfk__name",
        ).distinct()
        return [(did, dname) for did, dname in districts if did is not None]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(facility_name__districtfk__id=self.value())
        return queryset

class AimpeeFacilityFilter(admin.SimpleListFilter):
    title = "Facility"
    parameter_name = "facility"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)
        prov_id = request.GET.get("province")
        if prov_id:
            qs = qs.filter(aimfacilityname__districtfk__provincefk__id=prov_id)

        facilities = qs.values_list("aimfacilityname__id", "aimfacilityname__name").distinct().order_by("aimfacilityname__name")
        return [(fid, fname) for fid, fname in facilities if fid]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(aimfacilityname__id=self.value())
        return queryset
    
class FacilityFilter(admin.SimpleListFilter):
    title = "Facility"
    parameter_name = "facility"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)
        prov_id = request.GET.get("province")
        if prov_id:
            qs = qs.filter(facilityname__districtfk__provincefk__id=prov_id)

        facilities = qs.values_list("facility_name__id", "facility_name__name").distinct().order_by("facility_name__name")
        return [(fid, fname) for fid, fname in facilities if fid]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(facilityname__id=self.value())
        return queryset
    
# ============================================================
# AIM-PEE DUPLICATE PROTECTION
# ============================================================

def normalize_aimpee_period_value(value):
    """
    Normalize Shamsi year/month values for comparison.

    Examples:
        " 10 " becomes "10"
        "01" becomes "1"
        Persian/Arabic numbers are converted to English numbers
        "HAMAL" and "hamal" are treated as the same value
    """
    if value in (None, ""):
        return ""

    text = " ".join(str(value).split()).casefold()

    normalized_characters = []

    for character in text:
        if character.isdecimal():
            try:
                normalized_characters.append(
                    str(unicodedata.decimal(character))
                )
            except (TypeError, ValueError):
                normalized_characters.append(character)
        else:
            normalized_characters.append(character)

    normalized_text = "".join(normalized_characters)

    # Treat values such as 01 and 1 as the same month
    if normalized_text.isdigit():
        return str(int(normalized_text))

    return normalized_text


class AimpeeDuplicateProtectedAdminForm(AimpeeAdminForm):
    """
    Extends the existing AIM-PEE admin form.

    Prevents duplicate monthly records using:
        Health Facility + Shamsi Year + Shamsi Month
    """

    def clean(self):
        cleaned_data = super().clean()

        facility = cleaned_data.get("aimfacilityname")
        shamsi_year = cleaned_data.get("shamsiyear")
        shamsi_month = cleaned_data.get("shamsimonth")

        # Remove accidental spaces before saving
        if shamsi_year not in (None, ""):
            shamsi_year = str(shamsi_year).strip()
            cleaned_data["shamsiyear"] = shamsi_year

        if shamsi_month not in (None, ""):
            shamsi_month = str(shamsi_month).strip()
            cleaned_data["shamsimonth"] = shamsi_month

        # Let the normal form validation handle missing required fields
        if not facility or not shamsi_year or not shamsi_month:
            return cleaned_data

        normalized_year = normalize_aimpee_period_value(shamsi_year)
        normalized_month = normalize_aimpee_period_value(shamsi_month)

        model = self._meta.model
        database = router.db_for_write(
            model,
            instance=self.instance,
        )

        duplicate_records = (
            model._default_manager.using(database)
            .filter(aimfacilityname_id=facility.pk)
            .only(
                "pk",
                "aimfacilityname_id",
                "shamsiyear",
                "shamsimonth",
            )
            .order_by("pk")
        )

        # Allow users to edit the current record
        if self.instance and self.instance.pk:
            duplicate_records = duplicate_records.exclude(
                pk=self.instance.pk
            )

        existing_record = None

        for record in duplicate_records:
            existing_year = normalize_aimpee_period_value(
                record.shamsiyear
            )
            existing_month = normalize_aimpee_period_value(
                record.shamsimonth
            )

            if (
                existing_year == normalized_year
                and existing_month == normalized_month
            ):
                existing_record = record
                break

        if existing_record:
            duplicate_message = (
                "Duplicate AIM-PEE entry was not saved. "
                f"Record ID {existing_record.pk} already exists for "
                f"{facility}, Shamsi year {shamsi_year}, and "
                f"Shamsi month {shamsi_month}. "
                "Please open and update the existing record instead."
            )

            if "shamsimonth" in self.fields:
                self.add_error(
                    "shamsimonth",
                    duplicate_message,
                )
            else:
                self.add_error(
                    None,
                    duplicate_message,
                )

        return cleaned_data


# ============================================================
# AIM-PEE ADMIN
# ============================================================

@admin.register(aimpee)
class AimpeeAdmin(
    ProvinceRestrictedAdminMixin,
    admin.ModelAdmin,
):
    form = AimpeeDuplicateProtectedAdminForm

    list_display = (
        "id",
        "get_province",
        "aimfacilityname",
        "shamsiyear",
        "shamsimonth",
        "bl_progress",
        "gre_year",
        "gre_month",

        # ANC Screening
        "anc_total_seen",
        "anc_bp_measured",
        "preeclampsia_diagnosed",

        # Severe cases and treatment
        "severe_pree_or_eclampsia",
        "severe_pree_antihypertensive_within_1hr",
        "magnesium_sulfate_within_1hr",

        # Admissions
        "spe_admissions_before_delivery",
        "eclampsia_admissions_before_delivery",

        # Outcomes
        "total_complications",
        "maternal_death",
    )

    list_filter = (
        DistrictFilter,
        AimpeeFacilityFilter,
    )

    search_fields = (
        "aimfacilityname__name",
        "aimfacilityname__hfcode",
    )

    list_select_related = (
        "aimfacilityname__districtfk__provincefk",
    )

    list_per_page = 10
    save_on_top = True

    actions = [
        "export_aimpee_to_excel",
    ]

    # --------------------------------------------------------
    # Province display
    # --------------------------------------------------------
    @admin.display(
        description="Province",
        ordering="aimfacilityname__districtfk__provincefk__name",
    )
    def get_province(self, obj):
        facility = getattr(obj, "aimfacilityname", None)

        if not facility:
            return ""

        district = getattr(facility, "districtfk", None)

        if not district:
            return ""

        province = getattr(district, "provincefk", None)

        if not province:
            return ""

        return getattr(province, "name", "")

    # --------------------------------------------------------
    # Province restriction
    # --------------------------------------------------------
    def province_filter_kwargs(self, request):
        return {
            "aimfacilityname__districtfk__provincefk": (
                user_province(request)
            )
        }

    def formfield_for_foreignkey(
        self,
        db_field,
        request,
        **kwargs,
    ):
        if (
            db_field.name == "aimfacilityname"
            and not request.user.is_superuser
        ):
            province = user_province(request)

            if province:
                kwargs["queryset"] = (
                    Facility.objects
                    .filter(districtfk__provincefk=province)
                    .order_by("name")
                )
            else:
                kwargs["queryset"] = Facility.objects.none()

        return super().formfield_for_foreignkey(
            db_field,
            request,
            **kwargs,
        )

    # --------------------------------------------------------
    # Concurrent-submission and double-click protection
    # --------------------------------------------------------
    def _lock_selected_facility(
        self,
        request,
        database,
    ):
        """
        Lock the selected facility during form validation and saving.

        This prevents two simultaneous admin submissions from
        creating duplicate monthly records.
        """
        raw_facility_id = request.POST.get("aimfacilityname")

        if not raw_facility_id:
            return

        try:
            facility_id = Facility._meta.pk.to_python(
                raw_facility_id
            )
        except (TypeError, ValueError, ValidationError):
            # The form will show the invalid facility error
            return

        try:
            (
                Facility._default_manager.using(database)
                .select_for_update()
                .get(pk=facility_id)
            )
        except Facility.DoesNotExist:
            # The form will handle a missing/invalid facility
            return

    def changeform_view(
        self,
        request,
        object_id=None,
        form_url="",
        extra_context=None,
    ):
        """
        Wrap AIM-PEE POST submissions in a database transaction.
        """
        if request.method != "POST":
            return super().changeform_view(
                request,
                object_id,
                form_url,
                extra_context,
            )

        database = router.db_for_write(self.model)

        with transaction.atomic(using=database):
            self._lock_selected_facility(
                request,
                database,
            )

            return super().changeform_view(
                request,
                object_id,
                form_url,
                extra_context,
            )

    # --------------------------------------------------------
    # Excel Export Action
    # --------------------------------------------------------
    @admin.action(
        description="Export selected AIM-PEE records to Excel"
    )
    def export_aimpee_to_excel(
        self,
        request,
        queryset,
    ):
        """
        Export selected AIM-PEE records to Excel.

        Includes:
        - Province
        - District
        - Facility code
        - Facility name
        - Every database field in the AIM-PEE model
        """

        queryset = queryset.select_related(
            "aimfacilityname",
            "aimfacilityname__districtfk",
            "aimfacilityname__districtfk__provincefk",
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "AIM-PEE Export"

        # ----------------------------------------------------
        # Helper functions
        # ----------------------------------------------------
        def clean_value(value):
            if value is None:
                return ""

            if isinstance(value, Decimal):
                return float(value)

            if isinstance(value, datetime):
                if timezone.is_aware(value):
                    value = timezone.localtime(value)

                return value.replace(tzinfo=None)

            if isinstance(value, date):
                return value

            if isinstance(value, bool):
                return "Yes" if value else "No"

            return value

        def safe_get(obj, attribute_path):
            current = obj

            for attribute in attribute_path.split("__"):
                current = getattr(
                    current,
                    attribute,
                    None,
                )

                if current is None:
                    return ""

            return current

        # ----------------------------------------------------
        # Build columns
        # ----------------------------------------------------
        columns = [
            (
                "Province",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__districtfk__provincefk__name",
                ),
            ),
            (
                "District",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__districtfk__name",
                ),
            ),
            (
                "HF Code",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__hfcode",
                ),
            ),
            (
                "Facility Name",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__name",
                ),
            ),
        ]

        # Add every real database field
        for field in self.model._meta.fields:
            if isinstance(field, ForeignKey):
                columns.append(
                    (
                        str(field.verbose_name).title(),
                        lambda obj, current_field=field: str(
                            getattr(
                                obj,
                                current_field.name,
                                "",
                            )
                            or ""
                        ),
                    )
                )

                columns.append(
                    (
                        f"{field.name}_id",
                        lambda obj, current_field=field: getattr(
                            obj,
                            current_field.attname,
                            "",
                        ),
                    )
                )

            else:
                columns.append(
                    (
                        str(field.verbose_name).title(),
                        lambda obj, current_field=field: clean_value(
                            getattr(
                                obj,
                                f"get_{current_field.name}_display",
                            )()
                            if current_field.choices
                            else getattr(
                                obj,
                                current_field.name,
                                "",
                            )
                        ),
                    )
                )

        # ----------------------------------------------------
        # Write headers
        # ----------------------------------------------------
        headers = [
            column_name
            for column_name, value_function in columns
        ]

        worksheet.append(headers)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # ----------------------------------------------------
        # Write data
        # ----------------------------------------------------
        for obj in queryset:
            row = [
                clean_value(value_function(obj))
                for column_name, value_function in columns
            ]

            worksheet.append(row)

        # ----------------------------------------------------
        # Excel formatting
        # ----------------------------------------------------
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_number, column_cells in enumerate(
            worksheet.columns,
            start=1,
        ):
            maximum_length = 0
            column_letter = get_column_letter(column_number)

            for cell in column_cells:
                if cell.value is None:
                    value_length = 0
                else:
                    value_length = len(str(cell.value))

                maximum_length = max(
                    maximum_length,
                    value_length,
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(maximum_length + 3, 45)

        # ----------------------------------------------------
        # Return Excel file
        # ----------------------------------------------------
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            "aimpee_export_"
            f"{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )

        return response

# ============================================================
# AIM-PPH DUPLICATE PROTECTION
# ============================================================

def normalize_aimpph_period_value(value):
    """
    Normalize Shamsi year and month values for comparison.

    Examples:
        " 10 " becomes "10"
        "01" becomes "1"
        Persian/Arabic digits are converted to English digits
        "HAMAL" and "hamal" are treated as the same value
    """
    if value in (None, ""):
        return ""

    text = unicodedata.normalize("NFKC", str(value))
    text = " ".join(text.split()).casefold()

    normalized_characters = []

    for character in text:
        if character.isdecimal():
            try:
                normalized_characters.append(
                    str(unicodedata.decimal(character))
                )
            except (TypeError, ValueError):
                normalized_characters.append(character)
        else:
            normalized_characters.append(character)

    normalized_text = "".join(normalized_characters)

    # Treat values such as 01 and 1 as the same month
    if normalized_text.isdigit():
        return str(int(normalized_text))

    return normalized_text


class AimpphDuplicateProtectedAdminForm(AimpphAdminForm):
    """
    Extends the existing AIM-PPH admin form.

    Prevents duplicate monthly records using:
        Health Facility + Shamsi Year + Shamsi Month
    """

    def clean(self):
        cleaned_data = super().clean()

        facility = cleaned_data.get("aimfacilityname")
        shamsi_year = cleaned_data.get("shamsiyear")
        shamsi_month = cleaned_data.get("shamsimonth")

        # Remove accidental spaces before saving
        if shamsi_year not in (None, ""):
            shamsi_year = str(shamsi_year).strip()
            cleaned_data["shamsiyear"] = shamsi_year

        if shamsi_month not in (None, ""):
            shamsi_month = str(shamsi_month).strip()
            cleaned_data["shamsimonth"] = shamsi_month

        # Let the normal form validation handle missing fields
        if not facility or not shamsi_year or not shamsi_month:
            return cleaned_data

        normalized_year = normalize_aimpph_period_value(
            shamsi_year
        )
        normalized_month = normalize_aimpph_period_value(
            shamsi_month
        )

        model = self._meta.model
        database = router.db_for_write(
            model,
            instance=self.instance,
        )

        duplicate_records = (
            model._default_manager.using(database)
            .filter(aimfacilityname_id=facility.pk)
            .only(
                "pk",
                "aimfacilityname_id",
                "shamsiyear",
                "shamsimonth",
            )
            .order_by("pk")
        )

        # Allow users to edit the current record
        if self.instance and self.instance.pk:
            duplicate_records = duplicate_records.exclude(
                pk=self.instance.pk
            )

        existing_record = None

        for record in duplicate_records:
            existing_year = normalize_aimpph_period_value(
                record.shamsiyear
            )
            existing_month = normalize_aimpph_period_value(
                record.shamsimonth
            )

            if (
                existing_year == normalized_year
                and existing_month == normalized_month
            ):
                existing_record = record
                break

        if existing_record:
            duplicate_message = (
                "Duplicate AIM-PPH entry was not saved. "
                f"Record ID {existing_record.pk} already exists for "
                f"{facility}, Shamsi year {shamsi_year}, and "
                f"Shamsi month {shamsi_month}. "
                "Please open and update the existing record instead."
            )

            if "shamsimonth" in self.fields:
                self.add_error(
                    "shamsimonth",
                    duplicate_message,
                )
            else:
                self.add_error(
                    None,
                    duplicate_message,
                )

        return cleaned_data


# ============================================================
# AIM-PPH ADMIN
# ============================================================

@admin.register(aimpph)
class AimpphAdmin(
    ProvinceRestrictedAdminMixin,
    admin.ModelAdmin,
):
    form = AimpphDuplicateProtectedAdminForm

    list_display = (
        "id",
        "get_province",
        "aimfacilityname",
        "shamsiyear",
        "shamsimonth",
        "bl_progress",
        "gre_year",
        "gre_month",
        "total_births",
        "births_vaginal",
        "births_csection",
        "pph_vaginal_501_999",
        "pph_cs_1000_plus",
        "maternal_death_pph_transfer",
        "ai_total",
    )

    list_filter = (
        DistrictFilter,
        AimpeeFacilityFilter,
    )

    search_fields = (
        "aimfacilityname__name",
        "aimfacilityname__hfcode",
    )

    list_select_related = (
        "aimfacilityname__districtfk__provincefk",
    )

    list_per_page = 10
    save_on_top = True

    actions = [
        "export_aimpph_to_excel",
    ]

    # --------------------------------------------------------
    # Province display
    # --------------------------------------------------------
    @admin.display(
        description="Province",
        ordering="aimfacilityname__districtfk__provincefk__name",
    )
    def get_province(self, obj):
        facility = getattr(obj, "aimfacilityname", None)

        if not facility:
            return ""

        district = getattr(facility, "districtfk", None)

        if not district:
            return ""

        province = getattr(district, "provincefk", None)

        if not province:
            return ""

        return getattr(province, "name", "")

    # --------------------------------------------------------
    # Province restriction
    # --------------------------------------------------------
    def province_filter_kwargs(self, request):
        return {
            "aimfacilityname__districtfk__provincefk": (
                user_province(request)
            )
        }

    def formfield_for_foreignkey(
        self,
        db_field,
        request,
        **kwargs,
    ):
        if (
            db_field.name == "aimfacilityname"
            and not request.user.is_superuser
        ):
            province = user_province(request)

            if province:
                kwargs["queryset"] = (
                    Facility.objects
                    .filter(districtfk__provincefk=province)
                    .order_by("name")
                )
            else:
                kwargs["queryset"] = Facility.objects.none()

        return super().formfield_for_foreignkey(
            db_field,
            request,
            **kwargs,
        )

    # --------------------------------------------------------
    # Concurrent submission and double-click protection
    # --------------------------------------------------------
    def _lock_selected_facility(
        self,
        request,
        database,
    ):
        """
        Lock the selected facility while the AIM-PPH form is
        being validated and saved.

        This prevents simultaneous admin submissions from
        creating the same monthly record.
        """
        raw_facility_id = request.POST.get("aimfacilityname")

        if not raw_facility_id:
            return

        try:
            facility_id = Facility._meta.pk.to_python(
                raw_facility_id
            )
        except (TypeError, ValueError, ValidationError):
            # The form will display the invalid facility error
            return

        try:
            (
                Facility._default_manager.using(database)
                .select_for_update()
                .get(pk=facility_id)
            )
        except Facility.DoesNotExist:
            # The form will handle the invalid facility
            return

    def changeform_view(
        self,
        request,
        object_id=None,
        form_url="",
        extra_context=None,
    ):
        """
        Wrap AIM-PPH POST submissions in a database transaction.
        """
        if request.method != "POST":
            return super().changeform_view(
                request,
                object_id,
                form_url,
                extra_context,
            )

        database = router.db_for_write(self.model)

        with transaction.atomic(using=database):
            self._lock_selected_facility(
                request,
                database,
            )

            return super().changeform_view(
                request,
                object_id,
                form_url,
                extra_context,
            )

    # --------------------------------------------------------
    # Excel Export Action
    # --------------------------------------------------------
    @admin.action(
        description="Export selected AIM-PPH records to Excel"
    )
    def export_aimpph_to_excel(
        self,
        request,
        queryset,
    ):
        """
        Export selected AIM-PPH records to Excel.

        Includes Province, District, Facility Code,
        Facility Name, and every AIM-PPH database field.
        """

        queryset = queryset.select_related(
            "aimfacilityname",
            "aimfacilityname__districtfk",
            "aimfacilityname__districtfk__provincefk",
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "AIM-PPH Export"

        # ----------------------------------------------------
        # Helper functions
        # ----------------------------------------------------
        def clean_value(value):
            if value is None:
                return ""

            if isinstance(value, Decimal):
                return float(value)

            if isinstance(value, datetime):
                if timezone.is_aware(value):
                    value = timezone.localtime(value)

                return value.replace(tzinfo=None)

            if isinstance(value, date):
                return value

            if isinstance(value, bool):
                return "Yes" if value else "No"

            return value

        def safe_get(obj, attribute_path):
            current = obj

            for attribute in attribute_path.split("__"):
                current = getattr(
                    current,
                    attribute,
                    None,
                )

                if current is None:
                    return ""

            return current

        # ----------------------------------------------------
        # Build export columns
        # ----------------------------------------------------
        columns = [
            (
                "Province",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__districtfk__provincefk__name",
                ),
            ),
            (
                "District",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__districtfk__name",
                ),
            ),
            (
                "HF Code",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__hfcode",
                ),
            ),
            (
                "Facility Name",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__name",
                ),
            ),
        ]

        # Add every real database field from AIM-PPH
        for field in self.model._meta.fields:
            if isinstance(field, ForeignKey):
                columns.append(
                    (
                        str(field.verbose_name).title(),
                        lambda obj, current_field=field: str(
                            getattr(
                                obj,
                                current_field.name,
                                "",
                            )
                            or ""
                        ),
                    )
                )

                columns.append(
                    (
                        f"{field.name}_id",
                        lambda obj, current_field=field: getattr(
                            obj,
                            current_field.attname,
                            "",
                        ),
                    )
                )

            else:
                columns.append(
                    (
                        str(field.verbose_name).title(),
                        lambda obj, current_field=field: clean_value(
                            getattr(
                                obj,
                                f"get_{current_field.name}_display",
                            )()
                            if current_field.choices
                            else getattr(
                                obj,
                                current_field.name,
                                "",
                            )
                        ),
                    )
                )

        # ----------------------------------------------------
        # Write headers
        # ----------------------------------------------------
        headers = [
            column_name
            for column_name, value_function in columns
        ]

        worksheet.append(headers)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # ----------------------------------------------------
        # Write data rows
        # ----------------------------------------------------
        for obj in queryset:
            row = [
                clean_value(value_function(obj))
                for column_name, value_function in columns
            ]

            worksheet.append(row)

        # ----------------------------------------------------
        # Excel formatting
        # ----------------------------------------------------
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_number, column_cells in enumerate(
            worksheet.columns,
            start=1,
        ):
            maximum_length = 0
            column_letter = get_column_letter(column_number)

            for cell in column_cells:
                if cell.value is None:
                    value_length = 0
                else:
                    value_length = len(str(cell.value))

                maximum_length = max(
                    maximum_length,
                    value_length,
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(maximum_length + 3, 45)

        # ----------------------------------------------------
        # Return Excel response
        # ----------------------------------------------------
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            "aimpph_export_"
            f"{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )

        return response

@admin.register(AimPPHDashboard)
class AimPPHDashboardAdmin(ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    change_list_template = "admin/aimpph/dashboard.html"

    # ============================================================
    # Permissions
    # ============================================================
    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return request.user.is_active and request.user.is_staff

    def get_model_perms(self, request):
        if self.has_view_permission(request):
            return {"view": True}
        return {}

    # ============================================================
    # Helpers
    # ============================================================
    def _province_id(self, request):
        province = user_province(request)
        return getattr(province, "id", province) if province else None

    def _safe_int(self, value):
        try:
            return int(value or 0)
        except Exception:
            return 0

    def _safe_rate(self, numerator, denominator, decimals=1):
        numerator = self._safe_int(numerator)
        denominator = self._safe_int(denominator)

        if denominator <= 0:
            return 0

        return round((numerator / denominator) * 100, decimals)

    def _sum_fields(self, queryset, fields):
        aggregation = {
            f"sum_{field}": Sum(field)
            for field in fields
        }

        result = queryset.aggregate(**aggregation)

        return {
            field: self._safe_int(result.get(f"sum_{field}"))
            for field in fields
        }

    def _normalize_sum_fields(self, row, fields):
        for field in fields:
            row[field] = self._safe_int(row.get(f"sum_{field}"))
        return row

    def _calculated_births(self, row_or_totals):
        """
        Temporary dashboard correction:
        Total Births = Vaginal Births + C-section Births
        """
        return (
            self._safe_int(row_or_totals.get("births_vaginal"))
            + self._safe_int(row_or_totals.get("births_csection"))
        )

    def _calculated_pph_total(self, row_or_totals):
        """
        Temporary dashboard correction:
        Indicator 10 = Indicator 6 + Indicator 7 + Indicator 8 + Indicator 9
        """
        return (
            self._safe_int(row_or_totals.get("pph_vaginal_501_999"))
            + self._safe_int(row_or_totals.get("pph_cs_1000_plus"))
            + self._safe_int(row_or_totals.get("pph_referral_in_outside_aim"))
            + self._safe_int(row_or_totals.get("pph_referral_in_aim"))
        )

    def _progress_filter(self, queryset, progress_type):
        """
        Supports common PRE-I / PRE-P spelling variations.
        """
        if progress_type == "PRE-I":
            return queryset.filter(
                Q(bl_progress__iexact="PRE-I")
                | Q(bl_progress__iexact="PRE I")
                | Q(bl_progress__iexact="PREI")
                | Q(bl_progress__iexact="PRE_I")
            )

        if progress_type == "PRE-P":
            return queryset.filter(
                Q(bl_progress__iexact="PRE-P")
                | Q(bl_progress__iexact="PRE P")
                | Q(bl_progress__iexact="PREP")
                | Q(bl_progress__iexact="PRE_P")
            )

        return queryset.none()

    def _percent_change_display(self, pre_i, pre_p):
        pre_i = self._safe_int(pre_i)
        pre_p = self._safe_int(pre_p)

        if pre_i == 0 and pre_p == 0:
            return "0.00"

        if pre_i == 0 and pre_p > 0:
            return "New"

        percent_change = ((pre_p - pre_i) / pre_i) * 100
        return f"{percent_change:.2f}"

    def _improvement_status(self, absolute_change, direction):
        """
        Direction rules:
        - Higher is better: positive change = improved.
        - Lower is better: negative change = improved.
        - Context review: any change requires interpretation.
        """

        if direction == "Higher is better":
            if absolute_change > 0:
                return "Improved", "Potential positive change to verify"
            if absolute_change < 0:
                return "Declined", "Requires review before story use"
            return "No change", ""

        if direction == "Lower is better":
            if absolute_change < 0:
                return "Improved", "Potential positive change to verify"
            if absolute_change > 0:
                return "Declined", "Requires review before story use"
            return "No change", ""

        if absolute_change == 0:
            return "No change", ""

        return "Review", "Requires technical interpretation"

    # ============================================================
    # Main view
    # ============================================================
    def changelist_view(self, request, extra_context=None):
        data = self._build_dashboard_data(request)

        if request.GET.get("export") == "1":
            return self._export_dashboard_excel(data)

        context = {
            **self.admin_site.each_context(request),
            "title": "AIM-PPH Dashboard",
            **data,
        }

        return TemplateResponse(
            request,
            self.change_list_template,
            context,
        )

    def _base_queryset(self, request):
        queryset = aimpph.objects.select_related(
            "aimfacilityname",
            "aimfacilityname__districtfk",
            "aimfacilityname__districtfk__provincefk",
        )

        province_id = self._province_id(request)

        if province_id and not request.user.is_superuser:
            queryset = queryset.filter(
                aimfacilityname__districtfk__provincefk_id=province_id
            )

        return queryset

    def _apply_filters(self, request, queryset):
        province_id = request.GET.get("province", "").strip()
        facility_id = request.GET.get("facility", "").strip()
        gre_year = request.GET.get("gre_year", "").strip()
        gre_month = request.GET.get("gre_month", "").strip()
        shamsiyear = request.GET.get("shamsiyear", "").strip()
        shamsimonth = request.GET.get("shamsimonth", "").strip()
        bl_progress = request.GET.get("bl_progress", "").strip()

        if province_id:
            queryset = queryset.filter(
                aimfacilityname__districtfk__provincefk_id=province_id
            )

        if facility_id:
            queryset = queryset.filter(aimfacilityname_id=facility_id)

        if gre_year:
            queryset = queryset.filter(gre_year=gre_year)

        if gre_month:
            queryset = queryset.filter(gre_month=gre_month)

        if shamsiyear:
            queryset = queryset.filter(shamsiyear=shamsiyear)

        if shamsimonth:
            queryset = queryset.filter(shamsimonth=shamsimonth)

        if bl_progress:
            queryset = queryset.filter(bl_progress=bl_progress)

        filters = {
            "province": province_id,
            "facility": facility_id,
            "gre_year": gre_year,
            "gre_month": gre_month,
            "shamsiyear": shamsiyear,
            "shamsimonth": shamsimonth,
            "bl_progress": bl_progress,
        }

        return queryset, filters

    # ============================================================
    # Facility-level PRE-I vs PRE-P indicator comparison
    # ============================================================
    def _build_indicator_comparison_rows(self, queryset, indicator_fields):
        indicator_definitions = [
            {
                "label": "1. Corrected Total Births = Vaginal Births + C-section Births",
                "field": "total_births",
                "direction": "Higher is better",
            },
            {
                "label": "2. Number of births - vaginal delivery",
                "field": "births_vaginal",
                "direction": "Higher is better",
            },
            {
                "label": "3. Number of births - C-section",
                "field": "births_csection",
                "direction": "Context review",
            },
            {
                "label": "4. Number of patients receiving oxytocin immediately after birth",
                "field": "oxytocin_immediate",
                "direction": "Higher is better",
            },
            {
                "label": "5. Number of Antepartum Hemorrhage",
                "field": "antepartum_hemorrhage",
                "direction": "Lower is better",
            },
            {
                "label": "6. PPH after vaginal delivery 501–999 cc",
                "field": "pph_vaginal_501_999",
                "direction": "Lower is better",
            },
            {
                "label": "7. PPH after C-section ≥1000 cc",
                "field": "pph_cs_1000_plus",
                "direction": "Lower is better",
            },
            {
                "label": "8. PPH referrals in from outside AIM facilities",
                "field": "pph_referral_in_outside_aim",
                "direction": "Context review",
            },
            {
                "label": "9. PPH referrals in from AIM facilities",
                "field": "pph_referral_in_aim",
                "direction": "Context review",
            },
            {
                "label": "10. Total number of PPH cases - calculated indicators 6–9",
                "field": "pph_total",
                "direction": "Lower is better",
            },
            {
                "label": "QBL 0–500 ml",
                "field": "qbl_0_500",
                "direction": "Context review",
            },
            {
                "label": "QBL 501–999 ml",
                "field": "qbl_501_999",
                "direction": "Lower is better",
            },
            {
                "label": "QBL 1000–1499 ml",
                "field": "qbl_1000_1499",
                "direction": "Lower is better",
            },
            {
                "label": "QBL 1500–1999 ml",
                "field": "qbl_1500_1999",
                "direction": "Lower is better",
            },
            {
                "label": "QBL 2000–2499 ml",
                "field": "qbl_2000_2499",
                "direction": "Lower is better",
            },
            {
                "label": "QBL >2500 ml",
                "field": "qbl_2500_plus",
                "direction": "Lower is better",
            },
            {
                "label": "QBL unknown",
                "field": "qbl_unknown",
                "direction": "Lower is better",
            },
            {
                "label": "QBL total",
                "field": "qbl_total",
                "direction": "Context review",
            },
            {
                "label": "Transfers out for PPH",
                "field": "transfers_out_pph",
                "direction": "Lower is better",
            },
            {
                "label": "Maternal deaths due to PPH transfer",
                "field": "maternal_death_pph_transfer",
                "direction": "Lower is better",
            },
            {
                "label": "Maternal deaths due to other transfer causes",
                "field": "maternal_death_other_transfer",
                "direction": "Lower is better",
            },
            {
                "label": "Total maternal deaths transfer",
                "field": "maternal_death_total_transfer",
                "direction": "Lower is better",
            },
            {
                "label": "Cause: uterine atony",
                "field": "cause_uterine_atony",
                "direction": "Lower is better",
            },
            {
                "label": "Cause: severe lacerations",
                "field": "cause_severe_lacerations",
                "direction": "Lower is better",
            },
            {
                "label": "Cause: retained products",
                "field": "cause_retained_products",
                "direction": "Lower is better",
            },
            {
                "label": "Cause: DIC",
                "field": "cause_dic",
                "direction": "Lower is better",
            },
            {
                "label": "Cause: ruptured uterus",
                "field": "cause_ruptured_uterus",
                "direction": "Lower is better",
            },
            {
                "label": "Cause: abruption placenta",
                "field": "cause_abruption",
                "direction": "Lower is better",
            },
            {
                "label": "Cause: placenta previa",
                "field": "cause_placenta_previa",
                "direction": "Lower is better",
            },
            {
                "label": "Cause: placenta accreta",
                "field": "cause_placenta_accreta",
                "direction": "Lower is better",
            },
            {
                "label": "Cause: other",
                "field": "cause_other",
                "direction": "Lower is better",
            },
            {
                "label": "Cause: unknown",
                "field": "cause_unknown",
                "direction": "Lower is better",
            },
            {
                "label": "Total causes of PPH",
                "field": "causes_total",
                "direction": "Lower is better",
            },
            {
                "label": "PPH management by medication - uterotonic",
                "field": "pph_medication_uterotonic",
                "direction": "Higher is better",
            },
            {
                "label": "AI: uterine compression",
                "field": "ai_uterine_compression",
                "direction": "Context review",
            },
            {
                "label": "AI: manual removal of placenta",
                "field": "ai_manual_placenta",
                "direction": "Context review",
            },
            {
                "label": "AI: aortic compression",
                "field": "ai_aortic_compression",
                "direction": "Context review",
            },
            {
                "label": "AI: UBT",
                "field": "ai_ubt",
                "direction": "Context review",
            },
            {
                "label": "AI: laceration repair",
                "field": "ai_lac_repair",
                "direction": "Context review",
            },
            {
                "label": "AI: B-Lynch / UAL",
                "field": "ai_blynch_ual",
                "direction": "Context review",
            },
            {
                "label": "AI: NASG",
                "field": "ai_nasg",
                "direction": "Context review",
            },
            {
                "label": "AI: ruptured uterus repair",
                "field": "ai_ruptured_uterus_repair",
                "direction": "Context review",
            },
            {
                "label": "AI: PPH hysterectomy",
                "field": "ai_pph_hysterectomy",
                "direction": "Lower is better",
            },
            {
                "label": "AI: hysterectomy other causes",
                "field": "ai_hysterectomy_other",
                "direction": "Lower is better",
            },
            {
                "label": "Total advanced interventions",
                "field": "ai_total",
                "direction": "Context review",
            },
        ]

        sum_annotations = {
            f"sum_{field}": Sum(field)
            for field in indicator_fields
        }

        def build_grouped_totals(progress_type):
            progress_qs = self._progress_filter(queryset, progress_type)

            grouped_rows = list(
                progress_qs.values(
                    province=F("aimfacilityname__districtfk__provincefk__name"),
                    facility_id=F("aimfacilityname_id"),
                    facility=F("aimfacilityname__name"),
                )
                .annotate(**sum_annotations)
                .order_by("province", "facility")
            )

            output = {}

            for row in grouped_rows:
                self._normalize_sum_fields(row, indicator_fields)

                row["total_births"] = self._calculated_births(row)
                row["pph_total"] = self._calculated_pph_total(row)

                facility_id = row.get("facility_id")

                output[facility_id] = {
                    "province": row.get("province") or "",
                    "facility_id": facility_id,
                    "facility": row.get("facility") or "",
                    "values": row,
                }

            return output

        pre_i_map = build_grouped_totals("PRE-I")
        pre_p_map = build_grouped_totals("PRE-P")

        facility_ids = sorted(
            set(pre_i_map.keys()) | set(pre_p_map.keys()),
            key=lambda facility_id: (
                pre_i_map.get(facility_id, pre_p_map.get(facility_id, {})).get("province", ""),
                pre_i_map.get(facility_id, pre_p_map.get(facility_id, {})).get("facility", ""),
            ),
        )

        rows = []

        for facility_id in facility_ids:
            pre_i_info = pre_i_map.get(facility_id)
            pre_p_info = pre_p_map.get(facility_id)

            info = pre_i_info or pre_p_info or {}

            province = info.get("province", "")
            facility = info.get("facility", "")

            pre_i_values = pre_i_info.get("values", {}) if pre_i_info else {}
            pre_p_values = pre_p_info.get("values", {}) if pre_p_info else {}

            for indicator in indicator_definitions:
                field = indicator["field"]
                direction = indicator["direction"]

                pre_i = self._safe_int(pre_i_values.get(field))
                pre_p = self._safe_int(pre_p_values.get(field))
                absolute_change = pre_p - pre_i
                percent_change_display = self._percent_change_display(pre_i, pre_p)

                improvement_status, story_use = self._improvement_status(
                    absolute_change,
                    direction,
                )

                rows.append({
                    "province": province,
                    "facility": facility,
                    "indicator": indicator["label"],
                    "pre_i": pre_i,
                    "pre_p": pre_p,
                    "absolute_change": absolute_change,
                    "percent_change_display": percent_change_display,
                    "direction": direction,
                    "improvement_status": improvement_status,
                    "story_use": story_use,
                    "notes": (
                        "Verify with facility team, source documents, and field narrative "
                        "before using as success story."
                    ),
                })

        return rows

    # ============================================================
    # Build dashboard data
    # ============================================================
    def _build_dashboard_data(self, request):
        base_qs = self._base_queryset(request)
        queryset, filters = self._apply_filters(request, base_qs)

        # ============================================================
        # Filter options
        # ============================================================
        province_options = [
            {
                "province_id": row["aimfacilityname__districtfk__provincefk_id"],
                "province": row["aimfacilityname__districtfk__provincefk__name"] or "",
            }
            for row in (
                base_qs
                .exclude(aimfacilityname__districtfk__provincefk_id__isnull=True)
                .values(
                    "aimfacilityname__districtfk__provincefk_id",
                    "aimfacilityname__districtfk__provincefk__name",
                )
                .distinct()
                .order_by("aimfacilityname__districtfk__provincefk__name")
            )
        ]

        facility_options = [
            {
                "facility_id": row["aimfacilityname_id"],
                "facility": row["aimfacilityname__name"] or "",
            }
            for row in (
                base_qs
                .exclude(aimfacilityname_id__isnull=True)
                .values("aimfacilityname_id", "aimfacilityname__name")
                .distinct()
                .order_by("aimfacilityname__name")
            )
        ]

        gre_year_options = list(
            base_qs.values_list("gre_year", flat=True)
            .exclude(gre_year__isnull=True)
            .exclude(gre_year="")
            .distinct()
            .order_by("gre_year")
        )

        gre_month_options = list(
            base_qs.values_list("gre_month", flat=True)
            .exclude(gre_month__isnull=True)
            .exclude(gre_month="")
            .distinct()
            .order_by("gre_month")
        )

        shamsiyear_options = list(
            base_qs.values_list("shamsiyear", flat=True)
            .exclude(shamsiyear__isnull=True)
            .exclude(shamsiyear="")
            .distinct()
            .order_by("shamsiyear")
        )

        shamsimonth_options = list(
            base_qs.values_list("shamsimonth", flat=True)
            .exclude(shamsimonth__isnull=True)
            .exclude(shamsimonth="")
            .distinct()
            .order_by("shamsimonth")
        )

        bl_progress_options = list(
            base_qs.values_list("bl_progress", flat=True)
            .exclude(bl_progress__isnull=True)
            .exclude(bl_progress="")
            .distinct()
            .order_by("bl_progress")
        )

        # ============================================================
        # Aggregation fields
        # Stored total_births and pph_total are intentionally not used.
        # They are recalculated for dashboard accuracy.
        # ============================================================
        indicator_fields = [
            "births_vaginal",
            "births_csection",
            "oxytocin_immediate",
            "antepartum_hemorrhage",
            "pph_vaginal_501_999",
            "pph_cs_1000_plus",
            "pph_referral_in_outside_aim",
            "pph_referral_in_aim",
            "qbl_0_500",
            "qbl_501_999",
            "qbl_1000_1499",
            "qbl_1500_1999",
            "qbl_2000_2499",
            "qbl_2500_plus",
            "qbl_unknown",
            "qbl_total",
            "transfers_out_pph",
            "maternal_death_pph_transfer",
            "maternal_death_other_transfer",
            "maternal_death_total_transfer",
            "cause_uterine_atony",
            "cause_severe_lacerations",
            "cause_retained_products",
            "cause_dic",
            "cause_ruptured_uterus",
            "cause_abruption",
            "cause_placenta_previa",
            "cause_placenta_accreta",
            "cause_other",
            "cause_unknown",
            "causes_total",
            "pph_medication_uterotonic",
            "ai_uterine_compression",
            "ai_manual_placenta",
            "ai_aortic_compression",
            "ai_ubt",
            "ai_lac_repair",
            "ai_blynch_ual",
            "ai_nasg",
            "ai_ruptured_uterus_repair",
            "ai_pph_hysterectomy",
            "ai_hysterectomy_other",
            "ai_total",
        ]

        sum_annotations = {
            f"sum_{field}": Sum(field)
            for field in indicator_fields
        }

        totals = self._sum_fields(queryset, indicator_fields)
        totals["total_births"] = self._calculated_births(totals)
        totals["pph_total"] = self._calculated_pph_total(totals)

        # ============================================================
        # KPI cards
        # ============================================================
        kpis = {
            "records": queryset.count(),
            "facilities": queryset.values("aimfacilityname_id").distinct().count(),
            "provinces": queryset.values("aimfacilityname__districtfk__provincefk_id").distinct().count(),
            "total_births": totals["total_births"],
            "births_vaginal": totals["births_vaginal"],
            "births_csection": totals["births_csection"],
            "oxytocin_immediate": totals["oxytocin_immediate"],
            "pph_total": totals["pph_total"],
            "qbl_total": totals["qbl_total"],
            "transfers_out_pph": totals["transfers_out_pph"],
            "maternal_death_pph_transfer": totals["maternal_death_pph_transfer"],
            "ai_total": totals["ai_total"],
        }

        kpis["csection_rate"] = self._safe_rate(kpis["births_csection"], kpis["total_births"])
        kpis["oxytocin_rate"] = self._safe_rate(kpis["oxytocin_immediate"], kpis["total_births"])
        kpis["pph_rate"] = self._safe_rate(kpis["pph_total"], kpis["total_births"])
        kpis["transfer_rate_among_pph"] = self._safe_rate(kpis["transfers_out_pph"], kpis["pph_total"])
        kpis["ai_rate_among_pph"] = self._safe_rate(kpis["ai_total"], kpis["pph_total"])
        kpis["pph_death_rate_among_pph"] = self._safe_rate(
            kpis["maternal_death_pph_transfer"],
            kpis["pph_total"],
        )

        # ============================================================
        # Baseline / progress summary
        # ============================================================
        progress_rows = list(
            queryset.values("bl_progress")
            .annotate(
                records=Count("id"),
                facilities=Count("aimfacilityname", distinct=True),
                **sum_annotations,
            )
            .order_by("bl_progress")
        )

        for r in progress_rows:
            self._normalize_sum_fields(r, indicator_fields)
            r["total_births"] = self._calculated_births(r)
            r["pph_total"] = self._calculated_pph_total(r)

            r["csection_rate"] = self._safe_rate(r["births_csection"], r["total_births"])
            r["oxytocin_rate"] = self._safe_rate(r["oxytocin_immediate"], r["total_births"])
            r["pph_rate"] = self._safe_rate(r["pph_total"], r["total_births"])
            r["transfer_rate"] = self._safe_rate(r["transfers_out_pph"], r["pph_total"])
            r["ai_rate"] = self._safe_rate(r["ai_total"], r["pph_total"])

        # ============================================================
        # Facility-level PRE-I vs PRE-P comparison
        # ============================================================
        indicator_comparison_rows = self._build_indicator_comparison_rows(
            queryset,
            indicator_fields,
        )

        # ============================================================
        # Province summary
        # ============================================================
        province_rows = list(
            queryset.values(
                province=F("aimfacilityname__districtfk__provincefk__name"),
            )
            .annotate(
                records=Count("id"),
                facilities=Count("aimfacilityname", distinct=True),
                **sum_annotations,
            )
            .order_by("province")
        )

        for r in province_rows:
            self._normalize_sum_fields(r, indicator_fields)
            r["total_births"] = self._calculated_births(r)
            r["pph_total"] = self._calculated_pph_total(r)

            r["csection_rate"] = self._safe_rate(r["births_csection"], r["total_births"])
            r["oxytocin_rate"] = self._safe_rate(r["oxytocin_immediate"], r["total_births"])
            r["pph_rate"] = self._safe_rate(r["pph_total"], r["total_births"])
            r["transfer_rate"] = self._safe_rate(r["transfers_out_pph"], r["pph_total"])
            r["ai_rate"] = self._safe_rate(r["ai_total"], r["pph_total"])

        # ============================================================
        # Facility summary
        # ============================================================
        facility_rows = list(
            queryset.values(
                province=F("aimfacilityname__districtfk__provincefk__name"),
                district=F("aimfacilityname__districtfk__name"),
                facility=F("aimfacilityname__name"),
                hfcode=F("aimfacilityname__hfcode"),
            )
            .annotate(
                records=Count("id"),
                **sum_annotations,
            )
            .order_by("province", "district", "facility")
        )

        for r in facility_rows:
            self._normalize_sum_fields(r, indicator_fields)
            r["total_births"] = self._calculated_births(r)
            r["pph_total"] = self._calculated_pph_total(r)

            r["csection_rate"] = self._safe_rate(r["births_csection"], r["total_births"])
            r["oxytocin_rate"] = self._safe_rate(r["oxytocin_immediate"], r["total_births"])
            r["pph_rate"] = self._safe_rate(r["pph_total"], r["total_births"])
            r["transfer_rate"] = self._safe_rate(r["transfers_out_pph"], r["pph_total"])
            r["ai_rate"] = self._safe_rate(r["ai_total"], r["pph_total"])

            if r["pph_total"] > 0 and r["ai_total"] == 0:
                r["interpretation"] = "PPH cases reported, but no advanced intervention recorded."
            elif r["pph_total"] == 0:
                r["interpretation"] = "No PPH cases reported for selected filters."
            elif r["ai_total"] > 0:
                r["interpretation"] = "PPH and advanced intervention activity recorded."
            else:
                r["interpretation"] = "Review facility records for completeness."

        facility_rows_top_pph_rate = sorted(
            [r for r in facility_rows if r["total_births"] > 0],
            key=lambda x: x["pph_rate"],
            reverse=True,
        )[:25]

        # ============================================================
        # Monthly trend
        # ============================================================
        monthly_rows = list(
            queryset.values("gre_year", "gre_month", "period")
            .annotate(
                records=Count("id"),
                facilities=Count("aimfacilityname", distinct=True),
                **sum_annotations,
            )
            .order_by("gre_year", "gre_month", "period")
        )

        for r in monthly_rows:
            self._normalize_sum_fields(r, indicator_fields)
            r["total_births"] = self._calculated_births(r)
            r["pph_total"] = self._calculated_pph_total(r)

            r["month_label"] = (
                f"{r.get('gre_month') or ''} {r.get('gre_year') or ''}".strip()
                or r.get("period")
                or "Unknown"
            )

            r["csection_rate"] = self._safe_rate(r["births_csection"], r["total_births"])
            r["oxytocin_rate"] = self._safe_rate(r["oxytocin_immediate"], r["total_births"])
            r["pph_rate"] = self._safe_rate(r["pph_total"], r["total_births"])
            r["transfer_rate"] = self._safe_rate(r["transfers_out_pph"], r["pph_total"])
            r["ai_rate"] = self._safe_rate(r["ai_total"], r["pph_total"])

        # ============================================================
        # Distributions
        # ============================================================
        qbl_rows = [
            {"category": "0–500 ml", "value": totals["qbl_0_500"]},
            {"category": "501–999 ml", "value": totals["qbl_501_999"]},
            {"category": "1000–1499 ml", "value": totals["qbl_1000_1499"]},
            {"category": "1500–1999 ml", "value": totals["qbl_1500_1999"]},
            {"category": "2000–2499 ml", "value": totals["qbl_2000_2499"]},
            {"category": ">2500 ml", "value": totals["qbl_2500_plus"]},
            {"category": "Unknown", "value": totals["qbl_unknown"]},
        ]

        cause_rows = [
            {"category": "Uterine atony", "value": totals["cause_uterine_atony"]},
            {"category": "Severe lacerations", "value": totals["cause_severe_lacerations"]},
            {"category": "Retained products", "value": totals["cause_retained_products"]},
            {"category": "DIC", "value": totals["cause_dic"]},
            {"category": "Ruptured uterus", "value": totals["cause_ruptured_uterus"]},
            {"category": "Abruption placenta", "value": totals["cause_abruption"]},
            {"category": "Placenta previa", "value": totals["cause_placenta_previa"]},
            {"category": "Placenta accreta", "value": totals["cause_placenta_accreta"]},
            {"category": "Other", "value": totals["cause_other"]},
            {"category": "Unknown", "value": totals["cause_unknown"]},
        ]

        ai_rows = [
            {"category": "Uterine compression", "value": totals["ai_uterine_compression"]},
            {"category": "Manual placenta removal", "value": totals["ai_manual_placenta"]},
            {"category": "Aortic compression", "value": totals["ai_aortic_compression"]},
            {"category": "UBT", "value": totals["ai_ubt"]},
            {"category": "Laceration repair", "value": totals["ai_lac_repair"]},
            {"category": "B-Lynch / UAL", "value": totals["ai_blynch_ual"]},
            {"category": "NASG", "value": totals["ai_nasg"]},
            {"category": "Ruptured uterus repair", "value": totals["ai_ruptured_uterus_repair"]},
            {"category": "PPH hysterectomy", "value": totals["ai_pph_hysterectomy"]},
            {"category": "Hysterectomy other causes", "value": totals["ai_hysterectomy_other"]},
        ]

        # ============================================================
        # Data quality checks
        # ============================================================
        data_quality_rows = []

        for obj in queryset:
            facility = obj.aimfacilityname
            district = facility.districtfk if facility else None
            province = district.provincefk if district else None

            issues = []

            birth_components = (
                self._safe_int(obj.births_vaginal)
                + self._safe_int(obj.births_csection)
            )

            pph_components = (
                self._safe_int(obj.pph_vaginal_501_999)
                + self._safe_int(obj.pph_cs_1000_plus)
                + self._safe_int(obj.pph_referral_in_outside_aim)
                + self._safe_int(obj.pph_referral_in_aim)
            )

            qbl_components = (
                self._safe_int(obj.qbl_0_500)
                + self._safe_int(obj.qbl_501_999)
                + self._safe_int(obj.qbl_1000_1499)
                + self._safe_int(obj.qbl_1500_1999)
                + self._safe_int(obj.qbl_2000_2499)
                + self._safe_int(obj.qbl_2500_plus)
                + self._safe_int(obj.qbl_unknown)
            )

            cause_components = (
                self._safe_int(obj.cause_uterine_atony)
                + self._safe_int(obj.cause_severe_lacerations)
                + self._safe_int(obj.cause_retained_products)
                + self._safe_int(obj.cause_dic)
                + self._safe_int(obj.cause_ruptured_uterus)
                + self._safe_int(obj.cause_abruption)
                + self._safe_int(obj.cause_placenta_previa)
                + self._safe_int(obj.cause_placenta_accreta)
                + self._safe_int(obj.cause_other)
                + self._safe_int(obj.cause_unknown)
            )

            ai_components = (
                self._safe_int(obj.ai_uterine_compression)
                + self._safe_int(obj.ai_manual_placenta)
                + self._safe_int(obj.ai_aortic_compression)
                + self._safe_int(obj.ai_ubt)
                + self._safe_int(obj.ai_lac_repair)
                + self._safe_int(obj.ai_blynch_ual)
                + self._safe_int(obj.ai_nasg)
                + self._safe_int(obj.ai_ruptured_uterus_repair)
                + self._safe_int(obj.ai_pph_hysterectomy)
                + self._safe_int(obj.ai_hysterectomy_other)
            )

            death_components = (
                self._safe_int(obj.maternal_death_pph_transfer)
                + self._safe_int(obj.maternal_death_other_transfer)
            )

            if self._safe_int(obj.total_births) != birth_components:
                issues.append("Stored total births does not equal vaginal + C-section births")

            if self._safe_int(obj.pph_total) != pph_components:
                issues.append("Stored PPH total does not equal indicators 6–9")

            if self._safe_int(obj.qbl_total) != qbl_components:
                issues.append("QBL total does not equal QBL category sum")

            if self._safe_int(obj.causes_total) != cause_components:
                issues.append("Causes total does not equal cause category sum")

            if self._safe_int(obj.ai_total) != ai_components:
                issues.append("AI total does not equal advanced intervention category sum")

            if self._safe_int(obj.maternal_death_total_transfer) != death_components:
                issues.append("Maternal death total does not equal PPH + other maternal deaths")

            if pph_components > birth_components and birth_components > 0:
                issues.append("Calculated PPH total is higher than calculated total births")

            if issues:
                data_quality_rows.append({
                    "id": obj.id,
                    "province": getattr(province, "name", ""),
                    "district": getattr(district, "name", ""),
                    "facility": getattr(facility, "name", ""),
                    "hfcode": getattr(facility, "hfcode", ""),
                    "period": obj.period,
                    "bl_progress": obj.bl_progress,
                    "issues": "; ".join(issues),
                })

        # ============================================================
        # Chart data
        # ============================================================
        chart_data = {
            "progress_rates": [
                {
                    "progress": r.get("bl_progress") or "Unknown",
                    "Oxytocin rate": float(r.get("oxytocin_rate") or 0),
                    "PPH rate": float(r.get("pph_rate") or 0),
                    "AI among PPH": float(r.get("ai_rate") or 0),
                }
                for r in progress_rows
            ],
            "monthly_trend": [
                {
                    "month": r.get("month_label") or "Unknown",
                    "Births": int(r.get("total_births") or 0),
                    "PPH": int(r.get("pph_total") or 0),
                    "Oxytocin": int(r.get("oxytocin_immediate") or 0),
                    "AI": int(r.get("ai_total") or 0),
                }
                for r in monthly_rows
            ],
            "province_pph": [
                {
                    "province": r.get("province") or "Unknown",
                    "Births": int(r.get("total_births") or 0),
                    "PPH": int(r.get("pph_total") or 0),
                    "AI": int(r.get("ai_total") or 0),
                }
                for r in province_rows
            ],
            "facility_pph_rate": [
                {
                    "facility": (r.get("facility") or "Unknown")[:38],
                    "PPH rate": float(r.get("pph_rate") or 0),
                }
                for r in facility_rows_top_pph_rate[:10]
            ],
            "qbl_distribution": [
                {
                    "category": r["category"],
                    "value": int(r["value"] or 0),
                }
                for r in qbl_rows
            ],
            "cause_distribution": [
                {
                    "category": r["category"],
                    "value": int(r["value"] or 0),
                }
                for r in sorted(cause_rows, key=lambda x: x["value"], reverse=True)[:10]
            ],
            "ai_distribution": [
                {
                    "category": r["category"],
                    "value": int(r["value"] or 0),
                }
                for r in sorted(ai_rows, key=lambda x: x["value"], reverse=True)[:10]
            ],
            "indicator_comparison": [
                {
                    "indicator": r["indicator"][:45],
                    "PRE-I": int(r["pre_i"] or 0),
                    "PRE-P": int(r["pre_p"] or 0),
                }
                for r in indicator_comparison_rows[:20]
            ],
        }

        export_query = request.GET.copy()
        export_query["export"] = "1"

        return {
            "filters": filters,
            "export_query": export_query.urlencode(),
            "province_options": province_options,
            "facility_options": facility_options,
            "gre_year_options": gre_year_options,
            "gre_month_options": gre_month_options,
            "shamsiyear_options": shamsiyear_options,
            "shamsimonth_options": shamsimonth_options,
            "bl_progress_options": bl_progress_options,
            "kpis": kpis,
            "progress_rows": progress_rows,
            "indicator_comparison_rows": indicator_comparison_rows,
            "province_rows": province_rows,
            "facility_rows": facility_rows,
            "facility_rows_top_pph_rate": facility_rows_top_pph_rate,
            "monthly_rows": monthly_rows,
            "qbl_rows": qbl_rows,
            "cause_rows": sorted(cause_rows, key=lambda x: x["value"], reverse=True),
            "ai_rows": sorted(ai_rows, key=lambda x: x["value"], reverse=True),
            "data_quality_rows": data_quality_rows,
            "chart_data": chart_data,
            "methodology_note": (
                "AIM-PPH dashboard indicators are aggregated from facility-month records. "
                "For temporary dashboard correction, Total Births is calculated as Vaginal Births + C-section Births because the stored total_births field is currently inconsistent. "
                "Total PPH cases is calculated as the sum of indicators 6–9. "
                "Rates are calculated using summed numerators and corrected denominators for the selected filters."
            ),
        }

    # ============================================================
    # Excel export
    # ============================================================
    def _export_dashboard_excel(self, data):
        wb = Workbook()

        def style_sheet(ws):
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style="thin", color="D9E2F3"),
                right=Side(style="thin", color="D9E2F3"),
                top=Side(style="thin", color="D9E2F3"),
                bottom=Side(style="thin", color="D9E2F3"),
            )

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)

                for cell in col:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 65)

        ws = wb.active
        ws.title = "Progress_Summary"
        ws.append([
            "Baseline / Progress",
            "Records",
            "Facilities",
            "Corrected Total Births",
            "Vaginal Births",
            "C-section Births",
            "C-section Rate %",
            "Oxytocin",
            "Oxytocin Rate %",
            "10. Total PPH Cases 6-9",
            "PPH Rate %",
            "Transfers Out PPH",
            "PPH Deaths",
            "Advanced Interventions",
            "AI among PPH %",
        ])

        for r in data["progress_rows"]:
            ws.append([
                r.get("bl_progress"),
                r.get("records"),
                r.get("facilities"),
                r.get("total_births"),
                r.get("births_vaginal"),
                r.get("births_csection"),
                r.get("csection_rate"),
                r.get("oxytocin_immediate"),
                r.get("oxytocin_rate"),
                r.get("pph_total"),
                r.get("pph_rate"),
                r.get("transfers_out_pph"),
                r.get("maternal_death_pph_transfer"),
                r.get("ai_total"),
                r.get("ai_rate"),
            ])

        ws2 = wb.create_sheet("Facility_Indicator_Comparison")
        ws2.append([
            "Province",
            "Facility Name",
            "Indicator",
            "PRE_I",
            "PRE_P",
            "Absolute Change",
            "Percent Change",
            "Direction",
            "Improvement Status",
            "Story Use",
            "Notes for Field Verification",
        ])

        for r in data["indicator_comparison_rows"]:
            ws2.append([
                r.get("province"),
                r.get("facility"),
                r.get("indicator"),
                r.get("pre_i"),
                r.get("pre_p"),
                r.get("absolute_change"),
                r.get("percent_change_display"),
                r.get("direction"),
                r.get("improvement_status"),
                r.get("story_use"),
                r.get("notes"),
            ])

        ws3 = wb.create_sheet("Province_Summary")
        ws3.append([
            "Province",
            "Records",
            "Facilities",
            "Corrected Total Births",
            "Vaginal Births",
            "C-section Births",
            "C-section Rate %",
            "Oxytocin",
            "Oxytocin Rate %",
            "10. Total PPH Cases 6-9",
            "PPH Rate %",
            "Transfers Out PPH",
            "PPH Deaths",
            "Advanced Interventions",
            "AI among PPH %",
        ])

        for r in data["province_rows"]:
            ws3.append([
                r.get("province"),
                r.get("records"),
                r.get("facilities"),
                r.get("total_births"),
                r.get("births_vaginal"),
                r.get("births_csection"),
                r.get("csection_rate"),
                r.get("oxytocin_immediate"),
                r.get("oxytocin_rate"),
                r.get("pph_total"),
                r.get("pph_rate"),
                r.get("transfers_out_pph"),
                r.get("maternal_death_pph_transfer"),
                r.get("ai_total"),
                r.get("ai_rate"),
            ])

        ws4 = wb.create_sheet("Facility_Summary")
        ws4.append([
            "Province",
            "District",
            "Facility",
            "HF Code",
            "Records",
            "Corrected Total Births",
            "Vaginal Births",
            "C-section Births",
            "C-section Rate %",
            "Oxytocin",
            "Oxytocin Rate %",
            "10. Total PPH Cases 6-9",
            "PPH Rate %",
            "Transfers Out PPH",
            "PPH Deaths",
            "Advanced Interventions",
            "AI among PPH %",
            "Interpretation",
        ])

        for r in data["facility_rows"]:
            ws4.append([
                r.get("province"),
                r.get("district"),
                r.get("facility"),
                r.get("hfcode"),
                r.get("records"),
                r.get("total_births"),
                r.get("births_vaginal"),
                r.get("births_csection"),
                r.get("csection_rate"),
                r.get("oxytocin_immediate"),
                r.get("oxytocin_rate"),
                r.get("pph_total"),
                r.get("pph_rate"),
                r.get("transfers_out_pph"),
                r.get("maternal_death_pph_transfer"),
                r.get("ai_total"),
                r.get("ai_rate"),
                r.get("interpretation"),
            ])

        ws5 = wb.create_sheet("Monthly_Trend")
        ws5.append([
            "Gregorian Year",
            "Gregorian Month",
            "Period",
            "Records",
            "Facilities",
            "Corrected Total Births",
            "Vaginal Births",
            "C-section Births",
            "C-section Rate %",
            "Oxytocin",
            "Oxytocin Rate %",
            "10. Total PPH Cases 6-9",
            "PPH Rate %",
            "Transfers Out PPH",
            "PPH Deaths",
            "Advanced Interventions",
            "AI among PPH %",
        ])

        for r in data["monthly_rows"]:
            ws5.append([
                r.get("gre_year"),
                r.get("gre_month"),
                r.get("period"),
                r.get("records"),
                r.get("facilities"),
                r.get("total_births"),
                r.get("births_vaginal"),
                r.get("births_csection"),
                r.get("csection_rate"),
                r.get("oxytocin_immediate"),
                r.get("oxytocin_rate"),
                r.get("pph_total"),
                r.get("pph_rate"),
                r.get("transfers_out_pph"),
                r.get("maternal_death_pph_transfer"),
                r.get("ai_total"),
                r.get("ai_rate"),
            ])

        ws6 = wb.create_sheet("QBL_Distribution")
        ws6.append(["QBL Category", "Count"])
        for r in data["qbl_rows"]:
            ws6.append([r.get("category"), r.get("value")])

        ws7 = wb.create_sheet("Cause_Distribution")
        ws7.append(["Cause of PPH", "Count"])
        for r in data["cause_rows"]:
            ws7.append([r.get("category"), r.get("value")])

        ws8 = wb.create_sheet("Advanced_Interventions")
        ws8.append(["Advanced Intervention", "Count"])
        for r in data["ai_rows"]:
            ws8.append([r.get("category"), r.get("value")])

        ws9 = wb.create_sheet("Data_Quality")
        ws9.append([
            "Record ID",
            "Province",
            "District",
            "Facility",
            "HF Code",
            "Period",
            "Baseline / Progress",
            "Issue",
        ])

        for r in data["data_quality_rows"]:
            ws9.append([
                r.get("id"),
                r.get("province"),
                r.get("district"),
                r.get("facility"),
                r.get("hfcode"),
                r.get("period"),
                r.get("bl_progress"),
                r.get("issues"),
            ])

        ws10 = wb.create_sheet("Methodology_Notes")
        ws10.append(["Topic", "Explanation"])
        ws10.append(["Dashboard methodology", data["methodology_note"]])
        ws10.append([
            "Temporary correction",
            "Total Births is calculated as Vaginal Births + C-section Births. Total PPH cases is calculated as indicators 6–9.",
        ])
        ws10.append([
            "PRE-I vs PRE-P comparison",
            "The Facility_Indicator_Comparison sheet compares summed facility-level values for each indicator between PRE-I and PRE-P.",
        ])

        for sheet in wb.worksheets:
            style_sheet(sheet)

        timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
        filename = f"AIM_PPH_Dashboard_{timestamp}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
# ============================================================
# WHO Childbirth Checklist Monthly
# ============================================================
@admin.register(WhoChildbirthChecklistMonthly)
class WhoChildbirthChecklistMonthlyAdmin(ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    save_on_top = True
    list_per_page = 10

    list_display = (
        "id",
        "get_province",
        "facility_name",
        "reporting_period",
        "total_deliveries",
        "files_selected",
        "who_indicator_summary",
        # "created_at",
    )

    readonly_fields = (
        "clinical_entry_note",
        "sample_guidance_note",
        "ratio_guidance_note",
        "sec1_completeness_ratio",
        "partograph_use_ge4_rate",
        "sec2_completeness_ratio",
        "newborn_supplies_5_ratio",
        "sec3_completeness_ratio",
        "bf_s2s_first_hour_ratio",
        "sec4_completeness_ratio",
        "abx_need_checked_ratio",
        "all4_sections_completeness_ratio",
        "created_at",
        "updated_at",
    )

    list_filter = (
        # WhodistrictFilter,
        FacilityFilter,
        "shamsi_year_fk",
        "shamsi_month_fk",
        # "period_fk",
        "bl_progress_fk",
        "gre_year_fk",
        "gre_month_fk",
    )

    search_fields = (
        "facility_name__name",
        "facility_name__hfcode",
        "facility_name__districtfk__name",
        "facility_name__districtfk__provincefk__name",
    )

    list_select_related = (
        "facility_name",
        "facility_name__districtfk",
        "facility_name__districtfk__provincefk",
        "shamsi_month_fk",
        "shamsi_year_fk",
        "period_fk",
        "bl_progress_fk",
        "gre_month_fk",
        "gre_year_fk",
    )

    actions = ["export_who_childbirth_to_excel"]

    fieldsets = (
        ("Clinical Data Entry Guidance", {
            "fields": (
                "clinical_entry_note",
                "sample_guidance_note",
                "ratio_guidance_note",
            ),
        }),

        ("1. Facility and Reporting Period", {
            "description": (
                "Select the facility and reporting period carefully. "
                "These fields are used for dashboards and reporting."
            ),
            "fields": (
                "facility_name",
                ("shamsi_month_fk", "shamsi_year_fk"),
                ("period_fk", "bl_progress_fk"),
                ("gre_month_fk", "gre_year_fk"),
            ),
        }),

        ("2. Monthly Deliveries and Random File Sample", {
            "description": (
                "Enter total monthly deliveries and the number of randomly selected "
                "patient files. The selected files should be up to 20."
            ),
            "fields": (
                "total_deliveries",
                "files_selected",
            ),
        }),

        ("3. WHO Checklist Section 1 and Partograph", {
            "classes": ("collapse",),
            "fields": (
                ("sec1_complete", "sec1_completeness_ratio"),
                ("cervix_ge4_admission", "partograph_started_ge4"),
                "partograph_use_ge4_rate",
            ),
        }),

        ("4. WHO Checklist Section 2 and Newborn Supplies", {
            "classes": ("collapse",),
            "fields": (
                ("sec2_complete", "sec2_completeness_ratio"),
                ("newborn_supplies_5_available", "newborn_supplies_5_ratio"),
            ),
        }),

        ("5. WHO Checklist Section 3 and Early Newborn Care", {
            "classes": ("collapse",),
            "fields": (
                ("sec3_complete", "sec3_completeness_ratio"),
                ("bf_s2s_first_hour", "bf_s2s_first_hour_ratio"),
            ),
        }),

        ("6. WHO Checklist Section 4 and Discharge Checks", {
            "classes": ("collapse",),
            "fields": (
                ("sec4_complete", "sec4_completeness_ratio"),
                ("abx_need_checked_newborn", "abx_need_checked_ratio"),
            ),
        }),

        ("7. Full Checklist Completion", {
            "classes": ("collapse",),
            "fields": (
                ("all4_sections_complete", "all4_sections_completeness_ratio"),
            ),
        }),

        ("Audit Information", {
            "classes": ("collapse",),
            "fields": (
                "created_at",
                "updated_at",
            ),
        }),
    )

    # ------------------------------------------------------------
    # Admin guidance notes
    # ------------------------------------------------------------
    @admin.display(description="")
    def clinical_entry_note(self, obj):
        return format_html(
            """
            <div style="padding:12px 14px; border-left:5px solid #0072bc;
                        background:#f4f9fc; border-radius:8px; margin-bottom:8px;">
                <b style="font-size:14px;">WHO Childbirth Checklist Monthly Reporting</b><br>
                Enter monthly facility-level data based on selected patient files.
                Ratios are calculated automatically and shown as read-only values.
            </div>
            """
        )

    @admin.display(description="")
    def sample_guidance_note(self, obj):
        return format_html(
            """
            <div style="padding:12px 14px; border-left:5px solid #1b7f5c;
                        background:#f6fbf8; border-radius:8px; margin-bottom:8px;">
                <b>Sampling guidance:</b> Randomly select up to <b>20 patient files</b>
                from total monthly deliveries.
            </div>
            """
        )

    @admin.display(description="")
    def ratio_guidance_note(self, obj):
        return format_html(
            """
            <div style="padding:12px 14px; border-left:5px solid #8a6d3b;
                        background:#fffaf0; border-radius:8px;">
                <b>Ratio interpretation:</b>
                80% and above = good progress,
                50–79% = needs attention,
                below 50% = priority follow-up.
            </div>
            """
        )

    # ------------------------------------------------------------
    # Display reporting period
    # ------------------------------------------------------------
    @admin.display(description="Reporting Period")
    def reporting_period(self, obj):
        return format_html(
            "<b>{}</b> / {}<br><span style='color:#6b7280;'>{} - {}</span>",
            obj.shamsi_month_fk or "-",
            obj.shamsi_year_fk or "-",
            obj.period_fk or "-",
            obj.bl_progress_fk or "-",
        )

    # ------------------------------------------------------------
    # Percentage helper
    # ------------------------------------------------------------
    def _pct(self, num, den):
        try:
            if num is None or den in (None, 0):
                return None
            return (Decimal(num) / Decimal(den)) * Decimal("100.0")
        except (InvalidOperation, ZeroDivisionError, TypeError):
            return None

    # ------------------------------------------------------------
    # Mini indicator badge
    # ------------------------------------------------------------
    def _mini_badge(self, label, value):
        if value is None:
            return format_html(
                "<span style='display:inline-block;margin:2px;padding:4px 8px;"
                "border-radius:14px;background:#e5e7eb;color:#374151;font-size:11px;'>"
                "{}: N/A</span>",
                label,
            )

        try:
            value = Decimal(value)
        except Exception:
            return "-"

        if value >= Decimal("80"):
            bg, color = "#d1fae5", "#065f46"
        elif value >= Decimal("50"):
            bg, color = "#fef3c7", "#92400e"
        else:
            bg, color = "#fee2e2", "#991b1b"

        return format_html(
            "<span style='display:inline-block;margin:2px;padding:4px 8px;"
            "border-radius:14px;background:{};color:{};font-size:11px;font-weight:600;'>"
            "{}: {}%</span>",
            bg,
            color,
            label,
            round(value, 1),
        )

    # ------------------------------------------------------------
    # WHO indicator summary in list display
    # ------------------------------------------------------------
    @admin.display(description="WHO Indicators")
    def who_indicator_summary(self, obj):
        return format_html(
            "{} {} {} {} {} {} {} {} {}",
            self._mini_badge("Sec 1", obj.sec1_completeness_ratio),
            self._mini_badge("Partograph", obj.partograph_use_ge4_rate),
            self._mini_badge("Sec 2", obj.sec2_completeness_ratio),
            self._mini_badge("Supplies", obj.newborn_supplies_5_ratio),
            self._mini_badge("Sec 3", obj.sec3_completeness_ratio),
            self._mini_badge("BF/S2S", obj.bf_s2s_first_hour_ratio),
            self._mini_badge("Sec 4", obj.sec4_completeness_ratio),
            self._mini_badge("ABX", obj.abx_need_checked_ratio),
            self._mini_badge("All 4", obj.all4_sections_completeness_ratio),
        )

    # ------------------------------------------------------------
    # Optimized queryset
    # ------------------------------------------------------------
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related(
            "facility_name",
            "facility_name__districtfk",
            "facility_name__districtfk__provincefk",
            "shamsi_month_fk",
            "shamsi_year_fk",
            "period_fk",
            "bl_progress_fk",
            "gre_month_fk",
            "gre_year_fk",
        )

    # ------------------------------------------------------------
    # Province display
    # ------------------------------------------------------------
    @admin.display(description="Province")
    def get_province(self, obj):
        try:
            return obj.facility_name.districtfk.provincefk.name
        except Exception:
            return "-"

    # ------------------------------------------------------------
    # Province restriction
    # ------------------------------------------------------------
    def province_filter_kwargs(self, request):
        prov = user_province(request)
        if not prov:
            return {}
        return {
            "facility_name__districtfk__provincefk_id": prov.id
        }

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "facility_name" and not request.user.is_superuser:
            prov = user_province(request)
            kwargs["queryset"] = (
                Facility.objects.filter(districtfk__provincefk=prov)
                if prov else Facility.objects.none()
            )

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # ------------------------------------------------------------
    # Save FK dropdown values into old text fields
    # ------------------------------------------------------------
    def save_model(self, request, obj, form, change):
        # Copy FK dropdown values into old text fields.
        # This keeps old database fields, unique_together, ordering, and __str__ safe.

        if obj.shamsi_month_fk:
            obj.shamsi_month = str(obj.shamsi_month_fk)

        if obj.shamsi_year_fk:
            obj.shamsi_year = str(obj.shamsi_year_fk)

        if obj.period_fk:
            obj.period = str(obj.period_fk)

        if obj.bl_progress_fk:
            obj.bl_progress = str(obj.bl_progress_fk)

        if obj.gre_month_fk:
            obj.gre_month = str(obj.gre_month_fk)

        if obj.gre_year_fk:
            obj.gre_year = str(obj.gre_year_fk)

        super().save_model(request, obj, form, change)

    # ------------------------------------------------------------
    # Excel Export Action
    # ------------------------------------------------------------
    @admin.action(description="Export selected WHO Childbirth Checklist records to Excel")
    def export_who_childbirth_to_excel(self, request, queryset):
        """
        Export all WHO Childbirth Checklist Monthly fields into Excel.
        Includes:
        - Province
        - District
        - HF Code
        - Facility Name
        - All database fields
        - Foreign key display values
        - Foreign key IDs
        - Calculated ratio properties
        """

        queryset = queryset.select_related(
            "facility_name",
            "facility_name__districtfk",
            "facility_name__districtfk__provincefk",
            "shamsi_month_fk",
            "shamsi_year_fk",
            "period_fk",
            "bl_progress_fk",
            "gre_month_fk",
            "gre_year_fk",
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "WHO Childbirth Export"

        # -----------------------------
        # Helper functions
        # -----------------------------
        def clean_value(value):
            if value is None:
                return ""

            if isinstance(value, Decimal):
                return float(value)

            if isinstance(value, datetime):
                if timezone.is_aware(value):
                    value = timezone.localtime(value)
                return value.replace(tzinfo=None)

            if isinstance(value, date):
                return value

            if isinstance(value, bool):
                return "Yes" if value else "No"

            return value

        def safe_get(obj, attr_path):
            """
            Example:
            safe_get(obj, "facility_name__districtfk__provincefk__name")
            """
            current = obj

            for attr in attr_path.split("__"):
                current = getattr(current, attr, None)
                if current is None:
                    return ""

            return current

        # -----------------------------
        # Main export columns
        # -----------------------------
        columns = [
            (
                "Province",
                lambda obj: safe_get(
                    obj,
                    "facility_name__districtfk__provincefk__name"
                ),
            ),
            (
                "District",
                lambda obj: safe_get(
                    obj,
                    "facility_name__districtfk__name"
                ),
            ),
            (
                "HF Code",
                lambda obj: safe_get(
                    obj,
                    "facility_name__hfcode"
                ),
            ),
            (
                "Facility Name",
                lambda obj: safe_get(
                    obj,
                    "facility_name__name"
                ),
            ),
        ]

        # Add every real database field from WhoChildbirthChecklistMonthly model
        for field in self.model._meta.fields:
            if isinstance(field, ForeignKey):
                columns.append(
                    (
                        str(field.verbose_name),
                        lambda obj, f=field: str(getattr(obj, f.name, "") or "")
                    )
                )

                columns.append(
                    (
                        f"{field.name}_id",
                        lambda obj, f=field: getattr(obj, f.attname, "")
                    )
                )

            else:
                columns.append(
                    (
                        str(field.verbose_name),
                        lambda obj, f=field: clean_value(
                            getattr(obj, f"get_{f.name}_display")()
                            if f.choices
                            else getattr(obj, f.name, "")
                        )
                    )
                )

        # Add calculated ratio properties.
        # These are not database fields, so they must be added manually.
        ratio_columns = [
            ("Section 1 Completeness Ratio (%)", "sec1_completeness_ratio"),
            ("Partograph Use at Cervix ≥4cm Rate (%)", "partograph_use_ge4_rate"),
            ("Section 2 Completeness Ratio (%)", "sec2_completeness_ratio"),
            ("Newborn 5 Essential Supplies Ratio (%)", "newborn_supplies_5_ratio"),
            ("Section 3 Completeness Ratio (%)", "sec3_completeness_ratio"),
            ("Breastfeeding and Skin-to-Skin First Hour Ratio (%)", "bf_s2s_first_hour_ratio"),
            ("Section 4 Completeness Ratio (%)", "sec4_completeness_ratio"),
            ("Newborn Antibiotic Need Checked Ratio (%)", "abx_need_checked_ratio"),
            ("All 4 Sections Completeness Ratio (%)", "all4_sections_completeness_ratio"),
        ]

        for header, attr_name in ratio_columns:
            columns.append(
                (
                    header,
                    lambda obj, a=attr_name: clean_value(getattr(obj, a, ""))
                )
            )

        # -----------------------------
        # Write headers
        # -----------------------------
        headers = [col[0] for col in columns]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="D9EAF7")

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # -----------------------------
        # Write data rows
        # -----------------------------
        for obj in queryset:
            row = [clean_value(func(obj)) for _, func in columns]
            ws.append(row)

        # -----------------------------
        # Excel formatting
        # -----------------------------
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col_num, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_num)

            for cell in column_cells:
                try:
                    value_length = len(str(cell.value)) if cell.value is not None else 0
                    max_length = max(max_length, value_length)
                except Exception:
                    pass

            ws.column_dimensions[col_letter].width = min(max_length + 3, 45)

        # -----------------------------
        # Return Excel response
        # -----------------------------
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"who_childbirth_checklist_export_"
            f"{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

# ============================================================
# Safe Surgery (C-Section clinical)
# ============================================================
@admin.register(safesurgeryclinical)
class CSectionSafeSurgeryAdmin(ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    form = AimpeeAdminForm  # Replace with SafeSurgery form if you have one

    list_display = (
        "id",
        "get_province",
        "aimfacilityname",
        "shamsiyear",
        "shamsimonth",
        "total_cs",
        "total_deliv",
        "cs_rate",
        "who_ssc_rate",
        "safe_tracker_rate",
        "pph_cs_rate",
        "qbl_cs_rate",
        "postop_fever_rate",
        "hyst_rate",
        "mat_death_total",
    )

    readonly_fields = (
        "cs_rate",
        "who_ssc_rate",
        "safe_tracker_rate",
        "pph_cs_rate",
        "qbl_cs_rate",
        "postop_fever_rate",
        "bladder_injury_rate",
        "bowel_injury_rate",
        "hyst_rate",
        "vag_clean_rate",
        "foley_after_anes_rate",
        "abx_proph_rate",
        "skin_prep_rate",
    )

    list_filter = (AimpeeFacilityFilter,)
    search_fields = ("aimfacilityname__name", "aimfacilityname__hfcode")
    list_per_page = 10

    actions = ["export_safesurgeryclinical_to_excel"]

    # ------------------------------------------------------------
    # Improve queryset performance
    # ------------------------------------------------------------
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related(
            "aimfacilityname",
            "aimfacilityname__districtfk",
            "aimfacilityname__districtfk__provincefk",
        )

    # ------------------------------------------------------------
    # Percentage helper
    # ------------------------------------------------------------
    def _pct(self, num, den):
        try:
            if num is None or den in (None, 0):
                return Decimal("0.00")

            result = (Decimal(num) / Decimal(den)) * Decimal("100")
            return result.quantize(Decimal("0.01"))

        except (InvalidOperation, ZeroDivisionError, TypeError):
            return Decimal("0.00")

    # ------------------------------------------------------------
    # Auto-calculate rates before saving
    # ------------------------------------------------------------
    def save_model(self, request, obj, form, change):
        obj.cs_rate = self._pct(obj.total_cs, obj.total_deliv)
        obj.who_ssc_rate = self._pct(obj.who_ssc_completed, obj.total_cs)
        obj.safe_tracker_rate = self._pct(obj.safe_tracker_complete, obj.total_cs)
        obj.pph_cs_rate = self._pct(obj.pph_cs_num, obj.total_cs)
        obj.qbl_cs_rate = self._pct(obj.qbl_cs_num, obj.total_cs)
        obj.postop_fever_rate = self._pct(obj.postop_fever_num, obj.total_cs)
        obj.bladder_injury_rate = self._pct(obj.bladder_injury_num, obj.total_cs)
        obj.bowel_injury_rate = self._pct(obj.bowel_injury_num, obj.total_cs)
        obj.hyst_rate = self._pct(obj.hyst_num, obj.total_cs)
        obj.vag_clean_rate = self._pct(obj.vag_clean_num, obj.total_cs)
        obj.foley_after_anes_rate = self._pct(obj.foley_after_anes_num, obj.total_cs)
        obj.abx_proph_rate = self._pct(obj.abx_proph_num, obj.total_cs)
        obj.skin_prep_rate = self._pct(obj.skin_prep_num, obj.total_cs)

        super().save_model(request, obj, form, change)

    # ------------------------------------------------------------
    # Province display
    # ------------------------------------------------------------
    @admin.display(description="Province")
    def get_province(self, obj):
        try:
            return obj.aimfacilityname.districtfk.provincefk.name
        except AttributeError:
            return ""

    # ------------------------------------------------------------
    # Province restriction
    # ------------------------------------------------------------
    def province_filter_kwargs(self, request):
        return {
            "aimfacilityname__districtfk__provincefk": user_province(request)
        }

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "aimfacilityname" and not request.user.is_superuser:
            prov = user_province(request)
            kwargs["queryset"] = (
                Facility.objects.filter(districtfk__provincefk=prov)
                if prov else Facility.objects.none()
            )

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # ------------------------------------------------------------
    # Excel Export Action
    # ------------------------------------------------------------
    @admin.action(description="Export selected Safe Surgery records to Excel")
    def export_safesurgeryclinical_to_excel(self, request, queryset):
        """
        Export all fields from Safe Surgery Clinical model into Excel.
        Also includes Province, District, Facility Code, and Facility Name.
        """

        queryset = queryset.select_related(
            "aimfacilityname",
            "aimfacilityname__districtfk",
            "aimfacilityname__districtfk__provincefk",
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Safe Surgery Export"

        # -----------------------------
        # Helper functions
        # -----------------------------
        def clean_value(value):
            if value is None:
                return ""

            if isinstance(value, Decimal):
                return float(value)

            if isinstance(value, datetime):
                if timezone.is_aware(value):
                    value = timezone.localtime(value)
                return value.replace(tzinfo=None)

            if isinstance(value, date):
                return value

            if isinstance(value, bool):
                return "Yes" if value else "No"

            return value

        def safe_get(obj, attr_path):
            """
            Example:
            safe_get(obj, "aimfacilityname__districtfk__provincefk__name")
            """
            current = obj

            for attr in attr_path.split("__"):
                current = getattr(current, attr, None)
                if current is None:
                    return ""

            return current

        # -----------------------------
        # Build export columns
        # -----------------------------
        columns = [
            (
                "Province",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__districtfk__provincefk__name"
                ),
            ),
            (
                "District",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__districtfk__name"
                ),
            ),
            (
                "HF Code",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__hfcode"
                ),
            ),
            (
                "Facility Name",
                lambda obj: safe_get(
                    obj,
                    "aimfacilityname__name"
                ),
            ),
        ]

        # Add every real database field from safesurgeryclinical model
        for field in self.model._meta.fields:
            if isinstance(field, ForeignKey):
                columns.append(
                    (
                        str(field.verbose_name),
                        lambda obj, f=field: str(getattr(obj, f.name, "") or "")
                    )
                )

                columns.append(
                    (
                        f"{field.name}_id",
                        lambda obj, f=field: getattr(obj, f.attname, "")
                    )
                )

            else:
                columns.append(
                    (
                        str(field.verbose_name),
                        lambda obj, f=field: clean_value(
                            getattr(obj, f"get_{f.name}_display")()
                            if f.choices
                            else getattr(obj, f.name, "")
                        )
                    )
                )

        # -----------------------------
        # Write headers
        # -----------------------------
        headers = [col[0] for col in columns]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="D9EAF7")

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # -----------------------------
        # Write data rows
        # -----------------------------
        for obj in queryset:
            row = [clean_value(func(obj)) for _, func in columns]
            ws.append(row)

        # -----------------------------
        # Excel formatting
        # -----------------------------
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col_num, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_num)

            for cell in column_cells:
                try:
                    value_length = len(str(cell.value)) if cell.value is not None else 0
                    max_length = max(max_length, value_length)
                except Exception:
                    pass

            ws.column_dimensions[col_letter].width = min(max_length + 3, 45)

        # -----------------------------
        # Return Excel response
        # -----------------------------
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"safe_surgery_export_"
            f"{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

# ============================================================
# Facility Admin
# ============================================================

@admin.register(Facility)
class FacilityAdmin(ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    list_display = [
        "id", "get_province", "districtfk", "name", "hfcode",
        "facilitytypefk", "skilllab", "aim", "aimphase", "safesurgery",
        "ganc", "afiat", "nbcc", "sncu", "kmc",
    ]
    list_filter = ["districtfk__provincefk", "facilitytypefk"]
    search_fields = ["name", "districtfk__name", "districtfk__provincefk__name"]
    list_per_page = 15

    def province_filter_kwargs(self, request):
        return {"districtfk__provincefk": user_province(request)}

    @admin.display(description="Province")
    def get_province(self, obj):
        return obj.districtfk.provincefk.name


# ============================================================
# HQIP INLINE (details lines)
# ============================================================
class AssessmentLineInline(admin.TabularInline):
    model = HQIPAssessment
    extra = 0
    can_delete = False
    show_change_link = False

    fields = ("get_section", "get_standard", "get_criteria", "scorefk")
    readonly_fields = ("get_section", "get_standard", "get_criteria")

    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related(
            "header",
            "header__facilityfk__districtfk__provincefk",
            "criteriafk",
            "criteriafk__standardfk",
            "criteriafk__standardfk__sectionfk",
            "scorefk",
        ).order_by(
            "criteriafk__standardfk__sectionfk__id",
            "criteriafk__standardfk__id",
            "criteriafk__id",
        )

        if request.user.is_superuser:
            return qs

        prov = user_province(request)
        if prov is None:
            return qs.none()
        return qs.filter(header__facilityfk__districtfk__provincefk=prov)

    @admin.display(description="Section")
    def get_section(self, obj):
        return obj.criteriafk.standardfk.sectionfk.name if obj.criteriafk_id else "-"

    @admin.display(description="Standard")
    def get_standard(self, obj):
        return obj.criteriafk.standardfk.name if obj.criteriafk_id else "-"

    @admin.display(description="Criteria")
    def get_criteria(self, obj):
        return obj.criteriafk.name if obj.criteriafk_id else "-"

    def has_add_permission(self, request, obj=None):
        return False

# Score PK mapping:
SCORE_YES_ID = 1
SCORE_NO_ID = 2
SCORE_NA_ID = 3

# ============================================================
# HQIP HEADER ADMIN (creates details lines)
# ============================================================
@admin.register(HQIPAssessmentHeader)
class AssessmentHeaderAdmin(ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    actions = ["export_hqip_assessments_to_excel"]
    inlines = [AssessmentLineInline]

    list_display = (
        "facilityfk",
        "assessmenttype",
        "assessmentdate",
        "areafk",
        "assesorfk",
        "hqip_dashboard_button",
        "hqip_facility_button",
        "hqip_rca_button",
        "hqip_priority_button",
        "created_at",
        "id",
    )

    list_filter = (HQIPProvinceFilter, "areafk")
    search_fields = ("facilityfk__name", "facilityfk__hfcode")
    list_per_page = 10

    # ---------------------------
    # Province restriction (mix-in requirement)
    # ---------------------------
    def province_filter_kwargs(self, request):
        return {"facilityfk__districtfk__provincefk": user_province(request)}

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "facilityfk" and not request.user.is_superuser:
            prov = user_province(request)
            if prov:
                kwargs["queryset"] = Facility.objects.filter(districtfk__provincefk=prov).order_by("name")
            else:
                kwargs["queryset"] = Facility.objects.none()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # ---- helpers ----
    def _round2(self, x):
        if x is None:
            return None
        return float(Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

    def _pct(self, yes_count, applicable_count):
        if not applicable_count:
            return None
        return self._round2((Decimal(yes_count) / Decimal(applicable_count)) * Decimal(100))

    # ---- create missing detail lines AFTER header save ----
    def save_model(self, request, obj, form, change):
        if not obj.pk:
            if hasattr(obj, "created_by"):
                obj.created_by = request.user
            if hasattr(obj, "created_at"):
                obj.created_at = timezone.now()
        else:
            if hasattr(obj, "updated_by"):
                obj.updated_by = request.user
            if hasattr(obj, "updated_at"):
                obj.updated_at = timezone.now()

        super().save_model(request, obj, form, change)

        # Create all criteria lines for selected Area
        criteria_qs = Criteria.objects.filter(
            standardfk__sectionfk__areafk=obj.areafk
        ).order_by(
            "standardfk__sectionfk__id",
            "standardfk__id",
            "id",
        )

        existing_ids = set(obj.lines.values_list("criteriafk_id", flat=True))
        to_create = [
            HQIPAssessment(header=obj, criteriafk=c, scorefk=None)
            for c in criteria_qs
            if c.id not in existing_ids
        ]

        if to_create:
            with transaction.atomic():
                HQIPAssessment.objects.bulk_create(to_create)

    # ---- admin dashboards urls ----
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "hqip-standards-dashboard/",
                self.admin_site.admin_view(self.hqip_standards_dashboard),
                name="hqip_standards_dashboard",
            ),
            path(
                "hqip-facility-dashboard/",
                self.admin_site.admin_view(self.hqip_facility_dashboard),
                name="hqip_facility_dashboard",
            ),
            path(
                "hqip-rca-dashboard/",
                self.admin_site.admin_view(self.hqip_rca_dashboard),
                name="hqip_rca_dashboard",
            ),
            path(
                "hqip-priority-areas/",
                self.admin_site.admin_view(self.hqip_priority_areas),
                name="hqip_priority_areas",
            ),
        ]
        return custom_urls + urls
    
    @admin.action(description="Export selected HQIP assessments to Excel")
    def export_hqip_assessments_to_excel(self, request, queryset):
        """
        Export selected HQIP Assessment Headers and all related detail lines to Excel.

        Sheets:
        1. HQIP_Details
        2. Header_Summary
        3. Standard_Summary
        4. Section_Summary
        5. Area_Summary
        """

        queryset = queryset.select_related(
            "facilityfk",
            "facilityfk__districtfk",
            "facilityfk__districtfk__provincefk",
            "facilityfk__facilitytypefk",
            "assesorfk",
            "implementorfk",
            "assessmenttype",
            "areafk",
            "created_by",
            "updated_by",
        ).order_by(
            "facilityfk__districtfk__provincefk__name",
            "facilityfk__name",
            "assessmentdate",
            "areafk__name",
        )

        if not queryset.exists():
            messages.warning(request, "No HQIP assessments selected for export.")
            return None

        # ------------------------------------------------------------
        # Helper functions
        # ------------------------------------------------------------
        def safe_excel_value(value):
            """
            Converts values safely for Excel.
            Important: Excel/openpyxl does not support timezone-aware datetimes.
            """
            if value is None:
                return ""

            if isinstance(value, bool):
                return "Yes" if value else "No"

            if isinstance(value, py_datetime):
                if timezone.is_aware(value):
                    value = timezone.localtime(value)
                    value = timezone.make_naive(value)
                return value

            if isinstance(value, str):
                return ILLEGAL_CHARACTERS_RE.sub("", value)

            return value

        def safe_str(obj):
            if obj is None:
                return ""
            return safe_excel_value(str(obj))

        def get_name(obj):
            if obj is None:
                return ""
            return safe_excel_value(getattr(obj, "name", str(obj)))

        def get_short_or_name(obj):
            if obj is None:
                return ""
            return safe_excel_value(
                getattr(obj, "shortname", None)
                or getattr(obj, "name", str(obj))
            )

        def round2(value):
            if value is None:
                return None
            return float(
                Decimal(str(value)).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP
                )
            )

        def pct(yes_count, applicable_count):
            if not applicable_count:
                return None
            return round2(
                (Decimal(yes_count) / Decimal(applicable_count)) * Decimal(100)
            )

        def score_text(score_obj):
            if score_obj is None:
                return "Missing"
            return safe_str(score_obj)

        def score_category(score_id):
            if score_id == SCORE_YES_ID:
                return "YES"
            if score_id == SCORE_NO_ID:
                return "NO"
            if score_id == SCORE_NA_ID:
                return "N/A"
            return "Missing"

        def is_applicable(score_id):
            return "Yes" if score_id in [SCORE_YES_ID, SCORE_NO_ID] else "No"

        def style_worksheet(ws):
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style="thin", color="D9E2F3"),
                right=Side(style="thin", color="D9E2F3"),
                top=Side(style="thin", color="D9E2F3"),
                bottom=Side(style="thin", color="D9E2F3"),
            )

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                cell.border = border

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                    if isinstance(cell.value, py_datetime):
                        cell.number_format = "yyyy-mm-dd hh:mm:ss"
                    elif isinstance(cell.value, py_date):
                        cell.number_format = "yyyy-mm-dd"

            for column_cells in ws.columns:
                max_length = 0
                col_letter = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[col_letter].width = min(
                    max(max_length + 2, 12),
                    45
                )

        # ------------------------------------------------------------
        # Workbook setup
        # ------------------------------------------------------------
        wb = openpyxl.Workbook()

        ws_details = wb.active
        ws_details.title = "HQIP_Details"

        ws_header_summary = wb.create_sheet("Header_Summary")
        ws_standard_summary = wb.create_sheet("Standard_Summary")
        ws_section_summary = wb.create_sheet("Section_Summary")
        ws_area_summary = wb.create_sheet("Area_Summary")

        header_ids = list(queryset.values_list("id", flat=True))
        header_map = {h.id: h for h in queryset}

        # ------------------------------------------------------------
        # Sheet 1: HQIP Details
        # ------------------------------------------------------------
        ws_details.append([
            "Header ID",
            "Province",
            "District",
            "Facility",
            "HF Code",
            "Facility Type",
            "Implementor",
            "Assessor",
            "Assessment Type",
            "Assessment Start Date",
            "Assessment End Date",
            "Thematic Area",
            "Assessment Team",
            "RCA Conducted",
            "Section",
            "Standard",
            "Criteria",
            "Score",
            "Score Category",
            "Applicable",
            "Created By",
            "Created At",
            "Updated By",
            "Updated At",
        ])

        lines_qs = (
            HQIPAssessment.objects
            .filter(header_id__in=header_ids)
            .select_related(
                "header",
                "header__facilityfk",
                "header__facilityfk__districtfk",
                "header__facilityfk__districtfk__provincefk",
                "header__facilityfk__facilitytypefk",
                "header__assesorfk",
                "header__implementorfk",
                "header__assessmenttype",
                "header__areafk",
                "header__created_by",
                "header__updated_by",
                "criteriafk",
                "criteriafk__standardfk",
                "criteriafk__standardfk__sectionfk",
                "scorefk",
            )
            .order_by(
                "header__facilityfk__districtfk__provincefk__name",
                "header__facilityfk__name",
                "header__assessmentdate",
                "header__areafk__name",
                "criteriafk__standardfk__sectionfk__id",
                "criteriafk__standardfk__id",
                "criteriafk__id",
            )
        )

        for line in lines_qs:
            h = line.header
            facility = h.facilityfk
            district = facility.districtfk if facility else None
            province = district.provincefk if district else None
            facility_type = getattr(facility, "facilitytypefk", None) if facility else None

            criteria = line.criteriafk
            standard = criteria.standardfk if criteria else None
            section = standard.sectionfk if standard else None

            ws_details.append([
                h.id,
                get_name(province),
                get_name(district),
                get_name(facility),
                safe_excel_value(getattr(facility, "hfcode", "")),
                get_name(facility_type),
                get_name(h.implementorfk),
                get_name(h.assesorfk),
                get_name(h.assessmenttype),
                safe_excel_value(h.assessmentdate),
                safe_excel_value(h.assessmentend_date),
                get_name(h.areafk),
                safe_excel_value(h.assessmentteam),
                safe_excel_value(h.is_RCAduringtheassessment),
                get_short_or_name(section),
                get_short_or_name(standard),
                get_short_or_name(criteria),
                score_text(line.scorefk),
                score_category(line.scorefk_id),
                is_applicable(line.scorefk_id),
                safe_str(h.created_by),
                safe_excel_value(h.created_at),
                safe_str(h.updated_by),
                safe_excel_value(h.updated_at),
            ])

        # ------------------------------------------------------------
        # Sheet 2: Header Summary
        # ------------------------------------------------------------
        ws_header_summary.append([
            "Header ID",
            "Province",
            "District",
            "Facility",
            "HF Code",
            "Facility Type",
            "Implementor",
            "Assessor",
            "Assessment Type",
            "Assessment Start Date",
            "Assessment End Date",
            "Thematic Area",
            "RCA Conducted",
            "Total Criteria",
            "YES",
            "NO",
            "N/A",
            "Missing",
            "Applicable",
            "HQIP %",
        ])

        summary_qs = (
            HQIPAssessment.objects
            .filter(header_id__in=header_ids)
            .values("header_id")
            .annotate(
                total=Count("id"),
                yes=Count("id", filter=Q(scorefk_id=SCORE_YES_ID)),
                no=Count("id", filter=Q(scorefk_id=SCORE_NO_ID)),
                na=Count("id", filter=Q(scorefk_id=SCORE_NA_ID)),
                missing=Count("id", filter=Q(scorefk__isnull=True)),
                applicable=Count("id", filter=Q(scorefk_id__in=[SCORE_YES_ID, SCORE_NO_ID])),
            )
        )

        summary_map = {r["header_id"]: r for r in summary_qs}

        for h in queryset:
            facility = h.facilityfk
            district = facility.districtfk if facility else None
            province = district.provincefk if district else None
            facility_type = getattr(facility, "facilitytypefk", None) if facility else None

            r = summary_map.get(h.id, {})
            yes = r.get("yes", 0)
            applicable = r.get("applicable", 0)

            ws_header_summary.append([
                h.id,
                get_name(province),
                get_name(district),
                get_name(facility),
                safe_excel_value(getattr(facility, "hfcode", "")),
                get_name(facility_type),
                get_name(h.implementorfk),
                get_name(h.assesorfk),
                get_name(h.assessmenttype),
                safe_excel_value(h.assessmentdate),
                safe_excel_value(h.assessmentend_date),
                get_name(h.areafk),
                safe_excel_value(h.is_RCAduringtheassessment),
                r.get("total", 0),
                yes,
                r.get("no", 0),
                r.get("na", 0),
                r.get("missing", 0),
                applicable,
                pct(yes, applicable),
            ])

        # ------------------------------------------------------------
        # Sheet 3: Standard Summary
        # ------------------------------------------------------------
        ws_standard_summary.append([
            "Header ID",
            "Province",
            "District",
            "Facility",
            "Assessment Type",
            "Assessment Start Date",
            "Thematic Area",
            "Section",
            "Standard",
            "YES",
            "Applicable",
            "Standard %",
        ])

        std_rows = (
            HQIPAssessment.objects
            .filter(header_id__in=header_ids)
            .values(
                "header_id",
                "criteriafk__standardfk__id",
                "criteriafk__standardfk__name",
                "criteriafk__standardfk__sectionfk__id",
                "criteriafk__standardfk__sectionfk__name",
                "criteriafk__standardfk__sectionfk__areafk__name",
            )
            .annotate(
                yes=Count("id", filter=Q(scorefk_id=SCORE_YES_ID)),
                applicable=Count("id", filter=Q(scorefk_id__in=[SCORE_YES_ID, SCORE_NO_ID])),
            )
            .order_by(
                "header_id",
                "criteriafk__standardfk__sectionfk__name",
                "criteriafk__standardfk__name",
            )
        )

        section_percent_map = defaultdict(list)

        for r in std_rows:
            h = header_map.get(r["header_id"])
            if not h:
                continue

            facility = h.facilityfk
            district = facility.districtfk if facility else None
            province = district.provincefk if district else None

            yes = r["yes"]
            applicable = r["applicable"]
            standard_percent = pct(yes, applicable)

            area_name = safe_excel_value(r["criteriafk__standardfk__sectionfk__areafk__name"] or "")
            section_name = safe_excel_value(r["criteriafk__standardfk__sectionfk__name"] or "")
            standard_name = safe_excel_value(r["criteriafk__standardfk__name"] or "")

            ws_standard_summary.append([
                h.id,
                get_name(province),
                get_name(district),
                get_name(facility),
                get_name(h.assessmenttype),
                safe_excel_value(h.assessmentdate),
                area_name,
                section_name,
                standard_name,
                yes,
                applicable,
                standard_percent,
            ])

            if standard_percent is not None:
                section_key = (h.id, area_name, section_name)
                section_percent_map[section_key].append(standard_percent)

        # ------------------------------------------------------------
        # Sheet 4: Section Summary
        # Section % = Average of Standard %
        # ------------------------------------------------------------
        ws_section_summary.append([
            "Header ID",
            "Province",
            "District",
            "Facility",
            "Assessment Type",
            "Assessment Start Date",
            "Thematic Area",
            "Section",
            "Number of Standards Used",
            "Section %",
        ])

        area_percent_map = defaultdict(list)

        for section_key, percents in section_percent_map.items():
            header_id, area_name, section_name = section_key
            h = header_map.get(header_id)
            if not h:
                continue

            facility = h.facilityfk
            district = facility.districtfk if facility else None
            province = district.provincefk if district else None

            section_percent = round2(sum(percents) / len(percents)) if percents else None

            ws_section_summary.append([
                h.id,
                get_name(province),
                get_name(district),
                get_name(facility),
                get_name(h.assessmenttype),
                safe_excel_value(h.assessmentdate),
                area_name,
                section_name,
                len(percents),
                section_percent,
            ])

            if section_percent is not None:
                area_key = (h.id, area_name)
                area_percent_map[area_key].append(section_percent)

        # ------------------------------------------------------------
        # Sheet 5: Area Summary
        # Area % = Average of Section %
        # ------------------------------------------------------------
        ws_area_summary.append([
            "Header ID",
            "Province",
            "District",
            "Facility",
            "Assessment Type",
            "Assessment Start Date",
            "Thematic Area",
            "Number of Sections Used",
            "Area %",
        ])

        for area_key, percents in area_percent_map.items():
            header_id, area_name = area_key
            h = header_map.get(header_id)
            if not h:
                continue

            facility = h.facilityfk
            district = facility.districtfk if facility else None
            province = district.provincefk if district else None

            area_percent = round2(sum(percents) / len(percents)) if percents else None

            ws_area_summary.append([
                h.id,
                get_name(province),
                get_name(district),
                get_name(facility),
                get_name(h.assessmenttype),
                safe_excel_value(h.assessmentdate),
                area_name,
                len(percents),
                area_percent,
            ])

        # ------------------------------------------------------------
        # Formatting
        # ------------------------------------------------------------
        for ws in wb.worksheets:
            style_worksheet(ws)

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    header_name = ws.cell(row=1, column=cell.column).value
                    if header_name and "%" in str(header_name):
                        cell.number_format = "0.00"

        # ------------------------------------------------------------
        # Return Excel response
        # ------------------------------------------------------------
        timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
        filename = f"HQIP_Assessment_Export_{timestamp}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    # ---- buttons ----
    @admin.display(description="Score")
    def hqip_dashboard_button(self, obj):
        base_url = reverse("admin:hqip_standards_dashboard")
        qs = urlencode({"header_id": obj.id})
        return format_html('<a class="button" href="{}?{}">Score</a>', base_url, qs)

    @admin.display(description="Detail")
    def hqip_facility_button(self, obj):
        base_url = reverse("admin:hqip_facility_dashboard")
        qs = urlencode({"facility": obj.facilityfk_id, "header_id": obj.id})
        return format_html('<a class="button" href="{}?{}">View</a>', base_url, qs)

    @admin.display(description="RCA")
    def hqip_rca_button(self, obj):
        base_url = reverse("admin:hqip_rca_dashboard")
        qs = urlencode({"header_id": obj.id})
        return format_html('<a class="button" href="{}?{}">RCA</a>', base_url, qs)

    @admin.display(description="Priority")
    def hqip_priority_button(self, obj):
        base_url = reverse("admin:hqip_priority_areas")
        qs = urlencode({"facility_id": obj.facilityfk_id})
        return format_html('<a class="button" href="{}?{}">Priority</a>', base_url, qs)

    def _compute_hqip_rollups(self, headers_qs):
        """
        Shared HQIP rollup calculator:
        Criteria -> Standard %  (YES / (YES+NO)) ignoring NA/NULL
        Section % = average(Standard %)
        Area %    = average(Section %)

        Returns:
        standard_results: list[dict]
        section_results:  list[dict]
        area_results:     list[dict]
        """

        std_rows = (
            HQIPAssessment.objects
            .filter(header__in=headers_qs)
            .values(
                "criteriafk__standardfk__id",
                "criteriafk__standardfk__name",
                "criteriafk__standardfk__sectionfk__id",
                "criteriafk__standardfk__sectionfk__name",
                "criteriafk__standardfk__sectionfk__areafk__id",
                "criteriafk__standardfk__sectionfk__areafk__name",
            )
            .annotate(
                yes=Count("id", filter=Q(scorefk_id=SCORE_YES_ID)),
                applicable=Count("id", filter=Q(scorefk_id__in=[SCORE_YES_ID, SCORE_NO_ID])),
            )
            .order_by(
                "criteriafk__standardfk__sectionfk__areafk__name",
                "criteriafk__standardfk__sectionfk__name",
                "criteriafk__standardfk__name",
            )
        )

        from collections import defaultdict

        standard_results = []
        section_results = []
        area_results = []

        section_to_standard_percents = defaultdict(list)
        area_to_section_percents = defaultdict(list)

        # name maps
        sec_name_map = {}
        area_name_map = {}

        # ---------- Standard level ----------
        for r in std_rows:
            area_id = r["criteriafk__standardfk__sectionfk__areafk__id"]
            sec_id = r["criteriafk__standardfk__sectionfk__id"]
            sec_key = (area_id, sec_id)

            area_name = r["criteriafk__standardfk__sectionfk__areafk__name"] or "-"
            sec_name = r["criteriafk__standardfk__sectionfk__name"] or "-"
            std_name = r["criteriafk__standardfk__name"] or "-"

            sec_name_map[sec_key] = (area_name, sec_name)
            area_name_map[area_id] = area_name

            den = r["applicable"]
            num = r["yes"]
            std_percent = self._pct(num, den)  # <- YOUR SAME helper (YES/(YES+NO))*100

            standard_results.append({
                "area": area_name,
                "section": sec_name,
                "standard": std_name,
                "yes": num,
                "applicable": den,
                "percent": std_percent,
            })

            if std_percent is not None:
                section_to_standard_percents[sec_key].append(std_percent)

        # ---------- Section level = average(Standard %) ----------
        for sec_key, percents in section_to_standard_percents.items():
            area_id, _sec_id = sec_key
            area_name, sec_name = sec_name_map.get(sec_key, ("-", "-"))

            sec_percent = self._round2(sum(percents) / len(percents)) if percents else None

            section_results.append({
                "area": area_name,
                "section": sec_name,
                "num_standards_used": len(percents),
                "percent": sec_percent,
            })

            if sec_percent is not None:
                area_to_section_percents[area_id].append(sec_percent)

        section_results.sort(key=lambda x: (x["area"], x["section"]))

        # ---------- Area level = average(Section %) ----------
        for area_id, percents in area_to_section_percents.items():
            area_percent = self._round2(sum(percents) / len(percents)) if percents else None
            area_results.append({
                "area": area_name_map.get(area_id, "-"),
                "num_sections_used": len(percents),
                "percent": area_percent,
            })

        area_results.sort(key=lambda x: x["area"])

        return standard_results, section_results, area_results

    # ==========================================================
    # A) STANDARDS DASHBOARD
    # ==========================================================
    def hqip_standards_dashboard(self, request):
        header_id = request.GET.get("header_id")
        headers_qs = self.get_queryset(request)

        standard_results, section_results, area_results = self._compute_hqip_rollups(headers_qs)

        header_obj = None
        error_message = None

        # Row-level: lock to exact header
        if header_id:
            headers_qs = headers_qs.filter(id=header_id).select_related(
                "facilityfk__districtfk__provincefk", "areafk", "assessmenttype"
            )
            header_obj = headers_qs.first()
            selected_province = None  # irrelevant when header_id is set

            if not header_obj:
                error_message = "No header selected (missing header_id) or you don’t have access to that header."
                headers_qs = headers_qs.none()
        else:
            selected_province = request.GET.get("province")
            if request.user.is_superuser and selected_province:
                headers_qs = headers_qs.filter(
                    facilityfk__districtfk__provincefk_id=selected_province
                )

        std_rows = (
            HQIPAssessment.objects
            .filter(header__in=headers_qs)
            .values(
                "criteriafk__standardfk__id",
                "criteriafk__standardfk__name",
                "criteriafk__standardfk__sectionfk__id",
                "criteriafk__standardfk__sectionfk__name",
                "criteriafk__standardfk__sectionfk__areafk__id",
                "criteriafk__standardfk__sectionfk__areafk__name",
            )
            .annotate(
                yes=Count("id", filter=Q(scorefk_id=SCORE_YES_ID)),
                applicable=Count("id", filter=Q(scorefk_id__in=[SCORE_YES_ID, SCORE_NO_ID])),
            )
            .order_by(
                "criteriafk__standardfk__sectionfk__areafk__name",
                "criteriafk__standardfk__sectionfk__name",
                "criteriafk__standardfk__name",
            )
        )

        standard_results = []
        section_to_standard_percents = defaultdict(list)
        area_to_section_percents = defaultdict(list)
        sec_name_map = {}
        area_name_map = {}

        for r in std_rows:
            area_id = r["criteriafk__standardfk__sectionfk__areafk__id"]
            sec_id = r["criteriafk__standardfk__sectionfk__id"]
            sec_key = (area_id, sec_id)

            area_name = r["criteriafk__standardfk__sectionfk__areafk__name"] or "-"
            sec_name = r["criteriafk__standardfk__sectionfk__name"] or "-"

            sec_name_map[sec_key] = (area_name, sec_name)
            area_name_map[area_id] = area_name

            den = r["applicable"]
            num = r["yes"]
            std_percent = self._pct(num, den)

            standard_results.append({
                "area": area_name,
                "section": sec_name,
                "standard": r["criteriafk__standardfk__name"] or "-",
                "yes": num,
                "applicable": den,
                "percent": std_percent,
            })

            if std_percent is not None:
                section_to_standard_percents[sec_key].append(std_percent)

        section_results = []
        for sec_key, percents in section_to_standard_percents.items():
            area_id, _sec_id = sec_key
            area_name, sec_name = sec_name_map.get(sec_key, ("-", "-"))
            sec_percent = self._round2(sum(percents) / len(percents)) if percents else None

            section_results.append({
                "area": area_name,
                "section": sec_name,
                "num_standards_used": len(percents),
                "percent": sec_percent,
            })

            if sec_percent is not None:
                area_to_section_percents[area_id].append(sec_percent)

        section_results.sort(key=lambda x: (x["area"], x["section"]))

        area_results = []
        for area_id, percents in area_to_section_percents.items():
            area_percent = self._round2(sum(percents) / len(percents)) if percents else None
            area_results.append({
                "area": area_name_map.get(area_id, "-"),
                "num_sections_used": len(percents),
                "percent": area_percent,
            })
        area_results.sort(key=lambda x: x["area"])

        # show province dropdown only in global mode (no header_id)
        provinces = Province.objects.all().order_by("name") if (request.user.is_superuser and not header_id) else None

        context = dict(
            self.admin_site.each_context(request),
            title="HQIP Dashboard",
            header_obj=header_obj,              # ✅ add
            error_message=error_message,        # ✅ add
            standard_results=standard_results,
            section_results=section_results,
            area_results=area_results,
            provinces=provinces,
            selected_province=selected_province,
            header_id=header_id,
        )
        return TemplateResponse(request, "admin/hiva/hqip_dashboard_full.html", context)

    # ==========================================================
    # B) FACILITY DRILL-DOWN DASHBOARD
    # ==========================================================
    def hqip_facility_dashboard(self, request):
        selected_province = request.GET.get("province")
        selected_facility = request.GET.get("facility")
        selected_area = request.GET.get("area")
        selected_type = request.GET.get("assessmenttype")
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        header_id = request.GET.get("header_id")

        facilities_qs = Facility.objects.all().select_related("districtfk__provincefk")

        if not request.user.is_superuser:
            prov = user_province(request)
            facilities_qs = facilities_qs.filter(districtfk__provincefk=prov) if prov else Facility.objects.none()
        if request.user.is_superuser and selected_province:
            facilities_qs = facilities_qs.filter(districtfk__provincefk_id=selected_province)

        facilities = facilities_qs.order_by("name")
        areas = Area.objects.all().order_by("name")
        types = Assessmenttype.objects.all().order_by("name")
        provinces = Province.objects.all().order_by("name") if request.user.is_superuser else None

        facility_obj = None
        header_list = []
        standard_results, section_results, area_results = [], [], []
        std_chart_json, sec_chart_json = [], []
        error_message = None

        # If header_id is present, force facility from that header (prevents mismatch)
        header_obj = None
        if header_id:
            header_obj = self.get_queryset(request).filter(id=header_id).select_related(
                "facilityfk__districtfk__provincefk", "areafk", "assessmenttype"
            ).first()
            if not header_obj:
                error_message = "No header selected (missing header_id) or you don’t have access to that header."
            else:
                facility_obj = header_obj.facilityfk
                selected_facility = str(facility_obj.id)

        # Normal mode: use selected facility
        if (not facility_obj) and selected_facility:
            facility_obj = Facility.objects.filter(pk=selected_facility).select_related("districtfk__provincefk").first()
            if facility_obj and (not request.user.is_superuser):
                prov = user_province(request)
                if not prov or facility_obj.districtfk.provincefk_id != prov.id:
                    facility_obj = None
                    error_message = "You don’t have access to this facility."

        if facility_obj and not error_message:
            headers_qs = self.get_queryset(request).filter(facilityfk=facility_obj)

            # Row lock
            if header_id:
                headers_qs = headers_qs.filter(id=header_id)
            else:
                # Optional filters only in global mode
                if selected_area:
                    headers_qs = headers_qs.filter(areafk_id=selected_area)
                if selected_type:
                    headers_qs = headers_qs.filter(assessmenttype_id=selected_type)
                if date_from:
                    headers_qs = headers_qs.filter(assessmentdate__gte=date_from)
                if date_to:
                    headers_qs = headers_qs.filter(assessmentdate__lte=date_to)

            header_list = list(headers_qs.values("id", "assessmentdate"))

            std_rows = (
                HQIPAssessment.objects
                .filter(header__in=headers_qs)
                .values(
                    "criteriafk__standardfk__id",
                    "criteriafk__standardfk__name",
                    "criteriafk__standardfk__sectionfk__id",
                    "criteriafk__standardfk__sectionfk__name",
                    "criteriafk__standardfk__sectionfk__areafk__id",
                    "criteriafk__standardfk__sectionfk__areafk__name",
                )
                .annotate(
                    yes=Count("id", filter=Q(scorefk_id=SCORE_YES_ID)),
                    applicable=Count("id", filter=Q(scorefk_id__in=[SCORE_YES_ID, SCORE_NO_ID])),
                )
                .order_by(
                    "criteriafk__standardfk__sectionfk__areafk__name",
                    "criteriafk__standardfk__sectionfk__name",
                    "criteriafk__standardfk__name",
                )
            )

            section_to_standard_percents = defaultdict(list)
            area_to_section_percents = defaultdict(list)
            sec_name_map = {}
            area_name_map = {}

            for r in std_rows:
                area_name = r["criteriafk__standardfk__sectionfk__areafk__name"] or "-"
                sec_name = r["criteriafk__standardfk__sectionfk__name"] or "-"
                std_name = r["criteriafk__standardfk__name"] or "-"

                area_id = r["criteriafk__standardfk__sectionfk__areafk__id"]
                sec_id = r["criteriafk__standardfk__sectionfk__id"]
                sec_key = (area_id, sec_id)

                sec_name_map[sec_key] = (area_name, sec_name)
                area_name_map[area_id] = area_name

                den = r["applicable"]
                num = r["yes"]
                std_percent = self._pct(num, den)

                standard_results.append({
                    "area": area_name,
                    "section": sec_name,
                    "standard": std_name,
                    "yes": num,
                    "applicable": den,
                    "percent": std_percent,
                })

                if std_percent is not None:
                    section_to_standard_percents[sec_key].append(std_percent)

            for sec_key, percents in section_to_standard_percents.items():
                area_id, _sec_id = sec_key
                area_name, sec_name = sec_name_map.get(sec_key, ("-", "-"))
                sec_percent = self._round2(sum(percents) / len(percents)) if percents else None

                section_results.append({
                    "area": area_name,
                    "section": sec_name,
                    "num_standards_used": len(percents),
                    "percent": sec_percent,
                })

                if sec_percent is not None:
                    area_to_section_percents[area_id].append(sec_percent)

            section_results.sort(key=lambda x: (x["area"], x["section"]))

            for area_id, percents in area_to_section_percents.items():
                area_percent = self._round2(sum(percents) / len(percents)) if percents else None
                area_results.append({
                    "area": area_name_map.get(area_id, "-"),
                    "num_sections_used": len(percents),
                    "percent": area_percent,
                })
            area_results.sort(key=lambda x: x["area"])

            std_chart_json = [{"label": r["standard"], "value": r["percent"]} for r in standard_results if r.get("percent") is not None]
            sec_chart_json = [{"label": r["section"], "value": r["percent"]} for r in section_results if r.get("percent") is not None]

        context = dict(
            self.admin_site.each_context(request),
            title="HQIP Facility Drill-Down",
            error_message=error_message,     # ✅ add
            header_obj=header_obj,           # ✅ add
            provinces=provinces,
            selected_province=selected_province,
            facilities=facilities,
            selected_facility=selected_facility,
            facility_obj=facility_obj,
            areas=areas,
            selected_area=selected_area,
            types=types,
            selected_type=selected_type,
            date_from=date_from,
            date_to=date_to,
            header_id=header_id,
            header_list=header_list,
            standard_results=standard_results,
            section_results=section_results,
            area_results=area_results,
            std_chart_json=std_chart_json,
            sec_chart_json=sec_chart_json,
        )
        return TemplateResponse(request, "admin/hiva/hqip_facility_dashboard.html", context)

    # ==========================================================
    # C) RCA DASHBOARD (row-level only)
    # ==========================================================
    def hqip_rca_dashboard(self, request):
        header_id = request.GET.get("header_id")
        headers_qs = self.get_queryset(request)

        error_message = None

        if not header_id:
            error_message = "No header selected (missing header_id) or you don’t have access to that header."
            context = dict(
                self.admin_site.each_context(request),
                title="HQIP RCA – Failed Criteria (NO only)",
                header_obj=None,
                rca_rows=[],
                error_message=error_message,  # ✅ add
            )
            return TemplateResponse(request, "admin/hiva/hqip_rca_dashboard.html", context)

        header_obj = (
            headers_qs
            .filter(id=header_id)
            .select_related("facilityfk__districtfk__provincefk", "areafk", "assessmenttype")
            .first()
        )

        if not header_obj:
            error_message = "No header selected (missing header_id) or you don’t have access to that header."
            context = dict(
                self.admin_site.each_context(request),
                title="HQIP RCA – Failed Criteria (NO only)",
                header_obj=None,
                rca_rows=[],
                error_message=error_message,  # ✅ add
            )
            return TemplateResponse(request, "admin/hiva/hqip_rca_dashboard.html", context)

        rca_rows = (
            HQIPAssessment.objects
            .filter(header_id=header_obj.id, scorefk_id=SCORE_NO_ID)
            .select_related(
                "criteriafk",
                "criteriafk__standardfk",
                "criteriafk__standardfk__sectionfk",
                "criteriafk__standardfk__sectionfk__areafk",
                "header",
                "header__facilityfk",
            )
            .order_by(
                "criteriafk__standardfk__sectionfk__areafk__name",
                "criteriafk__standardfk__sectionfk__name",
                "criteriafk__standardfk__name",
                "criteriafk__id",
            )
        )

        context = dict(
            self.admin_site.each_context(request),
            title="HQIP RCA – Failed Criteria (NO only)",
            header_obj=header_obj,
            rca_rows=rca_rows,
            error_message=None,
        )
        return TemplateResponse(request, "admin/hiva/hqip_rca_dashboard.html", context)

    def hqip_priority_areas(self, request):
        """
        Facility-level HQIP Priority Thematic Areas by assessment round.

        This page:
        - Opens by facility_id
        - Shows Baseline, 2nd Round, 3rd Round, etc. separately
        - Calculates the lowest 3 priority thematic areas within each round
        - Does NOT mix baseline and follow-up rounds together
        - Uses the same HQIP calculation logic as Score/View through _compute_hqip_rollups()
        """

        from collections import defaultdict
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        facility_id = request.GET.get("facility_id")
        export = request.GET.get("export") == "1"

        headers_base = self.get_queryset(request)

        facility_obj = None
        rounds = []
        error_message = None

        # -----------------------------------------------------
        # 1. Validate facility
        # -----------------------------------------------------
        if not facility_id:
            error_message = "No facility selected."
            context = dict(
                self.admin_site.each_context(request),
                title="HQIP Priority Thematic Areas",
                error_message=error_message,
                facility_obj=None,
                rounds=[],
            )
            return TemplateResponse(request, "admin/hiva/hqip_priority_areas.html", context)

        facility_obj = (
            Facility.objects
            .select_related("districtfk__provincefk")
            .filter(pk=facility_id)
            .first()
        )

        if not facility_obj:
            error_message = "Invalid facility."
            context = dict(
                self.admin_site.each_context(request),
                title="HQIP Priority Thematic Areas",
                error_message=error_message,
                facility_obj=None,
                rounds=[],
            )
            return TemplateResponse(request, "admin/hiva/hqip_priority_areas.html", context)

        # -----------------------------------------------------
        # 2. Province access restriction
        # -----------------------------------------------------
        if not request.user.is_superuser:
            prov = user_province(request)
            if not prov or facility_obj.districtfk.provincefk_id != prov.id:
                error_message = "You don’t have access to this facility."
                context = dict(
                    self.admin_site.each_context(request),
                    title="HQIP Priority Thematic Areas",
                    error_message=error_message,
                    facility_obj=None,
                    rounds=[],
                )
                return TemplateResponse(request, "admin/hiva/hqip_priority_areas.html", context)

        # -----------------------------------------------------
        # 3. Load all HQIP headers for this facility
        # -----------------------------------------------------
        headers_all = list(
            headers_base
            .filter(facilityfk_id=facility_obj.id)
            .select_related(
                "facilityfk",
                "facilityfk__districtfk",
                "facilityfk__districtfk__provincefk",
                "assessmenttype",
                "areafk",
            )
            .order_by(
                "assessmenttype__name",
                "assessmentdate",
                "areafk__name",
                "id",
            )
        )

        if not headers_all:
            error_message = "No HQIP assessments found for this facility."
            context = dict(
                self.admin_site.each_context(request),
                title="HQIP Priority Thematic Areas",
                error_message=error_message,
                facility_obj=facility_obj,
                rounds=[],
            )
            return TemplateResponse(request, "admin/hiva/hqip_priority_areas.html", context)

        # -----------------------------------------------------
        # 4. Group headers by assessment round
        # IMPORTANT:
        # This assumes Assessment Type means Baseline, 2nd Round, 3rd Round, etc.
        # -----------------------------------------------------
        grouped_headers = defaultdict(list)

        for h in headers_all:
            round_key = h.assessmenttype_id or 0
            grouped_headers[round_key].append(h)

        # -----------------------------------------------------
        # 5. Calculate priority areas separately for each round
        # -----------------------------------------------------
        for round_key, header_list in grouped_headers.items():
            header_ids = [h.id for h in header_list]
            sample_header = header_list[0]

            round_label = (
                sample_header.assessmenttype.name
                if sample_header.assessmenttype
                else "Unknown Assessment Round"
            )

            assessment_dates = [
                h.assessmentdate for h in header_list
                if h.assessmentdate
            ]
            assessment_end_dates = [
                h.assessmentend_date for h in header_list
                if h.assessmentend_date
            ]

            date_from = min(assessment_dates) if assessment_dates else None
            date_to = max(assessment_end_dates) if assessment_end_dates else None

            thematic_area_count = len(
                set(h.areafk_id for h in header_list if h.areafk_id)
            )

            round_headers_qs = headers_base.filter(id__in=header_ids)

            # Same calculation logic as Score/View dashboard
            _std, _sec, area_results = self._compute_hqip_rollups(round_headers_qs)

            # Raw YES / Applicable counts for display only
            raw_counts = (
                HQIPAssessment.objects
                .filter(header_id__in=header_ids)
                .values("criteriafk__standardfk__sectionfk__areafk__name")
                .annotate(
                    yes=Count("id", filter=Q(scorefk_id=SCORE_YES_ID)),
                    applicable=Count("id", filter=Q(scorefk_id__in=[SCORE_YES_ID, SCORE_NO_ID])),
                )
            )

            raw_map = {
                r["criteriafk__standardfk__sectionfk__areafk__name"] or "-": {
                    "yes": r["yes"],
                    "applicable": r["applicable"],
                }
                for r in raw_counts
            }

            rows = []
            for r in area_results:
                area_name = r["area"]
                rows.append({
                    "area": area_name,
                    "percent": r["percent"],
                    "num_sections_used": r["num_sections_used"],
                    "yes": raw_map.get(area_name, {}).get("yes", 0),
                    "applicable": raw_map.get(area_name, {}).get("applicable", 0),
                    "is_priority": False,
                })

            # Lowest 3 thematic areas for this round only
            scored = [x for x in rows if x["percent"] is not None]
            scored_sorted = sorted(scored, key=lambda x: x["percent"])
            priority_set = set(x["area"] for x in scored_sorted[:3])

            for x in rows:
                x["is_priority"] = x["area"] in priority_set

            # User-friendly order: priority first, then lowest score
            rows = sorted(
                rows,
                key=lambda x: (
                    0 if x["is_priority"] else 1,
                    x["percent"] is None,
                    x["percent"] if x["percent"] is not None else 999999,
                    x["area"],
                )
            )

            rounds.append({
                "round_key": round_key,
                "round_label": round_label,
                "date_from": date_from,
                "date_to": date_to,
                "headers_count": len(header_list),
                "thematic_area_count": thematic_area_count,
                "rows": rows,
            })

        # Sort rounds by assessment date
        rounds = sorted(
            rounds,
            key=lambda x: (
                x["date_from"] is None,
                x["date_from"],
                x["round_label"],
            )
        )

        # -----------------------------------------------------
        # 6. Optional Excel export: all rounds in one Excel
        # -----------------------------------------------------
        if export:
            wb = openpyxl.Workbook()

            ws = wb.active
            ws.title = "All Rounds"

            ws.append([
                "Facility",
                "Province",
                "District",
                "HF Code",
                "Assessment Round",
                "Assessment Date From",
                "Assessment Date To",
                "Headers Included",
                "Thematic Areas Included",
                "Thematic Area",
                "HQIP % Achievement",
                "Priority Status",
                "YES",
                "Applicable",
                "Sections Used",
            ])

            for rd in rounds:
                for x in rd["rows"]:
                    ws.append([
                        facility_obj.name,
                        facility_obj.districtfk.provincefk.name,
                        facility_obj.districtfk.name,
                        facility_obj.hfcode,
                        rd["round_label"],
                        rd["date_from"],
                        rd["date_to"],
                        rd["headers_count"],
                        rd["thematic_area_count"],
                        x["area"],
                        x["percent"] if x["percent"] is not None else "",
                        "PRIORITY" if x["is_priority"] else "NON-PRIORITY",
                        x["yes"],
                        x["applicable"],
                        x["num_sections_used"],
                    ])

            ws_summary = wb.create_sheet("Round Summary")
            ws_summary.append([
                "Facility",
                "Province",
                "District",
                "HF Code",
                "Assessment Round",
                "Assessment Date From",
                "Assessment Date To",
                "Headers Included",
                "Thematic Areas Included",
                "Priority Thematic Areas",
            ])

            for rd in rounds:
                priority_areas = [
                    x["area"] for x in rd["rows"]
                    if x["is_priority"]
                ]

                ws_summary.append([
                    facility_obj.name,
                    facility_obj.districtfk.provincefk.name,
                    facility_obj.districtfk.name,
                    facility_obj.hfcode,
                    rd["round_label"],
                    rd["date_from"],
                    rd["date_to"],
                    rd["headers_count"],
                    rd["thematic_area_count"],
                    ", ".join(priority_areas),
                ])

            # Optional: separate sheet for each round
            for rd in rounds:
                sheet_name = str(rd["round_label"])[:25]
                invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
                for ch in invalid_chars:
                    sheet_name = sheet_name.replace(ch, "-")

                if sheet_name in wb.sheetnames:
                    sheet_name = f"{sheet_name[:20]} {rd['round_key']}"

                ws_round = wb.create_sheet(sheet_name)
                ws_round.append([
                    "Thematic Area",
                    "HQIP % Achievement",
                    "Priority Status",
                    "YES",
                    "Applicable",
                    "Sections Used",
                ])

                for x in rd["rows"]:
                    ws_round.append([
                        x["area"],
                        x["percent"] if x["percent"] is not None else "",
                        "PRIORITY" if x["is_priority"] else "NON-PRIORITY",
                        x["yes"],
                        x["applicable"],
                        x["num_sections_used"],
                    ])

            # Formatting
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style="thin", color="D9E2F3"),
                right=Side(style="thin", color="D9E2F3"),
                top=Side(style="thin", color="D9E2F3"),
                bottom=Side(style="thin", color="D9E2F3"),
            )

            for sheet in wb.worksheets:
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = sheet.dimensions

                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )
                    cell.border = border

                for row in sheet.iter_rows():
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

                for col in sheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)

                    for cell in col:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))

                    sheet.column_dimensions[col_letter].width = min(
                        max(max_len + 2, 12),
                        45,
                    )

            resp = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"HQIP_Priority_Areas_By_Round_Facility_{facility_obj.id}.xlsx"
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            wb.save(resp)
            return resp

        # -----------------------------------------------------
        # 7. Page context
        # -----------------------------------------------------
        context = dict(
            self.admin_site.each_context(request),
            title="HQIP Priority Thematic Areas",
            error_message=error_message,
            facility_obj=facility_obj,
            rounds=rounds,
        )

        return TemplateResponse(request, "admin/hiva/hqip_priority_areas.html", context)

@admin.register(HQIPAssessmentDashboard)
class HQIPAssessmentDashboardAdmin(
    ProvinceRestrictedAdminMixin,
    admin.ModelAdmin,
):
    """Read-only, audience-facing HQIP periodic assessment dashboard."""

    hqip_periodic_template = "admin/hiva/hqip_periodic_dashboard.html"

    def province_filter_kwargs(self, request):
        return {
            "facilityfk__districtfk__provincefk": user_province(request)
        }

    def changelist_view(self, request, extra_context=None):
        return self.hqip_periodic_dashboard(request)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        opts = HQIPAssessmentHeader._meta
        return (
            request.user.is_superuser
            or request.user.has_perm(
                f"{opts.app_label}.view_{opts.model_name}"
            )
            or request.user.has_perm(
                f"{opts.app_label}.change_{opts.model_name}"
            )
        )

    def has_module_permission(self, request):
        return self.has_view_permission(request)

    def _round2(self, value):
        if value is None:
            return None
        return float(
            Decimal(str(value)).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )
        )

    def _pct(self, yes_count, applicable_count):
        if not applicable_count:
            return None
        return self._round2(
            Decimal(yes_count)
            / Decimal(applicable_count)
            * Decimal("100")
        )

    # ------------------------------------------------------------------
    # Small calculation/display helpers
    # ------------------------------------------------------------------
    def _hqip_mean(self, values):
        clean = [Decimal(str(value)) for value in values if value is not None]
        if not clean:
            return None
        return self._round2(sum(clean) / Decimal(len(clean)))

    def _hqip_change(self, newer, older):
        if newer is None or older is None:
            return None
        return self._round2(Decimal(str(newer)) - Decimal(str(older)))

    def _hqip_percent_change(self, newer, older):
        if newer is None or older in (None, 0, 0.0):
            return None
        value = (
            (Decimal(str(newer)) - Decimal(str(older)))
            / Decimal(str(older))
            * Decimal("100")
        )
        return self._round2(value)

    def _hqip_change_status(self, change, has_comparison=True):
        if not has_comparison or change is None:
            return "Reference round", "neutral"
        if change >= 10:
            return "Strong improvement", "strong"
        if change >= 5:
            return "Improved", "improved"
        if change > 0:
            return "Slight improvement", "slight"
        if change == 0:
            return "No change", "neutral"
        if change > -5:
            return "Slight decline", "warning"
        return "Declined", "declined"

    def _hqip_rollup_from_standard_rows(self, rows):
        """Apply the existing HQIP hierarchy to pre-aggregated standard rows.

        Formula preserved from ``_compute_hqip_rollups``:
        - Standard % = YES / (YES + NO) * 100
        - Section % = mean of its scored Standard percentages
        - Thematic Area % = mean of its scored Section percentages
        - Overall HQIP % = mean of its scored Thematic Area percentages

        N/A and missing scores do not enter the Standard denominator.
        """

        standard_totals = defaultdict(
            lambda: {
                "yes": 0,
                "no": 0,
                "na": 0,
                "missing": 0,
                "applicable": 0,
            }
        )
        names = {}

        total_yes = 0
        total_no = 0
        total_na = 0
        total_missing = 0

        for row in rows:
            area_id = row["area_id"]
            section_id = row["section_id"]
            standard_id = row["standard_id"]
            key = (area_id, section_id, standard_id)

            names[key] = {
                "area_id": area_id,
                "area": row["area"],
                "section_id": section_id,
                "section": row["section"],
                "standard_id": standard_id,
                "standard": row["standard"],
            }

            standard_totals[key]["yes"] += row["yes"]
            standard_totals[key]["no"] += row["no"]
            standard_totals[key]["na"] += row["na"]
            standard_totals[key]["missing"] += row["missing"]
            standard_totals[key]["applicable"] += row["applicable"]

            total_yes += row["yes"]
            total_no += row["no"]
            total_na += row["na"]
            total_missing += row["missing"]

        standard_results = []
        section_percent_lists = defaultdict(list)
        section_names = {}

        for key, counts in standard_totals.items():
            meta = names[key]
            standard_percent = self._pct(counts["yes"], counts["applicable"])
            standard_results.append(
                {
                    **meta,
                    **counts,
                    "percent": standard_percent,
                }
            )

            section_key = (meta["area_id"], meta["section_id"])
            section_names[section_key] = {
                "area_id": meta["area_id"],
                "area": meta["area"],
                "section_id": meta["section_id"],
                "section": meta["section"],
            }
            if standard_percent is not None:
                section_percent_lists[section_key].append(standard_percent)

        section_results = []
        area_percent_lists = defaultdict(list)
        area_names = {}

        for section_key, percentages in section_percent_lists.items():
            meta = section_names[section_key]
            section_percent = self._hqip_mean(percentages)
            section_results.append(
                {
                    **meta,
                    "num_standards_used": len(percentages),
                    "percent": section_percent,
                }
            )

            area_names[meta["area_id"]] = meta["area"]
            if section_percent is not None:
                area_percent_lists[meta["area_id"]].append(section_percent)

        area_results = []
        for area_id, percentages in area_percent_lists.items():
            area_results.append(
                {
                    "area_id": area_id,
                    "area": area_names.get(area_id, "-"),
                    "num_sections_used": len(percentages),
                    "percent": self._hqip_mean(percentages),
                }
            )

        standard_results.sort(
            key=lambda item: (item["area"], item["section"], item["standard"])
        )
        section_results.sort(key=lambda item: (item["area"], item["section"]))
        area_results.sort(key=lambda item: item["area"])

        total_lines = total_yes + total_no + total_na + total_missing
        scored_lines = total_yes + total_no + total_na

        return {
            "standard_results": standard_results,
            "section_results": section_results,
            "area_results": area_results,
            "overall_percent": self._hqip_mean(
                [item["percent"] for item in area_results]
            ),
            "yes": total_yes,
            "no": total_no,
            "na": total_na,
            "missing": total_missing,
            "applicable": total_yes + total_no,
            "scored": scored_lines,
            "total_lines": total_lines,
            "completeness": self._pct(scored_lines, total_lines),
        }

    def _compute_hqip_rollups(self, headers_qs):
        """Backward-compatible wrapper for older dashboard copies.

        The standalone proxy dashboard now calculates its result through
        ``_hqip_rollup_from_standard_rows``.  Keeping this small wrapper means
        that an older leftover call in a large existing ``admin.py`` cannot
        raise ``AttributeError``.  Both paths use exactly the same hierarchy
        and N/A/missing-score rules.
        """

        rows = self._hqip_standard_rows(headers_qs)
        rollup = self._hqip_rollup_from_standard_rows(rows)
        return (
            rollup["standard_results"],
            rollup["section_results"],
            rollup["area_results"],
        )

    def _hqip_standard_rows(self, headers_qs):
        """Return one compact aggregate row per facility/round/standard."""

        queryset = (
            HQIPAssessment.objects
            .filter(
                header__in=headers_qs,
                criteriafk__standardfk__sectionfk__areafk_id=F(
                    "header__areafk_id"
                ),
            )
            .values(
                "header__facilityfk_id",
                "header__facilityfk__name",
                "header__facilityfk__hfcode",
                "header__facilityfk__districtfk_id",
                "header__facilityfk__districtfk__name",
                "header__facilityfk__districtfk__provincefk_id",
                "header__facilityfk__districtfk__provincefk__name",
                "header__assessmenttype_id",
                "header__assessmenttype__name",
                "criteriafk__standardfk__sectionfk__areafk_id",
                "criteriafk__standardfk__sectionfk__areafk__name",
                "criteriafk__standardfk__sectionfk_id",
                "criteriafk__standardfk__sectionfk__name",
                "criteriafk__standardfk_id",
                "criteriafk__standardfk__name",
            )
            .annotate(
                yes=Count("id", filter=Q(scorefk_id=SCORE_YES_ID)),
                no=Count("id", filter=Q(scorefk_id=SCORE_NO_ID)),
                na=Count("id", filter=Q(scorefk_id=SCORE_NA_ID)),
                missing=Count("id", filter=Q(scorefk__isnull=True)),
                applicable=Count(
                    "id",
                    filter=Q(scorefk_id__in=[SCORE_YES_ID, SCORE_NO_ID]),
                ),
            )
            .order_by()
        )

        rows = []
        for row in queryset:
            rows.append(
                {
                    "facility_id": row["header__facilityfk_id"],
                    "facility": row["header__facilityfk__name"] or "Unknown",
                    "hfcode": row["header__facilityfk__hfcode"] or "",
                    "district_id": row["header__facilityfk__districtfk_id"],
                    "district": (
                        row["header__facilityfk__districtfk__name"] or "Unknown"
                    ),
                    "province_id": (
                        row[
                            "header__facilityfk__districtfk__provincefk_id"
                        ]
                    ),
                    "province": (
                        row[
                            "header__facilityfk__districtfk__provincefk__name"
                        ]
                        or "Unknown"
                    ),
                    "round_id": row["header__assessmenttype_id"],
                    "round": row["header__assessmenttype__name"] or "Unknown",
                    "area_id": (
                        row["criteriafk__standardfk__sectionfk__areafk_id"]
                    ),
                    "area": (
                        row[
                            "criteriafk__standardfk__sectionfk__areafk__name"
                        ]
                        or "Unknown"
                    ),
                    "section_id": row["criteriafk__standardfk__sectionfk_id"],
                    "section": (
                        row["criteriafk__standardfk__sectionfk__name"]
                        or "Unknown"
                    ),
                    "standard_id": row["criteriafk__standardfk_id"],
                    "standard": row["criteriafk__standardfk__name"] or "Unknown",
                    "yes": row["yes"],
                    "no": row["no"],
                    "na": row["na"],
                    "missing": row["missing"],
                    "applicable": row["applicable"],
                }
            )
        return rows

    # ------------------------------------------------------------------
    # Main dashboard view
    # ------------------------------------------------------------------
    def hqip_periodic_dashboard(self, request):
        filters = {
            "province": request.GET.get("province", "").strip(),
            "district": request.GET.get("district", "").strip(),
            "facility": request.GET.get("facility", "").strip(),
            "assessmenttype": request.GET.get("assessmenttype", "").strip(),
            "area": request.GET.get("area", "").strip(),
        }

        # Permission-safe base queryset. Do not replace with .objects.all().
        allowed_headers = self.get_queryset(request).select_related(
            "facilityfk__districtfk__provincefk",
            "assessmenttype",
            "areafk",
        )

        # Filter choices are generated only from records this user may access.
        option_headers = allowed_headers
        province_options = list(
            option_headers
            .values(
                "facilityfk__districtfk__provincefk_id",
                "facilityfk__districtfk__provincefk__name",
            )
            .distinct()
            .order_by("facilityfk__districtfk__provincefk__name")
        )

        location_options = option_headers
        if filters["province"]:
            location_options = location_options.filter(
                facilityfk__districtfk__provincefk_id=filters["province"]
            )

        district_options = list(
            location_options
            .values("facilityfk__districtfk_id", "facilityfk__districtfk__name")
            .distinct()
            .order_by("facilityfk__districtfk__name")
        )

        facility_options_qs = location_options
        if filters["district"]:
            facility_options_qs = facility_options_qs.filter(
                facilityfk__districtfk_id=filters["district"]
            )
        facility_options = list(
            facility_options_qs
            .values("facilityfk_id", "facilityfk__name", "facilityfk__hfcode")
            .distinct()
            .order_by("facilityfk__name")
        )

        assessmenttype_options = list(
            option_headers
            .values("assessmenttype_id", "assessmenttype__name")
            .distinct()
            .order_by("assessmenttype__name")
        )
        area_options = list(
            option_headers
            .values("areafk_id", "areafk__name")
            .distinct()
            .order_by("areafk__name")
        )

        headers = allowed_headers
        if filters["province"]:
            headers = headers.filter(
                facilityfk__districtfk__provincefk_id=filters["province"]
            )
        if filters["district"]:
            headers = headers.filter(
                facilityfk__districtfk_id=filters["district"]
            )
        if filters["facility"]:
            headers = headers.filter(facilityfk_id=filters["facility"])
        if filters["assessmenttype"]:
            headers = headers.filter(
                assessmenttype_id=filters["assessmenttype"]
            )
        if filters["area"]:
            headers = headers.filter(areafk_id=filters["area"])

        # Evaluate compact header metadata once. A header represents one
        # facility/round/thematic-area assessment record.
        header_rows = list(
            headers.values(
                "id",
                "facilityfk_id",
                "facilityfk__name",
                "facilityfk__districtfk__provincefk_id",
                "assessmenttype_id",
                "assessmenttype__name",
                "assessmentdate",
                "assessmentend_date",
                "areafk_id",
            )
        )

        round_meta_query = (
            headers
            .values("assessmenttype_id", "assessmenttype__name")
            .annotate(
                date_from=Min("assessmentdate"),
                date_to=Max("assessmentend_date"),
                header_count=Count("id"),
                facility_count=Count("facilityfk_id", distinct=True),
            )
        )
        round_meta = {
            item["assessmenttype_id"]: {
                "round_id": item["assessmenttype_id"],
                "round": item["assessmenttype__name"] or "Unknown",
                "date_from": item["date_from"],
                "date_to": item["date_to"],
                "header_count": item["header_count"],
                "facility_count": item["facility_count"],
            }
            for item in round_meta_query
        }

        round_order = sorted(
            round_meta,
            key=lambda round_id: (
                round_meta[round_id]["date_from"] is None,
                round_meta[round_id]["date_from"] or py_date.max,
                round_meta[round_id]["round"],
            ),
        )
        round_position = {
            round_id: index for index, round_id in enumerate(round_order)
        }

        raw_rows = self._hqip_standard_rows(headers)

        by_round = defaultdict(list)
        by_province_round = defaultdict(list)
        by_area_round = defaultdict(list)
        by_facility_round = defaultdict(list)

        for row in raw_rows:
            by_round[row["round_id"]].append(row)
            by_province_round[(row["province_id"], row["round_id"])].append(row)
            by_area_round[(row["area_id"], row["round_id"])].append(row)
            by_facility_round[(row["facility_id"], row["round_id"])].append(row)

        # Use the standalone proxy dashboard's own hierarchical rollup helper.
        overall_counts = self._hqip_rollup_from_standard_rows(raw_rows)
        overall_score = overall_counts["overall_percent"]

        # Header/event counts avoid counting seven thematic headers as seven
        # separate facility assessment events.
        assessment_events = {
            (
                row["facilityfk_id"],
                row["assessmenttype_id"],
                row["assessmentdate"],
            )
            for row in header_rows
        }
        distinct_facilities = {row["facilityfk_id"] for row in header_rows}
        distinct_provinces = {
            row["facilityfk__districtfk__provincefk_id"] for row in header_rows
        }

        round_event_sets = defaultdict(set)
        round_area_sets = defaultdict(set)
        for row in header_rows:
            round_event_sets[row["assessmenttype_id"]].add(
                (row["facilityfk_id"], row["assessmentdate"])
            )
            round_area_sets[row["assessmenttype_id"]].add(row["areafk_id"])

        # --------------------------------------------------------------
        # Round trend
        # --------------------------------------------------------------
        round_rows = []
        previous_score = None
        for index, round_id in enumerate(round_order):
            rollup = self._hqip_rollup_from_standard_rows(by_round[round_id])
            score = rollup["overall_percent"]
            change_previous = self._hqip_change(score, previous_score)
            status, status_class = self._hqip_change_status(
                change_previous,
                has_comparison=index > 0 and previous_score is not None,
            )
            meta = round_meta[round_id]
            round_rows.append(
                {
                    **meta,
                    "assessment_events": len(round_event_sets[round_id]),
                    "thematic_areas": len(round_area_sets[round_id]),
                    "score": score,
                    "change_previous": change_previous,
                    "status": status,
                    "status_class": status_class,
                    "yes": rollup["yes"],
                    "no": rollup["no"],
                    "na": rollup["na"],
                    "missing": rollup["missing"],
                    "applicable": rollup["applicable"],
                    "completeness": rollup["completeness"],
                }
            )
            if score is not None:
                previous_score = score

        scored_round_rows = [row for row in round_rows if row["score"] is not None]
        first_round_row = scored_round_rows[0] if scored_round_rows else None
        latest_round_row = scored_round_rows[-1] if scored_round_rows else None
        overall_change = None
        if first_round_row and latest_round_row and first_round_row is not latest_round_row:
            overall_change = self._hqip_change(
                latest_round_row["score"], first_round_row["score"]
            )

        # --------------------------------------------------------------
        # Province by round
        # --------------------------------------------------------------
        province_rows = []
        province_round_scores = {}
        for (province_id, round_id), grouped_rows in by_province_round.items():
            rollup = self._hqip_rollup_from_standard_rows(grouped_rows)
            province = grouped_rows[0]["province"]
            facility_count = len({row["facility_id"] for row in grouped_rows})
            province_round_scores[(province_id, round_id)] = rollup["overall_percent"]
            province_rows.append(
                {
                    "province_id": province_id,
                    "province": province,
                    "round_id": round_id,
                    "round": round_meta[round_id]["round"],
                    "round_order": round_position[round_id],
                    "facilities": facility_count,
                    "score": rollup["overall_percent"],
                    "yes": rollup["yes"],
                    "no": rollup["no"],
                    "na": rollup["na"],
                    "missing": rollup["missing"],
                    "completeness": rollup["completeness"],
                }
            )

        province_baselines = {}
        for row in sorted(
            province_rows,
            key=lambda item: (item["province"], item["round_order"]),
        ):
            baseline = province_baselines.get(row["province_id"])
            if baseline is None and row["score"] is not None:
                province_baselines[row["province_id"]] = row["score"]
                baseline = row["score"]
            row["change_from_first"] = self._hqip_change(row["score"], baseline)
            row["status"], row["status_class"] = self._hqip_change_status(
                row["change_from_first"],
                has_comparison=(
                    baseline is not None
                    and row["score"] is not None
                    and row["round_order"] > 0
                ),
            )
        province_rows.sort(key=lambda item: (item["province"], item["round_order"]))

        # --------------------------------------------------------------
        # Thematic area by round
        # --------------------------------------------------------------
        area_rows = []
        for (area_id, round_id), grouped_rows in by_area_round.items():
            rollup = self._hqip_rollup_from_standard_rows(grouped_rows)
            area_rows.append(
                {
                    "area_id": area_id,
                    "area": grouped_rows[0]["area"],
                    "round_id": round_id,
                    "round": round_meta[round_id]["round"],
                    "round_order": round_position[round_id],
                    "score": rollup["overall_percent"],
                    "yes": rollup["yes"],
                    "no": rollup["no"],
                    "na": rollup["na"],
                    "missing": rollup["missing"],
                    "completeness": rollup["completeness"],
                }
            )

        area_baselines = {}
        for row in sorted(area_rows, key=lambda item: (item["area"], item["round_order"])):
            baseline = area_baselines.get(row["area_id"])
            if baseline is None and row["score"] is not None:
                area_baselines[row["area_id"]] = row["score"]
                baseline = row["score"]
            row["change_from_first"] = self._hqip_change(row["score"], baseline)
            row["status"], row["status_class"] = self._hqip_change_status(
                row["change_from_first"],
                has_comparison=(
                    baseline is not None
                    and row["score"] is not None
                    and row["round_order"] > 0
                ),
            )
        area_rows.sort(key=lambda item: (item["area"], item["round_order"]))

        # --------------------------------------------------------------
        # Facility by round and first-vs-latest comparison
        # --------------------------------------------------------------
        facility_round_rows = []
        facility_round_map = defaultdict(list)

        for (facility_id, round_id), grouped_rows in by_facility_round.items():
            rollup = self._hqip_rollup_from_standard_rows(grouped_rows)
            first = grouped_rows[0]
            item = {
                "facility_id": facility_id,
                "province": first["province"],
                "district": first["district"],
                "facility": first["facility"],
                "hfcode": first["hfcode"],
                "round_id": round_id,
                "round": round_meta[round_id]["round"],
                "round_order": round_position[round_id],
                "score": rollup["overall_percent"],
                "yes": rollup["yes"],
                "no": rollup["no"],
                "na": rollup["na"],
                "missing": rollup["missing"],
                "completeness": rollup["completeness"],
                "area_results": rollup["area_results"],
            }
            facility_round_rows.append(item)
            facility_round_map[facility_id].append(item)

        facility_round_rows.sort(
            key=lambda item: (
                item["province"],
                item["facility"],
                item["round_order"],
            )
        )

        facility_comparison_rows = []
        for facility_id, records in facility_round_map.items():
            scored = sorted(
                [row for row in records if row["score"] is not None],
                key=lambda item: item["round_order"],
            )
            if not scored:
                continue

            first = scored[0]
            latest = scored[-1]
            has_comparison = len(scored) >= 2 and first["round_id"] != latest["round_id"]
            change = (
                self._hqip_change(latest["score"], first["score"])
                if has_comparison
                else None
            )
            percent_change = (
                self._hqip_percent_change(latest["score"], first["score"])
                if has_comparison
                else None
            )
            status, status_class = self._hqip_change_status(
                change, has_comparison=has_comparison
            )

            first_areas = {
                area["area_id"]: area for area in first["area_results"]
            }
            latest_areas = {
                area["area_id"]: area for area in latest["area_results"]
            }

            improved_areas = 0
            declined_areas = 0
            unchanged_areas = 0
            for area_id in set(first_areas).intersection(latest_areas):
                area_change = self._hqip_change(
                    latest_areas[area_id]["percent"],
                    first_areas[area_id]["percent"],
                )
                if area_change is None:
                    continue
                if area_change > 0:
                    improved_areas += 1
                elif area_change < 0:
                    declined_areas += 1
                else:
                    unchanged_areas += 1

            lowest_latest = sorted(
                [
                    area for area in latest["area_results"]
                    if area["percent"] is not None
                ],
                key=lambda area: (area["percent"], area["area"]),
            )[:3]
            priority_areas = ", ".join(area["area"] for area in lowest_latest) or "-"

            if has_comparison and change is not None and change >= 10 and latest["score"] >= 70:
                story_use = "Strong success-story candidate"
            elif has_comparison and change is not None and change >= 5:
                story_use = "Potential success-story candidate"
            elif has_comparison and change is not None and change > 0:
                story_use = "Positive progress; verify before story use"
            elif has_comparison and change is not None and change < 0:
                story_use = "Learning and corrective-action case"
            else:
                story_use = "Routine monitoring"

            notes = (
                "Verify source checklists, assessment dates, scorer consistency, "
                "the improvement actions implemented, and supporting facility evidence."
                if has_comparison
                else "A second comparable assessment round is required for change analysis."
            )

            facility_comparison_rows.append(
                {
                    "facility_id": facility_id,
                    "province": latest["province"],
                    "district": latest["district"],
                    "facility": latest["facility"],
                    "hfcode": latest["hfcode"],
                    "first_round": first["round"],
                    "latest_round": latest["round"],
                    "first_score": first["score"],
                    "latest_score": latest["score"],
                    "absolute_change": change,
                    "percent_change": percent_change,
                    "status": status,
                    "status_class": status_class,
                    "improved_areas": improved_areas,
                    "declined_areas": declined_areas,
                    "unchanged_areas": unchanged_areas,
                    "priority_areas": priority_areas,
                    "story_use": story_use,
                    "notes": notes,
                    "has_comparison": has_comparison,
                }
            )

        facility_comparison_rows.sort(
            key=lambda item: (
                item["province"],
                item["facility"],
            )
        )

        success_story_candidates = sorted(
            [
                row for row in facility_comparison_rows
                if row["has_comparison"]
                and row["absolute_change"] is not None
                and row["absolute_change"] > 0
            ],
            key=lambda item: (
                item["absolute_change"],
                item["latest_score"],
                item["improved_areas"],
            ),
            reverse=True,
        )[:10]

        # Latest selected round's lowest thematic scores.
        current_gap_rows = []
        if latest_round_row:
            current_gap_rows = sorted(
                [
                    row for row in area_rows
                    if row["round_id"] == latest_round_row["round_id"]
                    and row["score"] is not None
                ],
                key=lambda item: (item["score"], item["area"]),
            )
            for rank, row in enumerate(current_gap_rows, start=1):
                row["priority_rank"] = rank
                row["is_top_priority"] = rank <= 3

        # --------------------------------------------------------------
        # KPI cards and chart payload
        # --------------------------------------------------------------
        kpis = {
            "overall_score": overall_score,
            "overall_change": overall_change,
            "facilities": len(distinct_facilities),
            "assessment_events": len(assessment_events),
            "provinces": len(distinct_provinces),
            "rounds": len(round_order),
            "failed_criteria": overall_counts["no"],
            "missing_scores": overall_counts["missing"],
            "scoring_completeness": overall_counts["completeness"],
        }

        province_names = sorted({row["province"] for row in province_rows})
        area_names = sorted({row["area"] for row in area_rows})

        chart_data = {
            "round_trend": [
                {
                    "round": row["round"],
                    "score": row["score"],
                    "completeness": row["completeness"],
                }
                for row in round_rows
            ],
            "province_comparison": {
                "labels": province_names,
                "datasets": [
                    {
                        "label": round_meta[round_id]["round"],
                        "data": [
                            next(
                                (
                                    row["score"] for row in province_rows
                                    if row["province"] == province
                                    and row["round_id"] == round_id
                                ),
                                None,
                            )
                            for province in province_names
                        ],
                    }
                    for round_id in round_order
                ],
            },
            "area_comparison": {
                "labels": area_names,
                "datasets": [
                    {
                        "label": round_meta[round_id]["round"],
                        "data": [
                            next(
                                (
                                    row["score"] for row in area_rows
                                    if row["area"] == area
                                    and row["round_id"] == round_id
                                ),
                                None,
                            )
                            for area in area_names
                        ],
                    }
                    for round_id in round_order
                ],
            },
            "facility_improvement": [
                {
                    "facility": row["facility"],
                    "province": row["province"],
                    "change": row["absolute_change"],
                }
                for row in sorted(
                    [
                        item for item in facility_comparison_rows
                        if item["has_comparison"]
                        and item["absolute_change"] is not None
                    ],
                    key=lambda item: item["absolute_change"],
                    reverse=True,
                )[:12]
            ],
            "current_gaps": [
                {"area": row["area"], "score": row["score"]}
                for row in current_gap_rows
            ],
            "response_mix": {
                "labels": ["YES", "NO", "N/A", "Missing"],
                "values": [
                    overall_counts["yes"],
                    overall_counts["no"],
                    overall_counts["na"],
                    overall_counts["missing"],
                ],
            },
        }

        methodology_note = (
            "Standard % = YES / (YES + NO) × 100. N/A and blank scores are "
            "excluded from the denominator. Section % is the average of scored "
            "Standard percentages; Thematic Area % is the average of scored "
            "Section percentages; Overall HQIP % is the average of scored "
            "Thematic Area percentages. Change is shown in percentage points."
        )

        export_params = request.GET.copy()
        export_params["export"] = "1"
        export_query = export_params.urlencode()

        if request.GET.get("export") == "1":
            return self._export_hqip_periodic_dashboard(
                kpis=kpis,
                filters=filters,
                methodology_note=methodology_note,
                round_rows=round_rows,
                province_rows=province_rows,
                area_rows=area_rows,
                facility_comparison_rows=facility_comparison_rows,
                facility_round_rows=facility_round_rows,
                success_story_candidates=success_story_candidates,
            )

        context = dict(
            self.admin_site.each_context(request),
            title="HQIP Periodic Assessment Dashboard",
            filters=filters,
            province_options=province_options,
            district_options=district_options,
            facility_options=facility_options,
            assessmenttype_options=assessmenttype_options,
            area_options=area_options,
            kpis=kpis,
            methodology_note=methodology_note,
            round_rows=round_rows,
            province_rows=province_rows,
            area_rows=area_rows,
            facility_comparison_rows=facility_comparison_rows,
            facility_round_rows=facility_round_rows,
            success_story_candidates=success_story_candidates,
            current_gap_rows=current_gap_rows,
            first_round_row=first_round_row,
            latest_round_row=latest_round_row,
            chart_data=chart_data,
            export_query=export_query,
            records_found=bool(raw_rows),
        )
        return TemplateResponse(request, self.hqip_periodic_template, context)

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------
    def _export_hqip_periodic_dashboard(
        self,
        *,
        kpis,
        filters,
        methodology_note,
        round_rows,
        province_rows,
        area_rows,
        facility_comparison_rows,
        facility_round_rows,
        success_story_candidates,
    ):
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Dashboard Summary"

        ws_summary.append(["HQIP Periodic Assessment Dashboard"])
        ws_summary.append(["Generated", timezone.localtime(timezone.now()).replace(tzinfo=None)])
        ws_summary.append(["Methodology", methodology_note])
        ws_summary.append([])
        ws_summary.append(["Selected Filter", "Value"])
        for key, value in filters.items():
            ws_summary.append([key.replace("_", " ").title(), value or "All"])
        ws_summary.append([])
        ws_summary.append(["KPI", "Value"])
        for key, value in kpis.items():
            ws_summary.append([key.replace("_", " ").title(), value if value is not None else "N/A"])

        ws_round = wb.create_sheet("Round Trend")
        ws_round.append(
            [
                "Assessment Round",
                "Date From",
                "Date To",
                "Facilities",
                "Assessment Events",
                "Thematic Areas",
                "HQIP %",
                "Change vs Previous (pp)",
                "Status",
                "YES",
                "NO",
                "N/A",
                "Missing",
                "Scoring Completeness %",
            ]
        )
        for row in round_rows:
            ws_round.append(
                [
                    row["round"],
                    row["date_from"],
                    row["date_to"],
                    row["facility_count"],
                    row["assessment_events"],
                    row["thematic_areas"],
                    row["score"],
                    row["change_previous"],
                    row["status"],
                    row["yes"],
                    row["no"],
                    row["na"],
                    row["missing"],
                    row["completeness"],
                ]
            )

        ws_province = wb.create_sheet("Province by Round")
        ws_province.append(
            [
                "Province",
                "Assessment Round",
                "Facilities",
                "HQIP %",
                "Change vs First (pp)",
                "Status",
                "YES",
                "NO",
                "N/A",
                "Missing",
                "Scoring Completeness %",
            ]
        )
        for row in province_rows:
            ws_province.append(
                [
                    row["province"],
                    row["round"],
                    row["facilities"],
                    row["score"],
                    row["change_from_first"],
                    row["status"],
                    row["yes"],
                    row["no"],
                    row["na"],
                    row["missing"],
                    row["completeness"],
                ]
            )

        ws_area = wb.create_sheet("Thematic by Round")
        ws_area.append(
            [
                "Thematic Area",
                "Assessment Round",
                "HQIP %",
                "Change vs First (pp)",
                "Status",
                "YES",
                "NO",
                "N/A",
                "Missing",
                "Scoring Completeness %",
            ]
        )
        for row in area_rows:
            ws_area.append(
                [
                    row["area"],
                    row["round"],
                    row["score"],
                    row["change_from_first"],
                    row["status"],
                    row["yes"],
                    row["no"],
                    row["na"],
                    row["missing"],
                    row["completeness"],
                ]
            )

        ws_facility = wb.create_sheet("Facility Comparison")
        ws_facility.append(
            [
                "Province",
                "District",
                "Facility",
                "HF Code",
                "First Round",
                "Latest Round",
                "First HQIP %",
                "Latest HQIP %",
                "Absolute Change (pp)",
                "Percent Change",
                "Status",
                "Improved Areas",
                "Declined Areas",
                "Unchanged Areas",
                "Latest Priority Areas",
                "Story Use",
                "Notes for Field Verification",
            ]
        )
        for row in facility_comparison_rows:
            ws_facility.append(
                [
                    row["province"],
                    row["district"],
                    row["facility"],
                    row["hfcode"],
                    row["first_round"],
                    row["latest_round"],
                    row["first_score"],
                    row["latest_score"],
                    row["absolute_change"],
                    row["percent_change"],
                    row["status"],
                    row["improved_areas"],
                    row["declined_areas"],
                    row["unchanged_areas"],
                    row["priority_areas"],
                    row["story_use"],
                    row["notes"],
                ]
            )

        ws_facility_round = wb.create_sheet("Facility Round Scores")
        ws_facility_round.append(
            [
                "Province",
                "District",
                "Facility",
                "HF Code",
                "Assessment Round",
                "HQIP %",
                "YES",
                "NO",
                "N/A",
                "Missing",
                "Scoring Completeness %",
            ]
        )
        for row in facility_round_rows:
            ws_facility_round.append(
                [
                    row["province"],
                    row["district"],
                    row["facility"],
                    row["hfcode"],
                    row["round"],
                    row["score"],
                    row["yes"],
                    row["no"],
                    row["na"],
                    row["missing"],
                    row["completeness"],
                ]
            )

        ws_story = wb.create_sheet("Story Candidates")
        ws_story.append(
            [
                "Province",
                "Facility",
                "First Round",
                "Latest Round",
                "First HQIP %",
                "Latest HQIP %",
                "Change (pp)",
                "Improved Areas",
                "Declined Areas",
                "Story Use",
                "Field Verification",
            ]
        )
        for row in success_story_candidates:
            ws_story.append(
                [
                    row["province"],
                    row["facility"],
                    row["first_round"],
                    row["latest_round"],
                    row["first_score"],
                    row["latest_score"],
                    row["absolute_change"],
                    row["improved_areas"],
                    row["declined_areas"],
                    row["story_use"],
                    row["notes"],
                ]
            )

        self._style_hqip_dashboard_workbook(wb)

        timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = (
            f'attachment; filename="HQIP_Periodic_Dashboard_{timestamp}.xlsx"'
        )
        wb.save(response)
        return response

    def _style_hqip_dashboard_workbook(self, workbook):
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )

        for worksheet in workbook.worksheets:
            if worksheet.title != "Dashboard Summary":
                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = worksheet.dimensions

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = border

            if worksheet.title == "Dashboard Summary":
                worksheet["A1"].font = Font(
                    color="FFFFFF", bold=True, size=16
                )

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if isinstance(cell.value, py_date):
                        cell.number_format = "yyyy-mm-dd"

            for column_cells in worksheet.columns:
                column_letter = get_column_letter(column_cells[0].column)
                max_length = max(
                    [
                        len(str(cell.value))
                        for cell in column_cells
                        if cell.value is not None
                    ]
                    or [10]
                )
                worksheet.column_dimensions[column_letter].width = min(
                    max(max_length + 2, 12), 45
                )

# ============================================================
# Other Admins (keep simple / safe)
# ============================================================
class QICMonthFilter(admin.SimpleListFilter):
    title = _('QI Committee Date (Month + Year)')
    parameter_name = 'qic_month'

    def lookups(self, request, model_admin):
        dates = (Qicdataset.objects.exclude(qiccommdate__isnull=True).dates('qiccommdate', 'month'))
        return [(d.strftime("%Y-%m"), d.strftime("%B %Y")) for d in dates]

    def queryset(self, request, queryset):
        if self.value():
            year, month = self.value().split('-')
            return queryset.filter(qiccommdate__year=year, qiccommdate__month=month)
        return queryset
    
@admin.register(QICommittee)
class QICommitteeAdmin(ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    list_display = (
        "facility",
        "facility_staff",
        "role",
        "created_at",
        "record_status",
    )

    list_filter = ("role", "facility")
    search_fields = ("facility__name", "facility_staff__name")

    def province_filter_kwargs(self, request):
        return {"facility__districtfk__provincefk": user_province(request)}

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "facility" and not request.user.is_superuser:
            prov = user_province(request)
            kwargs["queryset"] = Facility.objects.filter(districtfk__provincefk=prov) if prov else Facility.objects.none()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

@admin.register(FacilityStaff)
class FacilityStaffAdmin(ProvinceRestrictedAdminMixin, admin.ModelAdmin):

    list_display = (
        "first_name",
        "last_name",
        "father_name",
        "facility",
        "position",
        "gender",
        "phone",
        "is_active",
        "verified",
    )

    list_filter = (
        "facility",
        "position",
        "gender",
        "verified",
    )

    search_fields = (
        "first_name",
        "last_name",
        "phone",
        "tazkira_number",
    )

    def province_filter_kwargs(self, request):
        return {"facility__districtfk__provincefk": user_province(request)}

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "facility" and not request.user.is_superuser:
            prov = user_province(request)
            kwargs["queryset"] = Facility.objects.filter(districtfk__provincefk=prov) if prov else Facility.objects.none()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

from django.contrib import admin
from django.utils.html import format_html

@admin.register(Qicdataset)
class MyModelqicdataset(ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    actions = ["export_qic_dataset_to_excel"]
    
    list_display = [
        #"id",
        "qicfacility",
        "yearqic",
        "monthqic",
        "qiccommdate",
        "progress_bar",
        "qictotalquestions",
        "tor_status",
        "meeting_status",
        "minutes_status",
        "pip_status",
        "ngo_status",
        "peer_status",
        "shura_status",
    ]

    list_filter = [
        "qicfacility__districtfk__provincefk",
        "qicfacility",
        QICMonthFilter,
        "qictoravail_bool",
        "qiclastmonth_bool",
        "qicmmavial_bool",
        "qicpipavail_bool",
        "qicngoinvolved_bool",
        "qicpeertopeeravail_bool",
        "qicmetwithhealthshura_bool",
        "qicmeeting_quorum_met",
        "qicfollowup_required",
        "qicvalidated_by_supervisor",
    ]

    search_fields = [
        "qicfacility__name",
        "qicfacility__districtfk__name",
        "qicfacility__districtfk__provincefk__name",
        "remarks",
        "qicdata_quality_issue_note",
    ]

    readonly_fields = [
        "qictotalquestions",
        "qicpercentscore",
        "created_at",
        "updated_at",
        # legacy integer fields
        "qictoravailvalue",
        "qiclastmonthvalue",
        "qicmmavialvalue",
        "qicmmsignedvalue",
        "qicmmdatausevalue",
        "qichqiptollavailvalue",
        "qicpipavailvalue",
        "qicpipupdatedvalue",
        "qicngoinvolvedvalue",
        "qicpeertopeeravailvalue",
        "qicmenteelogbookavialvalue",
        "qicmenteelogbookupdatedvalue",
        "qicmetwithhealthshuravalue",
        "qichealthshurainvolvedincorractvalue",
    ]

    fieldsets = (
        ("Basic Information", {
            "fields": (
                "qiccommdate",
                "yearqic",
                "monthqic",
                "qicfacility",
                "qicdatacollector",
                "qicimplementor",
            )
        }),

        ("Current Boolean Data Entry", {
            "fields": (
                "qictoravail_bool",
                "qiclastmonth_bool",
                "qicmmavial_bool",
                "qicmmsigned_bool",
                "qicmmdatause_bool",
                "qichqiptollavail_bool",
                "qicpipavail_bool",
                "qicpipupdated_bool",
                "qicngoinvolved_bool",
                "qicpeertopeeravail_bool",
                "qicmenteelogbookavial_bool",
                "qicmenteelogbookupdated_bool",
                "qicmetwithhealthshura_bool",
                "qichealthshurainvolvedincorract_bool",
            )
        }),

        ("QIC Score Summary", {
            "fields": (
                "qictotalquestions",
                "qicpercentscore",
            )
        }),

        ("Additional QIC Monitoring Fields", {
            "classes": ("collapse",),
            "fields": (
                "qiccommittee_members_count",
                "qicmeeting_quorum_met",
                "qicaction_points_count",
                "qicactions_completed_count",
                "qicnext_meeting_date",
                "qicfollowup_required",
                "qicvalidated_by_supervisor",
                "qicdata_quality_issue_note",
            )
        }),

        ("Remarks", {
            "fields": (
                "remarks",
            )
        }),

        ("System Information", {
            "classes": ("collapse",),
            "fields": (
                "created_at",
                "updated_at",
            )
        }),

        ("Legacy Integer Values (Read Only)", {
            "classes": ("collapse",),
            "fields": (
                "qictoravailvalue",
                "qiclastmonthvalue",
                "qicmmavialvalue",
                "qicmmsignedvalue",
                "qicmmdatausevalue",
                "qichqiptollavailvalue",
                "qicpipavailvalue",
                "qicpipupdatedvalue",
                "qicngoinvolvedvalue",
                "qicpeertopeeravailvalue",
                "qicmenteelogbookavialvalue",
                "qicmenteelogbookupdatedvalue",
                "qicmetwithhealthshuravalue",
                "qichealthshurainvolvedincorractvalue",
            )
        }),
    )

    list_per_page = 20
    ordering = ["-created_at", "-id"]

    class Media:
        css = {
            "all": ("admin/css/qic_admin.css",)
        }

    def province_filter_kwargs(self, request):
        return {"qicfacility__districtfk__provincefk": user_province(request)}

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if not request.user.is_superuser:
            prov = user_province(request)

            if db_field.name == "qicfacility":
                kwargs["queryset"] = Facility.objects.filter(
                    districtfk__provincefk=prov
                ) if prov else Facility.objects.none()

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def progress_bar(self, obj):
        percent = float(obj.qicpercentscore or 0)

        if percent >= 80:
            color = "#2e7d32"
        elif percent >= 50:
            color = "#f9a825"
        else:
            color = "#c62828"

        return format_html(
            """
            <div style="width:140px;">
                <div style="
                    background:#e0e0e0;
                    border-radius:10px;
                    height:18px;
                    overflow:hidden;
                    position:relative;
                ">
                    <div style="
                        width:{}%;
                        background:{};
                        height:18px;
                        border-radius:10px;
                    "></div>
                </div>
                <div style="font-size:11px; margin-top:2px; text-align:center;">
                    {}%
                </div>
            </div>
            """,
            percent,
            color,
            round(percent, 1)
        )
    progress_bar.short_description = "Progress"

    def yes_no_badge(self, value):
        if value:
            return format_html(
                '<span style="color:#fff; background:#2e7d32; padding:2px 8px; border-radius:10px; font-size:11px;">Yes</span>'
            )
        return format_html(
            '<span style="color:#fff; background:#c62828; padding:2px 8px; border-radius:10px; font-size:11px;">No</span>'
        )

    def tor_status(self, obj):
        return self.yes_no_badge(obj.qictoravail_bool)
    tor_status.short_description = "TOR"

    def meeting_status(self, obj):
        return self.yes_no_badge(obj.qiclastmonth_bool)
    meeting_status.short_description = "Meeting"

    def minutes_status(self, obj):
        return self.yes_no_badge(obj.qicmmavial_bool)
    minutes_status.short_description = "Minutes"

    def pip_status(self, obj):
        return self.yes_no_badge(obj.qicpipavail_bool)
    pip_status.short_description = "PIP"

    def ngo_status(self, obj):
        return self.yes_no_badge(obj.qicngoinvolved_bool)
    ngo_status.short_description = "NGO"

    def peer_status(self, obj):
        return self.yes_no_badge(obj.qicpeertopeeravail_bool)
    peer_status.short_description = "Peer Learning"

    def shura_status(self, obj):
        return self.yes_no_badge(obj.qicmetwithhealthshura_bool)
    shura_status.short_description = "Shura"

     # ============================================================
    # EXPORT TO EXCEL
    # ============================================================
    @admin.action(description="Export selected QIC records to Excel")
    def export_qic_dataset_to_excel(self, request, queryset):
        """
        Exports selected QIC records to Excel for analysis and visualization.

        Sheets:
        1. QIC_Data
        2. Summary_By_Province
        3. Summary_By_Facility
        4. Summary_By_Month
        5. Data_Quality_Issues
        6. Question_Summary
        """

        queryset = queryset.select_related(
            "qicfacility",
            "qicfacility__districtfk",
            "qicfacility__districtfk__provincefk",
            "qicfacility__facilitytypefk",
            "qicdatacollector",
            "qicimplementor",
        ).order_by(
            "qicfacility__districtfk__provincefk__name",
            "qicfacility__districtfk__name",
            "qicfacility__name",
            "yearqic",
            "monthqic",
            "qiccommdate",
        )

        if not queryset.exists():
            messages.warning(request, "No QIC records selected for export.")
            return None

        month_map = dict(Qicdataset.MONTH_CHOICES)

        question_fields = [
            ("qictoravail_bool", "1. TOR of QI focal point and QI committee available"),
            ("qiclastmonth_bool", "2. QI committee meeting conducted last month"),
            ("qicmmavial_bool", "3. Meeting minutes available"),
            ("qicmmsigned_bool", "4. Meeting minutes signed by participants"),
            ("qicmmdatause_bool", "5. Data use discussed in QI committee meeting"),
            ("qichqiptollavail_bool", "6. HQIP tool available and accessible"),
            ("qicpipavail_bool", "7. PIP available"),
            ("qicpipupdated_bool", "8. PIP updated in last month meeting"),
            ("qicngoinvolved_bool", "9. NGO involved in completed corrective actions"),
            ("qicpeertopeeravail_bool", "10. Peer-to-peer learning conducted"),
            ("qicmenteelogbookavial_bool", "11. Mentee logbook available"),
            ("qicmenteelogbookupdated_bool", "12. Mentee logbook updated and signed"),
            ("qicmetwithhealthshura_bool", "13. QI committee met HF Shura-e-Sihie"),
            ("qichealthshurainvolvedincorract_bool", "14. HF Shura involved in corrective actions"),
        ]

        total_questions = len(question_fields)

        def safe_excel_value(value):
            if value is None:
                return ""

            if isinstance(value, bool):
                return "Yes" if value else "No"

            if isinstance(value, py_datetime):
                if timezone.is_aware(value):
                    value = timezone.localtime(value)
                    value = timezone.make_naive(value)
                return value

            if isinstance(value, Decimal):
                return float(value)

            if isinstance(value, str):
                return ILLEGAL_CHARACTERS_RE.sub("", value)

            return value

        def safe_str(obj):
            if obj is None:
                return ""
            return safe_excel_value(str(obj))

        def get_name(obj):
            if obj is None:
                return ""
            return safe_excel_value(getattr(obj, "name", str(obj)))

        def yes_no(value):
            return "Yes" if value else "No"

        def score_category(percent):
            percent = float(percent or 0)
            if percent >= 80:
                return "Good"
            if percent >= 50:
                return "Moderate"
            return "Low"

        def month_sort_value(value):
            try:
                return int(value)
            except Exception:
                return 0

        def detect_dq_issues(obj):
            issues = []

            if not obj.qiccommdate:
                issues.append("Missing QIC meeting date")

            if not obj.qicdatacollector_id:
                issues.append("Missing data collector")

            if not obj.qicimplementor_id:
                issues.append("Missing implementor")

            if float(obj.qicpercentscore or 0) < 50:
                issues.append("Low QIC score below 50%")

            if not obj.qiclastmonth_bool:
                issues.append("QI committee meeting not conducted last month")

            if not obj.qicmmavial_bool:
                issues.append("Meeting minutes not available")

            if not obj.qicpipavail_bool:
                issues.append("PIP not available")

            if obj.qicpipavail_bool and not obj.qicpipupdated_bool:
                issues.append("PIP available but not updated")

            if obj.qicaction_points_count and obj.qicactions_completed_count:
                if obj.qicactions_completed_count > obj.qicaction_points_count:
                    issues.append("Completed actions greater than total action points")

            if obj.qicnext_meeting_date and obj.qiccommdate:
                if obj.qicnext_meeting_date < obj.qiccommdate:
                    issues.append("Next meeting date is earlier than current QIC date")

            if obj.qicfollowup_required:
                issues.append("Follow-up required")

            if not obj.qicvalidated_by_supervisor:
                issues.append("Not validated by supervisor")

            if obj.qicdata_quality_issue_note:
                issues.append("Data quality issue note recorded")

            return issues

        def style_worksheet(ws):
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style="thin", color="D9E2F3"),
                right=Side(style="thin", color="D9E2F3"),
                top=Side(style="thin", color="D9E2F3"),
                bottom=Side(style="thin", color="D9E2F3"),
            )

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = border

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                    if isinstance(cell.value, py_datetime):
                        cell.number_format = "yyyy-mm-dd hh:mm:ss"
                    elif isinstance(cell.value, py_date):
                        cell.number_format = "yyyy-mm-dd"

            for column_cells in ws.columns:
                max_length = 0
                col_letter = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[col_letter].width = min(
                    max(max_length + 2, 12),
                    45,
                )

        wb = openpyxl.Workbook()

        ws_data = wb.active
        ws_data.title = "QIC_Data"

        ws_province = wb.create_sheet("Summary_By_Province")
        ws_facility = wb.create_sheet("Summary_By_Facility")
        ws_month = wb.create_sheet("Summary_By_Month")
        ws_dq = wb.create_sheet("Data_Quality_Issues")
        ws_question = wb.create_sheet("Question_Summary")

        # ============================================================
        # Sheet 1: Full QIC data
        # ============================================================
        main_headers = [
            "Record ID",
            "Province",
            "District",
            "Facility Type",
            "Facility",
            "HF Code",
            "Year",
            "Month Number",
            "Month Name",
            "QIC Meeting Date",
            "Data Collector",
            "Implementor",
            "Total YES",
            "Total Questions",
            "QIC Percent Score",
            "Score Category",
        ]

        question_headers = [label for _field, label in question_fields]

        additional_headers = [
            "Committee Members Count",
            "Quorum Met",
            "Action Points Count",
            "Actions Completed Count",
            "Action Completion %",
            "Next Meeting Date",
            "Follow-up Required",
            "Validated By Supervisor",
            "Data Quality Issue Note",
            "Remarks",
            "Created At",
            "Updated At",
        ]

        ws_data.append(main_headers + question_headers + additional_headers)

        province_summary = defaultdict(lambda: {
            "records": 0,
            "score_sum": 0,
            "yes_sum": 0,
            "questions_sum": 0,
            "followup_required": 0,
            "validated": 0,
        })

        facility_summary = defaultdict(lambda: {
            "records": 0,
            "score_sum": 0,
            "yes_sum": 0,
            "questions_sum": 0,
            "followup_required": 0,
            "validated": 0,
        })

        month_summary = defaultdict(lambda: {
            "records": 0,
            "score_sum": 0,
            "yes_sum": 0,
            "questions_sum": 0,
            "followup_required": 0,
            "validated": 0,
        })

        question_summary = {
            field: {
                "question": label,
                "yes": 0,
                "no": 0,
            }
            for field, label in question_fields
        }

        dq_rows = []

        for obj in queryset:
            facility = obj.qicfacility
            district = facility.districtfk if facility else None
            province = district.provincefk if district else None
            facility_type = getattr(facility, "facilitytypefk", None) if facility else None

            month_name = month_map.get(str(obj.monthqic), obj.monthqic)
            percent_score = float(obj.qicpercentscore or 0)
            total_yes = int(obj.qictotalquestions or 0)

            action_completion_pct = ""
            if obj.qicaction_points_count:
                action_completion_pct = round(
                    (float(obj.qicactions_completed_count or 0) / float(obj.qicaction_points_count)) * 100,
                    2,
                )

            question_values = []
            for field, _label in question_fields:
                value = bool(getattr(obj, field))
                question_values.append(yes_no(value))

                if value:
                    question_summary[field]["yes"] += 1
                else:
                    question_summary[field]["no"] += 1

            ws_data.append([
                obj.id,
                get_name(province),
                get_name(district),
                get_name(facility_type),
                get_name(facility),
                safe_excel_value(getattr(facility, "hfcode", "")),
                obj.yearqic,
                month_sort_value(obj.monthqic),
                month_name,
                safe_excel_value(obj.qiccommdate),
                safe_str(obj.qicdatacollector),
                safe_str(obj.qicimplementor),
                total_yes,
                total_questions,
                percent_score,
                score_category(percent_score),
                *question_values,
                obj.qiccommittee_members_count or 0,
                yes_no(obj.qicmeeting_quorum_met),
                obj.qicaction_points_count or 0,
                obj.qicactions_completed_count or 0,
                action_completion_pct,
                safe_excel_value(obj.qicnext_meeting_date),
                yes_no(obj.qicfollowup_required),
                yes_no(obj.qicvalidated_by_supervisor),
                safe_excel_value(obj.qicdata_quality_issue_note),
                safe_excel_value(obj.remarks),
                safe_excel_value(obj.created_at),
                safe_excel_value(obj.updated_at),
            ])

            province_key = get_name(province) or "Unknown"
            facility_key = (
                get_name(province) or "Unknown",
                get_name(district) or "Unknown",
                get_name(facility) or "Unknown",
                safe_excel_value(getattr(facility, "hfcode", "")),
            )
            month_key = (
                obj.yearqic,
                month_sort_value(obj.monthqic),
                month_name,
            )

            for bucket_key, bucket in [
                (province_key, province_summary[province_key]),
                (facility_key, facility_summary[facility_key]),
                (month_key, month_summary[month_key]),
            ]:
                bucket["records"] += 1
                bucket["score_sum"] += percent_score
                bucket["yes_sum"] += total_yes
                bucket["questions_sum"] += total_questions
                bucket["followup_required"] += 1 if obj.qicfollowup_required else 0
                bucket["validated"] += 1 if obj.qicvalidated_by_supervisor else 0

            issues = detect_dq_issues(obj)
            if issues:
                dq_rows.append([
                    obj.id,
                    get_name(province),
                    get_name(district),
                    get_name(facility),
                    safe_excel_value(getattr(facility, "hfcode", "")),
                    obj.yearqic,
                    month_name,
                    safe_excel_value(obj.qiccommdate),
                    percent_score,
                    "; ".join(issues),
                    safe_excel_value(obj.qicdata_quality_issue_note),
                    safe_excel_value(obj.remarks),
                ])

        # ============================================================
        # Sheet 2: Summary by province
        # ============================================================
        ws_province.append([
            "Province",
            "Records",
            "Average QIC %",
            "Total YES",
            "Total Questions",
            "Overall Achievement %",
            "Follow-up Required Count",
            "Validated Count",
            "Validation %",
        ])

        for province_name, s in sorted(province_summary.items()):
            avg_score = round(s["score_sum"] / s["records"], 2) if s["records"] else 0
            overall_achievement = round((s["yes_sum"] / s["questions_sum"]) * 100, 2) if s["questions_sum"] else 0
            validation_pct = round((s["validated"] / s["records"]) * 100, 2) if s["records"] else 0

            ws_province.append([
                province_name,
                s["records"],
                avg_score,
                s["yes_sum"],
                s["questions_sum"],
                overall_achievement,
                s["followup_required"],
                s["validated"],
                validation_pct,
            ])

        # ============================================================
        # Sheet 3: Summary by facility
        # ============================================================
        ws_facility.append([
            "Province",
            "District",
            "Facility",
            "HF Code",
            "Records",
            "Average QIC %",
            "Total YES",
            "Total Questions",
            "Overall Achievement %",
            "Follow-up Required Count",
            "Validated Count",
            "Validation %",
        ])

        for facility_key, s in sorted(facility_summary.items()):
            province_name, district_name, facility_name, hfcode = facility_key
            avg_score = round(s["score_sum"] / s["records"], 2) if s["records"] else 0
            overall_achievement = round((s["yes_sum"] / s["questions_sum"]) * 100, 2) if s["questions_sum"] else 0
            validation_pct = round((s["validated"] / s["records"]) * 100, 2) if s["records"] else 0

            ws_facility.append([
                province_name,
                district_name,
                facility_name,
                hfcode,
                s["records"],
                avg_score,
                s["yes_sum"],
                s["questions_sum"],
                overall_achievement,
                s["followup_required"],
                s["validated"],
                validation_pct,
            ])

        # ============================================================
        # Sheet 4: Summary by month
        # ============================================================
        ws_month.append([
            "Year",
            "Month Number",
            "Month Name",
            "Records",
            "Average QIC %",
            "Total YES",
            "Total Questions",
            "Overall Achievement %",
            "Follow-up Required Count",
            "Validated Count",
            "Validation %",
        ])

        for month_key, s in sorted(month_summary.items()):
            year, month_no, month_name = month_key
            avg_score = round(s["score_sum"] / s["records"], 2) if s["records"] else 0
            overall_achievement = round((s["yes_sum"] / s["questions_sum"]) * 100, 2) if s["questions_sum"] else 0
            validation_pct = round((s["validated"] / s["records"]) * 100, 2) if s["records"] else 0

            ws_month.append([
                year,
                month_no,
                month_name,
                s["records"],
                avg_score,
                s["yes_sum"],
                s["questions_sum"],
                overall_achievement,
                s["followup_required"],
                s["validated"],
                validation_pct,
            ])

        # ============================================================
        # Sheet 5: Data quality issues
        # ============================================================
        ws_dq.append([
            "Record ID",
            "Province",
            "District",
            "Facility",
            "HF Code",
            "Year",
            "Month",
            "QIC Meeting Date",
            "QIC Percent Score",
            "Detected Data Quality / Follow-up Issues",
            "Data Quality Issue Note",
            "Remarks",
        ])

        if dq_rows:
            for row in dq_rows:
                ws_dq.append(row)
        else:
            ws_dq.append([
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "No major data quality issues detected in selected records.",
                "",
                "",
            ])

        # ============================================================
        # Sheet 6: Question summary
        # ============================================================
        ws_question.append([
            "Question Field",
            "Question",
            "YES Count",
            "NO Count",
            "Total Responses",
            "YES %",
            "NO %",
        ])

        for field, s in question_summary.items():
            yes_count = s["yes"]
            no_count = s["no"]
            total = yes_count + no_count
            yes_pct = round((yes_count / total) * 100, 2) if total else 0
            no_pct = round((no_count / total) * 100, 2) if total else 0

            ws_question.append([
                field,
                s["question"],
                yes_count,
                no_count,
                total,
                yes_pct,
                no_pct,
            ])

        # ============================================================
        # Formatting
        # ============================================================
        for ws in wb.worksheets:
            style_worksheet(ws)

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    header_name = ws.cell(row=1, column=cell.column).value

                    if header_name and "%" in str(header_name):
                        cell.number_format = "0.00"

        timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
        filename = f"QIC_Dataset_Export_{timestamp}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

class Trainingdetails(admin.StackedInline):
    model = Training
    extra = 1

@admin.register(Trainingheader)
class TrainingAdmin(admin.ModelAdmin):
    inlines = [Trainingdetails]
    list_display = (
        "trainingname", "trainingvenue", "trainingstartdate", "trainingenddate",
        "remarks", "expectednumberofparticipant", "traingfocalpoint"
    )
    search_fields = ("trainingname",)

class MpdsrProvinceFilter(ProvinceFromFacilityFilter):
    province_path = "facilityname__districtfk__provincefk"

@admin.register(ShamsiMonth)
class ShamsiMonthAdmin(admin.ModelAdmin):
    list_display = ("id", "name")
    search_fields = ("name",)


@admin.register(ShamsiYear)
class ShamsiYearAdmin(admin.ModelAdmin):
    list_display = ("id", "year")
    search_fields = ("year",)


@admin.register(Period)
class PeriodAdmin(admin.ModelAdmin):
    list_display = ("id", "name")


@admin.register(BaselineProgress)
class BaselineProgressAdmin(admin.ModelAdmin):
    list_display = ("id", "name")


@admin.register(GregorianMonth)
class GregorianMonthAdmin(admin.ModelAdmin):
    list_display = ("id", "name")


@admin.register(GregorianYear)
class GregorianYearAdmin(admin.ModelAdmin):
    list_display = ("id", "year")

class AutoUserAdminMixin:
    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.created_by = request.user
        obj.updated_by = request.user
        super().save_model(request, obj, form, change)


def clean_excel_value(value):
    """
    Prevent Excel export errors caused by illegal characters.
    """
    if value is None:
        return ""
    value = str(value)
    return ILLEGAL_CHARACTERS_RE.sub("", value)


def safe_attr(obj, attr_path, default=""):
    """
    Safely get nested object attributes.

    Example:
    safe_attr(obj, "facilityname__districtfk__provincefk__name")
    """
    try:
        current = obj
        for attr in attr_path.split("__"):
            current = getattr(current, attr, None)
            if current is None:
                return default
        return current
    except Exception:
        return default


def percent(reviewed, reported):
    """
    Calculate review rate safely.
    """
    reviewed = reviewed or 0
    reported = reported or 0

    if reported > 0:
        return round((reviewed / reported) * 100, 1)

    return 0


@admin.action(description="Export selected MPDSR records to Excel")
def export_mpdsr_to_excel(modeladmin, request, queryset):
    queryset = queryset.select_related(
        "facilityname",
        "facilityname__districtfk",
        "facilityname__districtfk__provincefk",
        "facilityname__facilitytypefk",
        "created_by",
        "updated_by",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "MPDSR Export"

    generated_on = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")

    headers = [
        "ID",
        "Year",
        "Month",
        "Province",
        "District",
        "Health Facility",
        "Facility Type",
        "Facility Code",

        "Staff Participated in MPDSR Committee",

        "Maternal Deaths Reported",
        "Maternal Deaths Reviewed",
        "Maternal Death Review Rate (%)",
        "Legacy Cause of Maternal Deaths",

        "Maternal Death Cause Category",
        "Maternal Death Specific Cause",
        "Maternal Death Contributing Factor",
        "Maternal Death Preventability",
        "Timing of Maternal Death",
        "Place of Maternal Death",

        "Antepartum Stillbirths Reported",
        "Antepartum Stillbirths Reviewed",
        "Antepartum Stillbirth Review Rate (%)",

        "Intrapartum Stillbirths Reported",
        "Intrapartum Stillbirths Reviewed",
        "Intrapartum Stillbirth Review Rate (%)",

        "Neonatal Deaths After Live Birth Reported",
        "Neonatal Deaths After Live Birth Reviewed",
        "Neonatal Death Review Rate (%)",
        "Cause of Neonatal Death",

        "Intervention Performed",
        "Committee Recommendation",
        "Remarks",

        "Created By",
        "Created At",
        "Updated By",
        "Updated At",
    ]

    numeric_headers = {
        "ID",
        "Year",
        "Facility Code",
        "Staff Participated in MPDSR Committee",

        "Maternal Deaths Reported",
        "Maternal Deaths Reviewed",
        "Maternal Death Review Rate (%)",

        "Antepartum Stillbirths Reported",
        "Antepartum Stillbirths Reviewed",
        "Antepartum Stillbirth Review Rate (%)",

        "Intrapartum Stillbirths Reported",
        "Intrapartum Stillbirths Reviewed",
        "Intrapartum Stillbirth Review Rate (%)",

        "Neonatal Deaths After Live Birth Reported",
        "Neonatal Deaths After Live Birth Reviewed",
        "Neonatal Death Review Rate (%)",
    }

    total_columns = len(headers)
    last_column = get_column_letter(total_columns)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "MPDSR Monthly Reporting Export"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E78")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = f"Generated on: {generated_on}"
    subtitle_cell.font = Font(italic=True)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    header_row = 4

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = header_row + 1

    for obj in queryset:
        created_by = ""
        if obj.created_by:
            created_by = obj.created_by.get_full_name() or obj.created_by.username

        updated_by = ""
        if obj.updated_by:
            updated_by = obj.updated_by.get_full_name() or obj.updated_by.username

        created_at = ""
        if obj.created_at:
            created_at = timezone.localtime(obj.created_at).strftime("%Y-%m-%d %H:%M")

        updated_at = ""
        if obj.updated_at:
            updated_at = timezone.localtime(obj.updated_at).strftime("%Y-%m-%d %H:%M")

        row = [
            obj.id,
            obj.yearmpdsr,
            obj.get_monthmpdsr_display(),

            safe_attr(obj, "facilityname__districtfk__provincefk__name"),
            safe_attr(obj, "facilityname__districtfk__name"),
            safe_attr(obj, "facilityname__name"),
            safe_attr(obj, "facilityname__facilitytypefk__name"),
            safe_attr(obj, "facilityname__hfcode"),

            obj.n_mpdsrcommittee,

            obj.n_maternaldeathreported,
            obj.n_maternaldeathreviewed,
            percent(obj.n_maternaldeathreviewed, obj.n_maternaldeathreported),
            obj.causeofmaternaldeaths_m,

            obj.get_maternal_death_cause_category_display() if obj.maternal_death_cause_category else "",
            obj.get_maternal_death_specific_cause_display() if obj.maternal_death_specific_cause else "",
            obj.get_maternal_death_contributing_factor_display() if obj.maternal_death_contributing_factor else "",
            obj.get_maternal_death_preventability_display() if obj.maternal_death_preventability else "",
            obj.get_maternal_death_timing_display() if obj.maternal_death_timing else "",
            obj.get_maternal_death_place_display() if obj.maternal_death_place else "",

            obj.nastillbirthreportedreported,
            obj.nastillbirthreportedreviewed,
            percent(obj.nastillbirthreportedreviewed, obj.nastillbirthreportedreported),

            obj.nistillbirthreported,
            obj.nistillbirthreviewed,
            percent(obj.nistillbirthreviewed, obj.nistillbirthreported),

            obj.nndeath_afteralivebirth_reported,
            obj.nndeath_afteralivebirth_reviewed,
            percent(obj.nndeath_afteralivebirth_reviewed, obj.nndeath_afteralivebirth_reported),
            obj.causeofneonataldeath_n,

            obj.interventionperformed,
            obj.recfromMPDSRcommittee,
            obj.remarks,

            created_by,
            created_at,
            updated_by,
            updated_at,
        ]

        for col_num, value in enumerate(row, start=1):
            header_name = headers[col_num - 1]
            cell = ws.cell(row=row_num, column=col_num)

            if header_name in numeric_headers:
                if value in [None, ""]:
                    cell.value = None
                else:
                    try:
                        if "Rate" in header_name:
                            cell.value = float(value)
                            cell.number_format = "0.0"
                        else:
                            cell.value = int(value)
                            cell.number_format = "0"
                    except (TypeError, ValueError):
                        cell.value = None
            else:
                cell.value = clean_excel_value(value)

            cell.alignment = Alignment(vertical="top", wrap_text=True)

        row_num += 1

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row_cells in ws.iter_rows(
        min_row=header_row,
        max_row=ws.max_row,
        min_col=1,
        max_col=total_columns,
    ):
        for cell in row_cells:
            cell.border = thin_border

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:{last_column}{ws.max_row}"

    for col_num, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_num)

        if header in [
            "Legacy Cause of Maternal Deaths",
            "Cause of Neonatal Death",
            "Intervention Performed",
            "Committee Recommendation",
            "Remarks",
        ]:
            ws.column_dimensions[column_letter].width = 35
        elif header in [
            "Health Facility",
            "Maternal Death Specific Cause",
            "Maternal Death Contributing Factor",
        ]:
            ws.column_dimensions[column_letter].width = 28
        elif "Rate" in header:
            ws.column_dimensions[column_letter].width = 20
        else:
            ws.column_dimensions[column_letter].width = 16

    filename = f"mpdsr_export_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@admin.register(Mpdsr)
class mpdsrshow(ProvinceRestrictedAdminMixin, AutoUserAdminMixin, admin.ModelAdmin):
    list_display = [
        "id",
        "yearmpdsr",
        "monthmpdsr",
        "facilityname",
        "n_mpdsrcommittee",
        "n_maternaldeathreported",
        "n_maternaldeathreviewed",
        "maternal_death_cause_category_display",
        "maternal_death_specific_cause_display",
        "nastillbirthreportedreported",
        "nastillbirthreportedreviewed",
        "nistillbirthreported",
        "nistillbirthreviewed",
        "nndeath_afteralivebirth_reported",
        "nndeath_afteralivebirth_reviewed",
    ]

    list_filter = [
        MpdsrProvinceFilter,
        "monthmpdsr",
        "maternal_death_cause_category",
        "maternal_death_specific_cause",
        "maternal_death_contributing_factor",
        "maternal_death_preventability",
        "maternal_death_timing",
        "maternal_death_place",
    ]

    search_fields = (
        "facilityname__name",
        "facilityname__districtfk__name",
        "facilityname__districtfk__provincefk__name",
        "causeofmaternaldeaths_m",
        "causeofneonataldeath_n",
        "interventionperformed",
        "recfromMPDSRcommittee",
    )

    fieldsets = (
        ("Basic Information", {
            "fields": (
                "yearmpdsr",
                "monthmpdsr",
                "facilityname",
            )
        }),
        ("MPDSR Committee", {
            "fields": (
                "n_mpdsrcommittee",
            )
        }),
        ("Maternal Death Summary", {
            "fields": (
                "n_maternaldeathreported",
                "n_maternaldeathreviewed",
            )
        }),
        ("Structured Maternal Death Review (For New Records)", {
            "fields": (
                "maternal_death_cause_category",
                "maternal_death_specific_cause",
                "maternal_death_contributing_factor",
                "maternal_death_preventability",
                "maternal_death_timing",
                "maternal_death_place",
            ),
            "description": (
                "Use these structured dropdown fields for new maternal death records. "
                "Leave them blank when no maternal death is reported."
            ),
        }),
        ("Legacy Maternal Death Cause", {
            "fields": (
                "causeofmaternaldeaths_m",
            ),
            "description": (
                "Existing free-text field kept for historical data. "
                "For new records, prefer the structured maternal death review fields above."
            ),
            "classes": ("collapse",),
        }),
        ("Stillbirth and Neonatal Death Summary", {
            "fields": (
                "nastillbirthreportedreported",
                "nastillbirthreportedreviewed",
                "nistillbirthreported",
                "nistillbirthreviewed",
                "nndeath_afteralivebirth_reported",
                "nndeath_afteralivebirth_reviewed",
                "causeofneonataldeath_n",
            )
        }),
        ("Response and Recommendations", {
            "fields": (
                "interventionperformed",
                "recfromMPDSRcommittee",
                "remarks",
            )
        }),
    )

    actions = [export_mpdsr_to_excel]

    ordering = (
        "-yearmpdsr",
        "-monthmpdsr",
        "facilityname",
    )

    list_per_page = 50

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related(
            "facilityname",
            "facilityname__districtfk",
            "facilityname__districtfk__provincefk",
            "facilityname__facilitytypefk",
            "created_by",
            "updated_by",
        )

    def province_filter_kwargs(self, request):
        return {"facilityname__districtfk__provincefk": user_province(request)}

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "facilityname" and not request.user.is_superuser:
            prov = user_province(request)
            kwargs["queryset"] = Facility.objects.filter(
                districtfk__provincefk=prov
            ) if prov else Facility.objects.none()

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    @admin.display(description="Maternal cause category")
    def maternal_death_cause_category_display(self, obj):
        return obj.get_maternal_death_cause_category_display() if obj.maternal_death_cause_category else "-"

    @admin.display(description="Maternal specific cause")
    def maternal_death_specific_cause_display(self, obj):
        return obj.get_maternal_death_specific_cause_display() if obj.maternal_death_specific_cause else "-"

# ============================================================
# User admin with Profile inline
# ============================================================
User = get_user_model()

class UserProfileInline(admin.StackedInline):
    model = UserProfile
    fk_name = "user"
    can_delete = False
    extra = 0
    max_num = 1

class UserAdmin(BaseUserAdmin):
    inlines = (UserProfileInline,)

admin.site.unregister(User)
admin.site.register(User, UserAdmin)
admin.site.register(Score)
admin.site.register(Criteria)
admin.site.register(Section)
admin.site.register(Standards)
admin.site.register(Area)
admin.site.register(Assessmenttype)
admin.site.register(Province)
admin.site.register(District)
admin.site.register(Facilitytype)
admin.site.register(Implementor)
admin.site.register(Assessor)
admin.site.register(Participationtype)
admin.site.register(Participantposition)
admin.site.register(Participanteducation)
admin.site.register(Position)
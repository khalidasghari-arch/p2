import re
from django.contrib import admin
from django import forms
from django.core.exceptions import ValidationError
from .models import (
    ThematicMentorship, MentorshipTopics, MenteeTopicStatus,
    Mentorshipvisit, Mentorshipdetails, Staff, MentorshipDashboard,
)
from hiva.admin_utils import ProvinceRestrictedAdminMixin, user_province
from django.utils.html import format_html
from mentorship.recommender import recommend_next_for_staff_in_facility
from django.db.models import Count, Q
from django.utils.safestring import mark_safe
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.db.models import Count, Sum, Case, When, Value, IntegerField, Min, F, Q
from django.db.models.functions import TruncMonth
from django.template.response import TemplateResponse
from django.http import HttpResponse
from django.utils import timezone
from django.utils.html import format_html
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================
# Helpers
# =====================================================

def _prov_id(request):
    """
    Returns province id (int) safely.
    Works whether user_province returns Province object or id.
    """
    prov = user_province(request)
    if prov is None:
        return None
    return getattr(prov, "id", prov)

# =====================================================
# BASIC ADMINS
# =====================================================

@admin.register(ThematicMentorship)
class ThematicMentorshipAdmin(admin.ModelAdmin):
    list_display = ("id", "name", "shortname", "hqip_area")
    search_fields = ("name", "shortname")

@admin.register(MentorshipTopics)
class MentorshipTopicsAdmin(admin.ModelAdmin):
    list_display = ("id", "thematicfk", "shortname", "name", "nameeng", "namedari", "namepashto")
    list_filter = ("thematicfk",)
    search_fields = ("name", "shortname", "namedari", "namepashto", "nameeng")

# =====================================================
# STAFF ADMIN (Province restriction + popup facility prefill)
# =====================================================

@admin.register(Staff)
class StaffAdmin(admin.ModelAdmin):
    list_display = ("id", "hfname", "tazkiranumber", "firstname", "lastname", "fathername", "position", "gender", "status")
    list_filter = ("hfname", "status", "position")
    search_fields = ("firstname", "lastname", "tazkiranumber")

    def _facility_id_from_popup_context(self, request):
        """
        Priority:
          1) ?facility=<id> if present
          2) if popup opened from Mentorshipvisit change page, parse HTTP_REFERER
        """
        facility_id = request.GET.get("facility")
        if facility_id:
            return facility_id

        if request.GET.get("_popup") != "1":
            return None

        ref = request.META.get("HTTP_REFERER", "")
        m = re.search(r"/admin/mentorship/mentorshipvisit/(\d+)/change/?", ref)
        if not m:
            return None

        try:
            visit = Mentorshipvisit.objects.only("facilityfk_id").get(pk=m.group(1))
            return str(visit.facilityfk_id)
        except Mentorshipvisit.DoesNotExist:
            return None

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs

        prov_id = _prov_id(request)
        if prov_id is None:
            return qs.none()

        return qs.filter(hfname__districtfk__provincefk_id=prov_id)

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        facility_id = self._facility_id_from_popup_context(request)
        if facility_id:
            initial["hfname"] = facility_id
        return initial

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "hfname" and not request.user.is_superuser:
            Facility = db_field.remote_field.model

            facility_id = self._facility_id_from_popup_context(request)

            if facility_id:
                kwargs["queryset"] = Facility.objects.filter(pk=facility_id)
                field = super().formfield_for_foreignkey(db_field, request, **kwargs)
                field.initial = facility_id
                return field

            prov_id = _prov_id(request)
            kwargs["queryset"] = Facility.objects.filter(
                districtfk__provincefk_id=prov_id
            ).order_by("name") if prov_id else Facility.objects.none()

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

# =====================================================
# INLINE FORM VALIDATION:
# - Only ONE of LS/PC/MC
# (REMOVED: topic-thematic matching rule)
# =====================================================

# =====================================================
# INLINE FORM VALIDATION:
# - Required fields
# - Only ONE of LS/PC/MC
# =====================================================

class MentorshipdetailsInlineForm(forms.ModelForm):
    class Meta:
        model = Mentorshipdetails
        fields = "__all__"

    def clean(self):
        cleaned = super().clean()

        # Required FK fields
        required_fields = [
            "mentor",
            "menteename",
            "thematicname",
            "topicname",
        ]

        for field in required_fields:
            if not cleaned.get(field):
                self.add_error(field, "This field is required.")

        # Only ONE of LS / PC / MC
        selected = sum([
            bool(cleaned.get("ls")),
            bool(cleaned.get("pc")),
            bool(cleaned.get("mc")),
        ])

        if selected == 0:
            raise ValidationError("Select ONE of (LS, PC, MC).")

        if selected > 1:
            raise ValidationError("Only ONE of (LS, PC, MC) can be selected.")

        return cleaned

# =====================================================
# DETAILS INLINE:
# - Mentee filtered by visit facility
# - Mentor filtered by user province
# - Topic dropdown shows ALL topics (NO FILTERING)
# =====================================================

class MentorshipdetailsInline(admin.TabularInline):
    model = Mentorshipdetails
    form = MentorshipdetailsInlineForm
    extra = 1
    min_num = 1
    validate_min = True

    def get_extra(self, request, obj=None, **kwargs):
        return 1 if obj else 0

    def has_add_permission(self, request, obj=None):
        if obj is None:
            return False
        return super().has_add_permission(request, obj=obj)

    def get_formset(self, request, obj=None, **kwargs):
        request._mentorship_parent_obj = obj
        return super().get_formset(request, obj, **kwargs)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        parent_obj = getattr(request, "_mentorship_parent_obj", None)

        if db_field.name == "menteename":
            if parent_obj:
                kwargs["queryset"] = Staff.objects.filter(
                    hfname=parent_obj.facilityfk
                ).order_by("firstname", "lastname")
            else:
                kwargs["queryset"] = Staff.objects.none()

        if db_field.name == "mentor" and not request.user.is_superuser:
            prov_id = _prov_id(request)
            Assessor = db_field.remote_field.model
            kwargs["queryset"] = Assessor.objects.filter(
                province_id=prov_id
            ).order_by("id") if prov_id else Assessor.objects.none()

        # ✅ IMPORTANT: Always show ALL topics, no filtering, no JS, no endpoint
        if db_field.name == "topicname":
            kwargs["queryset"] = MentorshipTopics.objects.all().order_by(
                "thematicfk_id", "shortname", "name"
            )

        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    class Media:
            js = ("admin/js/topic_filter.js",)

# =====================================================
# MENTORSHIP VISIT ADMIN:
# - Province restriction (Mixin)
# - Facility dropdown restricted by province
# - NO custom URLs
# - NO JS Media
# =====================================================

class MentorshipDashboardMixin:
    """
    Adds: /admin/mentorship/mentorshipvisit/facility-dashboard/<facility_id>/
    """

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "facility-dashboard/<int:facility_id>/",
                self.admin_site.admin_view(self.facility_dashboard_view),
                name="facility_mentorship_dashboard",
            ),
        ]
        return custom_urls + urls

    def facility_dashboard_view(self, request, facility_id):
        # Active mentees in this facility
        mentees = Staff.objects.filter(hfname_id=facility_id, status=True)
        total_mentees = mentees.count()

        # All topics
        total_topics = MentorshipTopics.objects.count()

        # All statuses for mentees in this facility
        statuses = MenteeTopicStatus.objects.filter(mentee__hfname_id=facility_id)

        competent_count = statuses.filter(status="COMPETENT").count()

        # Overall competency rate (competent topics / total possible)
        competency_rate = 0
        total_possible = total_mentees * total_topics
        if total_possible > 0:
            competency_rate = round((competent_count / total_possible) * 100)

        # Thematic breakdown
        thematic_stats = []
        thematics = ThematicMentorship.objects.all().order_by("name")

        for th in thematics:
            topic_count = MentorshipTopics.objects.filter(thematicfk=th).count()

            competent = MenteeTopicStatus.objects.filter(
                mentee__hfname_id=facility_id,
                topic__thematicfk=th,
                status="COMPETENT",
            ).count()

            total_possible_th = topic_count * total_mentees
            percent = 0
            if total_possible_th > 0:
                percent = round((competent / total_possible_th) * 100)

            thematic_stats.append({
                "name": th.name,
                "percent": percent,
            })

        # Escalation: LS >= 4 and not competent
        escalations = MenteeTopicStatus.objects.filter(
            mentee__hfname_id=facility_id,
            consecutive_ls__gte=4
        ).exclude(status="COMPETENT").select_related("mentee", "topic")

        context = dict(
            self.admin_site.each_context(request),
            title="Facility Mentorship Dashboard",
            total_mentees=total_mentees,
            competency_rate=competency_rate,
            thematic_stats=thematic_stats,
            escalations=escalations,
            facility_id=facility_id,
        )
        return TemplateResponse(request, "admin/mentorship/facility_dashboard.html", context)
    
def export_selected_mentorship_large(modeladmin, request, queryset):
    """
    Large-data safe export:
    - Uses write_only workbook (low memory)
    - Uses values_list + iterator (fast)
    - Exports one row per Mentorshipdetails line

    Calculation rule:
    One mentorship visit = unique Visit Date + Mentor.
    Same Visit Date + same Mentor = counted once as 1.
    Repeated rows for the same Visit Date + Mentor = 0.
    Same Visit Date + different Mentor = counted separately as 1.
    """

    # If user selects visits: export details under those visits
    visit_ids = list(queryset.values_list("id", flat=True))

    if not visit_ids:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Mentorship")
        ws.append(["No data selected"])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="mentorship_empty.xlsx"'
        wb.save(response)
        return response

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Mentorship")

    headers = [
        "Detail ID",
        "Visit ID",
        "visit by mentors",
        "Visit Date",
        "Month",
        "Visit Round",
        "Start Time",
        "End Time",
        "Province",
        "District",
        "Facility",
        "Facility Code",
        "Mentee",
        "Mentee Position",
        "Mentee Gender",
        "Mentee Tazkira",
        "Mentor",
        "Thematic Area",
        "Topic Code",
        "LS",
        "PC",
        "MC",
    ]
    ws.append(headers)

    # This set tracks unique Visit Date + Mentor combinations
    seen_visit_date_mentor = set()

    rows = (
        Mentorshipdetails.objects
        .filter(mentorshipvistfk_id__in=visit_ids)
        .values_list(
            "id",
            "mentorshipvistfk_id",
            "mentorshipvistfk__visitdate",
            "mentorshipvistfk__visitround",
            "mentorshipvistfk__mentorshipstarttime",
            "mentorshipvistfk__mentorshipendtime",
            "mentorshipvistfk__facilityfk__districtfk__provincefk__name",
            "mentorshipvistfk__facilityfk__districtfk__name",
            "mentorshipvistfk__facilityfk__name",
            "mentorshipvistfk__facilityfk__hfcode",
            "menteename__firstname",
            "menteename__lastname",
            "menteename__position__name",
            "menteename__gender",
            "menteename__tazkiranumber",
            "mentor__name",
            "thematicname__name",
            "topicname__name",
            "ls",
            "pc",
            "mc",
        )
        .order_by(
            "mentorshipvistfk__visitdate",
            "mentor__name",
            "mentorshipvistfk_id",
            "id",
        )
        .iterator(chunk_size=2000)
    )

    for r in rows:
        (
            detail_id,
            visit_id,
            visit_date,
            visit_round,
            start_time,
            end_time,

            province_name,
            district_name,
            facility_name,
            facility_code,

            mentee_first,
            mentee_last,
            mentee_position,
            mentee_gender,
            mentee_tazkira,

            mentor_name,
            thematic_name,
            topic_code,

            ls,
            pc,
            mc,
        ) = r

        # Mentee full name
        mentee_full = " ".join(
            [x for x in [mentee_first, mentee_last] if x]
        ) if (mentee_first or mentee_last) else ""

        # Gender text
        gender_txt = ""
        if mentee_gender is True:
            gender_txt = "Female"
        elif mentee_gender is False:
            gender_txt = "Male"

        # Month name from visit date
        month_name = ""
        if visit_date:
            try:
                month_name = visit_date.strftime("%B")
            except Exception:
                month_name = ""

        # Normalize visit date in case it includes hidden time
        clean_visit_date = visit_date
        if visit_date and hasattr(visit_date, "date") and callable(visit_date.date):
            clean_visit_date = visit_date.date()

        # Normalize mentor name
        clean_mentor_name = (mentor_name or "").strip().lower()

        # Unique key: Visit Date + Mentor only
        unique_key = (clean_visit_date, clean_mentor_name)

        if clean_visit_date and clean_mentor_name:
            if unique_key not in seen_visit_date_mentor:
                visit_by_mentors = 1
                seen_visit_date_mentor.add(unique_key)
            else:
                visit_by_mentors = 0
        else:
            visit_by_mentors = 0

        ws.append([
            detail_id,
            visit_id,
            visit_by_mentors,
            visit_date,
            month_name,
            visit_round,
            start_time,
            end_time,
            province_name or "",
            district_name or "",
            facility_name or "",
            facility_code or "",
            mentee_full,
            mentee_position or "",
            gender_txt,
            mentee_tazkira or "",
            mentor_name or "",
            thematic_name or "",
            topic_code or "",
            1 if ls else 0,
            1 if pc else 0,
            1 if mc else 0,
        ])

    filename = f"mentorship_export_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


export_selected_mentorship_large.short_description = (
    "Export selected visits with mentor visit count"
)
# =====================================================
# VISIT ADMIN FORM (ADMIN-LEVEL REQUIRED VALIDATION)
# =====================================================

class MentorshipVisitAdminForm(forms.ModelForm):
    class Meta:
        model = Mentorshipvisit
        fields = "__all__"

    def clean(self):
        cleaned = super().clean()

        required_fields = [
            "facilityfk",
            "visitdate",
            "visitround",
            "mentorshipstarttime",
            "mentorshipendtime",
        ]

        for field in required_fields:
            if not cleaned.get(field):
                self.add_error(field, "This field is required.")

        # Validate time logic
        start = cleaned.get("mentorshipstarttime")
        end = cleaned.get("mentorshipendtime")

        if start and end and end <= start:
            raise ValidationError("End time must be after start time.")

        return cleaned
    
@admin.register(Mentorshipvisit)
class MentorshipvisitAdmin(ProvinceRestrictedAdminMixin, 
MentorshipDashboardMixin, admin.ModelAdmin):
    form = MentorshipVisitAdminForm   # 👈 ADD THIS
    actions = [export_selected_mentorship_large]
    list_display = (
        "facilityfk",
        "visitdate",
        #"visitround",
        "mentorshipstarttime",
        "mentorshipendtime",
        "get_mentors",
        "ls_count",
        "pc_count",
        "mc_count",
        "mentees_count",
        "facility_dashboard_btn",
        "ai_recommendation",
        "id",
    )

    #list_filter = ("visitdate", "facilityfk")
    search_fields = ("facilityfk__name", "facilityfk__hfcode")
    readonly_fields = ("ai_recommendation",)
    list_per_page = 10
    inlines = (MentorshipdetailsInline,)

    @admin.display(description="Dashboard")
    def facility_dashboard_btn(self, obj):
        url = reverse("admin:facility_mentorship_dashboard", args=[obj.facilityfk_id])
        return format_html('<a class="button" href="{}">Dashboard</a>', url)

    # -------------------------------------------------
    # AI RECOMMENDATION (FIXED)
    # -------------------------------------------------
    @admin.display(description="Recommendation")
    def ai_recommendation(self, obj):

        # Get unique mentees in this visit
        mentee_ids = (
            obj.items
            .values_list("menteename_id", flat=True)
            .distinct()
        )

        if not mentee_ids:
            return "—"

        rows = []

        for mentee_id in mentee_ids:

            rec = recommend_next_for_staff_in_facility(
                staff_id=mentee_id,
                facility_id=obj.facilityfk_id
            )

            topic = rec["topic"]
            track = rec["track"]
            session = rec["session_type"]
            support = rec["support_flag"]

            # --- progress calculation ---
            total_topics = MentorshipTopics.objects.filter(
                thematicfk__name=track
            ).count()

            competent_count = MenteeTopicStatus.objects.filter(
                mentee_id=mentee_id,
                status="COMPETENT"
            ).count()

            progress = 0
            if total_topics > 0:
                progress = round((competent_count / total_topics) * 100)

            # --- session badge ---
            session_badge = ""
            if session == "LS":
                session_badge = '<span style="background:#1f77b4;color:white;padding:2px 6px;border-radius:4px;">LS</span>'
            elif session == "PC":
                session_badge = '<span style="background:#ff7f0e;color:white;padding:2px 6px;border-radius:4px;">PC</span>'
            elif session == "MC":
                session_badge = '<span style="background:#2ca02c;color:white;padding:2px 6px;border-radius:4px;">MC</span>'

            support_badge = ""
            if support:
                support_badge = '<span style="background:#d62728;color:white;padding:2px 6px;border-radius:4px;">YES</span>'
            else:
                support_badge = '<span style="background:#2ca02c;color:white;padding:2px 6px;border-radius:4px;">NO</span>'

            progress_color = "#d62728"
            if progress >= 70:
                progress_color = "#2ca02c"
            elif progress >= 40:
                progress_color = "#ff7f0e"

            progress_badge = f'<span style="background:{progress_color};color:white;padding:2px 6px;border-radius:4px;">{progress}%</span>'

            mentee_name = str(
                obj.items.filter(menteename_id=mentee_id).first().menteename
            )

            rows.append(f"""
                <tr>
                    <td><strong>{mentee_name}</strong></td>
                    <td>{track or '-'}</td>
                    <td>{topic.name if topic else '-'}</td>
                    <td>{session_badge}</td>
                    <td>{support_badge}</td>
                    <td>{progress_badge}</td>
                </tr>
            """)

        table = f"""
            <table style="border-collapse:collapse;width:100%;">
                <thead>
                    <tr style="background:#f5f5f5;">
                        <th style="padding:4px;border:1px solid #ddd;">Mentee</th>
                        <th style="padding:4px;border:1px solid #ddd;">Track</th>
                        <th style="padding:4px;border:1px solid #ddd;">Next Topic</th>
                        <th style="padding:4px;border:1px solid #ddd;">Session</th>
                        <th style="padding:4px;border:1px solid #ddd;">Support</th>
                        <th style="padding:4px;border:1px solid #ddd;">Progress</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(rows)}
                </tbody>
            </table>
        """

        return mark_safe(table)

    ai_recommendation.short_description = "AI Recommendation"

    # -------------------------------------------------
    # CLINICAL MENTOR DISPLAY
    # -------------------------------------------------
    @admin.display(description="Clinical Mentor")
    def get_mentors(self, obj):
        mentors = (
            obj.items
            .select_related("mentor")
            .values_list("mentor__name", flat=True)
            .distinct()
        )
        return ", ".join([m for m in mentors if m]) if mentors else "-"

    # -------------------------------------------------
    # LS / PC / MC COUNTS
    # -------------------------------------------------
    @admin.display(description="Tot-LS")
    def ls_count(self, obj):
        return obj.items.filter(ls=True).count()

    @admin.display(description="Tot-PC")
    def pc_count(self, obj):
        return obj.items.filter(pc=True).count()

    @admin.display(description="Tot-MC")
    def mc_count(self, obj):
        return obj.items.filter(mc=True).count()

    # -------------------------------------------------
    # DISTINCT COUNTS
    # -------------------------------------------------
    @admin.display(description="Tot-Mentee")
    def mentees_count(self, obj):
        return (
            obj.items
            .values("menteename")
            .exclude(menteename__isnull=True)
            .distinct()
            .count()
        )

    @admin.display(description="Total-Thematic Areas")
    def thematics_count(self, obj):
        return (
            obj.items
            .values("thematicname")
            .exclude(thematicname__isnull=True)
            .distinct()
            .count()
        )

    @admin.display(description="Total-Topics")
    def topics_count(self, obj):
        return (
            obj.items
            .values("topicname")
            .exclude(topicname__isnull=True)
            .distinct()
            .count()
        )

    # -------------------------------------------------
    # PERFORMANCE OPTIMIZATION (SINGLE VERSION)
    # -------------------------------------------------
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.prefetch_related(
            "items",
            "items__mentor",
            "items__menteename",
            "items__thematicname",
            "items__topicname",
        )

    # -------------------------------------------------
    # FIELDSETS
    # -------------------------------------------------
    fieldsets = (
        ("Mentorship Visit", {
            "fields": (
                "facilityfk",
                "visitdate",
                "visitround",
                "mentorshipstarttime",
                "mentorshipendtime",
                "ai_recommendation",
            )
        }),
    )

    # -------------------------------------------------
    # PROVINCE RESTRICTION
    # -------------------------------------------------
    def province_filter_kwargs(self, request):
        prov_id = _prov_id(request)
        return {"facilityfk__districtfk__provincefk_id": prov_id} if prov_id else {"pk__in": []}

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "facilityfk" and not request.user.is_superuser:
            Facility = db_field.remote_field.model
            prov_id = _prov_id(request)
            kwargs["queryset"] = Facility.objects.filter(
                districtfk__provincefk_id=prov_id
            ).order_by("name") if prov_id else Facility.objects.none()

        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    # -------------------------------------------------
    # DYNAMIC TOPIC FILTER ENDPOINT
    # -------------------------------------------------

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "get-topics/<int:thematic_id>/",
                self.admin_site.admin_view(self.get_topics_by_thematic),
                name="mentorship_get_topics",
            ),
        ]
        return custom_urls + urls

    def get_topics_by_thematic(self, request, thematic_id):
        from django.http import JsonResponse

        topics = MentorshipTopics.objects.filter(
            thematicfk_id=thematic_id
        ).values("id", "name").order_by("name")

        return JsonResponse(list(topics), safe=False)
@admin.register(MentorshipDashboard)
class MentorshipDashboardAdmin(ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    change_list_template = "admin/mentorship/dashboard.html"

    # ------------------------------------------------------------
    # Dashboard permissions
    # ------------------------------------------------------------
    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return request.user.is_active and request.user.is_staff

    def get_model_perms(self, request):
        if self.has_view_permission(request):
            return {"view": True}
        return {}

    # ------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------
    def _bool_sum(self, field_name):
        return Sum(
            Case(
                When(**{field_name: True}, then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            )
        )

    def changelist_view(self, request, extra_context=None):
        data = self._build_dashboard_data(request)

        if request.GET.get("export") == "1":
            return self._export_dashboard_excel(data)

        context = {
            **self.admin_site.each_context(request),
            "title": "Mentorship Dashboard",
            **data,
        }

        return TemplateResponse(
            request,
            self.change_list_template,
            context,
        )

    def _base_querysets(self, request):
        visits_qs = Mentorshipvisit.objects.select_related(
            "facilityfk",
            "facilityfk__districtfk",
            "facilityfk__districtfk__provincefk",
        )

        details_qs = Mentorshipdetails.objects.select_related(
            "mentorshipvistfk",
            "mentorshipvistfk__facilityfk",
            "mentorshipvistfk__facilityfk__districtfk",
            "mentorshipvistfk__facilityfk__districtfk__provincefk",
            "menteename",
            "menteename__hfname",
            "menteename__position",
            "thematicname",
            "topicname",
            "mentor",
        ).filter(
            mentorshipvistfk__isnull=False
        )

        prov_id = _prov_id(request)

        if prov_id and not request.user.is_superuser:
            visits_qs = visits_qs.filter(
                facilityfk__districtfk__provincefk_id=prov_id
            )
            details_qs = details_qs.filter(
                mentorshipvistfk__facilityfk__districtfk__provincefk_id=prov_id
            )

        return visits_qs, details_qs

    def _apply_filters(self, request, visits_qs, details_qs):
        date_from = request.GET.get("date_from", "").strip()
        date_to = request.GET.get("date_to", "").strip()
        province_id = request.GET.get("province", "").strip()
        facility_id = request.GET.get("facility", "").strip()
        mentor_id = request.GET.get("mentor", "").strip()
        thematic_id = request.GET.get("thematic", "").strip()
        visit_round = request.GET.get("visit_round", "").strip()

        if date_from:
            visits_qs = visits_qs.filter(visitdate__gte=date_from)
            details_qs = details_qs.filter(mentorshipvistfk__visitdate__gte=date_from)

        if date_to:
            visits_qs = visits_qs.filter(visitdate__lte=date_to)
            details_qs = details_qs.filter(mentorshipvistfk__visitdate__lte=date_to)

        if province_id:
            visits_qs = visits_qs.filter(
                facilityfk__districtfk__provincefk_id=province_id
            )
            details_qs = details_qs.filter(
                mentorshipvistfk__facilityfk__districtfk__provincefk_id=province_id
            )

        if facility_id:
            visits_qs = visits_qs.filter(facilityfk_id=facility_id)
            details_qs = details_qs.filter(mentorshipvistfk__facilityfk_id=facility_id)

        if mentor_id:
            details_qs = details_qs.filter(mentor_id=mentor_id)
            visits_qs = visits_qs.filter(items__mentor_id=mentor_id).distinct()

        if thematic_id:
            details_qs = details_qs.filter(thematicname_id=thematic_id)
            visits_qs = visits_qs.filter(items__thematicname_id=thematic_id).distinct()

        if visit_round:
            visits_qs = visits_qs.filter(visitround=visit_round)
            details_qs = details_qs.filter(mentorshipvistfk__visitround=visit_round)

        filters = {
            "date_from": date_from,
            "date_to": date_to,
            "province": province_id,
            "facility": facility_id,
            "mentor": mentor_id,
            "thematic": thematic_id,
            "visit_round": visit_round,
        }

        return visits_qs, details_qs, filters

    def _build_dashboard_data(self, request):
        visits_qs, details_qs = self._base_querysets(request)
        option_visits_qs, option_details_qs = self._base_querysets(request)

        visits_qs, details_qs, filters = self._apply_filters(
            request,
            visits_qs,
            details_qs,
        )

        # ------------------------------------------------------------
        # Filter options
        # ------------------------------------------------------------
        province_options = list(
            option_visits_qs.values(
                province_id=F("facilityfk__districtfk__provincefk_id"),
                province=F("facilityfk__districtfk__provincefk__name"),
            )
            .exclude(province_id__isnull=True)
            .distinct()
            .order_by("province")
        )

        facility_options = list(
            option_visits_qs.values(
                facility_id=F("facilityfk_id"),
                facility=F("facilityfk__name"),
            )
            .exclude(facility_id__isnull=True)
            .distinct()
            .order_by("facility")
        )

        mentor_options = [
        {
            "mentor_id": row["mentor_id"],
            "mentor": row["mentor__name"] or "",
        }
        for row in (
            option_details_qs
            .exclude(mentor_id__isnull=True)
            .values("mentor_id", "mentor__name")
            .distinct()
            .order_by("mentor__name")
            )
        ]

        thematic_options = [
        {
            "thematic_id": row["id"],
            "thematic": row["name"] or "",
        }
        for row in (
            ThematicMentorship.objects
            .values("id", "name")
            .order_by("name")
            )
        ]

        visit_round_options = list(
            option_visits_qs.values_list("visitround", flat=True)
            .exclude(visitround__isnull=True)
            .distinct()
            .order_by("visitround")
        )

        # ------------------------------------------------------------
        # KPI cards
        # ------------------------------------------------------------
        kpis = details_qs.aggregate(
            total_records=Count("id"),
            distinct_mentees=Count("menteename", distinct=True),
            distinct_visits=Count("mentorshipvistfk", distinct=True),
            distinct_facilities=Count("mentorshipvistfk__facilityfk", distinct=True),
            distinct_mentors=Count("mentor", distinct=True),
            distinct_thematics=Count("thematicname", distinct=True),
            distinct_topics=Count("topicname", distinct=True),
            total_ls=self._bool_sum("ls"),
            total_pc=self._bool_sum("pc"),
            total_mc=self._bool_sum("mc"),
        )

        for key in kpis:
            kpis[key] = int(kpis[key] or 0)

        kpis["competency_instances"] = kpis["total_pc"] + kpis["total_mc"]
        kpis["competency_minus_ls"] = kpis["competency_instances"] - kpis["total_ls"]

        mentor_visit_keys = set(
            details_qs.exclude(mentor_id__isnull=True)
            .exclude(mentorshipvistfk__visitdate__isnull=True)
            .values_list("mentorshipvistfk__visitdate", "mentor_id")
        )
        kpis["mentor_visit_instances"] = len(mentor_visit_keys)

        # ------------------------------------------------------------
        # Province summary
        # ------------------------------------------------------------
        province_rows = list(
            details_qs.values(
                province=F("mentorshipvistfk__facilityfk__districtfk__provincefk__name"),
            )
            .annotate(
                visits=Count("mentorshipvistfk", distinct=True),
                mentor_visit_instances=Count("mentor", distinct=True),
                facilities=Count("mentorshipvistfk__facilityfk", distinct=True),
                mentees=Count("menteename", distinct=True),
                mentors=Count("mentor", distinct=True),
                thematics=Count("thematicname", distinct=True),
                topics=Count("topicname", distinct=True),
                ls=self._bool_sum("ls"),
                pc=self._bool_sum("pc"),
                mc=self._bool_sum("mc"),
                records=Count("id"),
            )
            .order_by("province")
        )

        max_province_value = max(
            [
                max(
                    int(r["ls"] or 0),
                    int(r["pc"] or 0),
                    int(r["mc"] or 0),
                )
                for r in province_rows
            ] or [1]
        )

        for r in province_rows:
            r["ls"] = int(r["ls"] or 0)
            r["pc"] = int(r["pc"] or 0)
            r["mc"] = int(r["mc"] or 0)
            r["competency_instances"] = r["pc"] + r["mc"]
            r["competency_minus_ls"] = r["competency_instances"] - r["ls"]

            r["ls_width"] = round((r["ls"] / max_province_value) * 100, 1) if max_province_value else 0
            r["pc_width"] = round((r["pc"] / max_province_value) * 100, 1) if max_province_value else 0
            r["mc_width"] = round((r["mc"] / max_province_value) * 100, 1) if max_province_value else 0

            if r["mc"] > r["ls"]:
                r["interpretation"] = "MC is higher than LS; competency demonstration instances exceeded learning session instances."
            elif r["competency_instances"] > r["ls"]:
                r["interpretation"] = "PC + MC is higher than LS; competency activity was strong."
            elif r["ls"] > r["competency_instances"]:
                r["interpretation"] = "LS is higher than competency instances; mentees may need follow-up PC/MC."
            else:
                r["interpretation"] = "LS and competency activity are balanced."

        # ------------------------------------------------------------
        # Facility summary
        # ------------------------------------------------------------
        facility_rows = list(
            details_qs.values(
                province=F("mentorshipvistfk__facilityfk__districtfk__provincefk__name"),
                district=F("mentorshipvistfk__facilityfk__districtfk__name"),
                facility=F("mentorshipvistfk__facilityfk__name"),
                hfcode=F("mentorshipvistfk__facilityfk__hfcode"),
            )
            .annotate(
                visits=Count("mentorshipvistfk", distinct=True),
                mentees=Count("menteename", distinct=True),
                mentors=Count("mentor", distinct=True),
                thematics=Count("thematicname", distinct=True),
                topics=Count("topicname", distinct=True),
                ls=self._bool_sum("ls"),
                pc=self._bool_sum("pc"),
                mc=self._bool_sum("mc"),
                records=Count("id"),
            )
            .order_by("province", "district", "facility")
        )

        for r in facility_rows:
            r["ls"] = int(r["ls"] or 0)
            r["pc"] = int(r["pc"] or 0)
            r["mc"] = int(r["mc"] or 0)
            r["competency_instances"] = r["pc"] + r["mc"]
            r["competency_minus_ls"] = r["competency_instances"] - r["ls"]

            if r["mc"] > r["ls"]:
                r["interpretation"] = "MC activity was higher than LS."
            elif r["competency_instances"] > r["ls"]:
                r["interpretation"] = "Competency assessment activity was higher than LS."
            elif r["ls"] > r["competency_instances"]:
                r["interpretation"] = "LS activity was higher; follow-up PC/MC may be needed."
            else:
                r["interpretation"] = "Balanced LS and competency activity."

        facility_driver_rows = sorted(
            facility_rows,
            key=lambda x: abs(x["competency_minus_ls"]),
            reverse=True,
        )[:25]

        # ------------------------------------------------------------
        # Monthly trend
        # ------------------------------------------------------------
        monthly_total_rows = list(
            details_qs.annotate(
                month=TruncMonth("mentorshipvistfk__visitdate")
            )
            .values("month")
            .annotate(
                mentees=Count("menteename", distinct=True),
                visits=Count("mentorshipvistfk", distinct=True),
                topics=Count("topicname", distinct=True),
                ls=self._bool_sum("ls"),
                pc=self._bool_sum("pc"),
                mc=self._bool_sum("mc"),
            )
            .order_by("month")
        )

        for r in monthly_total_rows:
            r["month_label"] = r["month"].strftime("%b %Y") if r["month"] else "Unknown"
            r["ls"] = int(r["ls"] or 0)
            r["pc"] = int(r["pc"] or 0)
            r["mc"] = int(r["mc"] or 0)
            r["competency_instances"] = r["pc"] + r["mc"]
            r["total"] = r["ls"] + r["pc"] + r["mc"]

        trend_rows = list(
            details_qs.annotate(
                month=TruncMonth("mentorshipvistfk__visitdate")
            )
            .values(
                "month",
                province=F("mentorshipvistfk__facilityfk__districtfk__provincefk__name"),
            )
            .annotate(
                mentees=Count("menteename", distinct=True),
                visits=Count("mentorshipvistfk", distinct=True),
                topics=Count("topicname", distinct=True),
                ls=self._bool_sum("ls"),
                pc=self._bool_sum("pc"),
                mc=self._bool_sum("mc"),
            )
            .order_by("province", "month")
        )

        for r in trend_rows:
            r["month_label"] = r["month"].strftime("%b %Y") if r["month"] else "Unknown"
            r["ls"] = int(r["ls"] or 0)
            r["pc"] = int(r["pc"] or 0)
            r["mc"] = int(r["mc"] or 0)
            r["competency_instances"] = r["pc"] + r["mc"]

        # ------------------------------------------------------------
        # First mentorship visit by province
        # ------------------------------------------------------------
        first_visit_rows = []

        first_by_province = list(
            visits_qs.values(
                province_id=F("facilityfk__districtfk__provincefk_id"),
                province=F("facilityfk__districtfk__provincefk__name"),
            )
            .exclude(province_id__isnull=True)
            .annotate(first_date=Min("visitdate"))
            .order_by("province")
        )

        for row in first_by_province:
            first_visit = (
                visits_qs.filter(
                    facilityfk__districtfk__provincefk_id=row["province_id"],
                    visitdate=row["first_date"],
                )
                .order_by("visitdate", "id")
                .first()
            )

            facility = first_visit.facilityfk if first_visit else None
            district = facility.districtfk if facility else None

            first_visit_rows.append({
                "province": row["province"],
                "first_date": row["first_date"],
                "facility": facility.name if facility else "",
                "hfcode": facility.hfcode if facility else "",
                "district": district.name if district else "",
                "visit_round": first_visit.visitround if first_visit else "",
            })

        # ------------------------------------------------------------
        # Mentee profile table
        # ------------------------------------------------------------
        mentee_profiles = {}

        detail_records = details_qs.order_by(
            "mentorshipvistfk__facilityfk__districtfk__provincefk__name",
            "mentorshipvistfk__facilityfk__districtfk__name",
            "mentorshipvistfk__facilityfk__name",
            "menteename__firstname",
            "mentorshipvistfk__visitdate",
            "id",
        )

        mentee_ids = list(
            detail_records.exclude(menteename_id__isnull=True)
            .values_list("menteename_id", flat=True)
            .distinct()
        )

        topic_status_map = {}
        if mentee_ids:
            status_rows = MenteeTopicStatus.objects.filter(
                mentee_id__in=mentee_ids
            ).values(
                "mentee_id",
                "topic_id",
                "status",
                "consecutive_ls",
                "last_session_type",
                "last_date",
                "competent_date",
            )

            for s in status_rows:
                topic_status_map[(s["mentee_id"], s["topic_id"])] = s

        def safe_text(value):
            return str(value).strip() if value is not None else ""

        def get_gender_text(value):
            if value is True:
                return "Female"
            if value is False:
                return "Male"
            return ""

        for d in detail_records:
            visit = d.mentorshipvistfk
            facility = visit.facilityfk if visit else None
            district = facility.districtfk if facility else None
            province = district.provincefk if district else None
            mentee = d.menteename
            topic = d.topicname

            if not mentee:
                continue

            topic_name = safe_text(topic) or "Unknown topic"
            thematic_name = safe_text(d.thematicname) or "Unknown thematic area"

            profile_key = (
                facility.id if facility else None,
                mentee.id,
            )

            if profile_key not in mentee_profiles:
                mentee_profiles[profile_key] = {
                    "province": safe_text(getattr(province, "name", "")),
                    "district": safe_text(getattr(district, "name", "")),
                    "facility": safe_text(getattr(facility, "name", "")),
                    "hfcode": safe_text(getattr(facility, "hfcode", "")),
                    "mentee": safe_text(mentee),
                    "mentee_facility": safe_text(getattr(mentee.hfname, "name", "")) if mentee.hfname else "",
                    "position": safe_text(mentee.position),
                    "gender": get_gender_text(mentee.gender),
                    "visits": set(),
                    "topics_all": set(),
                    "ls_topics": set(),
                    "pc_topics": set(),
                    "mc_topics": set(),
                    "competent_topics": set(),
                    "needs_graduation_topics": set(),
                    "last_visit_date": None,
                }

            profile = mentee_profiles[profile_key]

            if visit:
                profile["visits"].add(visit.id)

                if visit.visitdate:
                    if not profile["last_visit_date"] or visit.visitdate > profile["last_visit_date"]:
                        profile["last_visit_date"] = visit.visitdate

            profile["topics_all"].add(topic_name)

            if d.ls:
                profile["ls_topics"].add(topic_name)

            if d.pc:
                profile["pc_topics"].add(topic_name)
                profile["competent_topics"].add(topic_name)

            if d.mc:
                profile["mc_topics"].add(topic_name)
                profile["competent_topics"].add(topic_name)

            status = topic_status_map.get((mentee.id, topic.id if topic else None))
            if status and status.get("status") == "COMPETENT":
                profile["competent_topics"].add(topic_name)

        mentee_profile_rows = []

        for _key, profile in mentee_profiles.items():
            active_topics = (
                profile["ls_topics"]
                | profile["pc_topics"]
                | profile["mc_topics"]
            )

            profile["needs_graduation_topics"] = active_topics - profile["competent_topics"]

            ls_topics = sorted(profile["ls_topics"])
            pc_topics = sorted(profile["pc_topics"])
            mc_topics = sorted(profile["mc_topics"])
            competent_topics = sorted(profile["competent_topics"])
            needs_topics = sorted(profile["needs_graduation_topics"])

            needs_count = len(needs_topics)
            competent_count = len(competent_topics)

            if needs_count > 0:
                overall_status = "Needs graduation / follow-up"
                status_badge = "warning"
            elif competent_count > 0:
                overall_status = "Competent / graduated"
                status_badge = "success"
            else:
                overall_status = "No competency progress recorded"
                status_badge = "neutral"

            mentee_profile_rows.append({
                "province": profile["province"],
                "district": profile["district"],
                "facility": profile["facility"],
                "hfcode": profile["hfcode"],
                "mentee": profile["mentee"],
                "mentee_facility": profile["mentee_facility"],
                "position": profile["position"],
                "gender": profile["gender"],
                "visit_count": len(profile["visits"]),
                "topic_count": len(profile["topics_all"]),
                "ls_count": len(ls_topics),
                "pc_count": len(pc_topics),
                "mc_count": len(mc_topics),
                "competent_count": competent_count,
                "needs_count": needs_count,
                "ls_topics_text": "\n".join(ls_topics),
                "pc_topics_text": "\n".join(pc_topics),
                "mc_topics_text": "\n".join(mc_topics),
                "competent_topics_text": "\n".join(competent_topics),
                "needs_topics_text": "\n".join(needs_topics),
                "last_visit_date": profile["last_visit_date"],
                "overall_status": overall_status,
                "status_badge": status_badge,
            })

        mentee_profile_rows = sorted(
            mentee_profile_rows,
            key=lambda x: (
                x["province"],
                x["district"],
                x["facility"],
                -x["needs_count"],
                x["mentee"],
            ),
        )

        # ------------------------------------------------------------
        # Story candidates
        # ------------------------------------------------------------
        story_candidates = []

        for r in facility_rows:
            story_score = (
                int(r["ls"] or 0)
                + int(r["pc"] or 0) * 2
                + int(r["mc"] or 0) * 3
                + int(r["topics"] or 0)
            )

            if r["ls"] > 0 or r["pc"] > 0 or r["mc"] > 0:
                story_candidates.append({
                    **r,
                    "story_score": story_score,
                    "story_angle": (
                        "Strong competency progress"
                        if r["competency_instances"] > r["ls"]
                        else "Strong learning session implementation with follow-up need"
                    ),
                })

        story_candidates = sorted(
            story_candidates,
            key=lambda x: x["story_score"],
            reverse=True,
        )[:15]

        # ------------------------------------------------------------
        # Chart data
        # ------------------------------------------------------------
        chart_data = {
            "province_lspcmc": [
                {
                    "province": r.get("province") or "Unknown",
                    "LS": int(r.get("ls") or 0),
                    "PC": int(r.get("pc") or 0),
                    "MC": int(r.get("mc") or 0),
                }
                for r in province_rows
            ],
            "province_competency_gap": [
                {
                    "province": r.get("province") or "Unknown",
                    "gap": int(r.get("competency_minus_ls") or 0),
                }
                for r in province_rows
            ],
            "monthly_lspcmc": [
                {
                    "month": r.get("month_label") or "Unknown",
                    "LS": int(r.get("ls") or 0),
                    "PC": int(r.get("pc") or 0),
                    "MC": int(r.get("mc") or 0),
                    "Total": int(r.get("total") or 0),
                }
                for r in monthly_total_rows
            ],
            "facility_drivers": [
                {
                    "facility": (r.get("facility") or "Unknown")[:38],
                    "LS": int(r.get("ls") or 0),
                    "PC": int(r.get("pc") or 0),
                    "MC": int(r.get("mc") or 0),
                }
                for r in facility_driver_rows[:10]
            ],
            "story_candidates": [
                {
                    "facility": (r.get("facility") or "Unknown")[:38],
                    "story_score": int(r.get("story_score") or 0),
                    "LS": int(r.get("ls") or 0),
                    "PC": int(r.get("pc") or 0),
                    "MC": int(r.get("mc") or 0),
                }
                for r in story_candidates[:10]
            ],
            "mentee_needs": [
                {
                    "mentee": (r.get("mentee") or "Unknown")[:38],
                    "needs": int(r.get("needs_count") or 0),
                    "competent": int(r.get("competent_count") or 0),
                }
                for r in sorted(
                    mentee_profile_rows,
                    key=lambda x: x["needs_count"],
                    reverse=True,
                )[:10]
            ],
        }

        export_query = request.GET.copy()
        export_query["export"] = "1"

        return {
            "filters": filters,
            "export_query": export_query.urlencode(),
            "province_options": province_options,
            "facility_options": facility_options,
            "mentor_options": mentor_options,
            "thematic_options": thematic_options,
            "visit_round_options": visit_round_options,
            "kpis": kpis,
            "province_rows": province_rows,
            "facility_rows": facility_rows,
            "facility_driver_rows": facility_driver_rows,
            "trend_rows": trend_rows,
            "monthly_total_rows": monthly_total_rows,
            "first_visit_rows": first_visit_rows,
            "mentee_profile_rows": mentee_profile_rows,
            "story_candidates": story_candidates,
            "chart_data": chart_data,
            "methodology_note": (
                "LS, PC, and MC figures represent mentorship detail instances, not unique mentees. "
                "A single mentee may appear multiple times across visits, thematic areas, topics, mentors, or competency assessments."
            ),
        }

    def _export_dashboard_excel(self, data):
        wb = Workbook()

        def style_sheet(ws):
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style="thin", color="D9E2F3"),
                right=Side(style="thin", color="D9E2F3"),
                top=Side(style="thin", color="D9E2F3"),
                bottom=Side(style="thin", color="D9E2F3"),
            )

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)

                for cell in col:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 50)

        ws = wb.active
        ws.title = "Province_Summary"
        ws.append([
            "Province", "Visits", "Facilities", "Mentees", "Mentors",
            "Thematics", "Topics", "LS", "PC", "MC",
            "PC + MC", "Competency - LS", "Interpretation",
        ])

        for r in data["province_rows"]:
            ws.append([
                r.get("province"),
                r.get("visits"),
                r.get("facilities"),
                r.get("mentees"),
                r.get("mentors"),
                r.get("thematics"),
                r.get("topics"),
                r.get("ls"),
                r.get("pc"),
                r.get("mc"),
                r.get("competency_instances"),
                r.get("competency_minus_ls"),
                r.get("interpretation"),
            ])

        ws2 = wb.create_sheet("Facility_Detail")
        ws2.append([
            "Province", "District", "Facility", "HF Code",
            "Visits", "Mentees", "Mentors", "Thematics", "Topics",
            "LS", "PC", "MC", "PC + MC", "Competency - LS",
            "Interpretation",
        ])

        for r in data["facility_rows"]:
            ws2.append([
                r.get("province"),
                r.get("district"),
                r.get("facility"),
                r.get("hfcode"),
                r.get("visits"),
                r.get("mentees"),
                r.get("mentors"),
                r.get("thematics"),
                r.get("topics"),
                r.get("ls"),
                r.get("pc"),
                r.get("mc"),
                r.get("competency_instances"),
                r.get("competency_minus_ls"),
                r.get("interpretation"),
            ])

        ws3 = wb.create_sheet("Monthly_Trend")
        ws3.append([
            "Province", "Month", "Mentees", "Visits", "Topics",
            "LS", "PC", "MC", "PC + MC",
        ])

        for r in data["trend_rows"]:
            ws3.append([
                r.get("province"),
                r.get("month_label"),
                r.get("mentees"),
                r.get("visits"),
                r.get("topics"),
                r.get("ls"),
                r.get("pc"),
                r.get("mc"),
                r.get("competency_instances"),
            ])

        ws4 = wb.create_sheet("First_Visits")
        ws4.append([
            "Province", "First Visit Date", "First Facility",
            "HF Code", "District", "Visit Round",
        ])

        for r in data["first_visit_rows"]:
            ws4.append([
                r.get("province"),
                r.get("first_date"),
                r.get("facility"),
                r.get("hfcode"),
                r.get("district"),
                r.get("visit_round"),
            ])

        ws5 = wb.create_sheet("Mentee_Profile")
        ws5.append([
            "Province", "District", "Facility", "HF Code",
            "Mentee", "Mentee Facility", "Position", "Gender",
            "Visit Count", "Topic Count", "LS Count", "PC Count", "MC Count",
            "Competent Count", "Needs Graduation Count",
            "LS Topics", "PC Topics", "MC Topics",
            "Competent Topics", "Topics Needing Graduation",
            "Last Visit Date", "Overall Status",
        ])

        for r in data["mentee_profile_rows"]:
            ws5.append([
                r.get("province"),
                r.get("district"),
                r.get("facility"),
                r.get("hfcode"),
                r.get("mentee"),
                r.get("mentee_facility"),
                r.get("position"),
                r.get("gender"),
                r.get("visit_count"),
                r.get("topic_count"),
                r.get("ls_count"),
                r.get("pc_count"),
                r.get("mc_count"),
                r.get("competent_count"),
                r.get("needs_count"),
                r.get("ls_topics_text"),
                r.get("pc_topics_text"),
                r.get("mc_topics_text"),
                r.get("competent_topics_text"),
                r.get("needs_topics_text"),
                r.get("last_visit_date"),
                r.get("overall_status"),
            ])

        ws6 = wb.create_sheet("Story_Candidates")
        ws6.append([
            "Province", "District", "Facility", "HF Code",
            "LS", "PC", "MC", "PC + MC", "Story Angle", "Story Score",
        ])

        for r in data["story_candidates"]:
            ws6.append([
                r.get("province"),
                r.get("district"),
                r.get("facility"),
                r.get("hfcode"),
                r.get("ls"),
                r.get("pc"),
                r.get("mc"),
                r.get("competency_instances"),
                r.get("story_angle"),
                r.get("story_score"),
            ])

        ws7 = wb.create_sheet("Methodology_Notes")
        ws7.append(["Topic", "Explanation"])
        ws7.append(["LS / PC / MC interpretation", data["methodology_note"]])
        ws7.append([
            "Graduation logic",
            "A topic is counted as competent/graduated when PC or MC is recorded, or when MenteeTopicStatus is COMPETENT.",
        ])

        for sheet in wb.worksheets:
            style_sheet(sheet)

        timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
        filename = f"Mentorship_Dashboard_{timestamp}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
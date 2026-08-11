"""Complete replacement for ``hmis/admin.py``.

This keeps the four existing HMIS admin registrations and adds a separate,
read-only HMIS performance dashboard through the ``HMISDashboard`` proxy model.
The dashboard uses ``HMISFact`` to draw one separate monthly trend chart for
every reported HMIS indicator and to build matching dynamic detailed-result
tables. ``HMISMonthlySummary`` supplies the summary KPI cards. Excel export
uses the active dashboard filters and creates only Summary and Details sheets.
"""

import csv
# Required by _export_excel() when the HMIS upload report JSONField is
# written into the "Upload Register" worksheet.
import json
import logging
import os
import tempfile
from calendar import month_name
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl.cell import WriteOnlyCell
from django.conf import settings
from django.contrib import admin, messages
from django.db.models import Avg, Count, Max, Min, Sum
from django.http import FileResponse, HttpResponseRedirect, StreamingHttpResponse
from django.template.response import TemplateResponse
from django.utils import timezone
from hmis.models import (
    HMISDashboard,
    HMISFact,
    HMISMonthlySummary,
    HMISRawUpload,
    IndicatorMetadata,
)
from hmis.services.pipeline import run_import
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)

# This identifier is written to the response header and production log so the
# deployed filtered Excel implementation can be distinguished from old builds.
HMIS_EXPORT_BUILD = "2026.08.09-xlsx-filtered-r1"


# ===========================================================================
# EXISTING HMIS ADMINS
# ===========================================================================


@admin.register(HMISRawUpload)
class HMISRawUploadAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "status",
        "uploaded_at",
        "uploaded_by",
        "row_count",
        "hf_count",
        "period_min",
        "period_max",
    )
    list_filter = ("status", "uploaded_at")
    actions = ("import_selected",)

    @admin.action(description="Import selected HMIS uploads")
    def import_selected(self, request, queryset):
        ok, failed = 0, 0
        for upload in queryset:
            try:
                run_import(upload)
                ok += 1
            except Exception as exc:
                upload.status = "FAILED"
                upload.report = {"error": str(exc)}
                upload.save(update_fields=["status", "report"])
                failed += 1
        self.message_user(
            request,
            f"Import: {ok} success, {failed} failed.",
            level=messages.INFO,
        )


@admin.register(HMISFact)
class HMISFactAdmin(admin.ModelAdmin):
    list_display = (
        "prov",
        "dist",
        "hf",
        "year",
        "month",
        "month_name",
        "indicator_name",
        "value",
        "hiva_hfs",
    )
    list_filter = (
        "hiva_hfs",
        "prov",
        "year",
        "month",
        "indicator_name",
    )
    search_fields = ("hf", "prov", "dist", "indicator_name")
    ordering = ("-year", "-month", "prov", "dist", "hf", "indicator_name")
    list_per_page = 50
    list_select_related = ("source_upload",)

    @admin.display(description="Period")
    def period_readable(self, obj):
        period = obj.periodcode or ""
        return f"{period[:4]}-{period[4:6]}" if len(period) == 6 else period


@admin.register(HMISMonthlySummary)
class HMISMonthlySummaryAdmin(admin.ModelAdmin):
    list_display = (
        "prov",
        "dist",
        "hf",
        "year",
        "month",
        "month_name",
        "hiva_hfs",
        "anc1",
        "anc2",
        "anc3",
        "anc4",
        "pnc1",
        "pnc2",
        "n_delivery",
        "a_delivery",
        "c_section",
        "lbw",
        "stillbirth",
    )
    list_filter = ("hiva_hfs", "prov", "year", "month")
    search_fields = ("hf", "prov", "dist")
    ordering = ("-year", "-month", "prov", "dist", "hf")
    list_per_page = 50
    list_select_related = ("source_upload",)

    @admin.display(description="Period")
    def period_readable(self, obj):
        period = obj.periodcode or ""
        return f"{period[:4]}-{period[4:6]}" if len(period) == 6 else period


@admin.register(IndicatorMetadata)
class IndicatorMetadataAdmin(admin.ModelAdmin):

    list_display = (
        "indicator_code",
        "indicator_name",
        "indicator_short_name",
        "indicator_group",
        "indicator_domain",
        "unit_of_measure",
        "reporting_level",
        "data_source",
        "is_active",
        "sort_order",
    )

    list_filter = (
        "indicator_group",
        "indicator_domain",
        "unit_of_measure",
        "reporting_level",
        "data_source",
        "is_active",
    )

    search_fields = (
        "indicator_code",
        "indicator_name",
        "indicator_short_name",
        "indicator_group",
        "indicator_domain",
        "indicator_description",
    )

    ordering = (
        "sort_order",
        "indicator_name",
    )

    list_per_page = 30
    save_on_top = True

    readonly_fields = (
        "created_at",
        "updated_at",
    )

    fieldsets = (
        ("① Indicator Identification", {
            "fields": (
                ("indicator_code", "sort_order"),
                "indicator_name",
                "indicator_short_name",
                ("indicator_group", "indicator_domain"),
                "is_active",
            )
        }),

        ("② Indicator Definition", {
            "fields": (
                "indicator_description",
                "numerator_definition",
                "denominator_definition",
            )
        }),

        ("③ Measurement and Reporting", {
            "fields": (
                ("unit_of_measure", "reporting_level"),
                "data_source",
            )
        }),

        ("④ Audit Trail", {
            "classes": ("collapse",),
            "fields": (
                ("created_at", "updated_at"),
            )
        }),
    )

# ===========================================================================
# NEW HMIS PERFORMANCE DASHBOARD
# ===========================================================================


@admin.register(HMISDashboard)
class HMISDashboardAdmin(admin.ModelAdmin):
    """Read-only dashboard built from HMISMonthlySummary and HMISFact."""

    change_list_template = "admin/hmis/hmis_dashboard.html"

    INDICATORS = (
        ("anc1", "ANC 1st Visit"),
        ("anc2", "ANC 2nd Visit"),
        ("anc3", "ANC 3rd Visit"),
        ("anc4", "ANC 4th Visit"),
        ("pnc1", "PNC 1st Visit"),
        ("pnc2", "PNC 2nd Visit"),
        ("n_delivery", "Normal Deliveries"),
        ("a_delivery", "Assisted Deliveries"),
        ("c_section", "C-sections"),
        ("lbw", "Low Birth Weight"),
        ("stillbirth", "Stillbirths"),
    )
    INDICATOR_LABELS = dict(INDICATORS)
    INDICATOR_FIELDS = tuple(field for field, _label in INDICATORS)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        opts = HMISMonthlySummary._meta
        return (
            request.user.is_superuser
            or request.user.has_perm(f"{opts.app_label}.view_{opts.model_name}")
            or request.user.has_perm(f"{opts.app_label}.change_{opts.model_name}")
        )

    def has_module_permission(self, request):
        return self.has_view_permission(request)

    @staticmethod
    def _decimal(value):
        return Decimal("0") if value is None else Decimal(str(value))

    @staticmethod
    def _round2(value):
        if value is None:
            return None
        return float(
            Decimal(str(value)).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )
        )

    def _percent(self, numerator, denominator):
        denominator = self._decimal(denominator)
        if denominator == 0:
            return None
        value = self._decimal(numerator) / denominator * Decimal("100")
        return self._round2(value)

    def _rate_per_1000(self, numerator, denominator):
        denominator = self._decimal(denominator)
        if denominator == 0:
            return None
        value = self._decimal(numerator) / denominator * Decimal("1000")
        return self._round2(value)

    @staticmethod
    def _period_label(periodcode):
        period = str(periodcode or "")
        if len(period) == 6 and period.isdigit():
            month_number = int(period[4:6])
            if 1 <= month_number <= 12:
                return f"{month_name[month_number]} {period[:4]}"
        return period or "Not available"

    @staticmethod
    def _number_display(value, decimals=0):
        if value is None:
            return "—"
        number = float(value)
        if decimals == 0:
            return f"{number:,.0f}"
        return f"{number:,.{decimals}f}"

    @staticmethod
    def _percent_display(value):
        return "—" if value is None else f"{float(value):,.2f}%"

    @staticmethod
    def _rate_display(value):
        return "—" if value is None else f"{float(value):,.2f}"

    def _read_filters(self, request):
        indicator = (request.GET.get("indicator") or "anc1").strip()
        if indicator not in self.INDICATOR_LABELS:
            indicator = "anc1"

        facility_group = (request.GET.get("facility_group") or "all").strip()
        if facility_group not in {"all", "hiva", "non_hiva"}:
            facility_group = "all"

        year_value = (request.GET.get("year") or "").strip()
        month_value = (request.GET.get("month") or "").strip()

        try:
            selected_year = int(year_value) if year_value else None
        except (TypeError, ValueError):
            selected_year = None

        try:
            selected_month = int(month_value) if month_value else None
        except (TypeError, ValueError):
            selected_month = None
        if selected_month not in range(1, 13):
            selected_month = None

        return {
            "prov": (request.GET.get("prov") or "").strip(),
            "dist": (request.GET.get("dist") or "").strip(),
            "hf": (request.GET.get("hf") or "").strip(),
            "year": selected_year,
            "month": selected_month,
            "facility_group": facility_group,
            "indicator": indicator,
        }

    def _filter_queryset(self, queryset, filters):
        if filters["prov"]:
            queryset = queryset.filter(prov=filters["prov"])
        if filters["dist"]:
            queryset = queryset.filter(dist=filters["dist"])
        if filters["hf"]:
            queryset = queryset.filter(hf=filters["hf"])
        if filters["year"] is not None:
            queryset = queryset.filter(year=filters["year"])
        if filters["month"] is not None:
            queryset = queryset.filter(month=filters["month"])
        if filters["facility_group"] == "hiva":
            queryset = queryset.filter(hiva_hfs=True)
        elif filters["facility_group"] == "non_hiva":
            queryset = queryset.filter(hiva_hfs=False)
        return queryset

    @staticmethod
    def _apply_facility_group(queryset, facility_group):
        """Limit either HMIS table to the selected facility group."""
        if facility_group == "hiva":
            return queryset.filter(hiva_hfs=True)
        if facility_group == "non_hiva":
            return queryset.filter(hiva_hfs=False)
        return queryset

    def _filter_options(self, base_queryset, filters):
        # Facility Group is the first level of the location cascade:
        # Facility Group -> Province -> District -> Health Facility.
        group_queryset = self._apply_facility_group(
            base_queryset,
            filters["facility_group"],
        )

        provinces = list(
            group_queryset.exclude(prov="")
            .values_list("prov", flat=True)
            .distinct()
            .order_by("prov")
        )

        district_queryset = group_queryset
        if filters["prov"]:
            district_queryset = district_queryset.filter(prov=filters["prov"])
        districts = list(
            district_queryset.exclude(dist="")
            .values_list("dist", flat=True)
            .distinct()
            .order_by("dist")
        )

        facility_queryset = district_queryset
        if filters["dist"]:
            facility_queryset = facility_queryset.filter(dist=filters["dist"])
        facilities = list(
            facility_queryset.exclude(hf="")
            .values_list("hf", flat=True)
            .distinct()
            .order_by("hf")
        )

        period_queryset = facility_queryset
        if filters["hf"]:
            period_queryset = period_queryset.filter(hf=filters["hf"])

        years = list(
            period_queryset.values_list("year", flat=True)
            .distinct()
            .order_by("-year")
        )

        month_queryset = period_queryset
        if filters["year"] is not None:
            month_queryset = month_queryset.filter(year=filters["year"])
        month_numbers = list(
            month_queryset.values_list("month", flat=True)
            .distinct()
            .order_by("month")
        )
        months = [
            {"value": value, "label": month_name[value]}
            for value in month_numbers
            if value in range(1, 13)
        ]

        # A compact, distinct hierarchy lets the browser update the dependent
        # dropdowns immediately without reloading the chart-heavy dashboard.
        location_cascade = [
            {
                "group": "hiva" if row["hiva_hfs"] else "non_hiva",
                "province": row["prov"],
                "district": row["dist"],
                "facility": row["hf"],
            }
            for row in (
                base_queryset.exclude(prov="")
                .exclude(dist="")
                .exclude(hf="")
                .values("hiva_hfs", "prov", "dist", "hf")
                .distinct()
                .order_by("hiva_hfs", "prov", "dist", "hf")
            )
        ]

        return {
            "provinces": provinces,
            "districts": districts,
            "facilities": facilities,
            "years": years,
            "months": months,
            "location_cascade": location_cascade,
        }

    def _filter_fact_queryset(self, queryset, filters):
        """Apply the dashboard filters to the long-format HMIS indicator table."""
        if filters["prov"]:
            queryset = queryset.filter(prov=filters["prov"])
        if filters["dist"]:
            queryset = queryset.filter(dist=filters["dist"])
        if filters["hf"]:
            queryset = queryset.filter(hf=filters["hf"])
        if filters["year"] is not None:
            queryset = queryset.filter(year=filters["year"])
        if filters["month"] is not None:
            queryset = queryset.filter(month=filters["month"])
        if filters["facility_group"] == "hiva":
            queryset = queryset.filter(hiva_hfs=True)
        elif filters["facility_group"] == "non_hiva":
            queryset = queryset.filter(hiva_hfs=False)
        return queryset

    def _all_hmis_indicator_trends(self, queryset):
        """Build one monthly trend chart for every indicator in HMISFact."""
        available_periods = list(
            queryset.exclude(periodcode="")
            .values_list("periodcode", flat=True)
            .distinct()
            .order_by("periodcode")
        )

        metadata = {}
        for item in (
            IndicatorMetadata.objects.all()
            .order_by("-is_active", "sort_order", "indicator_name")
            .values(
                "indicator_code",
                "indicator_name",
                "indicator_short_name",
                "indicator_group",
                "indicator_domain",
                "sort_order",
            )
        ):
            metadata.setdefault(item["indicator_code"], item)

        grouped = (
            queryset.values("indicator_code", "periodcode")
            .annotate(
                # Do not reuse a model-field name as an aggregate alias.
                # Django 5.1 can resolve a later expression to that alias
                # instead of to the original database column.
                indicator_name_max=Max("indicator_name"),
                total=Sum("value"),
                reported_values=Count("value"),
            )
            .order_by("indicator_code", "periodcode")
        )

        indicators = {}
        for item in grouped:
            code = item["indicator_code"] or "Not coded"
            meta = metadata.get(code, {})
            indicator = indicators.setdefault(
                code,
                {
                    "code": code,
                    "title": (
                        meta.get("indicator_short_name")
                        or meta.get("indicator_name")
                        or item["indicator_name_max"]
                        or code
                    ),
                    "full_name": (
                        meta.get("indicator_name")
                        or item["indicator_name_max"]
                        or code
                    ),
                    "group": meta.get("indicator_group") or "Not classified",
                    "domain": meta.get("indicator_domain") or "Not classified",
                    "sort_order": meta.get("sort_order"),
                    "values_by_period": {},
                },
            )
            indicator["values_by_period"][item["periodcode"]] = (
                float(item["total"])
                if item["reported_values"] and item["total"] is not None
                else None
            )

        ordered = sorted(
            indicators.values(),
            key=lambda item: (
                item["sort_order"] is None,
                item["sort_order"] if item["sort_order"] is not None else 0,
                item["title"].lower(),
                item["code"],
            ),
        )
        palette = (
            "#174f78",
            "#2e7d32",
            "#6a3d9a",
            "#00796b",
            "#ef6c00",
            "#c62828",
            "#455a64",
            "#8d6e63",
        )
        cards = []
        charts = []
        for index, indicator in enumerate(ordered, start=1):
            chart_id = f"allHmisIndicatorTrendChart{index}"
            cards.append(
                {
                    "chart_id": chart_id,
                    "indicator_code": indicator["code"],
                    "title": indicator["title"],
                    "full_name": indicator["full_name"],
                    "group": indicator["group"],
                    "domain": indicator["domain"],
                    "period_count": sum(
                        value is not None
                        for value in indicator["values_by_period"].values()
                    ),
                    "search_text": " ".join(
                        str(value)
                        for value in (
                            indicator["code"],
                            indicator["title"],
                            indicator["full_name"],
                            indicator["group"],
                            indicator["domain"],
                        )
                    ).lower(),
                }
            )
            charts.append(
                {
                    "id": chart_id,
                    "labels": [
                        self._period_label(periodcode)
                        for periodcode in available_periods
                    ],
                    "datasets": [
                        {
                            "label": indicator["title"],
                            "data": [
                                indicator["values_by_period"].get(periodcode)
                                for periodcode in available_periods
                            ],
                            "color": palette[(index - 1) % len(palette)],
                        }
                    ],
                }
            )
        return cards, charts

    def _export_indicator_columns(self, queryset):
        """Return the export catalogue without building unused chart arrays.

        The interactive dashboard needs a value for every indicator/month
        combination.  Excel export only needs the ordered indicator catalogue,
        so using the chart builder here wastes memory and CPU on large exports.
        """
        metadata = {}
        for item in (
            IndicatorMetadata.objects.all()
            .order_by("-is_active", "sort_order", "indicator_name")
            .values(
                "indicator_code",
                "indicator_name",
                "indicator_short_name",
                "indicator_group",
                "indicator_domain",
                "sort_order",
            )
        ):
            metadata.setdefault(item["indicator_code"], item)

        grouped = (
            queryset.values("indicator_code")
            .annotate(indicator_name_max=Max("indicator_name"))
            .order_by("indicator_code")
        )
        columns = []
        for item in grouped.iterator(chunk_size=2000):
            code = item["indicator_code"] or "Not coded"
            meta = metadata.get(code, {})
            columns.append(
                {
                    "indicator_code": code,
                    "title": (
                        meta.get("indicator_short_name")
                        or meta.get("indicator_name")
                        or item["indicator_name_max"]
                        or code
                    ),
                    "full_name": (
                        meta.get("indicator_name")
                        or item["indicator_name_max"]
                        or code
                    ),
                    "group": meta.get("indicator_group") or "Not classified",
                    "domain": meta.get("indicator_domain") or "Not classified",
                    "sort_order": meta.get("sort_order"),
                }
            )
        columns.sort(
            key=lambda item: (
                item["sort_order"] is None,
                item["sort_order"] if item["sort_order"] is not None else 0,
                item["title"].lower(),
                item["indicator_code"],
            )
        )
        return columns

    def _fact_detail_tables(self, queryset, indicator_cards, selected_indicator):
        """Build compact matrices using the exact indicator set shown in charts.

        The values lists are aligned with ``indicator_cards``. This lets the
        Django template render a fully dynamic set of indicator columns without
        custom template filters or hard-coded ANC/PNC fields.
        """
        indicator_codes = [card["indicator_code"] for card in indicator_cards]
        indicator_indexes = {
            code: index for index, code in enumerate(indicator_codes)
        }

        def matrix(group_fields):
            grouped = (
                queryset.values(*group_fields, "indicator_code")
                .annotate(
                    reported_total=Sum("value"),
                    reported_values=Count("value"),
                )
                .order_by(*group_fields, "indicator_code")
            )
            rows_by_key = {}
            for item in grouped:
                code = item["indicator_code"] or "Not coded"
                index = indicator_indexes.get(code)
                if index is None:
                    continue
                key = tuple(item[field] for field in group_fields)
                if key not in rows_by_key:
                    rows_by_key[key] = {
                        "key": key,
                        "raw_values": [None] * len(indicator_codes),
                    }
                value = (
                    float(item["reported_total"])
                    if item["reported_values"]
                    and item["reported_total"] is not None
                    else None
                )
                rows_by_key[key]["raw_values"][index] = value

            rows = []
            for row in rows_by_key.values():
                row["indicator_values"] = [
                    {
                        "raw": value,
                        "display": self._number_display(value),
                    }
                    for value in row["raw_values"]
                ]
                rows.append(row)
            return rows

        monthly_rows = matrix(("periodcode",))
        for row in monthly_rows:
            row["periodcode"] = row["key"][0]
            row["period"] = self._period_label(row["periodcode"])
        monthly_rows.sort(key=lambda row: row["periodcode"] or "")

        province_rows = matrix(("prov",))
        for row in province_rows:
            row["province"] = row["key"][0] or "Not specified"

        facility_rows = matrix(("prov", "hf", "hiva_hfs"))
        for row in facility_rows:
            row["province"] = row["key"][0] or "Not specified"
            row["facility"] = row["key"][1] or "Not specified"
            row["hiva_hfs"] = bool(row["key"][2])
            row["facility_group"] = (
                "HIVA HF" if row["hiva_hfs"] else "Non-HIVA HF"
            )

        # Preserve the existing Selected Summary Indicator control by using it
        # to order province and facility rows when the same code exists in the
        # HMISFact indicator catalogue. Every indicator remains visible.
        selected_index = next(
            (
                index
                for index, code in enumerate(indicator_codes)
                if str(code).strip().lower() == selected_indicator.lower()
            ),
            None,
        )

        def sort_geography_rows(rows, name_key):
            if selected_index is None:
                rows.sort(key=lambda row: str(row[name_key]).lower())
                return
            rows.sort(
                key=lambda row: (
                    row["raw_values"][selected_index] is None,
                    -(
                        row["raw_values"][selected_index]
                        if row["raw_values"][selected_index] is not None
                        else 0
                    ),
                    str(row[name_key]).lower(),
                )
            )

        sort_geography_rows(province_rows, "province")
        sort_geography_rows(facility_rows, "facility")

        # Do not return a second indicator-column catalogue.  The template
        # deliberately uses ``all_hmis_indicator_cards`` for both the trend
        # charts and every Detailed Results header.  This single source of
        # truth makes it impossible for the chart indicators and table
        # indicators to drift apart again.
        return {
            "detailed_monthly_rows": monthly_rows,
            "detailed_province_rows": province_rows,
            "detailed_facility_rows": facility_rows,
            "detailed_indicator_count": len(indicator_cards),
        }

    def _summary_aggregate(self, queryset):
        expressions = {
            "reporting_records": Count("id"),
            "facilities": Count("hf", distinct=True),
            "periods": Count("periodcode", distinct=True),
            "period_min": Min("periodcode"),
            "period_max": Max("periodcode"),
        }
        for field in self.INDICATOR_FIELDS:
            expressions[f"{field}_total"] = Sum(field)
            expressions[f"{field}_reported"] = Count(field)
        return queryset.aggregate(**expressions)

    def _monthly_rows(self, queryset, selected_indicator):
        annotations = {
            "reporting_records": Count("id"),
            "facility_count": Count("hf", distinct=True),
        }
        for field in self.INDICATOR_FIELDS:
            # Use an alias that cannot shadow the real model field.  Using
            # ``anc1=Sum("anc1")`` followed by ``Count("anc1")`` in this
            # same annotate() call makes Django treat Count's anc1 as the
            # aggregate alias, causing "anc1 is an aggregate".
            annotations[f"{field}_total"] = Sum(field)
            annotations[f"{field}_reported"] = Count(field)

        rows = []
        grouped = queryset.values("periodcode").annotate(**annotations).order_by(
            "periodcode"
        )
        for item in grouped:
            row = {
                "periodcode": item["periodcode"],
                "period": self._period_label(item["periodcode"]),
                "reporting_records": item["reporting_records"],
                "facility_count": item["facility_count"],
            }
            for field in self.INDICATOR_FIELDS:
                total_key = f"{field}_total"
                row[field] = self._decimal(item[total_key])
                row[f"{field}_display"] = self._number_display(item[total_key])
            row["selected_total"] = row[selected_indicator]
            row["selected_total_display"] = self._number_display(
                row["selected_total"]
            )
            row["total_deliveries"] = (
                row["n_delivery"] + row["a_delivery"] + row["c_section"]
            )
            row["total_deliveries_display"] = self._number_display(
                row["total_deliveries"]
            )
            row["c_section_rate"] = self._percent(
                row["c_section"], row["total_deliveries"]
            )
            row["c_section_rate_display"] = self._percent_display(
                row["c_section_rate"]
            )
            rows.append(row)
        return rows

    def _facility_monthly_export_rows(self, queryset):
        """Return facility-month rows for Excel without creating dashboard charts."""
        annotations = {}
        for field in self.INDICATOR_FIELDS:
            annotations[f"{field}_total"] = Sum(field)
            annotations[f"{field}_reported"] = Count(field)

        grouped = (
            queryset.values(
                "prov",
                "dist",
                "hf",
                "hfid",
                "hiva_hfs",
                "periodcode",
            )
            .annotate(**annotations)
            .order_by("prov", "dist", "hf", "periodcode")
        )

        rows = []
        for item in grouped:
            values = {
                field: (
                    float(item[f"{field}_total"])
                    if item[f"{field}_reported"]
                    and item[f"{field}_total"] is not None
                    else None
                )
                for field in self.INDICATOR_FIELDS
            }
            rows.append(
                {
                    "province": item["prov"] or "Not specified",
                    "district": item["dist"] or "Not specified",
                    "facility": item["hf"] or "Not specified",
                    "hfid": item["hfid"],
                    "hiva_hfs": item["hiva_hfs"],
                    "periodcode": item["periodcode"],
                    "period": self._period_label(item["periodcode"]),
                    **values,
                }
            )
        return rows

    def _province_rows(self, queryset, selected_indicator):
        grouped = (
            queryset.values("prov")
            .annotate(
                reporting_records=Count("id"),
                facility_count=Count("hf", distinct=True),
                selected_total=Sum(selected_indicator),
                selected_average=Avg(selected_indicator),
                anc1_total=Sum("anc1"),
                anc4_total=Sum("anc4"),
                normal_total=Sum("n_delivery"),
                assisted_total=Sum("a_delivery"),
                c_section_total=Sum("c_section"),
                lbw_total=Sum("lbw"),
                stillbirth_total=Sum("stillbirth"),
            )
            .order_by("prov")
        )

        rows = []
        for item in grouped:
            deliveries = (
                self._decimal(item["normal_total"])
                + self._decimal(item["assisted_total"])
                + self._decimal(item["c_section_total"])
            )
            selected_total = self._decimal(item["selected_total"])
            anc_ratio = self._percent(item["anc4_total"], item["anc1_total"])
            c_section_rate = self._percent(item["c_section_total"], deliveries)
            stillbirth_rate = self._rate_per_1000(
                item["stillbirth_total"], deliveries
            )
            rows.append(
                {
                    "province": item["prov"] or "Not specified",
                    "reporting_records": item["reporting_records"],
                    "facility_count": item["facility_count"],
                    "selected_total": selected_total,
                    "selected_total_display": self._number_display(selected_total),
                    "selected_average": self._round2(item["selected_average"]),
                    "selected_average_display": self._number_display(
                        item["selected_average"], 2
                    ),
                    "anc4_ratio": anc_ratio,
                    "anc4_ratio_display": self._percent_display(anc_ratio),
                    "total_deliveries": deliveries,
                    "total_deliveries_display": self._number_display(deliveries),
                    "c_section_rate": c_section_rate,
                    "c_section_rate_display": self._percent_display(
                        c_section_rate
                    ),
                    "lbw_total": self._decimal(item["lbw_total"]),
                    "lbw_total_display": self._number_display(item["lbw_total"]),
                    "stillbirth_total": self._decimal(item["stillbirth_total"]),
                    "stillbirth_total_display": self._number_display(
                        item["stillbirth_total"]
                    ),
                    "stillbirth_rate": stillbirth_rate,
                    "stillbirth_rate_display": self._rate_display(stillbirth_rate),
                }
            )
        rows.sort(key=lambda row: row["selected_total"], reverse=True)
        return rows

    def _facility_rows(self, queryset, selected_indicator):
        grouped = (
            queryset.values("prov", "dist", "hf", "hfid", "hiva_hfs")
            .annotate(
                reporting_records=Count("id"),
                period_min=Min("periodcode"),
                period_max=Max("periodcode"),
                selected_total=Sum(selected_indicator),
                selected_average=Avg(selected_indicator),
                anc1_total=Sum("anc1"),
                anc4_total=Sum("anc4"),
                normal_total=Sum("n_delivery"),
                assisted_total=Sum("a_delivery"),
                c_section_total=Sum("c_section"),
                lbw_total=Sum("lbw"),
                stillbirth_total=Sum("stillbirth"),
            )
            .order_by("prov", "dist", "hf")
        )

        rows = []
        for item in grouped:
            deliveries = (
                self._decimal(item["normal_total"])
                + self._decimal(item["assisted_total"])
                + self._decimal(item["c_section_total"])
            )
            selected_total = self._decimal(item["selected_total"])
            anc_ratio = self._percent(item["anc4_total"], item["anc1_total"])
            c_section_rate = self._percent(item["c_section_total"], deliveries)
            stillbirth_rate = self._rate_per_1000(
                item["stillbirth_total"], deliveries
            )
            rows.append(
                {
                    "province": item["prov"] or "Not specified",
                    "district": item["dist"] or "Not specified",
                    "facility": item["hf"] or "Not specified",
                    "hfid": item["hfid"],
                    "hiva_hfs": item["hiva_hfs"],
                    "facility_group": "HIVA HF"
                    if item["hiva_hfs"]
                    else "Non-HIVA HF",
                    "reporting_records": item["reporting_records"],
                    "period_min": item["period_min"] or "",
                    "period_max": item["period_max"] or "",
                    "selected_total": selected_total,
                    "selected_total_display": self._number_display(selected_total),
                    "selected_average": self._round2(item["selected_average"]),
                    "selected_average_display": self._number_display(
                        item["selected_average"], 2
                    ),
                    "anc4_ratio": anc_ratio,
                    "anc4_ratio_display": self._percent_display(anc_ratio),
                    "total_deliveries": deliveries,
                    "total_deliveries_display": self._number_display(deliveries),
                    "c_section_rate": c_section_rate,
                    "c_section_rate_display": self._percent_display(
                        c_section_rate
                    ),
                    "lbw_total": self._decimal(item["lbw_total"]),
                    "lbw_total_display": self._number_display(item["lbw_total"]),
                    "stillbirth_total": self._decimal(item["stillbirth_total"]),
                    "stillbirth_total_display": self._number_display(
                        item["stillbirth_total"]
                    ),
                    "stillbirth_rate": stillbirth_rate,
                    "stillbirth_rate_display": self._rate_display(stillbirth_rate),
                }
            )
        rows.sort(key=lambda row: row["selected_total"], reverse=True)
        for rank, row in enumerate(rows, start=1):
            row["rank"] = rank
        return rows

    def _build_dashboard_data(self, queryset, selected_indicator):
        summary = self._summary_aggregate(queryset)
        totals = {
            field: self._decimal(summary[f"{field}_total"])
            for field in self.INDICATOR_FIELDS
        }

        reporting_records = summary["reporting_records"] or 0
        facilities = summary["facilities"] or 0
        periods = summary["periods"] or 0
        possible_cells = reporting_records * len(self.INDICATOR_FIELDS)
        reported_cells = sum(
            summary[f"{field}_reported"] or 0 for field in self.INDICATOR_FIELDS
        )
        data_completeness = self._percent(reported_cells, possible_cells)

        total_deliveries = (
            totals["n_delivery"]
            + totals["a_delivery"]
            + totals["c_section"]
        )
        anc4_ratio = self._percent(totals["anc4"], totals["anc1"])
        pnc2_ratio = self._percent(totals["pnc2"], totals["pnc1"])
        c_section_rate = self._percent(totals["c_section"], total_deliveries)
        stillbirth_rate = self._rate_per_1000(
            totals["stillbirth"], total_deliveries
        )

        selected_label = self.INDICATOR_LABELS[selected_indicator]
        period_min = summary["period_min"]
        period_max = summary["period_max"]
        if period_min and period_max:
            period_range = (
                self._period_label(period_min)
                if period_min == period_max
                else f"{self._period_label(period_min)} to {self._period_label(period_max)}"
            )
        else:
            period_range = "No reporting period available"

        kpis = [
            {
                "label": "Reporting Records",
                "value": self._number_display(reporting_records),
                "help": "Facility-month summary records",
                "class": "primary",
            },
            {
                "label": "Health Facilities",
                "value": self._number_display(facilities),
                "help": "Distinct reporting facilities",
                "class": "blue",
            },
            {
                "label": "Reporting Periods",
                "value": self._number_display(periods),
                "help": period_range,
                "class": "blue",
            },
            {
                "label": "Data Completeness",
                "value": self._percent_display(data_completeness),
                "help": "Non-blank cells across the 11 summary indicators",
                "class": "green",
            },
            {
                "label": "ANC 1st Visit",
                "value": self._number_display(totals["anc1"]),
                "help": "Total reported ANC1 services",
                "class": "purple",
            },
            {
                "label": "ANC 4th Visit",
                "value": self._number_display(totals["anc4"]),
                "help": f"ANC4-to-ANC1 ratio: {self._percent_display(anc4_ratio)}",
                "class": "purple",
            },
            {
                "label": "PNC 1st Visit",
                "value": self._number_display(totals["pnc1"]),
                "help": f"PNC2-to-PNC1 ratio: {self._percent_display(pnc2_ratio)}",
                "class": "teal",
            },
            {
                "label": "Total Deliveries",
                "value": self._number_display(total_deliveries),
                "help": "Normal + assisted + C-section",
                "class": "green",
            },
            {
                "label": "C-section Rate",
                "value": self._percent_display(c_section_rate),
                "help": "C-sections divided by total reported deliveries",
                "class": "orange",
            },
            {
                "label": "Stillbirths",
                "value": self._number_display(totals["stillbirth"]),
                "help": f"{self._rate_display(stillbirth_rate)} per 1,000 reported deliveries",
                "class": "red",
            },
        ]

        chart_data = {"all_hmis_indicator_trends": []}

        return {
            "summary": summary,
            "totals": totals,
            "kpis": kpis,
            "period_range": period_range,
            "selected_indicator": selected_indicator,
            "selected_indicator_label": selected_label,
            "data_completeness": data_completeness,
            "total_deliveries": total_deliveries,
            "anc4_ratio": anc4_ratio,
            "pnc2_ratio": pnc2_ratio,
            "c_section_rate": c_section_rate,
            "stillbirth_rate": stillbirth_rate,
            "chart_data": chart_data,
        }

    @staticmethod
    def _style_worksheet(worksheet, freeze="A2"):
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top")
        worksheet.freeze_panes = freeze
        if worksheet.max_row >= 2 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[letter].width = min(max(length + 2, 12), 42)

    @staticmethod
    def _style_large_worksheet(worksheet, freeze="A2"):
        """Style large raw-data sheets without scanning/styling every cell."""
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            letter = get_column_letter(cell.column)
            worksheet.column_dimensions[letter].width = min(
                max(len(str(cell.value)) + 2, 12), 28
            )
        worksheet.freeze_panes = freeze
        if worksheet.max_row >= 2 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

    @staticmethod
    def _datetime_display(value):
        if value is None:
            return ""
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime("%Y-%m-%d %H:%M:%S")

    def _append_matrix_sheet(
        self,
        workbook,
        title,
        leading_headers,
        leading_keys,
        rows,
        indicator_columns,
    ):
        worksheet = workbook.create_sheet(title)
        worksheet.append(
            list(leading_headers)
            + [
                f"{column['title']} [{column['indicator_code']}]"
                for column in indicator_columns
            ]
        )
        for row in rows:
            worksheet.append(
                [row[key] for key in leading_keys]
                + [cell["raw"] for cell in row["indicator_values"]]
            )
        self._style_worksheet(worksheet)
        return worksheet

    EXCEL_MAX_ROWS = 1_048_576
    EXCEL_MAX_CELL_TEXT = 32_767

    @classmethod
    def _excel_value(cls, value):
        """Return an Excel-safe scalar without retaining extra cell objects."""
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, str) and len(value) > cls.EXCEL_MAX_CELL_TEXT:
            return value[: cls.EXCEL_MAX_CELL_TEXT]
        return value

    @staticmethod
    def _streaming_header(worksheet, headers, widths=None):
        """Append a styled header to an openpyxl write-only worksheet."""
        worksheet.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        header_cells = []
        for index, header in enumerate(headers, start=1):
            cell = WriteOnlyCell(worksheet, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cells.append(cell)
            width = widths[index - 1] if widths and index <= len(widths) else 18
            worksheet.column_dimensions[get_column_letter(index)].width = width
        worksheet.append(header_cells)

    @staticmethod
    def _finish_streaming_sheet(worksheet, row_count, column_count):
        if column_count:
            worksheet.auto_filter.ref = (
                f"A1:{get_column_letter(column_count)}{max(row_count + 1, 1)}"
            )

    def _append_streaming_matrix_sheet(
        self,
        workbook,
        title,
        fact_queryset,
        indicator_columns,
        group_fields,
        leading_headers,
        leading_values,
    ):
        """Write one dynamic dashboard matrix while holding only one row."""
        headers = list(leading_headers) + [
            f"{column['title']} [{column['indicator_code']}]"
            for column in indicator_columns
        ]
        if len(headers) > 16_384:
            raise ValueError(
                "The HMIS indicator catalogue exceeds Excel's 16,384-column "
                "worksheet limit. Reduce duplicate indicator codes first."
            )

        worksheet = workbook.create_sheet(title)
        self._streaming_header(worksheet, headers)
        indicator_indexes = {
            column["indicator_code"]: index
            for index, column in enumerate(indicator_columns)
        }
        grouped = (
            fact_queryset.values(*group_fields, "indicator_code")
            .annotate(
                reported_total=Sum("value"),
                reported_values=Count("value"),
            )
            .order_by(*group_fields, "indicator_code")
        )

        row_count = 0
        current_key = None
        current_values = None

        def append_current():
            nonlocal row_count
            if current_key is None:
                return
            worksheet.append(
                [self._excel_value(value) for value in leading_values(current_key)]
                + [self._excel_value(value) for value in current_values]
            )
            row_count += 1

        for item in grouped.iterator(chunk_size=5000):
            key = tuple(item[field] for field in group_fields)
            if key != current_key:
                append_current()
                current_key = key
                current_values = [None] * len(indicator_columns)
            code = item["indicator_code"] or "Not coded"
            index = indicator_indexes.get(code)
            if index is not None and item["reported_values"]:
                current_values[index] = item["reported_total"]
        append_current()
        self._finish_streaming_sheet(worksheet, row_count, len(headers))

    def _new_split_sheet(self, workbook, base_title, part, headers, widths=None):
        title = base_title if part == 1 else f"{base_title[:27]} {part}"
        worksheet = workbook.create_sheet(title)
        self._streaming_header(worksheet, headers, widths)
        return worksheet

    def _append_complete_fact_sheets(self, workbook, fact_queryset):
        headers = [
            "Fact ID",
            "Source Upload ID",
            "Province",
            "District",
            "Health Facility",
            "Period Code",
            "Year",
            "Month",
            "Month Name",
            "Cleaned Facility Name",
            "HF ID",
            "HIVA HF",
            "Indicator Code",
            "Indicator Name",
            "Value",
            "Created At",
        ]
        value_fields = (
            "id",
            "source_upload_id",
            "prov",
            "dist",
            "hf",
            "periodcode",
            "year",
            "month",
            "month_name",
            "hf_name_cleaned",
            "hfid",
            "hiva_hfs",
            "indicator_code",
            "indicator_name",
            "value",
            "created_at",
        )
        part = 1
        row_count = 0
        worksheet = self._new_split_sheet(
            workbook, "Complete HMIS Facts", part, headers
        )
        # Raw facts do not need presentation sorting.  Avoiding ORDER BY here
        # prevents PostgreSQL from sorting the complete fact table before the
        # first export row can be written.
        rows = fact_queryset.order_by().values(*value_fields).iterator(
            chunk_size=5000
        )
        for row in rows:
            if row_count >= self.EXCEL_MAX_ROWS - 1:
                self._finish_streaming_sheet(worksheet, row_count, len(headers))
                part += 1
                row_count = 0
                worksheet = self._new_split_sheet(
                    workbook, "Complete HMIS Facts", part, headers
                )
            worksheet.append(
                [
                    row["id"],
                    row["source_upload_id"],
                    self._excel_value(row["prov"]),
                    self._excel_value(row["dist"]),
                    self._excel_value(row["hf"]),
                    row["periodcode"],
                    row["year"],
                    row["month"],
                    row["month_name"],
                    self._excel_value(row["hf_name_cleaned"]),
                    row["hfid"],
                    "Yes" if row["hiva_hfs"] else "No",
                    self._excel_value(row["indicator_code"]),
                    self._excel_value(row["indicator_name"]),
                    self._excel_value(row["value"]),
                    self._datetime_display(row["created_at"]),
                ]
            )
            row_count += 1
        self._finish_streaming_sheet(worksheet, row_count, len(headers))

    def _append_complete_summary_sheets(self, workbook, summary_queryset):
        headers = [
            "Summary ID",
            "Source Upload ID",
            "Province",
            "District",
            "Health Facility",
            "Period Code",
            "Year",
            "Month",
            "Month Name",
            "HF ID",
            "HIVA HF",
        ] + [label for _field, label in self.INDICATORS] + ["Created At"]
        value_fields = (
            "id",
            "source_upload_id",
            "prov",
            "dist",
            "hf",
            "periodcode",
            "year",
            "month",
            "month_name",
            "hfid",
            "hiva_hfs",
            *self.INDICATOR_FIELDS,
            "created_at",
        )
        part = 1
        row_count = 0
        worksheet = self._new_split_sheet(
            workbook, "Complete Monthly Summary", part, headers
        )
        rows = summary_queryset.order_by().values(*value_fields).iterator(
            chunk_size=5000
        )
        for row in rows:
            if row_count >= self.EXCEL_MAX_ROWS - 1:
                self._finish_streaming_sheet(worksheet, row_count, len(headers))
                part += 1
                row_count = 0
                worksheet = self._new_split_sheet(
                    workbook, "Complete Monthly Summary", part, headers
                )
            worksheet.append(
                [
                    row["id"],
                    row["source_upload_id"],
                    self._excel_value(row["prov"]),
                    self._excel_value(row["dist"]),
                    self._excel_value(row["hf"]),
                    row["periodcode"],
                    row["year"],
                    row["month"],
                    row["month_name"],
                    row["hfid"],
                    "Yes" if row["hiva_hfs"] else "No",
                ]
                + [self._excel_value(row[field]) for field in self.INDICATOR_FIELDS]
                + [self._datetime_display(row["created_at"])]
            )
            row_count += 1
        self._finish_streaming_sheet(worksheet, row_count, len(headers))

    def _legacy_export_csv_unused(self, fact_queryset):
        """Stream every unfiltered HMIS fact row without building a file in RAM."""
        logger.info(
            "HMIS export build %s: complete CSV export started",
            HMIS_EXPORT_BUILD,
        )

        headers = (
            "Fact ID",
            "Source Upload ID",
            "Province",
            "District",
            "Health Facility",
            "Period Code",
            "Year",
            "Month",
            "Month Name",
            "Cleaned Facility Name",
            "HF ID",
            "HIVA HF",
            "Indicator Code",
            "Indicator Name",
            "Value",
            "Created At",
        )
        value_fields = (
            "id",
            "source_upload_id",
            "prov",
            "dist",
            "hf",
            "periodcode",
            "year",
            "month",
            "month_name",
            "hf_name_cleaned",
            "hfid",
            "hiva_hfs",
            "indicator_code",
            "indicator_name",
            "value",
            "created_at",
        )

        class CsvEcho:
            """Supply csv.writer's file interface while returning each row."""

            @staticmethod
            def write(value):
                return value

        writer = csv.writer(CsvEcho())

        def stream_rows():
            # The BOM makes Excel recognize UTF-8 province/facility names.
            yield "\ufeff"
            yield writer.writerow(headers)
            row_count = 0
            try:
                rows = (
                    fact_queryset.order_by()
                    .values_list(*value_fields)
                    .iterator(chunk_size=10000)
                )
                for row in rows:
                    output = list(row)
                    output[11] = "Yes" if output[11] else "No"
                    output[15] = self._datetime_display(output[15])
                    yield writer.writerow(output)
                    row_count += 1
                logger.info(
                    "HMIS export build %s: complete CSV export finished (%s rows)",
                    HMIS_EXPORT_BUILD,
                    row_count,
                )
            except Exception:
                logger.exception(
                    "HMIS export build %s: CSV row streaming failed after %s rows",
                    HMIS_EXPORT_BUILD,
                    row_count,
                )
                raise

        timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
        response = StreamingHttpResponse(
            stream_rows(),
            content_type="text/csv; charset=utf-8",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="HMIS_Complete_Facts_{timestamp}.csv"'
        )
        response["Cache-Control"] = "no-store"
        response["X-Accel-Buffering"] = "no"
        response["X-Content-Type-Options"] = "nosniff"
        response["X-HMIS-Export-Build"] = HMIS_EXPORT_BUILD
        return response

    def _legacy_complete_excel_unused(self, fact_queryset, summary_queryset):
        """Stream a complete, unfiltered workbook with bounded memory use."""
        logger.info(
            "HMIS export build %s: complete Excel export started",
            HMIS_EXPORT_BUILD,
        )
        data = self._build_dashboard_data(summary_queryset, "anc1")
        indicator_columns = self._export_indicator_columns(fact_queryset)
        logger.info(
            "HMIS export build %s: %s indicator columns loaded",
            HMIS_EXPORT_BUILD,
            len(indicator_columns),
        )
        workbook = openpyxl.Workbook(write_only=True)

        summary_sheet = workbook.create_sheet("Dashboard Summary")
        summary_sheet.column_dimensions["A"].width = 28
        summary_sheet.column_dimensions["B"].width = 30
        summary_sheet.column_dimensions["C"].width = 55
        title_cell = WriteOnlyCell(summary_sheet, value="HMIS PERFORMANCE DASHBOARD")
        title_cell.fill = PatternFill("solid", fgColor="1F4E78")
        title_cell.font = Font(color="FFFFFF", bold=True, size=16)
        summary_sheet.append([title_cell])
        summary_sheet.append(
            ["Generated", timezone.localtime().strftime("%Y-%m-%d %H:%M")]
        )
        summary_sheet.append(
            ["Export Scope", "Complete HMIS dataset — dashboard filters ignored"]
        )
        summary_sheet.append(["Reporting Period", data["period_range"]])
        summary_sheet.append([])
        kpi_cells = []
        for heading in ("KPI", "Result", "Definition"):
            cell = WriteOnlyCell(summary_sheet, value=heading)
            cell.fill = PatternFill("solid", fgColor="5B9BD5")
            cell.font = Font(color="FFFFFF", bold=True)
            kpi_cells.append(cell)
        summary_sheet.append(kpi_cells)
        for kpi in data["kpis"]:
            summary_sheet.append([kpi["label"], kpi["value"], kpi["help"]])

        self._append_streaming_matrix_sheet(
            workbook,
            "Monthly All Indicators",
            fact_queryset,
            indicator_columns,
            ("periodcode",),
            ("Period Code", "Period"),
            lambda key: (key[0], self._period_label(key[0])),
        )
        logger.info("HMIS export: monthly indicator matrix completed")
        self._append_streaming_matrix_sheet(
            workbook,
            "Province All Indicators",
            fact_queryset,
            indicator_columns,
            ("prov",),
            ("Province",),
            lambda key: (key[0] or "Not specified",),
        )
        logger.info("HMIS export: province indicator matrix completed")
        self._append_streaming_matrix_sheet(
            workbook,
            "Facility All Indicators",
            fact_queryset,
            indicator_columns,
            ("prov", "hf", "hiva_hfs"),
            ("Province", "Health Facility", "Facility Group"),
            lambda key: (
                key[0] or "Not specified",
                key[1] or "Not specified",
                "HIVA HF" if key[2] else "Non-HIVA HF",
            ),
        )
        logger.info("HMIS export: facility indicator matrix completed")

        trend_headers = [
            "Indicator Code",
            "Indicator Name",
            "Indicator Group",
            "Indicator Domain",
            "Reporting Month",
            "Reported Total",
        ]
        trends_sheet = workbook.create_sheet("All Indicator Trends")
        self._streaming_header(trends_sheet, trend_headers)
        cards_by_code = {
            card["indicator_code"]: card for card in indicator_columns
        }
        trend_rows = (
            fact_queryset.values("indicator_code", "periodcode")
            .annotate(
                indicator_name_max=Max("indicator_name"),
                reported_total=Sum("value"),
                reported_values=Count("value"),
            )
            .order_by("indicator_code", "periodcode")
        )
        trend_count = 0
        for row in trend_rows.iterator(chunk_size=5000):
            code = row["indicator_code"] or "Not coded"
            card = cards_by_code.get(code, {})
            trends_sheet.append(
                [
                    code,
                    card.get("full_name") or row["indicator_name_max"] or code,
                    card.get("group") or "Not classified",
                    card.get("domain") or "Not classified",
                    self._period_label(row["periodcode"]),
                    self._excel_value(row["reported_total"])
                    if row["reported_values"]
                    else None,
                ]
            )
            trend_count += 1
        self._finish_streaming_sheet(
            trends_sheet, trend_count, len(trend_headers)
        )
        logger.info("HMIS export: indicator trend sheet completed")

        self._append_complete_fact_sheets(workbook, fact_queryset)
        logger.info("HMIS export: complete fact sheets completed")
        self._append_complete_summary_sheets(workbook, summary_queryset)
        logger.info("HMIS export: complete summary sheets completed")

        metadata_fields = (
            "indicator_code",
            "indicator_name",
            "indicator_short_name",
            "indicator_group",
            "indicator_domain",
            "indicator_description",
            "numerator_definition",
            "denominator_definition",
            "unit_of_measure",
            "reporting_level",
            "data_source",
            "is_active",
            "sort_order",
            "created_at",
            "updated_at",
        )
        metadata_headers = [
            field.replace("_", " ").title() for field in metadata_fields
        ]
        metadata_sheet = workbook.create_sheet("Indicator Metadata")
        self._streaming_header(metadata_sheet, metadata_headers)
        metadata_count = 0
        metadata_rows = IndicatorMetadata.objects.values(*metadata_fields).order_by(
            "sort_order", "indicator_name"
        )
        for row in metadata_rows.iterator(chunk_size=2000):
            metadata_sheet.append(
                [
                    self._datetime_display(row[field])
                    if field in {"created_at", "updated_at"}
                    else self._excel_value(row[field])
                    for field in metadata_fields
                ]
            )
            metadata_count += 1
        self._finish_streaming_sheet(
            metadata_sheet, metadata_count, len(metadata_headers)
        )

        upload_headers = [
            "Upload ID",
            "Spreadsheet Name",
            "File Path",
            "Uploaded By",
            "Uploaded At",
            "Status",
            "Rows Count",
            "Health Facility Count",
            "Start Month",
            "End Month",
            "Import Report",
        ]
        upload_sheet = workbook.create_sheet("Upload Register")
        self._streaming_header(upload_sheet, upload_headers)
        upload_count = 0
        upload_rows = HMISRawUpload.objects.values(
            "id",
            "title",
            "file",
            "uploaded_by_id",
            "uploaded_at",
            "status",
            "row_count",
            "hf_count",
            "period_min",
            "period_max",
            "report",
        ).order_by("id")
        for row in upload_rows.iterator(chunk_size=1000):
            upload_sheet.append(
                [
                    row["id"],
                    self._excel_value(row["title"]),
                    self._excel_value(row["file"]),
                    row["uploaded_by_id"] or "",
                    self._datetime_display(row["uploaded_at"]),
                    row["status"],
                    row["row_count"],
                    row["hf_count"],
                    row["period_min"],
                    row["period_max"],
                    self._excel_value(
                        json.dumps(row["report"], ensure_ascii=False, default=str)
                    ),
                ]
            )
            upload_count += 1
        self._finish_streaming_sheet(
            upload_sheet, upload_count, len(upload_headers)
        )

        # Write directly to disk.  SpooledTemporaryFile initially keeps bytes
        # in RAM and can still create a memory spike while openpyxl finalizes
        # the ZIP container on a small production instance.
        temp_file = tempfile.NamedTemporaryFile(
            prefix="hmis_complete_export_",
            suffix=".xlsx",
            delete=False,
        )
        temp_path = temp_file.name
        temp_file.close()
        try:
            workbook.save(temp_path)
            content_length = os.path.getsize(temp_path)
            export_stream = open(temp_path, "rb")
            timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
            response = FileResponse(
                export_stream,
                as_attachment=True,
                filename=f"HMIS_Performance_Dashboard_{timestamp}.xlsx",
                content_type=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )
            response["Content-Length"] = str(content_length)
            response["X-Content-Type-Options"] = "nosniff"
            response["X-HMIS-Export-Build"] = HMIS_EXPORT_BUILD

            # FileResponse closes the open handle when transmission finishes.
            # Register a second closer to remove the temporary file afterward.
            response._resource_closers.append(
                lambda path=temp_path: os.path.exists(path) and os.remove(path)
            )
            logger.info(
                "HMIS export build %s: complete Excel export ready (%s bytes)",
                HMIS_EXPORT_BUILD,
                content_length,
            )
            return response
        except Exception:
            logger.exception(
                "HMIS export build %s: complete Excel export generation failed",
                HMIS_EXPORT_BUILD,
            )
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise

    def _filtered_excel_row_limit(self):
        """Return a deployment-safe limit, never exceeding Excel's hard cap."""
        configured_limit = getattr(
            settings,
            "HMIS_EXCEL_MAX_DETAIL_ROWS",
            250_000,
        )
        try:
            configured_limit = int(configured_limit)
        except (TypeError, ValueError):
            configured_limit = 250_000
        return min(max(configured_limit, 1), self.EXCEL_MAX_ROWS - 1)

    @staticmethod
    def _excel_text(value):
        """Keep user-supplied text as text instead of an Excel formula."""
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    def _export_filtered_excel(self, fact_queryset, filters, detail_row_count):
        """Create a two-sheet Excel workbook from the active dashboard filters."""
        logger.info(
            "HMIS export build %s: filtered Excel export started (%s rows)",
            HMIS_EXPORT_BUILD,
            detail_row_count,
        )

        workbook = openpyxl.Workbook(write_only=True)

        # ------------------------------------------------------------------
        # Sheet 1: export scope plus one aggregate row for every indicator.
        # ------------------------------------------------------------------
        summary_sheet = workbook.create_sheet("Summary")
        summary_sheet.freeze_panes = "A2"
        summary_widths = (25, 34, 24, 24, 20, 18, 18, 18)
        for index, width in enumerate(summary_widths, start=1):
            summary_sheet.column_dimensions[get_column_letter(index)].width = width

        title_cell = WriteOnlyCell(summary_sheet, value="HMIS FILTERED DATA EXPORT")
        title_cell.fill = PatternFill("solid", fgColor="1F4E78")
        title_cell.font = Font(color="FFFFFF", bold=True, size=15)
        summary_sheet.append([title_cell])
        summary_sheet.append(
            ["Generated", timezone.localtime().strftime("%Y-%m-%d %H:%M")]
        )
        summary_sheet.append(["Export Build", HMIS_EXPORT_BUILD])
        summary_sheet.append(["Filtered Detail Rows", detail_row_count])
        summary_sheet.append([])

        group_labels = {
            "all": "All Facilities",
            "hiva": "HIVA Health Facilities",
            "non_hiva": "Non-HIVA Health Facilities",
        }
        filter_rows = (
            (
                "Facility Group",
                group_labels.get(filters["facility_group"], "All Facilities"),
            ),
            ("Province", filters["prov"] or "All Provinces"),
            ("District", filters["dist"] or "All Districts"),
            ("Health Facility", filters["hf"] or "All Health Facilities"),
            ("Year", filters["year"] if filters["year"] is not None else "All Years"),
            (
                "Month",
                month_name[filters["month"]]
                if filters["month"] in range(1, 13)
                else "All Months",
            ),
            (
                "Selected Summary Indicator",
                self.INDICATOR_LABELS.get(filters["indicator"], filters["indicator"]),
            ),
        )
        filter_header = []
        for heading in ("Applied Filter", "Selection"):
            cell = WriteOnlyCell(summary_sheet, value=heading)
            cell.fill = PatternFill("solid", fgColor="5B9BD5")
            cell.font = Font(color="FFFFFF", bold=True)
            filter_header.append(cell)
        summary_sheet.append(filter_header)
        for label, value in filter_rows:
            summary_sheet.append([label, self._excel_text(value)])
        summary_sheet.append([])

        metadata = {
            row["indicator_code"]: row
            for row in IndicatorMetadata.objects.values(
                "indicator_code",
                "indicator_name",
                "indicator_short_name",
                "indicator_group",
                "indicator_domain",
            )
        }
        indicator_headers = (
            "Indicator Code",
            "Indicator Name",
            "Indicator Group",
            "Indicator Domain",
            "Total Reported Value",
            "Reporting Rows",
            "Health Facilities",
            "Reporting Months",
        )
        indicator_header_cells = []
        for heading in indicator_headers:
            cell = WriteOnlyCell(summary_sheet, value=heading)
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            indicator_header_cells.append(cell)
        summary_sheet.append(indicator_header_cells)

        indicator_summary = (
            fact_queryset.values("indicator_code")
            .annotate(
                indicator_name_max=Max("indicator_name"),
                reported_total=Sum("value"),
                reporting_rows=Count("value"),
                health_facilities=Count("hf", distinct=True),
                reporting_months=Count("periodcode", distinct=True),
            )
            .order_by("indicator_code")
        )
        indicator_count = 0
        for row in indicator_summary.iterator(chunk_size=2000):
            code = row["indicator_code"] or "Not coded"
            meta = metadata.get(code, {})
            indicator_name = (
                meta.get("indicator_short_name")
                or meta.get("indicator_name")
                or row["indicator_name_max"]
                or code
            )
            summary_sheet.append(
                [
                    self._excel_text(code),
                    self._excel_text(indicator_name),
                    self._excel_text(meta.get("indicator_group") or "Not classified"),
                    self._excel_text(meta.get("indicator_domain") or "Not classified"),
                    self._excel_value(row["reported_total"]),
                    row["reporting_rows"],
                    row["health_facilities"],
                    row["reporting_months"],
                ]
            )
            indicator_count += 1

        # ------------------------------------------------------------------
        # Sheet 2: filtered long-format facts behind the dashboard results.
        # ------------------------------------------------------------------
        detail_headers = (
            "Fact ID",
            "Source Upload ID",
            "Province",
            "District",
            "Health Facility",
            "Period Code",
            "Year",
            "Month",
            "Month Name",
            "Cleaned Facility Name",
            "HF ID",
            "HIVA HF",
            "Indicator Code",
            "Indicator Name",
            "Value",
            "Created At",
        )
        detail_fields = (
            "id",
            "source_upload_id",
            "prov",
            "dist",
            "hf",
            "periodcode",
            "year",
            "month",
            "month_name",
            "hf_name_cleaned",
            "hfid",
            "hiva_hfs",
            "indicator_code",
            "indicator_name",
            "value",
            "created_at",
        )
        details_sheet = workbook.create_sheet("Details")
        self._streaming_header(
            details_sheet,
            detail_headers,
            (12, 16, 18, 18, 34, 13, 10, 10, 14, 34, 14, 12, 18, 48, 14, 20),
        )
        detail_rows = (
            fact_queryset.order_by()
            .values_list(*detail_fields)
            .iterator(chunk_size=5000)
        )
        written_rows = 0
        for row in detail_rows:
            output = list(row)
            for index in (2, 3, 4, 5, 8, 9, 10, 12, 13):
                output[index] = self._excel_text(output[index])
            output[11] = "Yes" if output[11] else "No"
            output[14] = self._excel_value(output[14])
            output[15] = self._datetime_display(output[15])
            details_sheet.append(output)
            written_rows += 1
        self._finish_streaming_sheet(details_sheet, written_rows, len(detail_headers))

        temp_file = tempfile.NamedTemporaryFile(
            prefix="hmis_filtered_export_",
            suffix=".xlsx",
            delete=False,
        )
        temp_path = temp_file.name
        temp_file.close()
        try:
            workbook.save(temp_path)
            content_length = os.path.getsize(temp_path)
            export_stream = open(temp_path, "rb")
            timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
            response = FileResponse(
                export_stream,
                as_attachment=True,
                filename=f"HMIS_Filtered_Export_{timestamp}.xlsx",
                content_type=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )
            response["Content-Length"] = str(content_length)
            response["Cache-Control"] = "no-store"
            response["X-Content-Type-Options"] = "nosniff"
            response["X-HMIS-Export-Build"] = HMIS_EXPORT_BUILD
            response._resource_closers.append(
                lambda path=temp_path: os.path.exists(path) and os.remove(path)
            )
            logger.info(
                "HMIS export build %s: filtered Excel ready "
                "(%s detail rows, %s indicators, %s bytes)",
                HMIS_EXPORT_BUILD,
                written_rows,
                indicator_count,
                content_length,
            )
            return response
        except Exception:
            logger.exception(
                "HMIS export build %s: filtered Excel generation failed",
                HMIS_EXPORT_BUILD,
            )
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise

    def changelist_view(self, request, extra_context=None):
        requested_filters = self._read_filters(request)
        filtered_excel_export = request.GET.get("export") == "xlsx"
        complete_print = request.GET.get("print") == "all"

        # Export before building chart/table matrices. This keeps the request
        # lightweight and guarantees that the workbook uses the active filters.
        if filtered_excel_export:
            filtered_facts = self._filter_fact_queryset(
                HMISFact.objects.all(),
                requested_filters,
            )
            detail_row_count = filtered_facts.count()
            export_limit = self._filtered_excel_row_limit()
            redirect_params = request.GET.copy()
            redirect_params.pop("export", None)
            redirect_url = request.path
            if redirect_params:
                redirect_url = f"{request.path}?{redirect_params.urlencode()}"

            if detail_row_count == 0:
                self.message_user(
                    request,
                    (
                        "Excel export was not started because no HMIS detail rows "
                        "match the selected filters. Please change one or more "
                        "filters and try again."
                    ),
                    level=messages.WARNING,
                )
                return HttpResponseRedirect(redirect_url)

            if detail_row_count > export_limit:
                self.message_user(
                    request,
                    (
                        "Excel export was not started. The selected filters return "
                        f"{detail_row_count:,} detail rows, which exceeds the safe "
                        f"export limit of {export_limit:,} rows. Please narrow the "
                        "selection by Year, Month, Province, District, or Health "
                        "Facility, then try again."
                    ),
                    level=messages.WARNING,
                )
                return HttpResponseRedirect(redirect_url)

            try:
                return self._export_filtered_excel(
                    filtered_facts,
                    requested_filters,
                    detail_row_count,
                )
            except Exception:
                logger.exception(
                    "HMIS export build %s: filtered Excel request failed",
                    HMIS_EXPORT_BUILD,
                )
                self.message_user(
                    request,
                    (
                        "The filtered Excel file could not be created. Please "
                        "narrow the filters and try again. If the problem continues, "
                        "contact the system administrator and mention export build "
                        f"{HMIS_EXPORT_BUILD}."
                    ),
                    level=messages.ERROR,
                )
                return HttpResponseRedirect(redirect_url)

        # Print/PDF intentionally keeps its established complete-data scope.
        # Excel export above uses requested_filters instead.
        if complete_print:
            filters = {
                "prov": "",
                "dist": "",
                "hf": "",
                "year": None,
                "month": None,
                "facility_group": "all",
                "indicator": "anc1",
            }
        else:
            filters = requested_filters

        base_queryset = HMISMonthlySummary.objects.all()
        queryset = self._filter_queryset(base_queryset, filters)
        data = self._build_dashboard_data(queryset, filters["indicator"])
        fact_queryset = self._filter_fact_queryset(HMISFact.objects.all(), filters)
        all_indicator_cards, all_indicator_charts = (
            self._all_hmis_indicator_trends(fact_queryset)
        )
        data["all_hmis_indicator_cards"] = all_indicator_cards
        data["chart_data"]["all_hmis_indicator_trends"] = all_indicator_charts
        data.update(
            self._fact_detail_tables(
                fact_queryset,
                all_indicator_cards,
                filters["indicator"],
            )
        )

        options = self._filter_options(base_queryset, filters)

        export_params = request.GET.copy()
        export_params.pop("print", None)
        export_params["export"] = "xlsx"

        context = {
            **self.admin_site.each_context(request),
            "title": "HMIS Performance Dashboard",
            "opts": self.model._meta,
            "filters": filters,
            "filter_options": options,
            "indicator_options": self.INDICATORS,
            "export_url": f"?{export_params.urlencode()}",
            "excel_export_limit_display": f"{self._filtered_excel_row_limit():,}",
            "print_url": "?print=all",
            "print_all_mode": complete_print,
            **data,
        }
        if extra_context:
            context.update(extra_context)
        request.current_app = self.admin_site.name
        return TemplateResponse(request, self.change_list_template, context)

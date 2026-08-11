from django.contrib import admin
from .models import (
    Gancohort,
    Gancenrollment,
    Gancfirstsession,
    Gancsecondsession,
    Gancthirdsession,
    Gancfouthsession,
    Gancdelivery,
    GroupPncfirstSession,
    GroupPncsecondSession,
)
from hiva.models import Facility
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django import forms

# ============================================================
# CUSTOM ADMIN MODEL ORDER
# GANC / PNC CLINICAL WORKFLOW
# ============================================================

from django.contrib import admin

# Keep a reference to Django's original get_app_list method
_original_get_app_list = admin.AdminSite.get_app_list

def custom_get_app_list(self, request, app_label=None):
    """
    Custom ordering for models inside Django Admin.

    GANC/PNC workflow:
        1. Cohort
        2. Enrollment
        3. ANC First Session
        4. ANC Second Session
        5. ANC Third Session
        6. ANC Fourth Session
        7. Delivery
        8. PNC First Session
        9. PNC Second Session

    Other applications/models keep their normal ordering.
    """

    app_list = _original_get_app_list(
        self,
        request,
        app_label,
    )

    # --------------------------------------------------------
    # Desired GANC/PNC ordering
    # Use model class names, not verbose names.
    # --------------------------------------------------------

    ganc_model_order = {
        "Gancohort": 1,
        "Gancenrollment": 2,
        "Gancfirstsession": 3,
        "Gancsecondsession": 4,
        "Gancthirdsession": 5,
        "Gancfouthsession": 6,
        "Gancdelivery": 7,
        "GroupPncfirstSession": 8,
        "GroupPncsecondSession": 9,
    }

    # --------------------------------------------------------
    # Find GANC/PNC application
    # --------------------------------------------------------

    for app in app_list:

        if app.get("app_label") == "gancgpnc":

            app["models"].sort(
                key=lambda model: ganc_model_order.get(
                    model.get("object_name"),
                    999,
                )
            )

    return app_list

# Apply custom ordering
admin.AdminSite.get_app_list = custom_get_app_list

# ============================================================
# Province helper (same style as hiva admin.py)
# ============================================================

def user_province(request):
    if request.user.is_superuser:
        return None
    profile = getattr(request.user, "profile", None) or getattr(request.user, "userprofile", None)
    return getattr(profile, "province", None)

class ProvinceRestrictedAdminMixin:
    """
    Universal restriction for province-based access.
    Subclasses must implement:
      - province_filter_kwargs(request) -> dict of filters
    """

    def province_filter_kwargs(self, request):
        raise NotImplementedError

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        prov = user_province(request)
        if prov is None:
            return qs.none()
        return qs.filter(**self.province_filter_kwargs(request))

    def _obj_in_scope(self, request, obj):
        return self.get_queryset(request).filter(pk=obj.pk).exists()

    def has_view_permission(self, request, obj=None):
        base = super().has_view_permission(request, obj=obj)
        if not base:
            return False
        if obj is None or request.user.is_superuser:
            return True
        return self._obj_in_scope(request, obj)

    def has_change_permission(self, request, obj=None):
        base = super().has_change_permission(request, obj=obj)
        if not base:
            return False
        if obj is None or request.user.is_superuser:
            return True
        return self._obj_in_scope(request, obj)

    def has_delete_permission(self, request, obj=None):
        base = super().has_delete_permission(request, obj=obj)
        if not base:
            return False
        if obj is None or request.user.is_superuser:
            return True
        return self._obj_in_scope(request, obj)

# ============================================================
# Optional reusable province filters
# ============================================================

class GancCohortProvinceFilter(admin.SimpleListFilter):
    title = "Province"
    parameter_name = "province"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)
        rows = qs.values_list(
            "facility__districtfk__provincefk__id",
            "facility__districtfk__provincefk__name",
        ).distinct().order_by("facility__districtfk__provincefk__name")
        return [(pid, pname) for pid, pname in rows if pid]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(facility__districtfk__provincefk_id=self.value())
        return queryset


class EnrollmentProvinceFilter(admin.SimpleListFilter):
    title = "Province"
    parameter_name = "province"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)
        rows = qs.values_list(
            "cohortname__facility__districtfk__provincefk__id",
            "cohortname__facility__districtfk__provincefk__name",
        ).distinct().order_by("cohortname__facility__districtfk__provincefk__name")
        return [(pid, pname) for pid, pname in rows if pid]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(cohortname__facility__districtfk__provincefk_id=self.value())
        return queryset

class SessionProvinceFilter(admin.SimpleListFilter):
    title = "Province"
    parameter_name = "province"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)
        rows = qs.values_list(
            "registerid__cohortname__facility__districtfk__provincefk__id",
            "registerid__cohortname__facility__districtfk__provincefk__name",
        ).distinct().order_by("registerid__cohortname__facility__districtfk__provincefk__name")
        return [(pid, pname) for pid, pname in rows if pid]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(
                registerid__cohortname__facility__districtfk__provincefk_id=self.value()
            )
        return queryset

# ============================================================
# GANC Cohort
# ============================================================

@admin.register(Gancohort)
class GancohortAdmin(ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    list_display = (
        "get_province",
        "facility",
        "cohortname",
        "cohortnumber",
        "cohortstatus",
        "cohortchecklist",
        "target_size",
        "created_by",
        "created_at",
    )

    readonly_fields = (
        "created_at",
        "updated_at",
        "created_by",
        "updated_by",
    )

    list_filter = (
        GancCohortProvinceFilter,
        "facility",
        "cohortstatus",
    )

    search_fields = (
        "cohortname",
        "facility__name",
        "facility__hfcode",
    )

    list_per_page = 20

    fieldsets = (
        ("Cohort Information", {
            "fields": (
                "facility",
                "cohortname",
                "cohortnumber",
                "cohortstatus",
                "cohortchecklist",
                "target_size",
            )
        }),
        ("Other Information", {
            "fields": (
                "remarks",
            )
        }),
        ("Audit Information", {
            "fields": (
                "created_at",
                "updated_at",
                "created_by",
                "updated_by",
            ),
            "classes": ("collapse",),
        }),
    )

    def save_model(self, request, obj, form, change):
        if not change and not obj.created_by:
            obj.created_by = request.user
        obj.updated_by = request.user
        super().save_model(request, obj, form, change)

    @admin.display(description="Province")
    def get_province(self, obj):
        return obj.facility.districtfk.provincefk.name if obj.facility and obj.facility.districtfk else "-"

    def province_filter_kwargs(self, request):
        return {"facility__districtfk__provincefk": user_province(request)}

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "facility" and not request.user.is_superuser:
            prov = user_province(request)
            kwargs["queryset"] = (
                Facility.objects.filter(districtfk__provincefk=prov).order_by("name")
                if prov else Facility.objects.none()
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

# ============================================================
# GANC Enrollment
# ============================================================

@admin.register(Gancenrollment)
class GancenrollmentAdmin(ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    list_display = (
        "name",
        "fathername",
        "cohortname",
        "get_facility",
        "get_province",
        "enrollmentid",
        "contactnumber",
        "education_level",
        "gravida",
        "gafirstanc",
        "edd",
        "age_years",
        "transfer_in",
        "numerof_ancvisits",
    )

    list_filter = (
        EnrollmentProvinceFilter,
        "cohortname",
        "edd",
        "education_level",
        "transfer_in",
    )

    search_fields = (
        "name",
        "fathername",
        "contactnumber",
        "address",
        "education_level",
        "cohortname__cohortname",
        "cohortname__facility__name",
        "cohortname__facility__district__name",
        "cohortname__facility__district__province__name",
    )

    list_per_page = 20

    fieldsets = (
        ("Enrollment Information", {
            "fields": (
                "cohortname",
                "enrollmentid",
                "name",
                "fathername",
                "contactnumber",
                "address",
            )
        }),
        ("Pregnancy Information", {
            "fields": (
                "gravida",
                "gafirstanc",
                "edd",
                "age_years",
                "transfer_in",
                "numerof_ancvisits",
            )
        }),
        ("Background Information", {
            "fields": (
                "education_level",
            )
        }),
        ("Other Information", {
            "fields": (
                "remarks",
            )
        }),
    )

    @admin.display(description="Facility")
    def get_facility(self, obj):
        return obj.cohortname.facility.name if obj.cohortname and obj.cohortname.facility else "-"

    @admin.display(description="Province")
    def get_province(self, obj):
        cohort = obj.cohortname
        if cohort and cohort.facility and cohort.facility.districtfk:
            return cohort.facility.districtfk.provincefk.name
        return "-"

    def province_filter_kwargs(self, request):
        return {"cohortname__facility__districtfk__provincefk": user_province(request)}

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "cohortname" and not request.user.is_superuser:
            prov = user_province(request)
            kwargs["queryset"] = (
                Gancohort.objects.filter(facility__districtfk__provincefk=prov).order_by("cohortname")
                if prov else Gancohort.objects.none()
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


# ============================================================
# Shared helper for session-like admins
# ============================================================

class BaseSessionAdmin(ProvinceRestrictedAdminMixin, admin.ModelAdmin):
    list_per_page = 20

    def province_filter_kwargs(self, request):
        return {"registerid__cohortname__facility__districtfk__provincefk": user_province(request)}

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "registerid" and not request.user.is_superuser:
            prov = user_province(request)
            kwargs["queryset"] = (
                Gancenrollment.objects.filter(
                    cohortname__facility__districtfk__provincefk=prov
                ).order_by("name")
                if prov else Gancenrollment.objects.none()
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    @admin.display(description="Province")
    def get_province(self, obj):
        try:
            return obj.registerid.cohortname.facility.districtfk.provincefk.name
        except Exception:
            return "-"

    @admin.display(description="Facility")
    def get_facility(self, obj):
        try:
            return obj.registerid.cohortname.facility.name
        except Exception:
            return "-"

# ============================================================
# GANC ANC SESSION - ENRICHED REGISTER DROPDOWN
# ============================================================

class GancRegisterChoiceField(forms.ModelChoiceField):
    """
    Display enrollment as:
    Register ID | Woman Name | Father Name | Cohort
    """

    def label_from_instance(self, obj):
        register_id = obj.pk
        name = getattr(obj, "name", "") or ""
        father_name = getattr(obj, "fathername", "") or ""

        cohort = getattr(obj, "cohortname", None)
        cohort_name = str(cohort) if cohort else "-"

        return (
            f"{register_id} | "
            f"{name} | "
            f"{father_name} | "
            f"{cohort_name}"
        )


class BaseGancSessionAdminForm(forms.ModelForm):
    """
    Shared form behavior for all four GANC ANC sessions.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if "registerid" in self.fields:
            original_field = self.fields["registerid"]

            self.fields["registerid"] = GancRegisterChoiceField(
                queryset=original_field.queryset,
                required=original_field.required,
                label="Register Name",
                help_text=(
                    "Select the correct woman using Register ID, "
                    "Name, Father Name and Cohort."
                ),
            )


# ============================================================
# FIRST SESSION FORM
# ============================================================

class GancFirstSessionAdminForm(BaseGancSessionAdminForm):

    class Meta:
        model = Gancfirstsession
        fields = "__all__"


# ============================================================
# SECOND SESSION FORM
# ============================================================

class GancSecondSessionAdminForm(BaseGancSessionAdminForm):

    class Meta:
        model = Gancsecondsession
        fields = "__all__"


# ============================================================
# THIRD SESSION FORM
# ============================================================

class GancThirdSessionAdminForm(BaseGancSessionAdminForm):

    class Meta:
        model = Gancthirdsession
        fields = "__all__"


# ============================================================
# FOURTH SESSION FORM
# ============================================================

class GancFourthSessionAdminForm(BaseGancSessionAdminForm):

    class Meta:
        model = Gancfouthsession
        fields = "__all__"

# ============================================================
# GANC First Session
# ============================================================

@admin.register(Gancfirstsession)
class GancfirstsessionAdmin(BaseSessionAdmin):
    form = GancFirstSessionAdminForm
    list_display = (
        "registerid",
        "get_facility",
        "get_province",
        "sessiontype",
        "sessionround",
        "sessiondate",
        "attendance",
        "presentga",
        "bp",
        "anemia",
        "dangersign",
    )

    list_filter = (
        SessionProvinceFilter,
        "sessiontype",
        "sessionround",
        "sessiondate",
        "attendance",
        "dhypertension",
        "anemia",
        "dangersign",
    )

    search_fields = (
        "registerid__name",
        "registerid__fathername",
        "sessiontype",
        "sessionround",
        "typeofdangersign",
    )

    ordering = ("-sessiondate",)
    list_per_page = 25
    save_on_top = True

    actions = ["export_ganc_first_session_excel"]

    fieldsets = (
        ("Session Information", {
            "fields": (
                "registerid",
                "sessiontype",
                "sessionround",
                "sessiondate",
                "attendance",
                "presentga",
            )
        }),
        ("Maternal Assessment", {
            "fields": (
                "bp",
                "dhypertension",
                "rhypertensiontoMD",
                "weight",
                "anemia",
                "ironfolate",
                "ironfolatepluswomen",
                "pcalcium",
                "acalcium",
                "muac",
                "dmam",
                "rmam",
                "dsam",
                "rsam",
            )
        }),
        ("Laboratory and Screening", {
            "fields": (
                "clabexm",
                "hemoglobin",
                "urinexamcheck",
                "urinexam",
                "rpositivepuriatomd",
                "coughmorethantwoweeks",
                "rcough",
                "ttvaccine",
            )
        }),
        ("Danger Signs", {
            "fields": (
                "dangersign",
                "typeofdangersign",
            )
        }),
        ("Other Information", {
            "fields": ("remarks",)
        }),
    )

    def export_ganc_first_session_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "GANC First Session"

        headers = [
            "Register Name",
            "Facility",
            "Province",
            "Session Type",
            "Session Round",
            "Session Date",
            "Attendance",
            "Present GA",
            "BP",
            "Diagnosed Hypertension",
            "Referred Hypertension to MD",
            "Weight",
            "Anemia",
            "Iron Folate Routine Dose",
            "Iron Folate 30+ for Anemic Woman",
            "Prescribe Calcium",
            "Absorbed Calcium Last Month",
            "MUAC",
            "Diagnosed MAM",
            "Refer MAM to Nutrition Counsellor",
            "Diagnosed SAM",
            "Refer SAM to Higher Level",
            "Completing Laboratory Exam",
            "Hemoglobin",
            "Urine Exam / Protein Uria",
            "Referred Positive Protein Uria to MD",
            "Cough More Than Two Weeks",
            "Referred Cough to DOTS Room",
            "TT Vaccine",
            "Danger Sign",
            "Type of Danger Sign",
            "Remarks",
        ]

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="0F766E")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def yes_no(value):
            return "Yes" if value else "No"

        for obj in queryset:
            ws.append([
                str(obj.registerid) if obj.registerid else "",
                self.get_facility(obj),
                self.get_province(obj),
                obj.sessiontype,
                obj.sessionround,
                obj.sessiondate.strftime("%Y-%m-%d") if obj.sessiondate else "",
                obj.attendance,
                obj.presentga,
                obj.bp,
                yes_no(obj.dhypertension),
                yes_no(obj.rhypertensiontoMD),
                obj.weight,
                yes_no(obj.anemia),
                yes_no(obj.ironfolate),
                yes_no(obj.ironfolatepluswomen),
                yes_no(obj.pcalcium),
                yes_no(obj.acalcium),
                obj.muac,
                yes_no(obj.dmam),
                yes_no(obj.rmam),
                yes_no(obj.dsam),
                yes_no(obj.rsam),
                yes_no(obj.clabexm),
                obj.hemoglobin,
                obj.urinexamcheck,
                obj.urinexam,
                yes_no(obj.rpositivepuriatomd),
                yes_no(obj.coughmorethantwoweeks),
                yes_no(obj.rcough),
                yes_no(obj.ttvaccine),
                yes_no(obj.dangersign),
                obj.typeofdangersign,
                obj.remarks,
            ])

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = min(max_length + 3, 35)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="ganc_first_session.xlsx"'

        wb.save(response)
        return response

    export_ganc_first_session_excel.short_description = "Export selected GANC First Session records to Excel"

# ============================================================
# GANC Second Session
# ============================================================
@admin.register(Gancsecondsession)
class GancsecondsessionAdmin(BaseSessionAdmin):
    form = GancSecondSessionAdminForm
    list_display = (
        "registerid",
        "get_facility",
        "get_province",
        "sessiontype",
        "sessionround",
        "sessiondate",
        "attendance",
        "presentga",
        "bp",
        "anemia",
        "mebendazole",
        "dangersign",
    )

    list_filter = (
        SessionProvinceFilter,
        "sessiontype",
        "sessionround",
        "sessiondate",
        "attendance",
        "dhypertension",
        "anemia",
        "mebendazole",
        "dangersign",
    )

    search_fields = (
        "registerid__name",
        "registerid__fathername",
        "sessiontype",
        "sessionround",
        "typeofdangersign",
    )

    ordering = ("-sessiondate",)
    list_per_page = 25
    save_on_top = True

    actions = ["export_ganc_second_session_excel"]

    readonly_fields = (
        "session_help",
        "maternal_help",
        "lab_help",
        "danger_help",
    )

    fieldsets = (
        ("① Session Information", {
            "classes": ("ganc-section", "wide"),
            "description": "Basic session details. Please confirm the woman, session date, attendance type and gestational age.",
            "fields": (
                "session_help",
                ("registerid", "sessiondate"),
                ("sessiontype", "sessionround"),
                ("attendance", "presentga"),
            )
        }),
        ("② Maternal Assessment", {
            "classes": ("ganc-section", "wide"),
            "description": "Record blood pressure, nutrition, anemia, supplementation and maternal assessment findings.",
            "fields": (
                "maternal_help",
                ("bp", "weight", "muac"),
                ("dhypertension", "rhypertensiontoMD"),
                ("anemia", "ironfolate", "ironfolatepluswomen"),
                ("pcalcium", "acalcium", "mebendazole"),
                ("dmam", "rmam"),
                ("dsam", "rsam"),
            )
        }),
        ("③ Laboratory and Screening", {
            "classes": ("ganc-section", "collapse"),
            "description": "Record urine protein, cough screening, referral and TT vaccine status.",
            "fields": (
                "lab_help",
                ("urinexamcheck", "urinexam", "rpositivepuriatomd"),
                ("coughmorethantwoweeks", "rcough"),
                "ttvaccine",
            )
        }),
        ("④ Danger Signs", {
            "classes": ("ganc-section", "collapse"),
            "description": "If danger sign is Yes, clearly mention the type of danger sign.",
            "fields": (
                "danger_help",
                ("dangersign", "typeofdangersign"),
            )
        }),
        ("⑤ Remarks", {
            "classes": ("ganc-section", "collapse"),
            "fields": ("remarks",)
        }),
    )

    class Media:
        css = {
            "all": ("admin/css/ganc_admin.css",)
        }
        js = ("admin/js/ganc_admin.js",)

    def session_help(self, obj=None):
        return "Use this section to confirm the correct client and second session details."
    session_help.short_description = "Data entry guidance"

    def maternal_help(self, obj=None):
        return "Check BP, weight, MUAC, anemia, supplements, MAM/SAM and referrals carefully."
    maternal_help.short_description = "Maternal assessment guidance"

    def lab_help(self, obj=None):
        return "Urine protein, cough screening and TT vaccine should be completed where applicable."
    lab_help.short_description = "Laboratory guidance"

    def danger_help(self, obj=None):
        return "If danger sign is selected, type the specific danger sign in the next field."
    danger_help.short_description = "Danger sign guidance"

    def export_ganc_second_session_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "GANC Second Session"

        headers = [
            "Register Name",
            "Facility",
            "Province",
            "Session Type",
            "Session Round",
            "Session Date",
            "Attendance",
            "Present GA",
            "BP",
            "Diagnosed Hypertension",
            "Referred Hypertension to MD",
            "Weight",
            "Anemia",
            "Iron Folate Routine Dose",
            "Iron Folate 30+ for Anemic Woman",
            "Prescribe Calcium",
            "Absorbed Calcium Last Month",
            "Mebendazole",
            "MUAC",
            "Diagnosed MAM",
            "Refer MAM to Nutrition Counsellor",
            "Diagnosed SAM",
            "Refer SAM to Higher Level",
            "Urine Exam / Protein Uria",
            "Referred Positive Protein Uria to MD",
            "Cough More Than Two Weeks",
            "Referred Cough to DOTS Room",
            "TT Vaccine",
            "Danger Sign",
            "Type of Danger Sign",
            "Remarks",
        ]

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="0F766E")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def yes_no(value):
            return "Yes" if value else "No"

        for obj in queryset:
            ws.append([
                str(obj.registerid) if obj.registerid else "",
                self.get_facility(obj),
                self.get_province(obj),
                obj.sessiontype,
                obj.sessionround,
                obj.sessiondate.strftime("%Y-%m-%d") if obj.sessiondate else "",
                obj.attendance,
                obj.presentga,
                obj.bp,
                yes_no(obj.dhypertension),
                yes_no(obj.rhypertensiontoMD),
                obj.weight,
                yes_no(obj.anemia),
                yes_no(obj.ironfolate),
                yes_no(obj.ironfolatepluswomen),
                yes_no(obj.pcalcium),
                yes_no(obj.acalcium),
                yes_no(obj.mebendazole),
                obj.muac,
                yes_no(obj.dmam),
                yes_no(obj.rmam),
                yes_no(obj.dsam),
                yes_no(obj.rsam),
                yes_no(obj.urinexamcheck),
                obj.urinexam,
                yes_no(obj.rpositivepuriatomd),
                yes_no(obj.coughmorethantwoweeks),
                yes_no(obj.rcough),
                yes_no(obj.ttvaccine),
                yes_no(obj.dangersign),
                obj.typeofdangersign,
                obj.remarks,
            ])

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = min(max_length + 3, 35)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="ganc_second_session.xlsx"'

        wb.save(response)
        return response

    export_ganc_second_session_excel.short_description = "Export selected GANC Second Session records to Excel"

# ============================================================
# GANC Third Session
# ============================================================

@admin.register(Gancthirdsession)
class GancthirdsessionAdmin(BaseSessionAdmin):
    form = GancThirdSessionAdminForm
    list_display = (
        "registerid", "get_facility", "get_province",
        "sessiontype", "sessionround", "sessiondate",
        "attendance", "presentga", "bp", "anemia",
        "antedepressionscreening", "antedepressiondiagnosed",
        "birthplanningcounseling", "dangersign",
    )

    list_filter = (
        SessionProvinceFilter,
        "sessiontype", "sessionround", "sessiondate",
        "attendance", "dhypertension", "anemia",
        "antedepressionscreening", "antedepressiondiagnosed",
        "rpsychosocialcounselor", "birthplanningcounseling",
        "dangersign",
    )

    search_fields = (
        "registerid__name",
        "registerid__fathername",
        "sessiontype",
        "sessionround",
        "typeofdangersign",
    )

    ordering = ("-sessiondate",)
    list_per_page = 25
    save_on_top = True

    actions = ["export_ganc_third_session_excel"]

    readonly_fields = (
        "session_help",
        "maternal_help",
        "mental_health_help",
        "lab_help",
        "danger_counseling_help",
    )

    fieldsets = (
        ("① Session Information", {
            "classes": ("ganc-section", "wide"),
            "description": "Confirm client, session date, attendance status and gestational age.",
            "fields": (
                "session_help",
                ("registerid", "sessiondate"),
                ("sessiontype", "sessionround"),
                ("attendance", "presentga"),
            )
        }),
        ("② Maternal Assessment", {
            "classes": ("ganc-section", "wide"),
            "description": "Record BP, weight, MUAC, anemia, supplements, nutrition status and referrals.",
            "fields": (
                "maternal_help",
                ("bp", "weight", "muac"),
                ("dhypertension", "rhypertensiontoMD"),
                ("anemia", "ironfolate", "ironfolatepluswomen"),
                ("pcalcium", "acalcium"),
                ("dmam", "rmam"),
                ("dsam", "rsam"),
            )
        }),
        ("③ Mental Health", {
            "classes": ("ganc-section", "wide"),
            "description": "Record antenatal depression screening, diagnosis and referral.",
            "fields": (
                "mental_health_help",
                (
                    "antedepressionscreening",
                    "antedepressiondiagnosed",
                    "rpsychosocialcounselor",
                ),
            )
        }),
        ("④ Laboratory and Screening", {
            "classes": ("ganc-section", "collapse"),
            "description": "Record urine protein, cough screening, referral and TT vaccine.",
            "fields": (
                "lab_help",
                ("urinexam", "rpositivepuriatomd"),
                ("coughmorethantwoweeks", "rcough"),
                "ttvaccine",
            )
        }),
        ("⑤ Danger Signs and Counseling", {
            "classes": ("ganc-section", "collapse"),
            "description": "Record danger signs and birth planning counseling.",
            "fields": (
                "danger_counseling_help",
                ("dangersign", "typeofdangersign"),
                "birthplanningcounseling",
            )
        }),
        ("⑥ Remarks", {
            "classes": ("ganc-section", "collapse"),
            "fields": ("remarks",)
        }),
    )

    class Media:
        css = {
            "all": ("admin/css/ganc_admin.css",)
        }
        js = ("admin/js/ganc_admin.js",)

    def session_help(self, obj=None):
        return "Use this section to confirm the correct woman and third ANC session details."
    session_help.short_description = "Guidance"

    def maternal_help(self, obj=None):
        return "Complete BP, nutrition, anemia, supplements and referral-related fields carefully."
    maternal_help.short_description = "Guidance"

    def mental_health_help(self, obj=None):
        return "If depression is diagnosed, referral to psychosocial counselor should be reviewed."
    mental_health_help.short_description = "Guidance"

    def lab_help(self, obj=None):
        return "Complete urine protein, cough screening and TT vaccine fields where applicable."
    lab_help.short_description = "Guidance"

    def danger_counseling_help(self, obj=None):
        return "If danger sign is Yes, enter the type of danger sign. Also record birth planning counseling."
    danger_counseling_help.short_description = "Guidance"

    def export_ganc_third_session_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "GANC Third Session"

        headers = [
            "Register Name",
            "Facility",
            "Province",
            "Session Type",
            "Session Round",
            "Session Date",
            "Attendance",
            "Present GA",
            "BP",
            "Diagnosed Hypertension",
            "Referred Hypertension to MD",
            "Weight",
            "Anemia",
            "Iron Folate Routine Dose",
            "Iron Folate 30+ for Anemic Woman",
            "Prescribe Calcium",
            "Absorbed Calcium Last Month",
            "MUAC",
            "Diagnosed MAM",
            "Refer MAM to Nutrition Counsellor",
            "Diagnosed SAM",
            "Refer SAM to Higher Level",
            "Antenatal Depression Screening",
            "Antenatal Depression Diagnosed",
            "Refer to Psychosocial Counselor",
            "Urine Exam / Protein Uria",
            "Referred Positive Protein Uria to MD",
            "Cough More Than Two Weeks",
            "Referred Cough to DOTS Room",
            "TT Vaccine",
            "Danger Sign",
            "Type of Danger Sign",
            "Birth Planning Counseling",
            "Remarks",
        ]

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="0F766E")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def yes_no(value):
            return "Yes" if value else "No"

        for obj in queryset:
            ws.append([
                str(obj.registerid) if obj.registerid else "",
                self.get_facility(obj),
                self.get_province(obj),
                obj.sessiontype,
                obj.sessionround,
                obj.sessiondate.strftime("%Y-%m-%d") if obj.sessiondate else "",
                obj.attendance,
                obj.presentga,
                obj.bp,
                yes_no(obj.dhypertension),
                yes_no(obj.rhypertensiontoMD),
                obj.weight,
                yes_no(obj.anemia),
                yes_no(obj.ironfolate),
                yes_no(obj.ironfolatepluswomen),
                yes_no(obj.pcalcium),
                yes_no(obj.acalcium),
                obj.muac,
                yes_no(obj.dmam),
                yes_no(obj.rmam),
                yes_no(obj.dsam),
                yes_no(obj.rsam),
                yes_no(obj.antedepressionscreening),
                yes_no(obj.antedepressiondiagnosed),
                yes_no(obj.rpsychosocialcounselor),
                yes_no(obj.urinexamcheck),
                obj.urinexam,
                yes_no(obj.rpositivepuriatomd),
                yes_no(obj.coughmorethantwoweeks),
                yes_no(obj.rcough),
                yes_no(obj.ttvaccine),
                yes_no(obj.dangersign),
                obj.typeofdangersign,
                yes_no(obj.birthplanningcounseling),
                obj.remarks,
            ])

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = min(max_length + 3, 35)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="ganc_third_session.xlsx"'

        wb.save(response)
        return response

    export_ganc_third_session_excel.short_description = "Export selected GANC Third Session records to Excel"

# ============================================================
# GANC Fourth Session
# ============================================================

@admin.register(Gancfouthsession)
class GancfouthsessionAdmin(BaseSessionAdmin):
    form = GancFourthSessionAdminForm
    list_display = (
        "registerid",
        "get_facility",
        "get_province",
        "sessiontype",
        "sessionround",
        "sessiondate",
        "attendance",
        "presentga",
        "bp",
        "anemia",
        "antedepressionscreening",
        "antedepressiondiagnosed",
        "birthplanningcounseling",
        "dangersign",
    )

    list_filter = (
        SessionProvinceFilter,
        "sessiontype",
        "sessionround",
        "sessiondate",
        "attendance",
        "dhypertension",
        "anemia",
        "antedepressionscreening",
        "antedepressiondiagnosed",
        "rpsychosocialcounselor",
        "birthplanningcounseling",
        "dangersign",
    )

    search_fields = (
        "registerid__name",
        "registerid__fathername",
        "sessiontype",
        "sessionround",
        "typeofdangersign",
    )

    ordering = ("-sessiondate",)
    list_per_page = 25
    save_on_top = True

    actions = ["export_ganc_fourth_session_excel"]

    readonly_fields = (
        "session_help",
        "maternal_help",
        "mental_health_help",
        "lab_help",
        "danger_counseling_help",
    )

    fieldsets = (
        ("① Session Information", {
            "classes": ("ganc-section", "wide"),
            "description": (
                "Confirm the correct woman, fourth-session date, "
                "attendance status and present gestational age."
            ),
            "fields": (
                "session_help",
                ("registerid", "sessiondate"),
                ("sessiontype", "sessionround"),
                ("attendance", "presentga"),
            ),
        }),

        ("② Maternal Assessment", {
            "classes": ("ganc-section", "wide"),
            "description": (
                "Record blood pressure, weight, MUAC, anemia, supplements, "
                "nutritional status and required referrals."
            ),
            "fields": (
                "maternal_help",
                ("bp", "weight", "muac"),
                ("dhypertension", "rhypertensiontoMD"),
                ("anemia", "ironfolate", "ironfolatepluswomen"),
                ("pcalcium", "acalcium"),
                ("dmam", "rmam"),
                ("dsam", "rsam"),
            ),
        }),

        ("③ Mental Health", {
            "classes": ("ganc-section", "wide"),
            "description": (
                "Record antenatal depression screening, diagnosis "
                "and psychosocial referral."
            ),
            "fields": (
                "mental_health_help",
                (
                    "antedepressionscreening",
                    "antedepressiondiagnosed",
                    "rpsychosocialcounselor",
                ),
            ),
        }),

        ("④ Laboratory and Screening", {
            "classes": ("ganc-section", "collapse"),
            "description": (
                "Record urine protein findings, cough screening, "
                "referrals and TT vaccination status."
            ),
            "fields": (
                "lab_help",
                ("urinexamcheck", "urinexam", "rpositivepuriatomd"),
                ("coughmorethantwoweeks", "rcough"),
                "ttvaccine",
            ),
        }),

        ("⑤ Danger Signs and Counseling", {
            "classes": ("ganc-section", "collapse"),
            "description": (
                "Record pregnancy danger signs and birth-planning counseling. "
                "The type of danger sign remains available for data entry."
            ),
            "fields": (
                "danger_counseling_help",
                ("dangersign", "typeofdangersign"),
                "birthplanningcounseling",
            ),
        }),

        ("⑥ Remarks", {
            "classes": ("ganc-section", "collapse"),
            "fields": (
                "remarks",
            ),
        }),
    )

    class Media:
        css = {
            "all": (
                "admin/css/ganc_admin.css",
            )
        }
        js = (
            "admin/js/ganc_admin.js",
        )

    def session_help(self, obj=None):
        return (
            "Confirm the correct woman and complete all fourth ANC "
            "session information."
        )

    session_help.short_description = "Session guidance"

    def maternal_help(self, obj=None):
        return (
            "Carefully complete blood pressure, nutrition, anemia, "
            "supplementation and referral fields."
        )

    maternal_help.short_description = "Maternal assessment guidance"

    def mental_health_help(self, obj=None):
        return (
            "When depression is diagnosed, review whether referral "
            "to a psychosocial counselor is required."
        )

    mental_health_help.short_description = "Mental health guidance"

    def lab_help(self, obj=None):
        return (
            "Complete urine protein, cough screening, referral and "
            "TT vaccine fields where applicable."
        )

    lab_help.short_description = "Laboratory guidance"

    def danger_counseling_help(self, obj=None):
        return (
            "If a danger sign is present, enter the specific type of "
            "danger sign and record birth-planning counseling."
        )

    danger_counseling_help.short_description = "Danger sign guidance"

    def export_ganc_fourth_session_excel(self, request, queryset):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "GANC Fourth Session"

        headers = [
            "Register Name",
            "Facility",
            "Province",
            "Session Type",
            "Session Round",
            "Session Date",
            "Attendance",
            "Present GA",
            "BP",
            "Diagnosed Hypertension",
            "Referred Hypertension to MD",
            "Weight",
            "Anemia",
            "Iron Folate Routine Dose",
            "Iron Folate 30+ for Anemic Woman",
            "Prescribe Calcium",
            "Absorbed Calcium Last Month",
            "MUAC",
            "Diagnosed MAM",
            "Refer MAM to Nutrition Counsellor",
            "Diagnosed SAM",
            "Refer SAM to Higher Level",
            "Antenatal Depression Screening",
            "Antenatal Depression Diagnosed",
            "Refer to Psychosocial Counselor",
            "Urine Exam / Protein Uria",
            "Referred Positive Protein Uria to MD",
            "Cough More Than Two Weeks",
            "Referred Cough to DOTS Room",
            "TT Vaccine",
            "Danger Sign",
            "Type of Danger Sign",
            "Birth Planning Counseling",
            "Remarks",
        ]

        worksheet.append(headers)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="0F766E",
        )
        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        def yes_no(value):
            if value is True:
                return "Yes"
            if value is False:
                return "No"
            return ""

        for obj in queryset:
            worksheet.append([
                str(obj.registerid) if obj.registerid else "",
                self.get_facility(obj),
                self.get_province(obj),
                obj.sessiontype or "",
                obj.sessionround or "",
                (
                    obj.sessiondate.strftime("%Y-%m-%d")
                    if obj.sessiondate else ""
                ),
                obj.attendance or "",
                obj.presentga,
                obj.bp or "",
                yes_no(obj.dhypertension),
                yes_no(obj.rhypertensiontoMD),
                obj.weight,
                yes_no(obj.anemia),
                yes_no(obj.ironfolate),
                yes_no(obj.ironfolatepluswomen),
                yes_no(obj.pcalcium),
                yes_no(obj.acalcium),
                obj.muac,
                yes_no(obj.dmam),
                yes_no(obj.rmam),
                yes_no(obj.dsam),
                yes_no(obj.rsam),
                yes_no(obj.antedepressionscreening),
                yes_no(obj.antedepressiondiagnosed),
                yes_no(obj.rpsychosocialcounselor),
                yes_no(obj.urinexamcheck),
                obj.urinexam or "",
                yes_no(obj.rpositivepuriatomd),
                yes_no(obj.coughmorethantwoweeks),
                yes_no(obj.rcough),
                yes_no(obj.ttvaccine),
                yes_no(obj.dangersign),
                obj.typeofdangersign or "",
                yes_no(obj.birthplanningcounseling),
                obj.remarks or "",
            ])

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(
                column_cells[0].column
            )
            max_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            worksheet.column_dimensions[column_letter].width = min(
                max_length + 3,
                40,
            )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 35

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = (
            'attachment; filename="ganc_fourth_session.xlsx"'
        )

        workbook.save(response)
        return response

    export_ganc_fourth_session_excel.short_description = (
        "Export selected GANC Fourth Session records to Excel"
    )

# ============================================================
# GANC DELIVERY - CUSTOM REGISTER DROPDOWN
# ============================================================

class GancDeliveryRegisterChoiceField(forms.ModelChoiceField):
    """
    Displays enrollment records as:

    Register ID | Woman Name | Father Name | Cohort
    """

    def label_from_instance(self, obj):
        register_id = obj.pk

        name = getattr(obj, "name", "") or ""
        father_name = getattr(obj, "fathername", "") or ""

        cohort = getattr(obj, "cohortname", None)
        cohort_name = str(cohort) if cohort else "-"

        return (
            f"{register_id} | "
            f"{name} | "
            f"{father_name} | "
            f"{cohort_name}"
        )


class GancDeliveryAdminForm(forms.ModelForm):

    class Meta:
        model = Gancdelivery
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if "registerid" in self.fields:
            original_field = self.fields["registerid"]

            self.fields["registerid"] = GancDeliveryRegisterChoiceField(
                queryset=original_field.queryset,
                required=original_field.required,
                label=original_field.label,
                help_text=(
                    "Select the correct woman using Register ID, "
                    "Name, Father Name and Cohort."
                ),
            )


# ============================================================
# GANC DELIVERY ADMIN
# ============================================================

@admin.register(Gancdelivery)
class GancdeliveryAdmin(BaseSessionAdmin):

    form = GancDeliveryAdminForm

    # ========================================================
    # LIST VIEW
    # ========================================================

    list_display = (
        "registerid",
        "get_cohort",
        "get_facility",
        "get_province",
        "date_of_delivery",
        "gestational_age_at_delivery",
        "place_of_delivery",
        "type_of_delivery",
        "maternal_death",
        "number_of_newborn",
        "number_of_alive_newborn",
        "number_of_newborn_death",
        "number_of_fresh_still_birth",
    )

    # ========================================================
    # FILTERS
    # ========================================================

    list_filter = (
        SessionProvinceFilter,
        "registerid__cohortname",
        "date_of_delivery",
        "place_of_delivery",
        "type_of_delivery",
        "maternal_death",
        "early_breastfeeding",
        "newborn_vaccination_before_discharge",
        "counseled_on_postpartum_fp_before_discharge",
        "immediate_ppfp_before_discharge",
    )

    # ========================================================
    # SEARCH
    # ========================================================

    search_fields = (
        "registerid__name",
        "registerid__fathername",
        "types_of_complication",
        "how_complication_was_managed",
    )

    ordering = ("-date_of_delivery",)

    list_per_page = 25

    save_on_top = True

    # ========================================================
    # EXCEL EXPORT
    # ========================================================

    actions = (
        "export_delivery_to_excel",
    )

    # ========================================================
    # READ-ONLY GUIDANCE
    # ========================================================

    readonly_fields = (
        "delivery_help",
        "complication_help",
        "newborn_help",
        "ppfp_help",
    )

    # ========================================================
    # FORM LAYOUT
    # ========================================================

    fieldsets = (

        ("① Delivery Information", {

            "classes": (
                "ganc-section",
                "wide",
            ),

            "description": (
                "Select the correct registered woman and complete "
                "the main delivery information."
            ),

            "fields": (

                "delivery_help",

                "registerid",

                (
                    "date_of_delivery",
                    "gestational_age_at_delivery",
                ),

                (
                    "place_of_delivery",
                    "type_of_delivery",
                ),

                "immediate_uterotonic_for_amtsl",
            ),
        }),

        ("② Maternal Complications", {

            "classes": (
                "ganc-section",
                "wide",
            ),

            "description": (
                "Record maternal complications, management "
                "and maternal outcome."
            ),

            "fields": (

                "complication_help",

                "types_of_complication",

                "how_complication_was_managed",

                "maternal_death",
            ),
        }),

        ("③ Newborn Outcome", {

            "classes": (
                "ganc-section",
                "wide",
            ),

            "description": (
                "Record newborn numbers and outcomes, "
                "including breastfeeding and vaccination."
            ),

            "fields": (

                "newborn_help",

                (
                    "number_of_newborn",
                    "number_of_alive_newborn",
                ),

                (
                    "number_of_newborn_death",
                    "number_of_fresh_still_birth",
                ),

                (
                    "early_breastfeeding",
                    "newborn_vaccination_before_discharge",
                ),
            ),
        }),

        ("④ Postpartum Family Planning", {

            "classes": (
                "ganc-section",
                "wide",
            ),

            "description": (
                "Record postpartum family planning counseling "
                "and method uptake before discharge."
            ),

            "fields": (

                "ppfp_help",

                (
                    "counseled_on_postpartum_fp_before_discharge",
                    "immediate_ppfp_before_discharge",
                ),

                "ppfp_method_taken_before_discharge",
            ),
        }),

        ("⑤ Remarks", {

            "classes": (
                "ganc-section",
                "collapse",
            ),

            "fields": (
                "remark",
            ),
        }),
    )

    # ========================================================
    # CSS
    # ========================================================

    class Media:
        css = {
            "all": (
                "admin/css/ganc_admin.css",
            )
        }

    # ========================================================
    # COHORT DISPLAY
    # ========================================================

    @admin.display(
        description="Cohort",
        ordering="registerid__cohortname",
    )
    def get_cohort(self, obj):
        if not obj.registerid:
            return "-"

        cohort = getattr(
            obj.registerid,
            "cohortname",
            None,
        )

        return str(cohort) if cohort else "-"

    # ========================================================
    # GUIDANCE
    # ========================================================

    def delivery_help(self, obj=None):
        return (
            "Select the correct woman carefully. "
            "The Register Name dropdown displays Register ID, "
            "Woman Name, Father Name and Cohort."
        )

    delivery_help.short_description = "Delivery guidance"

    def complication_help(self, obj=None):
        return (
            "If a maternal complication occurred, record the "
            "type of complication and how it was managed."
        )

    complication_help.short_description = (
        "Complication guidance"
    )

    def newborn_help(self, obj=None):
        return (
            "Ensure the total number of newborns and all "
            "newborn outcomes are entered consistently."
        )

    newborn_help.short_description = (
        "Newborn outcome guidance"
    )

    def ppfp_help(self, obj=None):
        return (
            "Record whether postpartum family planning counseling "
            "was provided, whether immediate PPFP was accepted, "
            "and the method taken before discharge."
        )

    ppfp_help.short_description = (
        "Postpartum FP guidance"
    )

    # ========================================================
    # EXCEL EXPORT
    # ========================================================

    @admin.action(
        description=(
            "Export selected GANC Delivery records to Excel"
        )
    )
    def export_delivery_to_excel(
        self,
        request,
        queryset,
    ):

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "GANC Delivery"

        # ====================================================
        # HEADERS
        # ====================================================

        headers = [

            # Enrollment information
            "Register ID",
            "Woman Name",
            "Father Name",
            "Cohort",
            "Facility",
            "Province",

            # Delivery information
            "Date of Delivery",
            "Gestational Age at Delivery",
            "Place of Delivery",
            "Type of Delivery",
            "Immediate Uterotonic for AMTSL",

            # Maternal outcome
            "Types of Complication",
            "How Complication Was Managed",
            "Maternal Death",

            # Newborn outcome
            "Number of Newborn",
            "Number of Alive Newborn",
            "Number of Newborn Death",
            "Number of Fresh Still Birth",
            "Early Breastfeeding",
            "Newborn Vaccination Before Discharge",

            # PPFP
            "Counseled on Postpartum FP Before Discharge",
            "Immediate PPFP Before Discharge",
            "PPFP Method Taken Before Discharge",

            # Other
            "Remark",
        ]

        worksheet.append(headers)

        # ====================================================
        # HEADER STYLE
        # ====================================================

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="0F766E",
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in worksheet[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[1].height = 40

        # ====================================================
        # HELPER FUNCTIONS
        # ====================================================

        def yes_no(value):

            if value is True:
                return "Yes"

            if value is False:
                return "No"

            return ""

        def safe_text(value):

            if value is None:
                return ""

            return str(value)

        # ====================================================
        # EXPORT DATA
        # ====================================================

        for obj in queryset:

            enrollment = obj.registerid

            # -----------------------------------------------
            # Enrollment data
            # -----------------------------------------------

            register_id = ""

            woman_name = ""

            father_name = ""

            cohort_name = ""

            if enrollment:

                register_id = enrollment.pk

                woman_name = getattr(
                    enrollment,
                    "name",
                    "",
                ) or ""

                father_name = getattr(
                    enrollment,
                    "fathername",
                    "",
                ) or ""

                cohort = getattr(
                    enrollment,
                    "cohortname",
                    None,
                )

                cohort_name = (
                    safe_text(cohort)
                    if cohort
                    else ""
                )

            # -----------------------------------------------
            # Add Excel row
            # -----------------------------------------------

            worksheet.append([

                # Enrollment
                register_id,
                woman_name,
                father_name,
                cohort_name,
                self.get_facility(obj),
                self.get_province(obj),

                # Delivery
                (
                    obj.date_of_delivery.strftime(
                        "%Y-%m-%d"
                    )
                    if obj.date_of_delivery
                    else ""
                ),

                obj.gestational_age_at_delivery,

                safe_text(
                    obj.place_of_delivery
                ),

                safe_text(
                    obj.type_of_delivery
                ),

                yes_no(
                    obj.immediate_uterotonic_for_amtsl
                ),

                # Maternal complications
                safe_text(
                    obj.types_of_complication
                ),

                safe_text(
                    obj.how_complication_was_managed
                ),

                yes_no(
                    obj.maternal_death
                ),

                # Newborn
                obj.number_of_newborn,

                obj.number_of_alive_newborn,

                obj.number_of_newborn_death,

                obj.number_of_fresh_still_birth,

                yes_no(
                    obj.early_breastfeeding
                ),

                yes_no(
                    obj.newborn_vaccination_before_discharge
                ),

                # PPFP
                yes_no(
                    obj.counseled_on_postpartum_fp_before_discharge
                ),

                yes_no(
                    obj.immediate_ppfp_before_discharge
                ),

                safe_text(
                    obj.ppfp_method_taken_before_discharge
                ),

                # Remarks
                safe_text(
                    obj.remark
                ),
            ])

        # ====================================================
        # DATA FORMATTING
        # ====================================================

        for row in worksheet.iter_rows():

            for cell in row:

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        # ====================================================
        # AUTO COLUMN WIDTH
        # ====================================================

        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                if cell.value is not None:

                    value_length = len(
                        str(cell.value)
                    )

                    if value_length > max_length:
                        max_length = value_length

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                40,
            )

        # ====================================================
        # EXCEL ANALYSIS FEATURES
        # ====================================================

        worksheet.freeze_panes = "A2"

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        # ====================================================
        # DOWNLOAD RESPONSE
        # ====================================================

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

        response[
            "Content-Disposition"
        ] = (
            'attachment; '
            'filename="ganc_delivery_export.xlsx"'
        )

        workbook.save(response)

        return response

# ============================================================
# GROUP PNC - CUSTOM REGISTER DROPDOWN
# ============================================================

class GroupPncRegisterChoiceField(forms.ModelChoiceField):
    """
    Display:
    Register ID | Woman Name | Father Name | Cohort
    """

    def label_from_instance(self, obj):
        register_id = obj.pk
        name = getattr(obj, "name", "") or ""
        father_name = getattr(obj, "fathername", "") or ""

        cohort = getattr(obj, "cohortname", None)
        cohort_name = str(cohort) if cohort else "-"

        return (
            f"{register_id} | "
            f"{name} | "
            f"{father_name} | "
            f"{cohort_name}"
        )


class GroupPncFirstSessionAdminForm(forms.ModelForm):

    class Meta:
        model = GroupPncfirstSession
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if "registerid" in self.fields:
            original_field = self.fields["registerid"]

            self.fields["registerid"] = GroupPncRegisterChoiceField(
                queryset=original_field.queryset,
                required=original_field.required,
                label=original_field.label,
                help_text=(
                    "Select using Register ID, Woman Name, "
                    "Father Name and Cohort."
                ),
            )


class GroupPncSecondSessionAdminForm(forms.ModelForm):

    class Meta:
        model = GroupPncsecondSession
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if "registerid" in self.fields:
            original_field = self.fields["registerid"]

            self.fields["registerid"] = GroupPncRegisterChoiceField(
                queryset=original_field.queryset,
                required=original_field.required,
                label=original_field.label,
                help_text=(
                    "Select using Register ID, Woman Name, "
                    "Father Name and Cohort."
                ),
            )

# ============================================================
# PNC First Session
# ============================================================

# ============================================================
# PNC FIRST SESSION
# ============================================================

@admin.register(GroupPncfirstSession)
class GroupPncfirstSessionAdmin(BaseSessionAdmin):

    form = GroupPncFirstSessionAdminForm

    list_display = (
        "registerid",
        "get_cohort",
        "get_facility",
        "get_province",
        "session_type",
        "session_round",
        "session_date",
        "post_natal_day",
        "attendance",
        "diagnosed_with_hypertension",
        "anemia",
        "newborn_death",
        "maternal_death",
    )

    list_filter = (
        SessionProvinceFilter,
        "registerid__cohortname",
        "session_type",
        "session_round",
        "session_date",
        "attendance",
        "diagnosed_with_hypertension",
        "anemia",
        "newborn_death",
        "maternal_death",
        "exclusive_breast_feeding",
        "chosen_ppfp_method",
    )

    search_fields = (
        "registerid__name",
        "registerid__fathername",
        "type_of_maternal_danger_sign",
        "type_of_newborn_danger_sign",
    )

    ordering = ("-session_date",)
    list_per_page = 25
    save_on_top = True

    actions = (
        "export_pnc_first_session_excel",
    )

    readonly_fields = (
        "session_help",
        "maternal_help",
        "danger_help",
        "laboratory_help",
        "newborn_fp_help",
    )

    fieldsets = (

        ("① Session Information", {
            "classes": ("ganc-section", "wide"),
            "description": (
                "Select the correct registered woman and complete "
                "the first postnatal session information."
            ),
            "fields": (
                "session_help",
                "registerid",
                ("session_type", "session_round"),
                ("session_date", "post_natal_day"),
                "attendance",
            ),
        }),

        ("② Maternal Assessment", {
            "classes": ("ganc-section", "wide"),
            "description": (
                "Record maternal blood pressure, nutritional status, "
                "anemia, supplementation and referrals."
            ),
            "fields": (
                "maternal_help",
                ("bp", "muac"),
                (
                    "diagnosed_with_hypertension",
                    "referred_hypertension_to_md",
                ),
                (
                    "diagnosed_with_mam",
                    "refer_mam_to_nutrition_counselor",
                ),
                (
                    "diagnosed_with_sam",
                    "refer_sam_to_higher_level",
                ),
                (
                    "anemia",
                    "iron_folate_routine_dose",
                    "iron_folate_plus_for_anemic_woman",
                ),
            ),
        }),

        ("③ Danger Signs and Outcome", {
            "classes": ("ganc-section", "wide"),
            "description": (
                "Record maternal and newborn danger signs and outcomes."
            ),
            "fields": (
                "danger_help",
                "type_of_maternal_danger_sign",
                "type_of_newborn_danger_sign",
                ("newborn_death", "maternal_death"),
            ),
        }),

        ("④ Laboratory and Screening", {
            "classes": ("ganc-section", "collapse"),
            "description": (
                "Record urine examination, proteinuria, cough "
                "screening and referral information."
            ),
            "fields": (
                "laboratory_help",
                ("urine_exam", "protein_uria"),
                "referred_positive_protein_uria_to_md",
                (
                    "cough_more_than_two_weeks",
                    "referred_cough_to_dots_room",
                ),
            ),
        }),

        ("⑤ Newborn and Postpartum Family Planning", {
            "classes": ("ganc-section", "wide"),
            "description": (
                "Record newborn vaccination, breastfeeding and "
                "postpartum family planning information."
            ),
            "fields": (
                "newborn_fp_help",
                (
                    "newborn_vaccination_completed",
                    "exclusive_breast_feeding",
                ),
                ("chosen_ppfp_method", "ppfp_method_taken"),
            ),
        }),

        ("⑥ Remarks", {
            "classes": ("ganc-section", "collapse"),
            "fields": ("remark",),
        }),
    )

    class Media:
        css = {
            "all": ("admin/css/ganc_admin.css",)
        }

    # --------------------------------------------------------
    # Cohort
    # --------------------------------------------------------

    @admin.display(
        description="Cohort",
        ordering="registerid__cohortname",
    )
    def get_cohort(self, obj):
        if not obj.registerid:
            return "-"

        cohort = getattr(obj.registerid, "cohortname", None)
        return str(cohort) if cohort else "-"

    # --------------------------------------------------------
    # Guidance
    # --------------------------------------------------------

    def session_help(self, obj=None):
        return (
            "Select the correct woman using Register ID, Name, "
            "Father Name and Cohort."
        )
    session_help.short_description = "Session guidance"

    def maternal_help(self, obj=None):
        return (
            "Complete maternal assessment, nutrition, anemia "
            "and referral information carefully."
        )
    maternal_help.short_description = "Maternal assessment guidance"

    def danger_help(self, obj=None):
        return (
            "Record maternal and newborn danger signs and "
            "their respective outcomes."
        )
    danger_help.short_description = "Danger sign guidance"

    def laboratory_help(self, obj=None):
        return (
            "Complete urine, proteinuria and cough screening "
            "information where applicable."
        )
    laboratory_help.short_description = "Screening guidance"

    def newborn_fp_help(self, obj=None):
        return (
            "Complete newborn vaccination, exclusive breastfeeding "
            "and postpartum family planning information."
        )
    newborn_fp_help.short_description = "Newborn and FP guidance"

    # --------------------------------------------------------
    # Excel Export
    # --------------------------------------------------------

    @admin.action(
        description="Export selected PNC First Session records to Excel"
    )
    def export_pnc_first_session_excel(self, request, queryset):

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "PNC First Session"

        headers = [
            "Register ID",
            "Woman Name",
            "Father Name",
            "Cohort",
            "Facility",
            "Province",

            "Session Type",
            "Session Round",
            "Session Date",
            "Post Natal Day",
            "Attendance",

            "BP",
            "Diagnosed With Hypertension",
            "Referred Hypertension To MD",
            "MUAC",
            "Diagnosed With MAM",
            "Refer MAM To Nutrition Counselor",
            "Diagnosed With SAM",
            "Refer SAM To Higher Level",
            "Anemia",
            "Iron Folate Routine Dose",
            "Iron Folate Plus For Anemic Woman",

            "Type of Maternal Danger Sign",
            "Type of Newborn Danger Sign",
            "Newborn Death",
            "Maternal Death",

            "Urine Exam",
            "Protein Uria",
            "Referred Positive Protein Uria To MD",
            "Cough More Than Two Weeks",
            "Referred Cough To DOTS Room",

            "Newborn Vaccination Completed",
            "Exclusive Breast Feeding",
            "Chosen PPFP Method",
            "PPFP Method Taken",

            "Remark",
        ]

        worksheet.append(headers)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="0F766E",
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[1].height = 40

        def yes_no(value):
            if value is True:
                return "Yes"
            if value is False:
                return "No"
            return ""

        def safe_text(value):
            return "" if value is None else str(value)

        for obj in queryset:

            enrollment = obj.registerid

            register_id = ""
            woman_name = ""
            father_name = ""
            cohort_name = ""

            if enrollment:
                register_id = enrollment.pk
                woman_name = getattr(
                    enrollment, "name", ""
                ) or ""

                father_name = getattr(
                    enrollment, "fathername", ""
                ) or ""

                cohort = getattr(
                    enrollment, "cohortname", None
                )

                cohort_name = (
                    safe_text(cohort)
                    if cohort else ""
                )

            worksheet.append([
                register_id,
                woman_name,
                father_name,
                cohort_name,
                self.get_facility(obj),
                self.get_province(obj),

                safe_text(obj.session_type),
                safe_text(obj.session_round),

                (
                    obj.session_date.strftime("%Y-%m-%d")
                    if obj.session_date else ""
                ),

                obj.post_natal_day,
                safe_text(obj.attendance),

                safe_text(obj.bp),
                yes_no(obj.diagnosed_with_hypertension),
                yes_no(obj.referred_hypertension_to_md),
                obj.muac,

                yes_no(obj.diagnosed_with_mam),
                yes_no(obj.refer_mam_to_nutrition_counselor),
                yes_no(obj.diagnosed_with_sam),
                yes_no(obj.refer_sam_to_higher_level),

                yes_no(obj.anemia),
                yes_no(obj.iron_folate_routine_dose),
                yes_no(obj.iron_folate_plus_for_anemic_woman),

                safe_text(obj.type_of_maternal_danger_sign),
                safe_text(obj.type_of_newborn_danger_sign),

                yes_no(obj.newborn_death),
                yes_no(obj.maternal_death),

                safe_text(obj.urine_exam),
                safe_text(obj.protein_uria),
                yes_no(obj.referred_positive_protein_uria_to_md),

                yes_no(obj.cough_more_than_two_weeks),
                yes_no(obj.referred_cough_to_dots_room),

                yes_no(obj.newborn_vaccination_completed),
                yes_no(obj.exclusive_breast_feeding),

                yes_no(obj.chosen_ppfp_method),
                safe_text(obj.ppfp_method_taken),

                safe_text(obj.remark),
            ])

        self._format_pnc_excel(worksheet)

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

        response["Content-Disposition"] = (
            'attachment; filename="pnc_first_session.xlsx"'
        )

        workbook.save(response)
        return response

    def _format_pnc_excel(self, worksheet):

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(max_length + 3, 40)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions


# ============================================================
# PNC Second Session
# ============================================================

# ============================================================
# PNC SECOND SESSION
# ============================================================

@admin.register(GroupPncsecondSession)
class GroupPncsecondSessionAdmin(BaseSessionAdmin):

    form = GroupPncSecondSessionAdminForm

    list_display = (
        "registerid",
        "get_cohort",
        "get_facility",
        "get_province",
        "sessiontype",
        "sessionround",
        "sessiondate",
        "postnatalday",
        "attendance",
        "dhypertension",
        "anemia",
        "newborndeath",
        "maternaldeath",
        "birthspacingmethodchosen",
    )

    list_filter = (
        SessionProvinceFilter,
        "registerid__cohortname",
        "sessiontype",
        "sessionround",
        "sessiondate",
        "attendance",
        "dhypertension",
        "anemia",
        "newborndeath",
        "maternaldeath",
        "exclusivebreastfeeding",
        "birthspacingmethodchosen",
        "postnataldepressiondiagnosed",
    )

    search_fields = (
        "registerid__name",
        "registerid__fathername",
        "typeofmaternaldangersign",
        "typeofnewborndangersign",
    )

    ordering = ("-sessiondate",)
    list_per_page = 25
    save_on_top = True

    actions = (
        "export_pnc_second_session_excel",
    )

    readonly_fields = (
        "session_help",
        "maternal_help",
        "mental_health_help",
        "danger_help",
        "newborn_help",
        "birth_spacing_help",
    )

    fieldsets = (

        ("① Session Information", {
            "classes": ("ganc-section", "wide"),
            "description": (
                "Select the correct registered woman and complete "
                "the second postnatal session information."
            ),
            "fields": (
                "session_help",
                "registerid",
                ("sessiontype", "sessionround"),
                ("sessiondate", "postnatalday"),
                "attendance",
            ),
        }),

        ("② Maternal Assessment", {
            "classes": ("ganc-section", "wide"),
            "fields": (
                "maternal_help",
                ("bp", "muac"),
                ("dhypertension", "rhypertensiontomd"),
                ("dmam", "rmam"),
                ("dsam", "rsam"),
                (
                    "anemia",
                    "ironfolate",
                    "ironfolatepluswomen",
                ),
            ),
        }),

        ("③ Mental Health", {
            "classes": ("ganc-section", "wide"),
            "fields": (
                "mental_health_help",
                (
                    "postnataldepressiondiagnosed",
                    "rpsychosocialcounselor",
                ),
            ),
        }),

        ("④ Danger Signs and Outcome", {
            "classes": ("ganc-section", "wide"),
            "fields": (
                "danger_help",
                "typeofmaternaldangersign",
                "typeofnewborndangersign",
                ("newborndeath", "maternaldeath"),
            ),
        }),

        ("⑤ Other Health Information", {
            "classes": ("ganc-section", "collapse"),
            "fields": (
                "newborn_help",
                "newbornvaccinationcompleted",
                ("coughmorethantwoweeks", "rcough"),
                "exclusivebreastfeeding",
            ),
        }),

        ("⑥ Birth Spacing", {
            "classes": ("ganc-section", "wide"),
            "fields": (
                "birth_spacing_help",
                (
                    "birthspacingmethodchosen",
                    "birthspacingmethod",
                ),
            ),
        }),

        ("⑦ Remarks", {
            "classes": ("ganc-section", "collapse"),
            "fields": ("remark",),
        }),
    )

    class Media:
        css = {
            "all": ("admin/css/ganc_admin.css",)
        }

    @admin.display(
        description="Cohort",
        ordering="registerid__cohortname",
    )
    def get_cohort(self, obj):
        if not obj.registerid:
            return "-"

        cohort = getattr(
            obj.registerid,
            "cohortname",
            None,
        )

        return str(cohort) if cohort else "-"

    def session_help(self, obj=None):
        return (
            "Select the correct woman using Register ID, Name, "
            "Father Name and Cohort."
        )
    session_help.short_description = "Session guidance"

    def maternal_help(self, obj=None):
        return (
            "Complete maternal assessment, nutritional status, "
            "anemia and referral information."
        )
    maternal_help.short_description = "Maternal assessment guidance"

    def mental_health_help(self, obj=None):
        return (
            "Record postnatal depression diagnosis and "
            "psychosocial referral."
        )
    mental_health_help.short_description = "Mental health guidance"

    def danger_help(self, obj=None):
        return (
            "Record maternal and newborn danger signs "
            "and maternal/newborn outcomes."
        )
    danger_help.short_description = "Danger sign guidance"

    def newborn_help(self, obj=None):
        return (
            "Record vaccination, cough screening, referral "
            "and exclusive breastfeeding."
        )
    newborn_help.short_description = "Newborn guidance"

    def birth_spacing_help(self, obj=None):
        return (
            "Record whether a birth-spacing method was chosen "
            "and specify the selected method."
        )
    birth_spacing_help.short_description = "Birth spacing guidance"

    @admin.action(
        description="Export selected PNC Second Session records to Excel"
    )
    def export_pnc_second_session_excel(self, request, queryset):

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "PNC Second Session"

        headers = [
            "Register ID",
            "Woman Name",
            "Father Name",
            "Cohort",
            "Facility",
            "Province",

            "Session Type",
            "Session Round",
            "Session Date",
            "Post Natal Day",
            "Attendance",

            "BP",
            "Diagnosed With Hypertension",
            "Referred Hypertension To MD",
            "MUAC",
            "Diagnosed With MAM",
            "Refer MAM",
            "Diagnosed With SAM",
            "Refer SAM",
            "Anemia",
            "Iron Folate",
            "Iron Folate Plus For Anemic Woman",

            "Postnatal Depression Diagnosed",
            "Refer To Psychosocial Counselor",

            "Type of Maternal Danger Sign",
            "Type of Newborn Danger Sign",
            "Newborn Death",
            "Maternal Death",

            "Newborn Vaccination Completed",
            "Cough More Than Two Weeks",
            "Referred Cough To DOTS Room",
            "Exclusive Breastfeeding",

            "Birth Spacing Method Chosen",
            "Birth Spacing Method",

            "Remark",
        ]

        worksheet.append(headers)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="0F766E",
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[1].height = 40

        def yes_no(value):
            if value is True:
                return "Yes"
            if value is False:
                return "No"
            return ""

        def safe_text(value):
            return "" if value is None else str(value)

        for obj in queryset:

            enrollment = obj.registerid

            register_id = ""
            woman_name = ""
            father_name = ""
            cohort_name = ""

            if enrollment:
                register_id = enrollment.pk

                woman_name = getattr(
                    enrollment,
                    "name",
                    "",
                ) or ""

                father_name = getattr(
                    enrollment,
                    "fathername",
                    "",
                ) or ""

                cohort = getattr(
                    enrollment,
                    "cohortname",
                    None,
                )

                cohort_name = (
                    safe_text(cohort)
                    if cohort else ""
                )

            worksheet.append([
                register_id,
                woman_name,
                father_name,
                cohort_name,
                self.get_facility(obj),
                self.get_province(obj),

                safe_text(obj.sessiontype),
                safe_text(obj.sessionround),

                (
                    obj.sessiondate.strftime("%Y-%m-%d")
                    if obj.sessiondate else ""
                ),

                obj.postnatalday,
                safe_text(obj.attendance),

                safe_text(obj.bp),
                yes_no(obj.dhypertension),
                yes_no(obj.rhypertensiontomd),
                obj.muac,

                yes_no(obj.dmam),
                yes_no(obj.rmam),
                yes_no(obj.dsam),
                yes_no(obj.rsam),

                yes_no(obj.anemia),
                yes_no(obj.ironfolate),
                yes_no(obj.ironfolatepluswomen),

                yes_no(obj.postnataldepressiondiagnosed),
                yes_no(obj.rpsychosocialcounselor),

                safe_text(obj.typeofmaternaldangersign),
                safe_text(obj.typeofnewborndangersign),

                yes_no(obj.newborndeath),
                yes_no(obj.maternaldeath),

                yes_no(obj.newbornvaccinationcompleted),
                yes_no(obj.coughmorethantwoweeks),
                yes_no(obj.rcough),
                yes_no(obj.exclusivebreastfeeding),

                yes_no(obj.birthspacingmethodchosen),
                safe_text(obj.birthspacingmethod),

                safe_text(obj.remark),
            ])

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(max_length + 3, 40)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

        response["Content-Disposition"] = (
            'attachment; filename="pnc_second_session.xlsx"'
        )

        workbook.save(response)
        return response